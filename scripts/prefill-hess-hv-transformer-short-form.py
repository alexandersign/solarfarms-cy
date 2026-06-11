"""Prefill short producer questionnaire (POWER Transformer sheet) for HESS."""
from __future__ import annotations

import json
import shutil
from pathlib import Path

import openpyxl

_HVT = Path(
    r"L:\My Drive\LINYANG\BESS CLIENTS\Individual_60-120-standalone\HV Transformer"
)
# Supplier-2 = Chinese producer via 7sun (form: POWER Transformer). See _supplier-key.md.
SRC = _HVT / "Supplier-2" / "source" / "POWER-Transformer-blank.xlsx"
OUT = _HVT / "Supplier-2" / "T1-main" / "POWER-T1-main-63MVA-HESS-prefilled.xlsx"
RESULTS_JSON = (
    Path(__file__).resolve().parent.parent
    / "docs"
    / "dso"
    / "analysis"
    / "hess-pandapower-results.json"
)
CLIENT_COL = 5


def load_results() -> dict:
    return json.loads(RESULTS_JSON.read_text(encoding="utf-8"))


def index_rows(ws) -> list[tuple[int, str, str]]:
    rows = []
    for r in range(1, ws.max_row + 1):
        item = ws.cell(r, 1).value
        char = ws.cell(r, 2).value
        if item is None and not char:
            continue
        rows.append((r, str(item).strip() if item is not None else "", str(char or "").strip()))
    return rows


def set_char(rows, char_sub: str, value: str) -> None:
    for r, _item, char in rows:
        if char_sub.lower() in char.lower():
            ws_cell(r, value)
            return
    raise KeyError(f"Characteristic not found: {char_sub!r}")


def ws_cell(r: int, value: str) -> None:
    global _ws
    _ws.cell(r, CLIENT_COL).value = value


def main() -> None:
    global _ws
    res = load_results()
    hv = res["protection_indicative"]["hv_132"]
    lv = res["protection_indicative"]["lv_33"]

    shutil.copy2(SRC, OUT)
    wb = openpyxl.load_workbook(OUT)
    _ws = wb["POWER Transformer"]
    rows = index_rows(_ws)

    def fill(char_sub: str, value: str) -> None:
        set_char(rows, char_sub, value)

    fill("Altitude", "264")
    fill("Installation", "Outdoor")
    fill("Contamination", "III (heavy — composite ≥35 mm/kV)")
    ik_33 = round(res["short_circuit"]["33_kv_combined_ka"], 1)
    fill("Application", "Main step-up transformer (KYEA BESS) — EAE side per POS SLD")
    fill("Seismic", "CYS EN 1998-1 Zone II, agR = 0.23 g (Zone I/II boundary)")
    fill(
        "Special Conditions",
        "Mediterranean; ISO 12944 C3. "
        f"132 kV Ik PROVISIONAL 31.5 kA (T1.8.6 — confirm at connection bay). "
        f"33 kV Ik ~{ik_33} kA combined (BESS plant). CT ALF TBC — ISM bay protection.",
    )
    fill("Design Standards", "IEC / EN + TSOC T14 (not ANSI)")
    fill("Apparent power", "63 / 63 (ONAN / ONAF)")
    fill("Frequency", "50")
    fill("Connection Group", "YNd11 (132 kV Star / 33 kV Delta)")
    fill("Percentage impedance", "21 % @ 63 MVA CMR — per Requirements §17 (confirmed)")
    fill("Insulation fluid", "Mineral oil")
    fill("PCB", "Without PCBs")
    fill("Number of phases", "3")
    fill("Number of windings", "2")
    fill("Factor K", "1.1")
    fill("Noise level", "≤60 dB(A) outdoor")
    fill("temperature", "55 oil / 60 windings (°C rise)")
    fill("taps regulation", "OLTC +12.5% / −18.75%, 25 × 1.25%")
    fill("Losses core", "Per Tier 2 Ecodesign — state in offer")
    fill("Losses Windings", "Per Tier 2 Ecodesign — state in offer")

    fill("windings material", "Copper")
    fill("Configuration", "Star (132 kV HV)")  # first match — need separate for delta
    # Primary section — row-specific by order
    for r, item, char in rows:
        if item == "7.2":
            _ws.cell(r, CLIENT_COL).value = "Star (132 kV)"
        elif item == "7.3" and "Voltaje" in char:
            _ws.cell(r, CLIENT_COL).value = "132000"
        elif item == "7.3" and "Maximum" in char:
            _ws.cell(r, CLIENT_COL).value = "145000"
        elif item == "7.5":
            _ws.cell(r, CLIENT_COL).value = "550"
        elif item == "8.2":
            _ws.cell(r, CLIENT_COL).value = "Delta (33 kV)"
        elif item == "8.3" and "Voltaje" in char:
            _ws.cell(r, CLIENT_COL).value = "33000"
        elif item == "8.4":
            _ws.cell(r, CLIENT_COL).value = "36000"
        elif item == "8.5":
            # LV BIL — 36 kV class per client clarification Jun 2026 (was 125 = 24 kV class)
            _ws.cell(r, CLIENT_COL).value = "170 (LV BIL, 36 kV class); 70 kV AC 1 min"
        elif item == "8.1":
            _ws.cell(r, CLIENT_COL).value = "Copper"

    fill("Primary side", "Yes")
    for r, item, char in rows:
        if item == "9.2":
            _ws.cell(r, CLIENT_COL).value = "3"
        elif item == "9.3":
            _ws.cell(r, CLIENT_COL).value = f"{hv['ct_ratio']} A — 0.2, 30 VA (ALF TBC — ISM)"
        elif item == "9.4":
            _ws.cell(r, CLIENT_COL).value = f"{hv['ct_ratio']} A — 5P20, 20 VA (ALF TBC — ISM)"
        elif item == "9.5":
            _ws.cell(r, CLIENT_COL).value = "Yes"
        elif item == "9.6":
            _ws.cell(r, CLIENT_COL).value = "3"
        elif item == "9.7":
            _ws.cell(r, CLIENT_COL).value = f"{lv['ct_ratio']} A — 0.2, 30 VA"
        elif item == "9.8":
            _ws.cell(r, CLIENT_COL).value = f"{lv['ct_ratio']} A — 5P20, 20 VA"
        elif item == "9.9":
            _ws.cell(r, CLIENT_COL).value = f"{lv['ct_ratio']} A — 5P20, 20 VA (WTI)"
        elif item == "10.1":
            _ws.cell(r, CLIENT_COL).value = "No (KYEA AIS scope)"
        elif item == "10.4":
            _ws.cell(r, CLIENT_COL).value = "No (KYEA AIS scope)"

    for r, item, char in rows:
        if item == "11.1":
            _ws.cell(r, CLIENT_COL).value = "OTI with contacts (per EN 60076)"
        elif item == "11.3":
            _ws.cell(r, CLIENT_COL).value = "PRD with flag"
        elif item == "11.4":
            _ws.cell(r, CLIENT_COL).value = "WTI with contacts"
        elif item == "11.5":
            _ws.cell(r, CLIENT_COL).value = "Magnetic oil level indicator"

    wb.save(OUT)
    print(f"Wrote {OUT}")


if __name__ == "__main__":
    main()
