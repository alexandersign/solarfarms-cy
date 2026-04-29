"""
Generate PV + BESS scenario analysis Excel for:
  Plot 1: Softades, Larnaka (1,363 m²) — Παύλος Παύλου  (Plot 0/2150)
  Plot 2: Αρκατζιά, Αμμόχωστος (3,704 m²) — Παύλου Κούλλα (Plot 0/2257)

Scenarios: South-facing (25°) and East–West (12°) for each plot.
Panel:     AE Solar Meteor N-type TOPCon bifacial 620–630 Wp, η 22.9%, bifaciality 80%
Albedo:    White ground cover 0.70 → bifacial gain +10.3% south / +5.0% EW
BESS:      4-hour duration, €127,000/MWh installed (Lighthief confirmed — Agios Theodoros RTB)
PV EPC:    €720/kWp turnkey (Lighthief confirmed — Agios Theodoros RTB)
Financing: 65% LTV, 5.5% p.a. fixed, 12-year annuity
O&M:       PV €15/kWp/yr + BESS €2,500/MWh/yr + flat monitoring/insurance
Tariff:    €0.16/kWh net metering / PPA
PV deg:    0.4%/year (N-type TOPCon — low degradation)
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os, math

# ─────────────────────────────────────────────────────────
# BRAND PALETTE  (Lighthief: navy + gold)
# ─────────────────────────────────────────────────────────
NAVY      = "1A365D"
NAVY2     = "244F87"
NAVY3     = "EEF3FB"    # very light navy tint
GOLD      = "C9A432"
GOLD_DARK = "9C7D22"
WHITE     = "FFFFFF"
GREY_TXT  = "404040"
GREEN     = "1E7D4A"
GREEN_LT  = "E8F5ED"
INPUT_BG  = "FFF9E6"    # yellow — editable input
CALC_BG   = "EEF6EE"    # green tint — calculated
WARN_BG   = "FEF3C7"    # amber tint
RED_TXT   = "C00000"

def _fill(hex_c):
    return PatternFill(start_color=hex_c, end_color=hex_c, fill_type="solid")

def _border(color="C5D4E8"):
    s = Side(style="thin", color=color)
    return Border(left=s, right=s, top=s, bottom=s)

def _hfont(sz=10, bold=True):
    return Font(name="Calibri", bold=bold, size=sz, color=WHITE)

def _font(bold=False, sz=11, color=GREY_TXT, italic=False):
    return Font(name="Calibri", bold=bold, size=sz, color=color, italic=italic)

def _center():
    return Alignment(horizontal="center", vertical="center", wrap_text=True)

def _left():
    return Alignment(horizontal="left", vertical="center")

def _right():
    return Alignment(horizontal="right", vertical="center")

def hdr(ws, row, col, value, width=None, bg=NAVY, fg=WHITE, sz=10):
    c = ws.cell(row=row, column=col, value=value)
    c.font = Font(name="Calibri", bold=True, size=sz, color=fg)
    c.fill = _fill(bg)
    c.alignment = _center()
    c.border = _border()
    return c

def cell(ws, row, col, value, fmt=None, bold=False, color=GREY_TXT,
         bg=None, align="left", border=True, italic=False):
    c = ws.cell(row=row, column=col, value=value)
    c.font = Font(name="Calibri", bold=bold, size=11, color=color, italic=italic)
    if fmt:
        c.number_format = fmt
    if bg:
        c.fill = _fill(bg)
    if align == "left":
        c.alignment = _left()
    elif align == "right":
        c.alignment = _right()
    else:
        c.alignment = _center()
    if border:
        c.border = _border()
    return c

def merge_hdr(ws, row, col1, col2, value, bg=NAVY, fg=WHITE, sz=11):
    ws.merge_cells(start_row=row, start_column=col1, end_row=row, end_column=col2)
    c = ws.cell(row=row, column=col1, value=value)
    c.font = Font(name="Calibri", bold=True, size=sz, color=fg)
    c.fill = _fill(bg)
    c.alignment = _center()
    c.border = _border()
    return c

def section_title(ws, row, col, value):
    c = ws.cell(row=row, column=col, value=value)
    c.font = Font(name="Calibri", bold=True, size=11, color=NAVY)
    c.alignment = _left()
    return c

# ─────────────────────────────────────────────────────────
# SCENARIO DATA
# ─────────────────────────────────────────────────────────
RAW = [
    dict(id="P1S",  plot=1, location="Softades, Larnaka",       ref="0/2150", owner="Παύλος Παύλου",
         area=1363, orient="South-Facing 25°", kwp=95,  panels=44,  bess_mwh=0.38,
         yield_kwh=2150, om_fixed=2000),
    dict(id="P1EW", plot=1, location="Softades, Larnaka",       ref="0/2150", owner="Παύλος Παύλου",
         area=1363, orient="East–West 12°",    kwp=150, panels=69,  bess_mwh=0.60,
         yield_kwh=1450, om_fixed=2000),
    dict(id="P2S",  plot=2, location="Αρκατζιά, Αμμόχωστος",   ref="0/2257", owner="Παύλου Κούλλα",
         area=3704, orient="South-Facing 25°", kwp=270, panels=125, bess_mwh=1.08,
         yield_kwh=2150, om_fixed=2500),
    dict(id="P2EW", plot=2, location="Αρκατζιά, Αμμόχωστος",   ref="0/2257", owner="Παύλου Κούλλα",
         area=3704, orient="East–West 12°",    kwp=420, panels=194, bess_mwh=1.68,
         yield_kwh=1450, om_fixed=2500),
]

# Global assumptions
PV_EPC_PER_KWP  = 720       # €/kWp
BESS_PER_MWH    = 127_000   # €/MWh installed (4h)
PERMITTING      = 35_000    # €/park
TARIFF          = 0.16      # €/kWh
OM_PV_PER_KWP   = 15        # €/kWp/yr
OM_BESS_PER_MWH = 2_500     # €/MWh/yr
LTV             = 0.65
LOAN_RATE       = 0.055
LOAN_YEARS      = 12
PV_DEGRADATION  = 0.004     # 0.4% per year (N-type TOPCon)
PROJECT_LIFE    = 20        # years modelled

def annuity_factor(r, n):
    return r * (1 + r)**n / ((1 + r)**n - 1)

AF = annuity_factor(LOAN_RATE, LOAN_YEARS)

# Enrich scenarios
SCENARIOS = []
for s in RAW:
    s = dict(s)
    s["pv_epc"]       = s["kwp"] * PV_EPC_PER_KWP
    s["bess_cost"]    = s["bess_mwh"] * BESS_PER_MWH
    s["capex"]        = s["pv_epc"] + s["bess_cost"] + PERMITTING
    s["mwh_y1"]       = s["kwp"] * s["yield_kwh"] / 1000
    s["revenue_y1"]   = s["mwh_y1"] * 1000 * TARIFF
    s["om_yr"]        = s["kwp"] * OM_PV_PER_KWP + s["bess_mwh"] * OM_BESS_PER_MWH + s["om_fixed"]
    s["ebitda_y1"]    = s["revenue_y1"] - s["om_yr"]
    s["loan"]         = s["capex"] * LTV
    s["equity"]       = s["capex"] * (1 - LTV)
    s["debt_service"] = s["loan"] * AF
    s["net_cash_y1"]  = s["ebitda_y1"] - s["debt_service"]
    s["dscr"]         = s["ebitda_y1"] / s["debt_service"] if s["debt_service"] else 0
    s["unlevered_pb"] = s["capex"] / s["ebitda_y1"] if s["ebitda_y1"] > 0 else 99
    s["equity_pb"]    = s["equity"] / s["net_cash_y1"] if s["net_cash_y1"] > 0 else 99

    # 20-year cash flows
    flows = []
    cum = -s["equity"]   # investor puts in equity at t=0
    for yr in range(1, PROJECT_LIFE + 1):
        deg = (1 - PV_DEGRADATION) ** (yr - 1)
        rev = s["revenue_y1"] * deg
        om  = s["om_yr"]
        ebitda = rev - om
        ds = s["debt_service"] if yr <= LOAN_YEARS else 0
        net = ebitda - ds
        cum += net
        flows.append(dict(yr=yr, deg=deg, rev=rev, om=om,
                          ebitda=ebitda, ds=ds, net=net, cum=cum))
    s["flows"] = flows

    # Simple payback (equity basis, from flows)
    s["equity_pb_precise"] = next(
        (f["yr"] - 1 + (-flows[i-1]["cum"] if i > 0 else s["equity"]) / f["net"]
         for i, f in enumerate(flows) if f["cum"] >= 0),
        99
    )
    s["total_20yr_net"] = flows[-1]["cum"] + s["equity"]   # total net profit excl equity
    SCENARIOS.append(s)


# ─────────────────────────────────────────────────────────
# WORKBOOK
# ─────────────────────────────────────────────────────────
wb = openpyxl.Workbook()
wb.remove(wb.active)   # remove default sheet

# ══════════════════════════════════════════════════════════
# SHEET 1: DASHBOARD
# ══════════════════════════════════════════════════════════
ws = wb.create_sheet("Dashboard")
ws.sheet_properties.tabColor = "C9A432"
ws.freeze_panes = "A6"

# Set column widths
col_widths = [2, 30, 16, 16, 16, 16, 2]
for i, w in enumerate(col_widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w

# Title block
ws.merge_cells("B1:F1")
c = ws.cell(row=1, column=2, value="PV + BESS SCENARIO ANALYSIS — LIGHTHIEF CYPRUS")
c.font = Font(name="Calibri", bold=True, size=16, color=NAVY)
c.alignment = _center()
c.fill = _fill("F0F4F8")

ws.merge_cells("B2:F2")
c = ws.cell(row=2, column=2, value="Softades (Larnaka) · Αρκατζιά (Αμμόχωστος)  |  AE Solar Meteor TOPCon Bifacial · White Albedo 0.70 · 4h BESS · 65% Bank Finance")
c.font = Font(name="Calibri", size=10, color="5A7090", italic=True)
c.alignment = _center()
c.fill = _fill("F0F4F8")

ws.merge_cells("B3:F3")
c = ws.cell(row=3, column=2,
    value="Cost basis: Agios Theodoros PARK-RTB-2026 confirmed CAPEX stack  |  PV €720/kWp · BESS €127k/MWh · O&M €15/kWp+€2,500/MWh  |  April 2026")
c.font = Font(name="Calibri", size=9, color="7A8FA6", italic=True)
c.alignment = _center()
c.fill = _fill("F0F4F8")

# Column headers
r = 5
hdr(ws, r, 2, "Metric",              bg=NAVY, sz=10)
hdr(ws, r, 3, "Plot 1 · South 25°\n95 kWp · Softades",  bg=NAVY, sz=10)
hdr(ws, r, 4, "Plot 1 · EW 12°\n150 kWp · Softades",   bg=NAVY, sz=10)
hdr(ws, r, 5, "Plot 2 · South 25°\n270 kWp · Αρκατζιά", bg=NAVY2, sz=10)
hdr(ws, r, 6, "Plot 2 · EW 12°\n420 kWp · Αρκατζιά",   bg=NAVY2, sz=10)
ws.row_dimensions[r].height = 32

def dash_section(ws, r, label):
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)
    c = ws.cell(row=r, column=2, value=f"  {label}")
    c.font = Font(name="Calibri", bold=True, size=10, color=WHITE)
    c.fill = _fill(NAVY2)
    c.alignment = _left()
    return r + 1

def dash_row(ws, r, label, vals, fmts, bgs=None, bold_vals=False, bold_label=False):
    c = ws.cell(row=r, column=2, value=label)
    c.font = Font(name="Calibri", bold=bold_label, size=11, color=GREY_TXT)
    c.border = _border()
    c.alignment = _left()
    for i, (v, f) in enumerate(zip(vals, fmts)):
        bg = (bgs[i] if bgs else None)
        col_color = GREEN if (bold_vals and i == min(range(len(vals)), key=lambda x: vals[x] if isinstance(vals[x], (int, float)) else 99)) else GREY_TXT
        cc = ws.cell(row=r, column=3+i, value=v)
        cc.font = Font(name="Calibri", bold=bold_vals, size=11, color=GREY_TXT)
        cc.number_format = f
        cc.border = _border()
        cc.alignment = _right()
        if bg: cc.fill = _fill(bg)
    return r + 1

r = 6
r = dash_section(ws, r, "PLOT DETAILS")
dash_row(ws, r, "Plot Reference", [s["ref"] for s in SCENARIOS], ["@"]*4); r+=1
dash_row(ws, r, "Owner",          [s["owner"] for s in SCENARIOS], ["@"]*4); r+=1
dash_row(ws, r, "Land Area (m²)", [s["area"] for s in SCENARIOS], ["#,##0"]*4); r+=1
dash_row(ws, r, "Orientation",    [s["orient"] for s in SCENARIOS], ["@"]*4); r+=1
dash_row(ws, r, "Panel Count",    [s["panels"] for s in SCENARIOS], ["#,##0"]*4); r+=1
dash_row(ws, r, "System Size (kWp)", [s["kwp"] for s in SCENARIOS], ["#,##0"]*4); r+=1
dash_row(ws, r, "BESS Duration / Capacity", [f"4h / {s['bess_mwh']} MWh" for s in SCENARIOS], ["@"]*4); r+=1
dash_row(ws, r, "Specific Yield (kWh/kWp)", [s["yield_kwh"] for s in SCENARIOS], ["#,##0"]*4); r+=1
r += 1

r = dash_section(ws, r, "CAPITAL EXPENDITURE (CAPEX)")
dash_row(ws, r, "PV EPC",         [s["pv_epc"] for s in SCENARIOS], ["€#,##0"]*4); r+=1
dash_row(ws, r, "BESS (4h installed)", [s["bess_cost"] for s in SCENARIOS], ["€#,##0"]*4); r+=1
dash_row(ws, r, "Permitting & Docs", [PERMITTING]*4, ["€#,##0"]*4); r+=1
for i, s in enumerate(SCENARIOS):
    if i == 0:
        dash_row(ws, r, "TOTAL CAPEX",
                 [s["capex"] for s in SCENARIOS], ["€#,##0"]*4,
                 bold_vals=False)
        r += 1
        break
r += 1

r = dash_section(ws, r, "ANNUAL PERFORMANCE (Year 1)")
dash_row(ws, r, "Production (MWh/yr)", [round(s["mwh_y1"],1) for s in SCENARIOS], ["#,##0.0"]*4); r+=1
dash_row(ws, r, "Gross Revenue (€/yr)", [s["revenue_y1"] for s in SCENARIOS], ["€#,##0"]*4); r+=1
dash_row(ws, r, "O&M Cost (€/yr)",     [s["om_yr"] for s in SCENARIOS], ["€#,##0"]*4); r+=1
dash_row(ws, r, "EBITDA (€/yr)",       [s["ebitda_y1"] for s in SCENARIOS], ["€#,##0"]*4); r+=1
r += 1

r = dash_section(ws, r, "BANK FINANCING (65% LTV · 5.5% · 12yr)")
dash_row(ws, r, "Senior Loan (65%)",   [s["loan"] for s in SCENARIOS], ["€#,##0"]*4); r+=1
dash_row(ws, r, "Equity Required (35%)",[s["equity"] for s in SCENARIOS], ["€#,##0"]*4); r+=1
dash_row(ws, r, "Annual Debt Service", [s["debt_service"] for s in SCENARIOS], ["€#,##0"]*4); r+=1
dash_row(ws, r, "DSCR (Yr 1)",        [round(s["dscr"],2) for s in SCENARIOS], ["0.00\"×\""]*4); r+=1
dash_row(ws, r, "Net Cash / yr (after O&M + debt)", [s["net_cash_y1"] for s in SCENARIOS], ["€#,##0"]*4); r+=1
r += 1

r = dash_section(ws, r, "RETURNS SUMMARY")
dash_row(ws, r, "Unlevered Payback (yrs)", [round(s["unlevered_pb"],1) for s in SCENARIOS], ["0.0"]*4); r+=1
dash_row(ws, r, "Equity Payback (levered, yrs)", [round(s["equity_pb_precise"],1) for s in SCENARIOS], ["0.0"]*4); r+=1
dash_row(ws, r, "20-Year Net Profit (€)", [round(s["total_20yr_net"]) for s in SCENARIOS], ["€#,##0"]*4); r+=1

# Highlight best (P2 South) column
for row_idx in range(6, r):
    c = ws.cell(row=row_idx, column=5)  # column 5 = P2 South
    if c.fill.patternType == "solid" and c.fill.start_color.rgb in ("FF" + NAVY, "FF" + NAVY2):
        pass
    elif row_idx >= 9:
        c.fill = _fill("F0FBF5")  # very light green highlight
    ws.cell(row=row_idx, column=6).fill = _fill("F8F9FB")

# O&M breakdown note
r += 2
ws.merge_cells(f"B{r}:F{r}")
c = ws.cell(row=r, column=2, value="O&M Breakdown:  PV €15/kWp/yr  +  BESS €2,500/MWh/yr  +  Fixed monitoring/insurance €2,000–2,500/yr  |  Source: Agios Theodoros RTB opex stack (pvOm €15.15/kWp, bessOm €2,462/MWh)")
c.font = Font(name="Calibri", size=9, color="7A8FA6", italic=True)
c.alignment = _left()

r += 1
ws.merge_cells(f"B{r}:F{r}")
c = ws.cell(row=r, column=2, value="⚠  BESS sizing note: All systems (0.38–1.68 MWh) are below Linyang minimum container (5.015 MWh). Practical BESS: sub-5MWh commercial units (Huawei/Sungrow). €127k/MWh applied proportionally from confirmed Agios Theodoros stack.")
c.font = Font(name="Calibri", size=9, color=RED_TXT, italic=True)
c.alignment = _left()

r += 1
ws.merge_cells(f"B{r}:F{r}")
c = ws.cell(row=r, column=2, value="⚠  Plot 2 (0/2257) has Provisional title (Art.10, Law 44/84). Confirm upgrade to full title before permitting. Both plots agricultural — Town Planning Permission required.")
c.font = Font(name="Calibri", size=9, color=RED_TXT, italic=True)
c.alignment = _left()

# ══════════════════════════════════════════════════════════
# SHEET 2: INPUTS
# ══════════════════════════════════════════════════════════
wi = wb.create_sheet("Inputs")
wi.sheet_properties.tabColor = "C9A432"
wi.column_dimensions["A"].width = 2
wi.column_dimensions["B"].width = 42
wi.column_dimensions["C"].width = 18
wi.column_dimensions["D"].width = 12
wi.column_dimensions["E"].width = 55

wi.merge_cells("B1:E1")
c = wi.cell(row=1, column=2, value="INPUTS — PV + BESS Scenario Model · Softades & Αρκατζιά")
c.font = Font(name="Calibri", bold=True, size=14, color=NAVY)
wi.merge_cells("B2:E2")
c = wi.cell(row=2, column=2, value="Yellow cells are editable. Calculated cells update automatically.")
c.font = Font(name="Calibri", size=10, color="7A8FA6", italic=True)

def inp_row(ws, r, label, value, unit, note="", editable=True):
    ws.cell(row=r, column=2, value=label).font = _font(sz=11)
    c = ws.cell(row=r, column=3, value=value)
    c.font = Font(name="Calibri", bold=True, size=11,
                  color=RED_TXT if editable else "006100")
    c.fill = _fill(INPUT_BG if editable else CALC_BG)
    c.border = _border()
    c.alignment = _right()
    ws.cell(row=r, column=4, value=unit).font = _font(sz=10, color="7A8FA6")
    ws.cell(row=r, column=5, value=note).font = _font(sz=9, color="7A8FA6", italic=True)

r = 4
wi.cell(row=r, column=2, value="EPC & EQUIPMENT COSTS").font = _font(bold=True, sz=11, color=NAVY)
r += 1
inp_row(wi, r, "PV EPC Rate",         PV_EPC_PER_KWP,  "€/kWp",   "Lighthief confirmed — Agios Theodoros RTB (€0.72/Wp)"); r+=1
inp_row(wi, r, "BESS Installed Cost",  BESS_PER_MWH,    "€/MWh",   "4-hour LFP, all-in installed. Agios Theodoros RTB confirmed stack"); r+=1
inp_row(wi, r, "Permitting & Docs",    PERMITTING,       "€/park",  "CERA license, town planning, building permit, EAC study, consultant fees"); r+=1

r += 1
wi.cell(row=r, column=2, value="YIELD ASSUMPTIONS").font = _font(bold=True, sz=11, color=NAVY)
r += 1
inp_row(wi, r, "South-Facing Yield",  2150, "kWh/kWp/yr", "AE Solar Meteor + white albedo 0.70: monofacial 1,950 + 10.3% bifacial gain (80% bifaciality)"); r+=1
inp_row(wi, r, "East–West Yield",     1450, "kWh/kWp/yr", "AE Solar Meteor + white albedo 0.70 at 12° tilt: monofacial 1,380 + 5% bifacial gain"); r+=1
inp_row(wi, r, "PV Annual Degradation", 0.4, "%/yr",       "N-type TOPCon — low degradation rate"); r+=1
inp_row(wi, r, "Electricity Tariff",  0.16,  "€/kWh",     "Net metering avoided cost / PPA rate"); r+=1

r += 1
wi.cell(row=r, column=2, value="O&M COSTS (per year)").font = _font(bold=True, sz=11, color=NAVY)
r += 1
inp_row(wi, r, "PV O&M Rate",         OM_PV_PER_KWP,    "€/kWp/yr", "Cleaning, inspection, monitoring. Ref: Agios Theodoros pvOm €15.15/kWp"); r+=1
inp_row(wi, r, "BESS O&M Rate",       OM_BESS_PER_MWH,  "€/MWh/yr", "Maintenance, spares. Ref: Agios Theodoros bessOm €2,462/MWh"); r+=1
inp_row(wi, r, "Fixed Overhead",      "2,000 – 2,500",  "€/yr",     "SCADA monitoring + insurance (Plot 1: €2,000; Plot 2: €2,500)"); r+=1

r += 1
wi.cell(row=r, column=2, value="BANK FINANCING").font = _font(bold=True, sz=11, color=NAVY)
r += 1
inp_row(wi, r, "Loan-to-Value (LTV)", LTV * 100,        "%",        "Senior debt — Cyprus green / SME loan"); r+=1
inp_row(wi, r, "Interest Rate",       LOAN_RATE * 100,  "% p.a.",   "Fixed rate"); r+=1
inp_row(wi, r, "Loan Term",           LOAN_YEARS,        "years",    "Annuity (equal annual payments)"); r+=1
inp_row(wi, r, "Annuity Factor",      round(AF, 6),      "",         "r×(1+r)^n / ((1+r)^n−1) — auto-calculated", editable=False); r+=1

r += 1
wi.cell(row=r, column=2, value="PANEL SPECIFICATION — AE SOLAR METEOR").font = _font(bold=True, sz=11, color=NAVY)
r += 1
for lbl, val, unit in [
    ("Model",             "AE620CMER-132BDS / similar",  ""),
    ("Technology",        "N-type TOPCon, half-cut cells", ""),
    ("Power Range",       "620–630",                     "Wp"),
    ("Module Efficiency", 22.97,                         "%"),
    ("Bifaciality",       80,                            "%  (±5%)"),
    ("Module Dimensions", "2,383 × 1,133",               "mm"),
    ("Weight",            33.7,                          "kg"),
    ("Product Warranty",  30,                            "years"),
    ("Performance Warranty", 30,                         "years (linear)"),
]:
    inp_row(wi, r, lbl, val, unit, editable=False); r+=1

r += 1
wi.cell(row=r, column=2, value="GROUND COVER (WHITE ALBEDO)").font = _font(bold=True, sz=11, color=NAVY)
r += 1
for lbl, val, unit, note in [
    ("Albedo Coefficient", 0.70, "",       "White limestone gravel or coated concrete — matches Agios Theodoros RTB spec"),
    ("South Bifacial Gain", "10.3%", "",  "Bifaciality 80% × rear/front irradiance ratio at 25° tilt, GHI 1,900 kWh/m²"),
    ("EW Bifacial Gain",    "5.0%", "",   "Lower rear irradiance at 12° tilt — conservative estimate"),
]:
    inp_row(wi, r, lbl, val, unit, note, editable=False); r+=1

# ══════════════════════════════════════════════════════════
# SHEETS 3–6: 15-YEAR CASH FLOW PER SCENARIO
# ══════════════════════════════════════════════════════════
CF_HEADERS = [
    "Year", "PV Degrad.", "Production\n(MWh)", "Gross Revenue\n(€)",
    "O&M Cost\n(€)", "EBITDA\n(€)", "Debt Service\n(€)", "Net Cash\n(€)",
    "Cumulative\n(€)"
]

for s in SCENARIOS:
    ws_cf = wb.create_sheet(f"CF_{s['id']}")
    ws_cf.sheet_properties.tabColor = "2B5FA0"
    ws_cf.freeze_panes = "A5"

    # Column widths
    for i, w in enumerate([14, 11, 14, 16, 14, 16, 15, 16, 16], 1):
        ws_cf.column_dimensions[get_column_letter(i)].width = w

    # Title
    ws_cf.merge_cells("A1:I1")
    c = ws_cf.cell(row=1, column=1,
        value=f"15/20-YEAR CASH FLOW — Plot {s['plot']} · {s['orient']} · {s['kwp']} kWp · {s['location']}")
    c.font = Font(name="Calibri", bold=True, size=13, color=NAVY)
    c.alignment = _center()
    c.fill = _fill("F0F4F8")

    ws_cf.merge_cells("A2:I2")
    c = ws_cf.cell(row=2, column=1,
        value=f"Owner: {s['owner']}  |  Plot ref: {s['ref']}  |  CAPEX: €{s['capex']:,.0f}  |  Equity: €{s['equity']:,.0f}  |  Loan: €{s['loan']:,.0f} @ 5.5%/12yr  |  BESS: {s['bess_mwh']} MWh (4h)")
    c.font = Font(name="Calibri", size=10, color="5A7090", italic=True)
    c.alignment = _center()
    c.fill = _fill("F0F4F8")

    # Key metrics row
    ws_cf.merge_cells("A3:I3")
    metrics_str = (
        f"Y1 Revenue: €{s['revenue_y1']:,.0f}  |  Y1 EBITDA: €{s['ebitda_y1']:,.0f}  |  "
        f"DSCR: {s['dscr']:.2f}×  |  Net cash/yr (Y1): €{s['net_cash_y1']:,.0f}  |  "
        f"Equity payback: {s['equity_pb_precise']:.1f} yr  |  "
        f"Unlevered payback: {s['unlevered_pb']:.1f} yr"
    )
    ws_cf.merge_cells("A3:I3")
    c = ws_cf.cell(row=3, column=1, value=metrics_str)
    c.font = Font(name="Calibri", bold=True, size=10, color=GREEN)
    c.alignment = _center()
    c.fill = _fill(GREEN_LT)

    # Headers
    for col, h in enumerate(CF_HEADERS, 1):
        hdr(ws_cf, 4, col, h, bg=NAVY, sz=10)
    ws_cf.row_dimensions[4].height = 30

    # Year 0 (equity investment)
    r = 5
    for col, val in enumerate(["Year 0", "—", "—", "—", "—", "—", "—",
                                f"−€{s['equity']:,.0f}  (equity in)", f"−€{s['equity']:,.0f}"], 1):
        c = ws_cf.cell(row=r, column=col, value=val)
        c.font = Font(name="Calibri", bold=True, size=11, color=RED_TXT)
        c.fill = _fill("FEF3C7")
        c.border = _border()
        c.alignment = _right() if col > 1 else _left()

    for yr_idx, f in enumerate(s["flows"]):
        r = yr_idx + 6
        debt_yr = s["debt_service"] if f["yr"] <= LOAN_YEARS else 0

        row_bg = None
        if f["yr"] > LOAN_YEARS:
            row_bg = "F0FBF5"  # post-debt, light green
        if f["cum"] >= 0 and (yr_idx == 0 or s["flows"][yr_idx-1]["cum"] < 0):
            row_bg = "C6EFCE"  # equity recovered — bright green

        vals = [
            (f"Year {f['yr']}", "@", False, False),
            (f"{f['deg']*100:.1f}%", "@", False, False),
            (round(f["rev"]/TARIFF/1000, 1), "#,##0.0", False, False),
            (round(f["rev"]), "€#,##0", False, False),
            (round(f["om"]), "€#,##0", False, False),
            (round(f["ebitda"]), "€#,##0", True, False),
            (round(debt_yr), "€#,##0", False, debt_yr == 0),
            (round(f["net"]), "€#,##0", True, False),
            (round(f["cum"]), "€#,##0", True, f["cum"] >= 0),
        ]
        for col, (v, fmt, bold, highlight) in enumerate(vals, 1):
            c = ws_cf.cell(row=r, column=col, value=v)
            c.number_format = fmt
            c.font = Font(name="Calibri", bold=bold, size=11,
                         color=GREEN if highlight else GREY_TXT)
            c.border = _border()
            c.alignment = _right() if col > 1 else _left()
            if row_bg:
                c.fill = _fill(row_bg)

    # Totals row
    r = 6 + PROJECT_LIFE
    ws_cf.cell(row=r, column=1, value="TOTALS (Yr 1–20)").font = _font(bold=True, sz=11, color=WHITE)
    ws_cf.cell(row=r, column=1).fill = _fill(NAVY)
    ws_cf.cell(row=r, column=1).border = _border()
    ws_cf.cell(row=r, column=1).alignment = _left()

    total_rev = sum(f["rev"] for f in s["flows"])
    total_om  = sum(f["om"]  for f in s["flows"])
    total_ds  = sum(f["ds"]  for f in s["flows"])
    total_net = sum(f["net"] for f in s["flows"])

    for col, (v, fmt) in enumerate([
        ("", "@"), ("", "@"),
        ("", "@"), (round(total_rev), "€#,##0"),
        (round(total_om), "€#,##0"),   ("", "@"),
        (round(total_ds), "€#,##0"),   (round(total_net), "€#,##0"),
        (round(total_net - s["equity"]), "€#,##0"),
    ], 1):
        c = ws_cf.cell(row=r, column=col, value=v)
        c.number_format = fmt
        c.font = Font(name="Calibri", bold=True, size=11, color=WHITE)
        c.fill = _fill(NAVY)
        c.border = _border()
        c.alignment = _right() if col > 1 else _left()

    # Legend
    r += 2
    ws_cf.merge_cells(f"A{r}:I{r}")
    c = ws_cf.cell(row=r, column=1,
        value="🟢 Green rows = post-debt period (no more debt service).  "
              "✅ Bright green = year equity is fully recovered.  "
              "Note: PV revenue decreases 0.4%/yr (N-type TOPCon degradation).")
    c.font = Font(name="Calibri", size=9, color="5A7090", italic=True)
    c.alignment = _left()

# ══════════════════════════════════════════════════════════
# SHEET 7: O&M BREAKDOWN
# ══════════════════════════════════════════════════════════
wo = wb.create_sheet("O&M_Breakdown")
wo.sheet_properties.tabColor = "1E7D4A"
wo.column_dimensions["A"].width = 2
wo.column_dimensions["B"].width = 36
for i in range(3, 7):
    wo.column_dimensions[get_column_letter(i)].width = 18

wo.merge_cells("B1:F1")
c = wo.cell(row=1, column=2, value="O&M COST BREAKDOWN — ALL SCENARIOS")
c.font = Font(name="Calibri", bold=True, size=14, color=NAVY)
c.alignment = _center()

r = 3
hdr(wo, r, 2, "O&M Component", bg=NAVY)
for i, s in enumerate(SCENARIOS):
    hdr(wo, r, 3+i, f"Plot {s['plot']} {s['orient']}\n{s['kwp']} kWp", bg=NAVY if s['plot']==1 else NAVY2)
r += 1

components = [
    ("PV O&M  (€15/kWp/yr)",        [s["kwp"] * OM_PV_PER_KWP for s in SCENARIOS]),
    ("BESS O&M  (€2,500/MWh/yr)",   [round(s["bess_mwh"] * OM_BESS_PER_MWH) for s in SCENARIOS]),
    ("Monitoring + Insurance (flat)",[s["om_fixed"] for s in SCENARIOS]),
]
for lbl, vals in components:
    cell(wo, r, 2, lbl, align="left"); 
    for i, v in enumerate(vals):
        cell(wo, r, 3+i, v, fmt="€#,##0", align="right")
    r += 1

# Total row
cell(wo, r, 2, "TOTAL O&M / yr", bold=True, color=NAVY)
for i, s in enumerate(SCENARIOS):
    cell(wo, r, 3+i, s["om_yr"], fmt="€#,##0", bold=True, color=GREEN, align="right")
r += 1

r += 2
wo.merge_cells(f"B{r}:F{r}")
c = wo.cell(row=r, column=2,
    value="O&M basis: Agios Theodoros RTB confirmed opex (pvOm €15.15/kWp, bessOm €2,462/MWh). "
          "Monitoring/insurance flat fee covers SCADA access, site insurance, compliance renewals. "
          "O&M increases ~20% after Year 5 as plant ages — modelled as flat here (conservative upside).")
c.font = Font(name="Calibri", size=9, color="7A8FA6", italic=True)
c.alignment = _left()

# ══════════════════════════════════════════════════════════
# REORDER SHEETS
# ══════════════════════════════════════════════════════════
sheet_order = ["Dashboard", "Inputs", "CF_P1S", "CF_P1EW", "CF_P2S", "CF_P2EW", "O&M_Breakdown"]
for i, name in enumerate(sheet_order):
    if name in wb.sheetnames:
        wb.move_sheet(name, offset=i - wb.sheetnames.index(name))

# ── SAVE ──
for out_dir in [
    "docs/clients/Individual_Pavlos_Pavlou",
    "docs/clients/Individual_Pavlou_Koulla",
]:
    os.makedirs(out_dir, exist_ok=True)
    out_path = f"{out_dir}/softades-arkatzia-pv-bess-scenarios-apr2026.xlsx"
    wb.save(out_path)
    print(f"Saved: {out_path}")
print(f"Sheets: {', '.join(wb.sheetnames)}")
