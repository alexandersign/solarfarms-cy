import openpyxl
import sys

out = open('/Volumes/T7 Grey/solinvest/scripts/adders_output.txt', 'w')

# V4 Adders
wb4 = openpyxl.load_workbook('docs/internal/Lighthief-EPC-Confirmed-Adders-v4-Feb2026.xlsx', data_only=True)
out.write('=== V4 ADDERS ===\n')
out.write(f'Sheets: {wb4.sheetnames}\n')
for sn in wb4.sheetnames:
    ws = wb4[sn]
    out.write(f'\n--- {sn} (rows={ws.max_row}, cols={ws.max_column}) ---\n')
    for r in range(1, min(ws.max_row+1, 60)):
        vals = []
        for c in range(1, min(ws.max_column+1, 40)):
            cell = ws.cell(row=r, column=c)
            if cell.value is not None:
                vals.append((cell.column_letter + str(r), cell.value))
        if vals:
            out.write(str(vals) + '\n')

out.write('\n\n')

# V2 EPC System Cost
wb2 = openpyxl.load_workbook('docs/Bess - EPC System Cost v2.xlsx', data_only=True)
ws2 = wb2['Pricing_Model_All_Projects']

headers = {}
for c in range(1, ws2.max_column+1):
    v = ws2.cell(row=1, column=c).value
    if v:
        headers[c] = str(v)

out.write('=== V2 ADDER COLUMNS (S through AO) ===\n')
adder_cols = []
for c, h in sorted(headers.items()):
    if c >= 19 and c <= 41:
        adder_cols.append((c, h))
        out.write(f'  Col {c} ({ws2.cell(row=1, column=c).column_letter}): {h}\n')

out.write('\n=== V2 ADDER VALUES PER PARK ===\n')
for r in range(4, min(ws2.max_row+1, 60)):
    park = ws2.cell(row=r, column=7).value
    mwh = ws2.cell(row=r, column=9).value
    if not park:
        continue
    out.write(f'\nPark: {park} ({mwh} MWh)\n')
    for c, h in adder_cols:
        v = ws2.cell(row=r, column=c).value
        if v is not None and v != 0:
            out.write(f'  {h}: {v}\n')
    ao = ws2.cell(row=r, column=41).value
    out.write(f'  ADDERS TOTAL (excl EMS): {ao}\n')
    az = ws2.cell(row=r, column=52).value
    out.write(f'  EMS/SCADA Total: {az}\n')

out.close()
print('Done')
