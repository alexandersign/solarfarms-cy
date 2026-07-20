"""
Generate investor Excel models for Galascope 5MW and Christos 3.3MW parks for sale.
Formula-based — investors edit yellow INPUTS cells; CASHFLOW/SUMMARY update automatically.
Run: python3 scripts/gen-parks-for-sale-excel.py
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os

# ── Brand colours ───────────────────────────────────────────────────────────
NAVY  = "1A365D";  GOLD  = "C9A432";  WHITE = "FFFFFF"
YELL  = "FFF9C4"   # editable inputs
GRN   = "E8F5E9";  BLU   = "E3F2FD"
LGREY = "F0F0F0";  GREY  = "E8E8E8"

def _b(style="thin", color="CCCCCC"):
    return Side(style=style, color=color)

TBDR = Border(left=_b(), right=_b(), top=_b(), bottom=_b())
MBDR = Border(left=_b("medium","999999"), right=_b("medium","999999"),
              top=_b("medium","999999"), bottom=_b("medium","999999"))

NUM0 = "#,##0"; NUM1 = "#,##0.0"; EUR = "#,##0"
PCT  = "0.0%";  PCT2 = "0.00%"

def C(ws, row, col, value="", bold=False, sz=10, color=WHITE, bg=None,
      align="left", fmt=None, border=TBDR, wrap=False):
    c = ws.cell(row=row, column=col, value=value)
    c.font      = Font(bold=bold, size=sz, name="Calibri", color=color)
    c.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
    c.border    = border
    if bg:
        c.fill = PatternFill("solid", fgColor=bg)
    if fmt:
        c.number_format = fmt
    return c

def HDR(ws, r, c, txt, bg=NAVY, fg=WHITE, sz=10, wrap=True):
    return C(ws, r, c, txt, bold=True, sz=sz, color=fg, bg=bg,
             align="center", wrap=wrap, border=TBDR)

def LBL(ws, r, c, txt, bold=False, bg=None, indent=0):
    cell = C(ws, r, c, txt, bold=bold, sz=10, color="000000", bg=bg,
             align="left", border=TBDR)
    if indent:
        cell.alignment = Alignment(horizontal="left", vertical="center", indent=indent)
    return cell

def INP(ws, r, c, value, fmt=None):
    """Yellow editable input cell."""
    return C(ws, r, c, value, bold=True, sz=10, color="1A1A1A", bg=YELL,
             align="right", fmt=fmt, border=MBDR)

def VAL(ws, r, c, formula, fmt=None, bold=False, bg=None):
    color = "000000"
    return C(ws, r, c, formula, bold=bold, sz=10, color=color, bg=bg,
             align="right", fmt=fmt, border=TBDR)

def SHDR(ws, r, txt, ncols=4):
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=ncols)
    C(ws, r, 1, txt, bold=True, sz=10, color=WHITE, bg=NAVY, align="left")

def NOTE(ws, r, txt, ncols=3):
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=ncols)
    c = ws.cell(row=r, column=1, value=txt)
    c.font      = Font(italic=True, size=8, color="555555", name="Calibri")
    c.alignment = Alignment(horizontal="left", wrap_text=True)

def BANNER(ws, r, txt, ncols=4, sz=12):
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=ncols)
    ws.row_dimensions[r].height = 28
    C(ws, r, 1, txt, bold=True, sz=sz, color=WHITE, bg=NAVY, align="center", border=TBDR)

# ────────────────────────────────────────────────────────────────────────────
#  Helper: build INPUTS sheet, return dict of B-column row numbers
# ────────────────────────────────────────────────────────────────────────────

def build_inputs_galascope(ws):
    ws.sheet_view.showGridLines = False
    ws.column_dimensions['A'].width = 36
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 42

    BANNER(ws, 1, "PARK-REF-5001 — 5.01 MWp Famagusta Operational Solar + BESS Investor Model", ncols=3)
    NOTE(ws, 2, "Lighthief Cyprus Ltd · Jun 2026 · Edit YELLOW cells only — all other sheets auto-recalculate")

    ws.row_dimensions[1].height = 28
    HDR(ws, 3, 1, "Parameter"); HDR(ws, 3, 2, "Value  ← EDIT YELLOW", bg=GOLD, fg="1A1A1A"); HDR(ws, 3, 3, "Notes / Source")

    ref = {}  # {key: row_number}

    def add(r, key, label, default, fmt, note=""):
        LBL(ws, r, 1, label)
        INP(ws, r, 2, default, fmt)
        LBL(ws, r, 3, note)
        if key:
            ref[key] = r

    r = 4
    SHDR(ws, r, "PV PARK", ncols=3); r += 1
    add(r, "dc_cap",    "DC installed capacity (MWp)",       5.01,     NUM1, "Installed nameplate"); r+=1
    add(r, "ac_cap",    "AC export capacity (MW)",           4.62,     NUM1, "44 × Huawei SUN2000-105KTL-H1"); r+=1
    add(r, "gross",     "Annual gross yield (MWh/yr)",     10200,      NUM0, "2021 actual: 10,146 MWh @ 0% curtailment"); r+=1
    add(r, "curt",      "Curtailment rate — base case (%)", 0.50,      PCT,  "50% = 2025/26 reality. Try 35%–60%."); r+=1
    add(r, "pv_price",  "PV avg selling price (€/MWh)",     127,       EUR,  "DAM full-year daytime avg (TSOC verified)"); r+=1
    add(r, "pv_degr",   "PV panel degradation (%/yr)",     0.005,      PCT2, "Standard: 0.5%/yr LID + aging"); r+=1
    add(r, "pv_opex",   "PV OPEX total (€/yr)",          138000,       EUR,  "Land €25K + O&M €40K + Insurance €25K + TSOC/admin €48K"); r+=1
    add(r, "asking",    "Park asking price (€)",          9000000,      EUR,  "Indicative — subject to NDA and negotiation"); r+=1

    SHDR(ws, r, "BESS — Lighthief Turnkey (optional addition)", ncols=3); r += 1
    add(r, "bess_mw",   "BESS power (MW)",                   5.0,      NUM1, "4 × Kehua BCS1250K PCS"); r+=1
    add(r, "bess_mwh",  "BESS energy (MWh)",                  20,      NUM0, "4 × Linyang 5.015 MWh LFP containers"); r+=1
    add(r, "bess_epc",  "BESS turnkey EPC price (€)",    2380000,      EUR,  "€119,000/MWh × 20 MWh — fully installed + commissioned"); r+=1
    add(r, "capture",   "BESS effective capture rate, Yr 1", 0.85,      PCT,  "Tank overflow + 87.8% RTE (model). Avg 14 MWh/day curtailed."); r+=1
    add(r, "bess_px",   "BESS discharge price (€/MWh)",       175,     EUR,  "Conservative DAM evening (TSOC avg: €184/MWh). Adjust up/down."); r+=1
    add(r, "batt_degr", "Battery degradation (%/yr)",       0.025,     PCT2, "EVE LFP 7,000 cycles @ 70% EOL, ~2.5%/yr"); r+=1
    add(r, "bopex1",    "BESS OPEX — Yr 1–5 (€/yr)",       30000,      EUR,  "Standard warranty period: insurance + basic maintenance"); r+=1
    add(r, "bopex2",    "BESS OPEX — Yr 6–10 (€/yr)",      53000,      EUR,  "Linyang extended warranty + insurance (Yr 6–10)"); r+=1
    add(r, "bopex3",    "BESS OPEX — Yr 11–15 (€/yr)",     53000,      EUR,  "Linyang extended warranty + insurance (Yr 11–15)"); r+=1

    SHDR(ws, r, "FINANCIAL", ncols=3); r += 1
    add(r, "disc",      "Discount rate for NPV (%)",          0.08,     PCT,  "Adjust to your target hurdle rate"); r+=1

    NOTE(ws, r+1, "⚡ Edit yellow cells only. CASHFLOW and SUMMARY recalculate automatically. "
         "Base-case curtailment 50% reflects 2025 trend (45.8% in 2025, rising).", ncols=3)
    NOTE(ws, r+2, "📋 Sources: TSOC DAM 137 files Oct 2025–Feb 2026, SCADA production 2020–2025, "
         "Linyang specs, Lighthief EPC scope PARK-REF-5001.", ncols=3)

    ws.freeze_panes = "A4"
    return ref


def build_inputs_christos(ws):
    ws.sheet_view.showGridLines = False
    ws.column_dimensions['A'].width = 36
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 42

    BANNER(ws, 1, "PARK-OP-CHR-33-2026 — 3.3 MWp Nicosia Solar + BESS Investor Model", ncols=3)
    NOTE(ws, 2, "Lighthief Cyprus Ltd · Jun 2026 · Edit YELLOW cells only — all other sheets auto-recalculate")

    HDR(ws, 3, 1, "Parameter"); HDR(ws, 3, 2, "Value  ← EDIT YELLOW", bg=GOLD, fg="1A1A1A"); HDR(ws, 3, 3, "Notes / Source")

    ref = {}

    def add(r, key, label, default, fmt, note=""):
        LBL(ws, r, 1, label)
        INP(ws, r, 2, default, fmt)
        LBL(ws, r, 3, note)
        if key:
            ref[key] = r

    r = 4
    SHDR(ws, r, "PV PARK", ncols=3); r += 1
    add(r, "dc_cap",     "DC installed capacity (MWp)",         3.30,  NUM1, "Grid-connected Jul 2025, Margí, Nicosia District"); r+=1
    add(r, "sp_yield",   "Annual specific yield (kWh/kWp)",     2100,  NUM0, "Client-reported; Trina 620 Wp bifacial, fixed tilt"); r+=1
    add(r, "gross",      "Annual gross yield (MWh/yr)",         6930,  NUM0, "= 3.3 MWp × 2,100 kWh/kWp"); r+=1
    add(r, "curt",       "Curtailment — base case (%)",         0.50,  PCT,  "50% = 2025/26 reality. Client reported ~47%. Try 35–60%."); r+=1
    add(r, "pv_price",   "PV avg selling price (€/MWh)",        127,   EUR,  "DAM full-year daytime avg (TSOC verified)"); r+=1
    add(r, "pv_degr",    "PV panel degradation (%/yr)",        0.005,  PCT2, "Standard: 0.5%/yr"); r+=1
    add(r, "pv_opex",    "PV OPEX total (€/yr)",              88000,   EUR,  "O&M €20K + land €10K + insurance €8K + admin €50K"); r+=1
    add(r, "exp_mw",     "Expansion PV (MWp — no grid conn.)",   1.70, NUM1, "CERA permit + building permit — captive/BTM use only"); r+=1
    add(r, "exp_mwh",    "Expansion yield (MWh/yr, BTM)",       3230,  NUM0, "1.7 MWp × 1,900 kWh/kWp (bifacial, conservative)"); r+=1
    add(r, "asking",     "Park asking price (€)",                  0,  EUR,  "Available under NDA — enter price once received"); r+=1

    SHDR(ws, r, "BESS — Lighthief Turnkey (optional addition)", ncols=3); r += 1
    add(r, "bess_mw",    "BESS power (MW)",                      3.3,  NUM1, "Power-matched to park capacity"); r+=1
    add(r, "bess_mwh",   "BESS energy (MWh)",                     12,  NUM0, "12 MWh usable — Linyang LFP containers"); r+=1
    add(r, "bess_epc",   "BESS turnkey EPC price (€)",       1560000,  EUR,  "€130,000/MWh × 12 MWh — fully installed + commissioned"); r+=1
    add(r, "capture",    "BESS effective capture rate, Yr 1",   0.78,  PCT,  "Includes tank overflow + 86.3% AC-AC RTE. Avg 9.5 MWh/day."); r+=1
    add(r, "bess_px",    "BESS discharge price (€/MWh)",         183,  EUR,  "TSOC verified peak avg 17:00–21:00 (134 days, Oct–Feb 2026)"); r+=1
    add(r, "batt_degr",  "Battery degradation (%/yr)",         0.025,  PCT2, "EVE LFP 7,000 cycles @ 70% EOL, ~2.5%/yr"); r+=1
    add(r, "bopex1",     "BESS OPEX — Yr 1–5 (€/yr)",         38000,  EUR,  "Standard warranty + insurance (scaled to 12 MWh)"); r+=1
    add(r, "bopex2",     "BESS OPEX — Yr 6–10 (€/yr)",        55000,  EUR,  "Linyang extended warranty + insurance"); r+=1
    add(r, "bopex3",     "BESS OPEX — Yr 11–15 (€/yr)",       55000,  EUR,  "Linyang extended warranty + insurance"); r+=1

    SHDR(ws, r, "FINANCIAL", ncols=3); r += 1
    add(r, "disc",       "Discount rate for NPV (%)",            0.08, PCT,  "Adjust to your target hurdle rate"); r+=1

    NOTE(ws, r+1, "⚡ Edit yellow cells only. CASHFLOW and SUMMARY recalculate automatically. "
         "Park asking price available under NDA — enter once received from Lighthief.", ncols=3)
    NOTE(ws, r+2, "📋 Sources: TSOC DAM 134 files Oct 2025–Feb 2026, client-reported production 2025, "
         "Linyang BESS specs, Lighthief EPC scope PARK-OP-CHR-33-2026.", ncols=3)

    ws.freeze_panes = "A4"
    return ref


# ────────────────────────────────────────────────────────────────────────────
#  Shared: CASHFLOW sheet builder
# ────────────────────────────────────────────────────────────────────────────

def build_cashflow(ws, ref, park_id):
    """
    Build 15-year BESS cash flow. ref = dict of key→row in INPUTS.
    All formulas reference INPUTS!$B$<row>.
    """
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "B3"

    col_w = [5, 13, 13, 13, 13, 13, 13, 12, 12, 13, 13, 13, 14, 11, 14]
    for i, w in enumerate(col_w, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    BANNER(ws, 1, f"{park_id} — 15-Year Cash Flow (all formulas; edit INPUTS to update)", ncols=15)

    hdrs = ["Yr", "Gross PV\nMWh", "Curtailed\nMWh", "Net PV\nMWh",
            "BESS\nDisch. MWh", "PV Rev\n€", "BESS Rev\n€",
            "BESS\nOPEX €", "PV\nOPEX €", "BESS\nNet €",
            "PV FCF\n€", "Total FCF\n€", "Cum BESS\nNet €",
            "NPV\nFactor", "BESS Net\nDisc. €"]
    for ci, h in enumerate(hdrs, 1):
        HDR(ws, 2, ci, h, sz=8, wrap=True)
    ws.row_dimensions[2].height = 36

    I = "INPUTS"
    G  = ref["gross"];    CU  = ref["curt"];    PP  = ref["pv_price"]
    PD = ref["pv_degr"];  PO  = ref["pv_opex"]; CA  = ref["capture"]
    BP = ref["bess_px"];  BD  = ref["batt_degr"]
    O1 = ref["bopex1"];   O2  = ref["bopex2"];  O3  = ref["bopex3"]
    DR = ref["disc"]

    for yr in range(1, 16):
        rr  = yr + 2          # worksheet row
        yrm = yr - 1
        bg  = LGREY if yr % 2 == 0 else None
        hl  = GRN if yr in (1, 5, 10, 15) else bg

        # Gross PV (degrades each year by pv_degr)
        VAL(ws, rr, 1,  yr,   NUM0, bold=True, bg=LGREY)
        VAL(ws, rr, 2,  f"=INPUTS!$B${G}*(1-INPUTS!$B${PD})^{yrm}", NUM0, bg=bg)
        VAL(ws, rr, 3,  f"=B{rr}*INPUTS!$B${CU}",                    NUM0, bg=bg)
        VAL(ws, rr, 4,  f"=B{rr}*(1-INPUTS!$B${CU})",                NUM0, bg=bg)
        # BESS discharged = curtailed × capture × battery_remaining
        VAL(ws, rr, 5,  f"=C{rr}*INPUTS!$B${CA}*(1-INPUTS!$B${BD})^{yrm}", NUM0, bg=bg)
        VAL(ws, rr, 6,  f"=D{rr}*INPUTS!$B${PP}",                    EUR,  bg=bg)
        VAL(ws, rr, 7,  f"=E{rr}*INPUTS!$B${BP}",                    EUR,  bg=bg)
        # BESS OPEX: 3 brackets by year
        bopex_row = O1 if yr <= 5 else (O2 if yr <= 10 else O3)
        VAL(ws, rr, 8,  f"=INPUTS!$B${bopex_row}",                   EUR,  bg=bg)
        VAL(ws, rr, 9,  f"=INPUTS!$B${PO}",                          EUR,  bg=bg)
        VAL(ws, rr, 10, f"=G{rr}-H{rr}",                             EUR,  bold=(yr==1), bg=hl)
        VAL(ws, rr, 11, f"=F{rr}-I{rr}",                             EUR,  bg=bg)
        VAL(ws, rr, 12, f"=K{rr}+J{rr}",                             EUR,  bold=(yr==1), bg=hl)
        cum = f"=J{rr}" if yr == 1 else f"=M{rr-1}+J{rr}"
        VAL(ws, rr, 13, cum,                                          EUR,  bg=bg)
        VAL(ws, rr, 14, f"=1/(1+INPUTS!$B${DR})^A{rr}",             "0.0000", bg=bg)
        VAL(ws, rr, 15, f"=J{rr}*N{rr}",                             EUR,  bg=bg)

    # Totals row
    tr = 18
    LBL(ws, tr, 1, "TOTALS", bold=True, bg=LGREY)
    for col, rng in [(10, "J"), (11, "K"), (12, "L"), (13, "M"), (15, "O")]:
        VAL(ws, tr, col, f"=SUM({rng}3:{rng}17)", EUR, bold=True, bg=GRN)

    NOTE(ws, 19, "Battery 2.5%/yr decay · PV 0.5%/yr · BESS OPEX rises in Yr 6 (extended warranty) · Edit all rates in INPUTS", ncols=15)


# ────────────────────────────────────────────────────────────────────────────
#  Shared: SUMMARY sheet builder
# ────────────────────────────────────────────────────────────────────────────

def build_summary(ws, ref, park_id, has_bess_epc_key="bess_epc",
                  has_asking_key="asking", has_exp_key=None):
    ws.sheet_view.showGridLines = False
    ws.column_dimensions['A'].width = 34
    for col in ['B','C','D']:
        ws.column_dimensions[col].width = 20

    BANNER(ws, 1, f"{park_id} — Summary Dashboard", ncols=4)
    NOTE(ws, 2, "All numbers pull from INPUTS — edit there. Yellow = inputs; green = with BESS; blue = PV-only.", ncols=4)

    HDR(ws, 3, 1, "Metric")
    HDR(ws, 3, 2, "PV Only")
    HDR(ws, 3, 3, "PV + BESS", bg=GOLD, fg="1A1A1A")
    HDR(ws, 3, 4, "BESS Standalone")

    I = "INPUTS"
    G  = ref["gross"];   CU = ref["curt"];   PP = ref["pv_price"]
    PD = ref["pv_degr"]; PO = ref["pv_opex"]
    CA = ref["capture"]; BP = ref["bess_px"]
    O1 = ref["bopex1"];  EPC = ref[has_bess_epc_key]
    ASK = ref[has_asking_key]

    # Reusable formula fragments
    net_pv    = f"INPUTS!$B${G}*(1-INPUTS!$B${CU})"
    curt_mwh  = f"INPUTS!$B${G}*INPUTS!$B${CU}"
    bess_dis  = f"INPUTS!$B${G}*INPUTS!$B${CU}*INPUTS!$B${CA}"
    pv_rev    = f"{net_pv}*INPUTS!$B${PP}"
    bess_rev  = f"{bess_dis}*INPUTS!$B${BP}"
    bess_net  = f"({bess_rev})-INPUTS!$B${O1}"

    r2 = 4
    SHDR(ws, r2, "YEAR 1 RESULTS", ncols=4); r2 += 1

    def row(label, fPV, fBESS, fBESSonly, fmt=EUR, bold=False, bg_pv=None, bg_b=None):
        nonlocal r2
        LBL(ws, r2, 1, label, bold=bold)
        VAL(ws, r2, 2, fPV      if fPV      else "—", fmt=fmt if fPV      else None, bold=bold, bg=bg_pv)
        VAL(ws, r2, 3, fBESS    if fBESS    else "—", fmt=fmt if fBESS    else None, bold=bold, bg=bg_b)
        VAL(ws, r2, 4, fBESSonly if fBESSonly else "—", fmt=fmt if fBESSonly else None, bold=bold, bg=bg_b)
        r2 += 1

    row("Annual gross PV (MWh)",      f"=INPUTS!$B${G}", f"=INPUTS!$B${G}", None, NUM0)
    row("Net PV export (MWh)",         f"={net_pv}",      f"={net_pv}",     None, NUM0)
    row("Curtailed → BESS input (MWh)",None,              f"={curt_mwh}",   f"={curt_mwh}", NUM0)
    row("BESS discharged (MWh/yr)",    None,              f"={bess_dis}",    f"={bess_dis}", NUM0)
    row("PV revenue (€)",              f"={pv_rev}",      f"={pv_rev}",     None, EUR)
    row("BESS revenue (€)",            None,              f"={bess_rev}",    f"={bess_rev}", EUR)
    row("Total gross revenue (€)",     f"={pv_rev}",      f"={pv_rev}+{bess_rev}", f"={bess_rev}", EUR, bold=True, bg_pv=BLU, bg_b=GRN)
    row("Total OPEX (€)",              f"=INPUTS!$B${PO}", f"=INPUTS!$B${PO}+INPUTS!$B${O1}", f"=INPUTS!$B${O1}", EUR)
    row("Net FCF / Net income Y1 (€)", f"={pv_rev}-INPUTS!$B${PO}",
        f"={pv_rev}+{bess_rev}-INPUTS!$B${PO}-INPUTS!$B${O1}",
        f"={bess_net}", EUR, bold=True, bg_pv=BLU, bg_b=GRN)

    SHDR(ws, r2, "INVESTMENT & RETURNS", ncols=4); r2 += 1
    row("Park asking price (€)",       f"=INPUTS!$B${ASK}", f"=INPUTS!$B${ASK}", None, EUR, bold=True)
    row("BESS investment (€)",          None,               f"=INPUTS!$B${EPC}", f"=INPUTS!$B${EPC}", EUR, bold=True)
    row("Total investment (€)",         f"=INPUTS!$B${ASK}",
        f"=INPUTS!$B${ASK}+INPUTS!$B${EPC}", f"=INPUTS!$B${EPC}", EUR, bold=True, bg_pv=BLU, bg_b=GRN)
    fcf_pv  = f"({pv_rev}-INPUTS!$B${PO})"
    fcf_all = f"({pv_rev}+{bess_rev}-INPUTS!$B${PO}-INPUTS!$B${O1})"
    row("Y1 FCF yield (PV) / ROI (BESS) %",
        f"=IF(INPUTS!$B${ASK}>0,{fcf_pv}/INPUTS!$B${ASK},\"ASK N/A\")",
        f"=IF(INPUTS!$B${ASK}>0,{fcf_all}/(INPUTS!$B${ASK}+INPUTS!$B${EPC}),\"ASK N/A\")",
        f"={bess_net}/INPUTS!$B${EPC}", PCT, bold=True, bg_pv=BLU, bg_b=GRN)
    row("Simple payback (years)",
        f"=IF({fcf_pv}>0,INPUTS!$B${ASK}/{fcf_pv},\"n/a\")",
        f"=IF({fcf_all}>0,(INPUTS!$B${ASK}+INPUTS!$B${EPC})/{fcf_all},\"n/a\")",
        f"=INPUTS!$B${EPC}/{bess_net}", NUM1, bg_pv=BLU, bg_b=GRN)

    SHDR(ws, r2, "15-YEAR TOTALS (from CASHFLOW)", ncols=4); r2 += 1
    row("15-yr cumulative BESS net (€)", None, "=CASHFLOW!M17", "=CASHFLOW!M17", EUR, bold=True, bg_b=GRN)
    row("15-yr BESS profit (€)",         None,
        f"=CASHFLOW!M17-INPUTS!$B${EPC}",
        f"=CASHFLOW!M17-INPUTS!$B${EPC}", EUR, bold=True, bg_b=GRN)
    row("NPV of BESS net (€, disc. rate)", None, "=CASHFLOW!O17", "=CASHFLOW!O17", EUR, bold=True, bg_b=GRN)

    if has_exp_key:
        SHDR(ws, r2, "EXPANSION ASSET (1.7 MWp — behind-the-meter)", ncols=4); r2 += 1
        exp_r = ref[has_exp_key]
        LBL(ws, r2, 1, "Expansion potential (MWh/yr, BTM)", bold=True)
        VAL(ws, r2, 2, f"=INPUTS!$B${exp_r}", NUM0, bold=True, bg=BLU)
        VAL(ws, r2, 3, f"=INPUTS!$B${exp_r}", NUM0, bg=BLU)
        VAL(ws, r2, 4, "No grid licence required")
        r2 += 1

    NOTE(ws, r2, "⚡ Edit INPUTS sheet to change assumptions. All cells here are read-only formulas.", ncols=4)


# ────────────────────────────────────────────────────────────────────────────
#  Shared: SENSITIVITY sheet
# ────────────────────────────────────────────────────────────────────────────

def build_sensitivity(ws, ref, park_id):
    ws.sheet_view.showGridLines = False
    for col in ['A','B','C','D','E','F','G','H']:
        ws.column_dimensions[col].width = 16
    ws.column_dimensions['A'].width = 20

    BANNER(ws, 1, f"{park_id} — Sensitivity: BESS Net Y1 (€) by Curtailment & Discharge Price", ncols=8)
    NOTE(ws, 2, "Green = base case. Uses INPUTS for gross yield, capture rate, and BESS OPEX Y1. Edit INPUTS to update.", ncols=8)

    curtailments  = [0.30, 0.35, 0.40, 0.45, 0.50, 0.55, 0.60]
    prices        = [140, 155, 175, 183, 195, 210]

    HDR(ws, 3, 1, "Curt% ↓ / Price →")
    for ci, p in enumerate(prices, 2):
        HDR(ws, 3, ci, f"€{p}/MWh")

    I = "INPUTS"
    G  = ref["gross"]; CA = ref["capture"]; O1 = ref["bopex1"]
    BD = ref["batt_degr"]

    for ri, crt in enumerate(curtailments, 4):
        is_base = abs(crt - 0.50) < 0.001
        bg_r    = GRN if is_base else (LGREY if ri % 2 == 0 else None)
        LBL(ws, ri, 1, f"{int(crt*100)}%", bold=is_base, bg=bg_r)
        for ci, p in enumerate(prices, 2):
            f = f"=INPUTS!$B${G}*{crt}*INPUTS!$B${CA}*{p}-INPUTS!$B${O1}"
            VAL(ws, ri, ci, f, EUR, bold=is_base, bg=bg_r)

    # Detailed table
    r4 = 13
    SHDR(ws, r4, "BESS Returns by Curtailment (base discharge price from INPUTS)", ncols=7); r4 += 1
    for h, ci in [("Curtailment",1),("Curtailed\nMWh",2),("Discharged\nMWh",3),
                   ("BESS Rev\n€",4),("OPEX Y1\n€",5),("BESS Net\n€",6),("ROI\n%",7)]:
        HDR(ws, r4, ci, h, sz=8, wrap=True)
    r4 += 1
    BP = ref["bess_px"]; EPC = ref["bess_epc"]

    for crt in curtailments:
        is_base = abs(crt - 0.50) < 0.001
        bg_r    = GRN if is_base else None
        LBL(ws, r4, 1, f"{int(crt*100)}%", bold=is_base, bg=bg_r)
        curt_f  = f"=INPUTS!$B${G}*{crt}"
        disc_f  = f"=INPUTS!$B${G}*{crt}*INPUTS!$B${CA}"
        rev_f   = f"=INPUTS!$B${G}*{crt}*INPUTS!$B${CA}*INPUTS!$B${BP}"
        net_f   = f"={rev_f}-INPUTS!$B${O1}"
        roi_f   = f"=({rev_f}-INPUTS!$B${O1})/INPUTS!$B${EPC}"
        VAL(ws, r4, 2, curt_f, NUM0, bg=bg_r)
        VAL(ws, r4, 3, disc_f, NUM0, bg=bg_r)
        VAL(ws, r4, 4, rev_f,  EUR,  bg=bg_r)
        VAL(ws, r4, 5, f"=INPUTS!$B${O1}", EUR, bg=bg_r)
        VAL(ws, r4, 6, net_f,  EUR, bold=is_base, bg=bg_r)
        VAL(ws, r4, 7, roi_f,  PCT, bold=is_base, bg=bg_r)
        r4 += 1


# ────────────────────────────────────────────────────────────────────────────
#  Shared: DAM PRICES reference sheet
# ────────────────────────────────────────────────────────────────────────────

def build_dam(ws, source_note):
    ws.sheet_view.showGridLines = False
    ws.column_dimensions['A'].width = 32
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 48

    BANNER(ws, 1, "Verified TSOC Day-Ahead Market Data — Cyprus Competitive Market", ncols=3)
    HDR(ws, 2, 1, "Metric"); HDR(ws, 2, 2, "Value"); HDR(ws, 2, 3, "Notes")

    rows = [
        ("Source",           source_note,             "tsoc.org.cy — DAM Daily Activity Reports"),
        ("Period",           "Oct 2025 – Feb 2026",   "Full market dataset since Cyprus DAM launch"),
        ("",                 "",                      ""),
        ("Average MCP (all hours)",          "€158/MWh",   "Full market average"),
        ("Solar hours avg (06:00–17:00)",    "€141/MWh",   "PV export price proxy — TSOC verified"),
        ("Midday avg (10:00–14:00)",         "€102/MWh",   "Curtailment window — very low prices"),
        ("Peak evening avg (17:00–21:00)",   "€184/MWh",   "BESS discharge window — TSOC verified"),
        ("Peak–midday spread",               "€82/MWh",    "Value BESS creates every day"),
        ("Days with positive spread",        "100%",        "Every day has arbitrage opportunity"),
        ("Days with zero midday price",      "42%",         "Curtailment most likely on these days"),
        ("BESS charge cost",                 "€0/MWh",      "Own curtailment — zero charge under Cyprus regs"),
        ("",                 "",                      ""),
        ("Monthly breakdown",  "",              ""),
        ("Oct 2025",   "Peak €178  /  Midday €79",    "31 days"),
        ("Nov 2025",   "Peak €185  /  Midday €82",    "30 days"),
        ("Dec 2025",   "Peak €179  /  Midday €111",   "31 days"),
        ("Jan 2026",   "Peak €188  /  Midday €138",   "31 days"),
        ("Feb 2026",   "Peak €192  /  Midday €85",    "14 days"),
        ("",                 "",                      ""),
        ("Model notes",    "",    ""),
        ("Galascope model discharge price",  "€175/MWh",  "Conservative (€9 below TSOC avg) — change in INPUTS"),
        ("Christos model discharge price",   "€183/MWh",  "TSOC verified peak average — change in INPUTS"),
        ("Revenue per MWh stored (Gala)",    "€154/MWh",  "€175 × 87.8% RTE"),
        ("Revenue per MWh stored (Christos)","€158/MWh",  "€183 × 86.3% RTE"),
    ]

    for ri, (a, b, c2) in enumerate(rows, 3):
        is_hdr = a in ("Source","Monthly breakdown","Model notes")
        bg = NAVY if is_hdr else (LGREY if ri % 2 == 0 else None)
        fg = WHITE if is_hdr else "000000"
        LBL(ws, ri, 1, a, bold=is_hdr, bg=bg, color=fg) if not is_hdr else C(ws, ri, 1, a, bold=True, sz=10, color=fg, bg=bg, align="left", border=TBDR)
        LBL(ws, ri, 2, b, bold=is_hdr, bg=bg)
        LBL(ws, ri, 3, c2, bg=bg)


# ugly workaround for `color` arg in LBL
def LBL(ws, r, c, txt, bold=False, bg=None, indent=0, color="000000"):
    cell = C(ws, r, c, txt, bold=bold, sz=10, color=color, bg=bg,
             align="left", border=TBDR)
    if indent:
        cell.alignment = Alignment(horizontal="left", vertical="center", indent=indent)
    return cell


# ════════════════════════════════════════════════════════════════════════════
#  BUILD GALASCOPE
# ════════════════════════════════════════════════════════════════════════════

def build_galascope():
    wb   = openpyxl.Workbook(); wb.remove(wb.active)
    ws_i = wb.create_sheet("INPUTS")
    ref  = build_inputs_galascope(ws_i)

    ws_s = wb.create_sheet("SUMMARY")
    build_summary(ws_s, ref, "PARK-REF-5001", has_asking_key="asking")

    ws_c = wb.create_sheet("CASHFLOW")
    build_cashflow(ws_c, ref, "PARK-REF-5001")

    ws_n = wb.create_sheet("SENSITIVITY")
    build_sensitivity(ws_n, ref, "PARK-REF-5001")

    ws_d = wb.create_sheet("DAM PRICES")
    build_dam(ws_d, "137 TSOC DAM files")

    return wb


# ════════════════════════════════════════════════════════════════════════════
#  BUILD CHRISTOS
# ════════════════════════════════════════════════════════════════════════════

def build_christos():
    wb   = openpyxl.Workbook(); wb.remove(wb.active)
    ws_i = wb.create_sheet("INPUTS")
    ref  = build_inputs_christos(ws_i)

    ws_s = wb.create_sheet("SUMMARY")
    build_summary(ws_s, ref, "PARK-OP-CHR-33-2026",
                  has_asking_key="asking", has_exp_key="exp_mwh")

    ws_c = wb.create_sheet("CASHFLOW")
    build_cashflow(ws_c, ref, "PARK-OP-CHR-33-2026")

    ws_n = wb.create_sheet("SENSITIVITY")
    build_sensitivity(ws_n, ref, "PARK-OP-CHR-33-2026")

    ws_d = wb.create_sheet("DAM PRICES")
    build_dam(ws_d, "134 TSOC DAM files")

    return wb


# ════════════════════════════════════════════════════════════════════════════
#  MAIN
# ════════════════════════════════════════════════════════════════════════════

GDRIVE = os.path.expanduser(
    "~/Library/CloudStorage/GoogleDrive-lighthiefcyprus@gmail.com/My Drive"
    "/SOLARFARMS/parks-for-sale"
)

gala_path  = os.path.join(GDRIVE, "galascope-5mw-trackers",
                           "galascope-5mw-investor-model-jun2026.xlsx")
chris_path = os.path.join(GDRIVE, "christos-nicosia-3.3",
                           "christos-nicosia-3.3-investor-model-jun2026.xlsx")

print("Building Galascope 5 MW model …")
build_galascope().save(gala_path)
print(f"  ✓ {gala_path}")

print("Building Christos 3.3 MW model …")
build_christos().save(chris_path)
print(f"  ✓ {chris_path}")

print("\n✅  Both models saved to Google Drive.")
