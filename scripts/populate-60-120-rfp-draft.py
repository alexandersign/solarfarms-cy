"""Populate Larnaca 50MW/120MWh RfP Excel drafts from Lighthief SSOT data."""
from __future__ import annotations

import shutil
from datetime import date
from pathlib import Path

import openpyxl

OUT_DIR = Path(__file__).resolve().parents[1] / "docs/clients/Individual_60-120-standalone"
RFP_DIR = OUT_DIR / "RfP docs"
DRAFT_DIR = OUT_DIR / "draft-populated-may2026"

CFG = {
    "project": "Larnaca Cyprus — Psevdas BESS (H.E.S.S.)",
    "poc_mw": 50,
    "export_mw": 60,
    "poc_kv": "132kV",
    "mv_kv": 33,
    "guaranteed_mwh_poc": 100,
    "nameplate_mwh": round(24 * 5.015, 2),
    "containers": 24,
    "container_kwh": 5015,
    "container_model": "Linyang Power Atlantic ME 5.015 MWh (20HC)",
    "skids": 5,
    "skid_model": "Kehua BCS10000K-C-HUD/T8",
    "skid_mw": 10,
    "pcs_model": "Kehua BCS1250K-C-HUD",
    "pcs_count": 40,
    "pcs_kw": 1250,
    "xfmr_model": "Kehua SL-10000",
    "xfmr_count": 5,
    "xfmr_kva": "10,000 kVA",
    "xfmr_ratio": "0.69/33 kV (custom)",
    "rmu": "Schneider RM AirSeT 24kV SF6-free",
    "system_type": (
        "Oversized + Augmentation — 5x T8 @ 10MW (50MW POC); "
        "24x 5MWh containers (120MWh nameplate); "
        "CONFIRM: 20 BESS @ 2:1 PCS ratio at COD + 4 aug, or 6th T8 skid"
    ),
    "cycles_day": 1.5,
    "delivery": "DDP Larnaka",
    "warranty_product_years": 3,
    "rte_system": 86.32,
    "contact": "Alexander Papacosta",
    "email": "office@lighthief.com",
    "offer_date": date.today().isoformat(),
    "validity_days": 30,
    "production_place": "Jiangsu, China (Linyang) / Xiamen, China (Kehua)",
}

BESS = {
    "technology": "LFP (LiFePO4) — EVE 314Ah prismatic cells",
    "dc_voltage_nom": 1331.2,
    "dc_voltage_range": "1164.8 – 1497.6",
    "weight_t": 43,
    "dims_mm": "6058 x 2438 x 2896",
    "ip": "IP54",
    "corrosion": "C5",
    "cooling": "Liquid cooling (45 kW unit)",
    "aux_kw": 45,
    "fire": "Aerosol + off-gas detection (H2, CO) + backup water sprinkler",
    "temp_amb": "-30 to +50 (container); -10 to +45 @ full power per TSOC",
    "humidity": "<95% RH non-condensing",
    "lifetime_years": 15,
    "lifetime_cycles": 6000,
    "noise_db": "TBC — request factory test report",
}

PCS = {
    "manufacturer": "Kehua (Xiamen Kehua Digital Energy Tech Co., Ltd)",
    "weight_kg": 950,
    "dims_mm": "735 x 2135 x 1300",
}


def set_val(ws, coord: str, val) -> None:
    ws[coord] = val


def populate_rfp(dst: Path) -> None:
    wb = openpyxl.load_workbook(dst)
    fin = wb["Financial"]
    wy = CFG["warranty_product_years"]

    set_val(fin, "E11", CFG["system_type"])
    set_val(fin, "E13", CFG["nameplate_mwh"])
    set_val(fin, "E14", "TBC — post-SAT measured (target >=100 MWh @ POC)")
    set_val(fin, "E15", CFG["guaranteed_mwh_poc"])
    set_val(fin, "E16", "TBC — degradation model + augmentation modules")
    set_val(fin, "E17", ">100 (contractual guarantee)")
    set_val(fin, "E18", CFG["containers"])
    set_val(fin, "E19", "YES")
    set_val(fin, "E20", CFG["skids"])
    for row in range(21, 37):
        set_val(fin, f"E{row}", "TBC — pending commercial offer")
    set_val(fin, "E37", "Linyang Energy (battery) + Kehua (PCS/skid) via Lighthief Cyprus Ltd")
    set_val(fin, "E38", "Lighthief Cyprus Ltd — LTSA available Y1-10")
    set_val(fin, "E39", CFG["production_place"])
    set_val(fin, "E40", "Larnaca, Cyprus")
    set_val(fin, "E41", CFG["delivery"])
    set_val(fin, "E42", "~90 days production (per OEM lead time)")
    set_val(fin, "E43", "~50 days shipping CIF + local transport")
    set_val(
        fin,
        "E44",
        f"{wy} years DC minimum per RfP (Linyang standard 5yr base available)",
    )
    set_val(fin, "E45", f"{wy} years AC (PCS + MV skid)")
    set_val(fin, "E46", "Performance warranty per Linyang terms — SOH curve attached")
    set_val(fin, "E47", "Yes — critical spares list available on request")
    set_val(fin, "E48", "TBC")
    set_val(fin, "E49", "Cyprus-based service team (Lighthief) + OEM remote support")
    set_val(
        fin,
        "E50",
        "Greece: Athalassa 40MW/80MWh, Anatoliko 40MW/160MWh, FIZ 40MW/160MWh",
    )
    set_val(
        fin,
        "E51",
        "Cyprus: Lighthief BESS portfolio — Linyang/Kehua (249 MW pipeline)",
    )
    set_val(fin, "E52", CFG["offer_date"])
    set_val(fin, "E53", f"{CFG['validity_days']} days")
    set_val(fin, "E54", CFG["contact"])
    set_val(fin, "E55", CFG["email"])

    tech = wb["Technical"]
    set_val(tech, "E11", CFG["system_type"])
    set_val(tech, "E12", CFG["nameplate_mwh"])
    set_val(tech, "E13", CFG["containers"])
    set_val(tech, "E14", BESS["technology"])
    set_val(tech, "E15", CFG["container_kwh"])
    set_val(tech, "E16", BESS["dc_voltage_nom"])
    set_val(tech, "E17", BESS["dc_voltage_range"])
    set_val(tech, "E18", CFG["cycles_day"])
    set_val(tech, "E19", "0.5C standard / 1C max (cell spec)")
    set_val(tech, "E20", BESS["noise_db"])
    set_val(tech, "E21", BESS["temp_amb"])
    set_val(tech, "E22", BESS["humidity"])
    set_val(tech, "E23", BESS["ip"])
    set_val(tech, "E24", BESS["corrosion"])
    set_val(tech, "E25", BESS["cooling"])
    set_val(tech, "E26", BESS["aux_kw"])
    set_val(tech, "E27", BESS["fire"])
    set_val(tech, "E28", BESS["lifetime_years"])
    set_val(tech, "E29", BESS["lifetime_cycles"])
    set_val(tech, "E30", BESS["weight_t"])
    set_val(tech, "E31", BESS["dims_mm"])
    set_val(tech, "E32", "YES")
    set_val(tech, "E33", CFG["skids"])
    set_val(tech, "E34", CFG["skid_model"])
    set_val(tech, "E35", 38)
    set_val(tech, "E36", "12192 x 2896 x 2438")
    set_val(tech, "E37", f"{PCS['manufacturer']} / {CFG['pcs_model']}")
    set_val(tech, "E38", CFG["pcs_count"])
    set_val(tech, "E39", CFG["pcs_kw"])
    set_val(tech, "E40", PCS["weight_kg"])
    set_val(tech, "E41", PCS["dims_mm"])
    set_val(tech, "E42", f"Kehua / {CFG['xfmr_model']}")
    set_val(tech, "E43", CFG["xfmr_count"])
    set_val(tech, "E44", CFG["xfmr_kva"])
    set_val(tech, "E45", CFG["xfmr_ratio"])
    set_val(tech, "E46", CFG["rmu"])
    set_val(tech, "E47", "TBC — qty per SLD")
    set_val(tech, "E48", CFG["mv_kv"])
    set_val(tech, "E49", "TBC — per protection study")
    set_val(tech, "E54", "On-site training included (commissioning scope)")

    oth = wb["Other"]
    set_val(oth, "E6", "EVE LF314")
    set_val(oth, "E7", "LFP (LiFePO4)")
    set_val(oth, "E8", 1004.8)
    set_val(oth, "E9", 3.2)
    set_val(oth, "E10", "2.5 – 3.65")
    set_val(oth, "E11", 5.5)
    set_val(oth, "E12", "72 x 174 x 207.7")
    set_val(oth, "E13", "1P104S Battery Pack")
    set_val(oth, "E14", "1P104S (104 cells series)")
    set_val(oth, "E15", 104.499)
    set_val(oth, "E16", 332.8)
    set_val(oth, "E17", "291.2 – 374.4")
    set_val(oth, "E18", 690)
    set_val(oth, "E19", "762.5 x 2180 x 252")
    set_val(oth, "E20", CFG["container_model"])
    set_val(oth, "E21", "12P416S (12 racks x 1P416S)")
    set_val(oth, "E22", CFG["container_kwh"])
    set_val(oth, "E23", BESS["dc_voltage_nom"])
    set_val(oth, "E24", BESS["dc_voltage_range"])
    set_val(oth, "E25", BESS["weight_t"] * 1000)
    set_val(oth, "E26", BESS["dims_mm"])
    set_val(oth, "E27", 48)
    set_val(oth, "E28", 12)
    set_val(oth, "E29", 4)

    wb.save(dst)


def fill_compliance_sheet(ws) -> int:
    filled = 0
    for r in range(4, ws.max_row + 1):
        item = ws.cell(r, 1).value
        req = ws.cell(r, 2).value
        if req is None or not str(req).strip():
            continue
        if isinstance(item, str) and item[0:2] in {
            "A.",
            "B.",
            "C.",
            "D.",
            "E.",
            "I.",
            "II",
            "III",
            "IV.",
            "V.",
        }:
            continue
        existing = ws.cell(r, 3).value
        note = ws.cell(r, 4).value or ""
        if existing not in (None, ""):
            continue

        text = str(req).lower()
        answer = None
        supplier_note = ""

        if "not in bess supplier" in text:
            answer, supplier_note = "N/A", note
        elif "ul 9540" in text:
            answer, supplier_note = (
                "Partial",
                "Cell + Container UL9540A on file; Pack report is DRAFT",
            )
        elif "62619" in text:
            answer, supplier_note = "Yes", "PACK IEC 62619.pdf"
        elif "63056" in text:
            answer, supplier_note = "Yes", "PACK IEC 63056.pdf"
        elif "62933" in text:
            answer, supplier_note = (
                "Partial",
                "IEC62933-5-2 notification letter — full cert TBC",
            )
        elif "61000-6-2" in text or "61000-6-4" in text or "emc levels" in text:
            answer, supplier_note = "Yes", "IEC/EN 61000 certs on file (PCS + pack)"
        elif "62485" in text or "60364" in text or "61936" in text:
            answer, supplier_note = (
                "Yes",
                "Design per IEC; equipment type test reports on file",
            )
        elif "fire suppression" in text or "off-gas" in text:
            answer, supplier_note = (
                "Yes",
                "Aerosol + H2/CO detection + sprinkler per 5MWh spec",
            )
        elif "water supply" in text or "water tank" in text:
            answer, supplier_note = "N/A", "Owner scope per Annex A"
        elif "hv/mv substation" in text or "hv bay" in text:
            answer, supplier_note = "N/A", "Out of BESS supplier scope"
        elif "round-trip efficiency" in text:
            answer, supplier_note = (
                "Yes",
                f"System RTE {CFG['rte_system']}% AC-AC — curve attached",
            )
        elif "availability" in text or "eaf" in text:
            answer, supplier_note = (
                "Yes",
                "LTSA 97% availability target — exceeds 92% minimum",
            )
        elif "standby" in text:
            answer, supplier_note = (
                "TBC",
                "Standby calc pending — target <=15% guaranteed capacity/day",
            )
        elif "frequency" in text or "fsm" in text or "lfsm" in text:
            answer, supplier_note = (
                "Yes",
                "Grid-following PCS + EMS FSM/LFSM-U/LFSM-O (Disperon) — config TBC",
            )
        elif "tier ii" in text or "ecodesign" in text:
            answer, supplier_note = (
                "Yes",
                "Kehua SL-series MV transformers — Tier 2 per datasheets",
            )
        elif "disconnecting switch" in text:
            answer, supplier_note = (
                "Yes",
                "DC/AC disconnectors at PCS and container per OEM design",
            )
        elif "type and routine tested" in text:
            answer, supplier_note = (
                "Yes",
                "PCS type tests on file; battery pack certs on file",
            )
        elif isinstance(item, (int, float)):
            answer, supplier_note = "TBC", "Pending detailed compliance review"

        if answer:
            ws.cell(r, 3).value = answer
            ws.cell(r, 5).value = supplier_note
            filled += 1

    return filled


def populate_compliance(dst: Path) -> int:
    wb = openpyxl.load_workbook(dst)
    total = 0
    for sn in wb.sheetnames:
        total += fill_compliance_sheet(wb[sn])
    wb.save(dst)
    return total


def main() -> None:
    DRAFT_DIR.mkdir(exist_ok=True)

    src_rfp = RFP_DIR / "260526_Batteries_RfP.xlsx"
    dst_rfp = DRAFT_DIR / "260526_Batteries_RfP-lighthief-draft-may2026.xlsx"
    shutil.copy2(src_rfp, dst_rfp)
    populate_rfp(dst_rfp)
    print(f"Saved {dst_rfp}")

    src_comp = next(RFP_DIR.glob("*Compliance*.xlsx"))
    dst_comp = DRAFT_DIR / "Annex-A-List-of-Compliance-lighthief-draft-may2026.xlsx"
    shutil.copy2(src_comp, dst_comp)
    n = populate_compliance(dst_comp)
    print(f"Saved {dst_comp} ({n} compliance rows populated)")


if __name__ == "__main__":
    main()
