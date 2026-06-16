import openpyxl, sys
sys.stdout.reconfigure(encoding='utf-8', errors='replace')
from pathlib import Path

wb = openpyxl.load_workbook(
    r"L:\My Drive\LINYANG\BESS CLIENTS\Individual_60-120-standalone\clarification.xlsx",
    data_only=True
)
print("Sheets:", wb.sheetnames)
for ws in wb.worksheets:
    print(f"\n{'='*80}")
    print(f"Sheet: {ws.title}  (max_row={ws.max_row}, max_col={ws.max_column})")
    print('='*80)
    for r in range(1, ws.max_row + 1):
        cells = []
        for c in range(1, ws.max_column + 1):
            v = ws.cell(r, c).value
            if v not in (None, ""):
                cleaned = str(v).strip().replace("\n", " ").replace("\r", "")
                cells.append(f"[col{c}] {cleaned}")
        if cells:
            print(f"  r{r}: " + " | ".join(cells))
