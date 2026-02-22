import openpyxl

out = open('/Volumes/T7 Grey/solinvest/scripts/adders_compare.txt', 'w')

# ===== V4 ADDERS =====
wb4 = openpyxl.load_workbook('docs/internal/Lighthief-EPC-Confirmed-Adders-v4-Feb2026.xlsx', data_only=True)

# Cost Stack sheet - portfolio totals
ws_cs = wb4['Cost Stack']
out.write('=== V4 COST STACK (Portfolio Totals) ===\n')
for r in range(1, ws_cs.max_row+1):
    b = ws_cs.cell(row=r, column=2).value
    c = ws_cs.cell(row=r, column=3).value
    d = ws_cs.cell(row=r, column=4).value
    if b:
        out.write(f'  {b}: {c}  {d if d else ""}\n')

# Park Cost Breakdown totals row
ws_pk = wb4['Park Cost Breakdown']
out.write('\n=== V4 PARK TOTALS (row 55) ===\n')
for c in range(1, ws_pk.max_column+1):
    h = ws_pk.cell(row=3, column=c).value
    v = ws_pk.cell(row=55, column=c).value
    if v is not None:
        out.write(f'  {h}: {v}\n')

# ===== V2 EPC SYSTEM COST =====
wb2 = openpyxl.load_workbook('docs/Bess - EPC System Cost v2.xlsx', data_only=True)
ws2 = wb2['Pricing_Model_All_Projects']

# Find totals row
out.write('\n=== V2 TOTALS ROW ===\n')
for r in range(50, ws2.max_row+1):
    a = ws2.cell(row=r, column=1).value
    b = ws2.cell(row=r, column=2).value
    g = ws2.cell(row=r, column=7).value
    if b and 'TOTAL' in str(b).upper():
        out.write(f'Row {r}: {b}\n')
        for c in range(1, ws2.max_column+1):
            h = ws2.cell(row=1, column=c).value
            v = ws2.cell(row=r, column=c).value
            if v is not None:
                out.write(f'  {h}: {v}\n')

# Per-park comparison: extract adder totals from both sheets
out.write('\n\n=== PER-PARK ADDER COMPARISON ===\n')
out.write(f'{"Park":<30} {"V4 CIF":>14} {"V4 PhysAdder":>14} {"V4 EMS/SCADA":>14} {"V4 Total":>14} | {"V2 CIF":>14} {"V2 PhysAdder":>14} {"V2 EMS/SCADA":>14} {"V2 Total":>14} | {"CIF Diff":>10} {"Adder Diff":>10} {"EMS Diff":>10}\n')
out.write('-' * 210 + '\n')

# Build V2 lookup by park name
v2_parks = {}
for r in range(4, ws2.max_row+1):
    park = ws2.cell(row=r, column=7).value  # G = Park Name
    if not park or 'TOTAL' in str(park).upper():
        continue
    cif = ws2.cell(row=r, column=16).value or 0   # P = CIF
    adders = ws2.cell(row=r, column=41).value or 0  # AO = Adders Total
    ems = ws2.cell(row=r, column=52).value or 0    # AZ = Total EMS/SCADA
    installed = ws2.cell(row=r, column=42).value or 0  # AP = Installed Cost
    v2_parks[park] = {'cif': cif, 'adders': adders, 'ems': ems, 'installed': installed}

# V4 parks
for r in range(4, 55):
    park = ws_pk.cell(row=r, column=2).value  # B = Park Name
    if not park or 'TOTAL' in str(park).upper():
        continue
    v4_cif = ws_pk.cell(row=r, column=12).value or 0   # L = CIF
    v4_adders = ws_pk.cell(row=r, column=32).value or 0  # AF = Physical Adders
    v4_ems = ws_pk.cell(row=r, column=36).value or 0    # AJ = Total EMS/SCADA
    v4_total = ws_pk.cell(row=r, column=37).value or 0   # AK = Installed Cost

    v2 = v2_parks.get(park, {})
    v2_cif = v2.get('cif', 0)
    v2_adders = v2.get('adders', 0)
    v2_ems = v2.get('ems', 0)
    v2_total = v2.get('installed', 0)

    cif_diff = v4_cif - v2_cif if v2_cif else 'N/A'
    adder_diff = v4_adders - v2_adders if v2_adders else 'N/A'
    ems_diff = v4_ems - v2_ems if v2_ems else 'N/A'

    out.write(f'{park:<30} {v4_cif:>14,.2f} {v4_adders:>14,.2f} {v4_ems:>14,.2f} {v4_total:>14,.2f} | {v2_cif:>14,.2f} {v2_adders:>14,.2f} {v2_ems:>14,.2f} {v2_total:>14,.2f} | {str(cif_diff):>10} {str(adder_diff):>10} {str(ems_diff):>10}\n')

# Category-level comparison
out.write('\n\n=== CATEGORY-LEVEL ADDER COMPARISON (Portfolio Totals) ===\n')

# V4 Cost Stack categories (already from Cost Stack sheet)
v4_cats = {}
for r in range(1, ws_cs.max_row+1):
    b = ws_cs.cell(row=r, column=2).value
    c = ws_cs.cell(row=r, column=3).value
    if b and c:
        v4_cats[b] = c

# V2 totals from the totals row - we need to find it
v2_totals = {}
for r in range(50, ws2.max_row+1):
    b = ws2.cell(row=r, column=2).value
    if b and 'TOTAL' in str(b).upper():
        for c in range(1, ws2.max_column+1):
            h = ws2.cell(row=1, column=c).value
            v = ws2.cell(row=r, column=c).value
            if h and v is not None:
                v2_totals[h] = v
        break

out.write('\nV4 Cost Stack:\n')
for k, v in v4_cats.items():
    if isinstance(v, (int, float)):
        out.write(f'  {k}: {v:,.2f}\n')
    else:
        out.write(f'  {k}: {v}\n')

out.write('\nV2 Totals:\n')
for k, v in v2_totals.items():
    if isinstance(v, (int, float)):
        out.write(f'  {k}: {v:,.2f}\n')
    else:
        out.write(f'  {k}: {v}\n')

# Map V4 categories to V2 columns where possible
out.write('\n=== MAPPED COMPARISON ===\n')
mapping = [
    ('Import Duty (2.66%)', 'ImportDuty'),
    ('Port Landing (ECTL)', 'PortCustoms'),
    ('Crane + Inland Transport', 'A Soulis / Crane + Inland Interfreight'),
    ('LV Cabling', 'LV_Cabling'),
    ('MV Cabling', 'MV_Cabling'),
    ('MV Terminations', 'MV_Terminations'),
    ('Protection Engineering', 'Protection_Eng'),
    ('Remote Trip / SCADA Comms', 'RemoteTrip'),
    ('UPS & Auxiliary', 'UPS_Aux'),
    ('LPS (Lightning)', 'LPS External lightning Protection'),
    ('SPD (Surge Protection)', 'SPDs Surge Protection (DEHN)'),
    ('Site Earthing', 'Earthing Grid (DEHN)'),
    ('Installation Labor', 'Protection Testing & DSO + StrikeRA Install'),
    ('Civil Works', 'Civil Works (Kamil €2000/MWh)'),
    ('Insurance (0.75%)', 'Insurance'),
    ('Docs / Compliance', 'Docs_Compliance'),
    ('Voltus EMS', 'VOLTUS EMS (Feb 16 Update)'),
]

out.write(f'{"V4 Category":<35} {"V4 Total":>14} | {"V2 Column":<50} {"V2 Total":>14} | {"Diff":>14}\n')
out.write('-' * 150 + '\n')
for v4_name, v2_name in mapping:
    v4_val = v4_cats.get(v4_name, 0)
    v2_val = v2_totals.get(v2_name, 0)
    if not isinstance(v4_val, (int, float)):
        v4_val = 0
    if not isinstance(v2_val, (int, float)):
        v2_val = 0
    diff = v4_val - v2_val
    out.write(f'{v4_name:<35} {v4_val:>14,.2f} | {v2_name:<50} {v2_val:>14,.2f} | {diff:>14,.2f}\n')

out.close()
print('Done')
