"""Generate the HESS T2 earthing & auxiliary transformer datasheet (one per supplier).

The earthing & auxiliary transformer is a SEPARATE unit from the 63 MVA main transformer
(TSOC connection terms list it as a distinct 'MT/0.4 kV earthing transformer'). A power-
transformer questionnaire does not fit it, so each supplier gets this purpose-built sheet:
zig-zag earthing winding (33 kV neutral) + auxiliary 400/230 V station-service winding.

Values confirmed by the client clarification (Jun 2026): solid earthing (no NER),
≥20 kA / 3 s (or 25 kA / 1 s), aux ≥315 kVA (final by designer load calc).
Identical for all suppliers (it is OUR specification) -> written into each Supplier-*/T2-earthing-aux/.
"""
from __future__ import annotations

import sys
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

try:
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
except Exception:  # noqa: BLE001
    pass

BASE = Path(
    r"L:\My Drive\LINYANG\BESS CLIENTS\Individual_60-120-standalone\HV Transformer"
)
SUPPLIERS = ["Supplier-1", "Supplier-2", "Supplier-3"]
FILENAME = "T2-earthing-aux-transformer-datasheet-HESS.xlsx"

# Lighthief brand
NAVY = "1A365D"
GOLD = "C9A432"
WHITE = "FFFFFF"
GREY_BG = "F0F4F8"

THIN = Side(style="thin", color="BFC9D4")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

# (item, parameter, unit, requirement, offer-by-producer?)
ROWS: list[tuple[str, str, str, str]] = [
    ("1", "GENERAL", "", ""),
    ("1.1", "Equipment", "-", "Combined earthing & auxiliary transformer (separate from 63 MVA main unit)"),
    ("1.2", "Application", "-", "Derive & solidly earth the 33 kV system neutral (delta LV of main TX) + station-service supply"),
    ("1.3", "Quantity", "off", "1"),
    ("1.4", "Installation", "-", "Outdoor, oil-immersed (dry-type acceptable if compliant — state)"),
    ("1.5", "Standards", "-", "EN/IEC 60076-1/-2/-3/-5, IEC 60076-6, EN 60289 (earthing transformers)"),
    ("1.6", "Country of origin / manufacturer", "-", "State in offer"),

    ("2", "SITE CONDITIONS (same as main transformer)", "", ""),
    ("2.1", "Altitude", "m", "264 (≤1000)"),
    ("2.2", "Ambient air temperature", "°C", "Max 50; design 45; daily avg 30; annual avg 20"),
    ("2.3", "Pollution", "-", "Heavy — IEC 60815 Class III; creepage ≥35 mm/kV (composite)"),
    ("2.4", "Seismic", "-", "CYS EN 1998-1 Zone II, agR 0.23 g; importance factor 1.4"),

    ("3", "EARTHING WINDING (zig-zag / interconnected-star)", "", ""),
    ("3.1", "Connection", "-", "Zig-zag (interconnected-star), ZN — neutral solidly earthed"),
    ("3.2", "System earthing method", "-", "SOLID earthing (no NER) — client confirmed Jun 2026"),
    ("3.3", "Nominal system voltage (MV)", "kV", "33 (highest equipment 36)"),
    ("3.4", "Rated short-time earth-fault current / duration", "kA / s", "≥20 kA for 3 s (or ≥25 kA for 1 s) per EN 60076-5"),
    ("3.5", "Continuous neutral current rating", "A", "State in offer (≥10% of short-time, or per design)"),
    ("3.6", "Zero-sequence impedance Z0", "Ω/phase", "State in offer (to limit earth-fault to system value)"),
    ("3.7", "Insulation level (36 kV class)", "kV", "170 BIL / 70 AC (1 min) — client clarification Jun 2026"),
    ("3.8", "Neutral / HVN insulation", "kV", "Per EN 60076-3 for solidly earthed neutral"),

    ("4", "AUXILIARY WINDING (station service)", "", ""),
    ("4.1", "Rated power", "kVA", "≥315 (minimum) — final per designer load calc; state upratable range"),
    ("4.2", "Voltage", "V", "400 / 230, 3-phase 4-wire (TN)"),
    ("4.3", "Vector group (overall unit)", "-", "ZNyn (zig-zag earthing + star auxiliary) or equivalent — state"),
    ("4.4", "Frequency", "Hz", "50"),
    ("4.5", "Auxiliary winding insulation level", "kV", "1.1 kV class (LV) per IEC"),
    ("4.6", "Price delta per kVA step", "EUR", "Quote 315 → 400 → 500 kVA options"),

    ("5", "CONSTRUCTION & ACCESSORIES", "", ""),
    ("5.1", "Cooling", "-", "ONAN"),
    ("5.2", "Insulating fluid", "-", "Mineral oil EN 60296 (or state alternative)"),
    ("5.3", "Tap changer", "-", "Off-circuit ±2×2.5% on auxiliary winding (state)"),
    ("5.4", "Temperature rise", "°C", "≤55 oil / ≤60 winding"),
    ("5.5", "Protection accessories", "-", "Buchholz, OTI/WTI, PRD, oil level, silica-gel breather"),
    ("5.6", "Neutral CT", "-", "Earth-fault / REF CT on neutral (ratio TBC with ISM protection)"),
    ("5.7", "Anti-vibration / base", "-", "Flat base, anti-vibration pads; EN 795 anchorage"),

    ("6", "TESTS & CERTIFICATION", "", ""),
    ("6.1", "Routine tests", "-", "Per EN 60076-1 (ratio, losses, insulation, etc.)"),
    ("6.2", "Type tests", "-", "Temperature rise (EN 60076-2); short-circuit withstand (EN 60076-5)"),
    ("6.3", "Short-circuit type-test certificate", "-", "Independent lab (KEMA/CESI/IPH) — required"),
    ("6.4", "Ecodesign / CE", "-", "EU DoC + CE marking; confirm applicability to earthing transformer"),
    ("6.5", "Material certs", "-", "Oil EN 60296; bushings EN 60137"),

    ("7", "COMMERCIAL", "", ""),
    ("7.1", "No-load loss", "kW", "STATE IN OFFER"),
    ("7.2", "Load loss", "kW", "STATE IN OFFER"),
    ("7.3", "Mass (total / oil)", "kg", "STATE IN OFFER"),
    ("7.4", "Overall dimensions (L×W×H)", "mm", "STATE IN OFFER"),
    ("7.5", "Manufacturing lead time (ex-works)", "weeks", "STATE IN OFFER"),
    ("7.6", "Delivery time to Limassol Port (DAP)", "weeks", "STATE IN OFFER (incl. transit + customs)"),
    ("7.7", "Incoterms 2020", "-", "Baseline DAP Limassol Port; option DAP/DDP Plot 26 Psevdas"),
    ("7.8", "Unit price (itemised separately from main TX)", "EUR", "STATE IN OFFER"),
    ("7.9", "Warranty", "months", "STATE IN OFFER (≥24)"),
]


def build_sheet(ws) -> None:
    ws.title = "T2 Earthing-Aux TX"
    widths = [8, 42, 12, 60, 28]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Title bar
    ws.merge_cells("A1:E1")
    c = ws["A1"]
    c.value = "HESS PSEVDAS BESS — EARTHING & AUXILIARY TRANSFORMER (T2) DATASHEET"
    c.font = Font(name="Calibri", size=13, bold=True, color=WHITE)
    c.fill = PatternFill("solid", fgColor=NAVY)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 26

    ws.merge_cells("A2:E2")
    s = ws["A2"]
    s.value = (
        "Separate unit from the 63 MVA 132/33 kV main transformer (T1). "
        "Client: H.E.S.S. Hybrid Energy Storage Systems Ltd. EPC: Lighthief Cyprus Ltd. "
        "Producer to complete the OFFER column."
    )
    s.font = Font(name="Calibri", size=9, italic=True, color="404040")
    s.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.row_dimensions[2].height = 30

    # Header row
    headers = ["Item", "Parameter", "Unit", "Requirement (Lighthief)", "OFFER (producer)"]
    hr = 4
    for col, h in enumerate(headers, 1):
        cell = ws.cell(hr, col, h)
        cell.font = Font(bold=True, color=WHITE, size=10)
        cell.fill = PatternFill("solid", fgColor=NAVY)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER

    r = hr + 1
    for item, param, unit, req in ROWS:
        is_section = unit == "" and req == ""
        cells = [item, param, unit, req, ""]
        for col, val in enumerate(cells, 1):
            cell = ws.cell(r, col, val)
            cell.border = BORDER
            cell.alignment = Alignment(
                horizontal="left", vertical="center", wrap_text=(col in (2, 4, 5))
            )
            if is_section:
                cell.fill = PatternFill("solid", fgColor=GOLD)
                cell.font = Font(bold=True, color=NAVY, size=10)
            else:
                cell.font = Font(size=9, color="000000")
                if col == 5:
                    cell.fill = PatternFill("solid", fgColor=GREY_BG)
                if col == 4 and "STATE IN OFFER" in str(val):
                    cell.font = Font(size=9, bold=True, color="9C7D22")
        if is_section:
            ws.row_dimensions[r].height = 18
        r += 1

    ws.freeze_panes = "A5"


def main() -> None:
    wb = openpyxl.Workbook()
    build_sheet(wb.active)
    for s in SUPPLIERS:
        out = BASE / s / "T2-earthing-aux" / FILENAME
        out.parent.mkdir(parents=True, exist_ok=True)
        wb.save(out)
        print(f"Wrote {out.relative_to(BASE)}")


if __name__ == "__main__":
    main()
