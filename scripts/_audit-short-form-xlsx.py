"""Audit short-form POWER Transformer sheet: blank vs filled."""
import sys
import io
from pathlib import Path
import openpyxl

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")


base = Path(r"L:\My Drive\LINYANG\BESS CLIENTS\Individual_60-120-standalone\HV Transformer")
filled = base / "POWER-Transformer-HESS-prefilled-jun2026.xlsx"

wb = openpyxl.load_workbook(filled, data_only=True)
ws = wb["POWER Transformer"]

blank, filled_rows = [], []
for r in range(1, ws.max_row + 1):
    item = ws.cell(r, 1).value
    char = ws.cell(r, 2).value
    val = ws.cell(r, 5).value
    if not char and not item:
        continue
    s_item = str(item).strip() if item else ""
    s_char = str(char or "").strip()
    if not s_char and not s_item:
        continue
    if val is None or str(val).strip() == "":
        blank.append((s_item, s_char))
    else:
        filled_rows.append((s_item, s_char, str(val)[:90]))

print(f"Filled: {len(filled_rows)}, Blank: {len(blank)}\n")
print("=== STILL BLANK (producer / client must supply) ===")
for item, char in blank:
    print(f"  [{item}] {char}")

print("\n=== FILLED ===")
for item, char, val in filled_rows:
    print(f"  [{item}] {char[:55]} => {val}")
