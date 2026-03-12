#!/usr/bin/env python3
"""
Generate CIF Cost Target Excel for prospective BESS providers.
Source data: lib/portfolio-data.ts + Linyang quotation analysis (Jan/Feb 2026)
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter
from datetime import date

wb = openpyxl.Workbook()

# ── Colour palette ──
DARK_BLUE = "1B2A4A"
MID_BLUE = "2E5090"
LIGHT_BLUE = "D6E4F0"
ACCENT_GREEN = "27AE60"
ACCENT_ORANGE = "E67E22"
WHITE = "FFFFFF"
LIGHT_GREY = "F2F2F2"
BORDER_GREY = "B0B0B0"

# ── Reusable styles ──
header_font = Font(name="Calibri", bold=True, color=WHITE, size=11)
header_fill = PatternFill(start_color=MID_BLUE, end_color=MID_BLUE, fill_type="solid")
subheader_font = Font(name="Calibri", bold=True, size=11)
subheader_fill = PatternFill(start_color=LIGHT_BLUE, end_color=LIGHT_BLUE, fill_type="solid")
title_font = Font(name="Calibri", bold=True, color=WHITE, size=14)
title_fill = PatternFill(start_color=DARK_BLUE, end_color=DARK_BLUE, fill_type="solid")
section_font = Font(name="Calibri", bold=True, color=DARK_BLUE, size=12)
normal_font = Font(name="Calibri", size=11)
small_font = Font(name="Calibri", size=10, italic=True, color="666666")
currency_fmt = '#,##0'
currency_kwh_fmt = '€#,##0.00'
pct_fmt = '0.0%'
thin_border = Border(
    left=Side(style="thin", color=BORDER_GREY),
    right=Side(style="thin", color=BORDER_GREY),
    top=Side(style="thin", color=BORDER_GREY),
    bottom=Side(style="thin", color=BORDER_GREY),
)
alt_fill = PatternFill(start_color=LIGHT_GREY, end_color=LIGHT_GREY, fill_type="solid")


def style_header_row(ws, row, cols, font=header_font, fill=header_fill):
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = font
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border


def style_data_row(ws, row, cols, alt=False):
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = normal_font
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center", vertical="center")
        if alt:
            cell.fill = alt_fill


def title_bar(ws, row, text, cols):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=cols)
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = title_font
    cell.fill = title_fill
    cell.alignment = Alignment(horizontal="left", vertical="center")
    cell.border = thin_border
    ws.row_dimensions[row].height = 30


def section_label(ws, row, text, cols):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=cols)
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = section_font
    cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[row].height = 24


def note_row(ws, row, text, cols):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=cols)
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = small_font
    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)


# ═══════════════════════════════════════════
# SHEET 1: CIF PRICE TARGETS
# ═══════════════════════════════════════════
ws1 = wb.active
ws1.title = "CIF Price Targets"
ws1.sheet_properties.tabColor = MID_BLUE
COLS1 = 9

title_bar(ws1, 1, "  BESS CIF PRICE TARGETS — Lighthief Cyprus Ltd", COLS1)
note_row(ws1, 2, f"  Confidential  |  Prepared: {date.today().strftime('%d %B %Y')}  |  All prices CIF Limassol (EUR)  |  Incoterms: CIF", COLS1)
ws1.row_dimensions[2].height = 20

# ── 2-HOUR SYSTEMS ──
r = 4
section_label(ws1, r, "2-Hour Duration Systems (C-rate = 0.5C)", COLS1)
r += 1
headers = ["Power (MW)", "Energy (MWh)", "Duration (h)", "Containers (est.)",
           "Reference CIF €/kWh", "Target CIF €/kWh", "Reference CIF Total", "Target CIF Total", "Notes"]
for i, h in enumerate(headers, 1):
    ws1.cell(row=r, column=i, value=h)
style_header_row(ws1, r, COLS1)

data_2h = [
    (1,    2,    2, 1,  125.00, 115.00, "Small system premium"),
    (2.5,  5,    2, 1,  118.00, 108.00, ""),
    (5,   10,    2, 2,  110.11, 100.00, "Based on 5MW/10MWh actual"),
    (10,  20,    2, 4,  100.00,  92.00, ""),
    (25,  50,    2, 10,  92.00,  84.00, "Volume pricing target"),
    (50, 100,    2, 20,  88.00,  80.00, "Large-scale benchmark"),
]

for idx, (mw, mwh, dur, cnt, ref, tgt, note) in enumerate(data_2h):
    r += 1
    ws1.cell(row=r, column=1, value=mw)
    ws1.cell(row=r, column=2, value=mwh)
    ws1.cell(row=r, column=3, value=dur)
    ws1.cell(row=r, column=4, value=cnt)
    ws1.cell(row=r, column=5, value=ref).number_format = currency_kwh_fmt
    ws1.cell(row=r, column=6, value=tgt).number_format = currency_kwh_fmt
    ws1.cell(row=r, column=7, value=ref * mwh * 1000).number_format = currency_fmt
    ws1.cell(row=r, column=8, value=tgt * mwh * 1000).number_format = currency_fmt
    ws1.cell(row=r, column=9, value=note)
    style_data_row(ws1, r, COLS1, alt=(idx % 2 == 1))

# ── 4-HOUR SYSTEMS ──
r += 2
section_label(ws1, r, "4-Hour Duration Systems (C-rate = 0.25C)", COLS1)
r += 1
for i, h in enumerate(headers, 1):
    ws1.cell(row=r, column=i, value=h)
style_header_row(ws1, r, COLS1)

data_4h = [
    (1,    4,    4, 1,  121.87, 112.00, "Based on 1MW/4MWh actual (€509,420)"),
    (2.5, 10,    4, 2,   97.14,  90.00, "Based on 2.5MW/10MWh actual (€974,457)"),
    (5,   20,    4, 4,   92.14,  85.00, "Based on 5MW/20MWh actual (€1,848,712)"),
    (10,  40,    4, 8,   94.25,  86.00, "Based on 12MW/40MWh actual (€3,781,992)"),
    (25, 100,    4, 20,  87.54,  80.00, "Based on 25MW/100MWh actual (€8,782,302)"),
    (50, 200,    4, 40,  84.00,  76.00, "Extrapolated volume target"),
]

for idx, (mw, mwh, dur, cnt, ref, tgt, note) in enumerate(data_4h):
    r += 1
    ws1.cell(row=r, column=1, value=mw)
    ws1.cell(row=r, column=2, value=mwh)
    ws1.cell(row=r, column=3, value=dur)
    ws1.cell(row=r, column=4, value=cnt)
    ws1.cell(row=r, column=5, value=ref).number_format = currency_kwh_fmt
    ws1.cell(row=r, column=6, value=tgt).number_format = currency_kwh_fmt
    ws1.cell(row=r, column=7, value=ref * mwh * 1000).number_format = currency_fmt
    ws1.cell(row=r, column=8, value=tgt * mwh * 1000).number_format = currency_fmt
    ws1.cell(row=r, column=9, value=note)
    style_data_row(ws1, r, COLS1, alt=(idx % 2 == 1))

# ── 8-HOUR SYSTEMS ──
r += 2
section_label(ws1, r, "8-Hour Duration Systems (C-rate = 0.125C)", COLS1)
r += 1
for i, h in enumerate(headers, 1):
    ws1.cell(row=r, column=i, value=h)
style_header_row(ws1, r, COLS1)

data_8h = [
    (1,    8,    8, 2,  108.00,  98.00, "Estimated from portfolio curve"),
    (2.5, 20,    8, 4,   94.00,  86.00, ""),
    (5,   40,    8, 8,   88.00,  80.00, ""),
    (8,   60,    8, 12,  82.03,  76.00, "Based on 8MW/60MWh actual (€4,937,363)"),
    (10,  80,    8, 16,  80.00,  74.00, "Volume target"),
    (25, 200,    8, 40,  78.00,  72.00, "Large-scale 8h target"),
    (50, 400,    8, 80,  76.00,  70.00, "Benchmark competitive pricing"),
]

for idx, (mw, mwh, dur, cnt, ref, tgt, note) in enumerate(data_8h):
    r += 1
    ws1.cell(row=r, column=1, value=mw)
    ws1.cell(row=r, column=2, value=mwh)
    ws1.cell(row=r, column=3, value=dur)
    ws1.cell(row=r, column=4, value=cnt)
    ws1.cell(row=r, column=5, value=ref).number_format = currency_kwh_fmt
    ws1.cell(row=r, column=6, value=tgt).number_format = currency_kwh_fmt
    ws1.cell(row=r, column=7, value=ref * mwh * 1000).number_format = currency_fmt
    ws1.cell(row=r, column=8, value=tgt * mwh * 1000).number_format = currency_fmt
    ws1.cell(row=r, column=9, value=note)
    style_data_row(ws1, r, COLS1, alt=(idx % 2 == 1))

# Notes
r += 2
note_row(ws1, r, "Reference CIF = Current portfolio pricing (887 MWh across 51 parks, Q1 2026).  Target CIF = Competitive pricing expectation for new supplier.", COLS1)
r += 1
note_row(ws1, r, "All prices EUR, CIF Limassol.  CIF scope: BESS containers (EVE LFP cells) + Grid-Forming PCS (Kehua C-HUD series or equivalent) + MV Transformer/Skid + MV Switchgear + BMS + Fire Suppression + Liquid Cooling + Commissioning.", COLS1)
r += 1
note_row(ws1, r, "Container count is approximate based on ~5 MWh per 20ft HC container. Actual count depends on supplier container capacity.", COLS1)

# Column widths
widths1 = [14, 14, 14, 18, 20, 20, 22, 22, 40]
for i, w in enumerate(widths1, 1):
    ws1.column_dimensions[get_column_letter(i)].width = w


# ═══════════════════════════════════════════
# SHEET 2: ACTUAL CIF DATA (from Linyang quotation)
# ═══════════════════════════════════════════
ws2 = wb.create_sheet("Actual CIF Data")
ws2.sheet_properties.tabColor = ACCENT_GREEN
COLS2 = 8

title_bar(ws2, 1, "  ACTUAL CIF PRICING — Current Portfolio (Q1 2026)", COLS2)
note_row(ws2, 2, f"  Portfolio: 51 parks, 249 MW, 887 MWh  |  Total CIF: €85,963,925  |  EVE LFP cells, Kehua C-HUD Grid-Forming PCS  |  CIF Limassol", COLS2)

r = 4
section_label(ws2, r, "CIF by System Configuration (sorted by €/kWh ascending)", COLS2)
r += 1
headers2 = ["Configuration", "Power (MW)", "Energy (MWh)", "Duration (h)",
            "CIF Total (€)", "CIF €/kWh", "# in Portfolio", "Containers"]
for i, h in enumerate(headers2, 1):
    ws2.cell(row=r, column=i, value=h)
style_header_row(ws2, r, COLS2)

configs = [
    ("8MW / 60MWh",    8.0, 60, 7.50, 4937363, 82.03, 1, 14),
    ("25MW / 100MWh", 25.0, 100, 4.00, 8782302, 87.54, 1, 22),
    ("7MW / 30MWh",    7.0, 30, 4.29, 2731668, 90.77, 1, 10),
    ("5MW / 20MWh",    5.0, 20, 4.00, 1848712, 92.14, 2, 5),
    ("8MW / 35MWh",    8.0, 35, 4.38, 3303521, 94.09, 1, 9),
    ("12MW / 40MWh",  12.0, 40, 3.33, 3781992, 94.25, 5, 12),
    ("6MW / 20MWh",    6.0, 20, 3.33, 1907319, 95.06, 1, 6),
    ("6.5MW / 20MWh",  6.5, 20, 3.08, 1951711, 97.27, 1, 5),
    ("2.5MW / 10MWh",  2.5, 10, 4.00,  974457, 97.14, 2, 3),
    ("7.5MW / 25MWh",  7.5, 25, 3.33, 2476650, 98.75, 1, 7),
    ("3MW / 10MWh",    3.0, 10, 3.33,  993203, 99.00, 3, 3),
    ("5MW / 15MWh",    5.0, 15, 3.00, 1592099, 99.00, 4, 5),  
    ("7MW / 20MWh",    7.0, 20, 2.86, 2001737, 99.77, 1, 6),
    ("7.7MW / 25MWh",  7.7, 25, 3.25, 2523653, 100.62, 2, 7),
    ("3.29MW / 10MWh", 3.29, 10, 3.04, 1007809, 100.46, 1, 3),
    ("4MW / 15MWh",    4.0, 15, 3.75, 1562533, 103.84, 2, 4),
    ("4.5MW / 15MWh",  4.5, 15, 3.33, 1562533, 103.84, 1, 4),
    ("2.5MW / 7.5MWh", 2.5, 7.5, 3.00,  871308, 104.22, 3, 3),
    ("5MW / 10MWh",    5.0, 10, 2.00, 1104571, 110.11, 3, 3),
    ("3.5MW / 10MWh",  3.5, 10, 2.86, 1064650, 106.13, 1, 3),
    ("3MW / 12MWh",    3.0, 12, 4.00, 1347385, 107.45, 1, 4),
    ("1MW / 5MWh",     1.0,  5, 5.00,  560994, 111.84, 1, 2),
    ("1.5MW / 5MWh",   1.5,  5, 3.33,  585977, 116.82, 4, 2),
    ("2MW / 5MWh",     2.0,  5, 2.50,  602846, 120.19, 1, 2),
    ("1MW / 4MWh",     1.0,  4, 4.00,  509420, 121.87, 2, 2),
    ("1.5MW / 5MWh (Paphos)", 1.5, 5, 3.33, 541047, 129.44, 1, 2),
]

for idx, (cfg, mw, mwh, dur, cif, cifkwh, count, cnt) in enumerate(configs):
    r += 1
    ws2.cell(row=r, column=1, value=cfg)
    ws2.cell(row=r, column=2, value=mw)
    ws2.cell(row=r, column=3, value=mwh)
    ws2.cell(row=r, column=4, value=round(dur, 2))
    ws2.cell(row=r, column=5, value=cif).number_format = currency_fmt
    ws2.cell(row=r, column=6, value=cifkwh).number_format = currency_kwh_fmt
    ws2.cell(row=r, column=7, value=count)
    ws2.cell(row=r, column=8, value=cnt)
    style_data_row(ws2, r, COLS2, alt=(idx % 2 == 1))

r += 2
section_label(ws2, r, "Portfolio CIF Summary", COLS2)
r += 1
summary_data = [
    ("Total Portfolio CIF", "€85,963,925"),
    ("Total Capacity", "887 MWh (249 MW)"),
    ("Weighted Average CIF", "€96.94/kWh (€96,940/MWh)"),
    ("CIF Range", "€82.03 - €129.44/kWh"),
    ("Total Parks", "51 (46 active + 5 future)"),
    ("Cell Supplier", "EVE Energy (LFP, 314Ah, Tier-1)"),
    ("PCS", "Kehua C-HUD Series (Grid-Forming Capable)"),
    ("Incoterms", "CIF Limassol"),
]
for idx, (label, val) in enumerate(summary_data):
    r += 1
    ws2.cell(row=r, column=1, value=label).font = Font(name="Calibri", bold=True, size=11)
    ws2.merge_cells(start_row=r, start_column=2, end_row=r, end_column=4)
    ws2.cell(row=r, column=2, value=val).font = normal_font
    for c in range(1, 5):
        ws2.cell(row=r, column=c).border = thin_border
        if idx % 2 == 1:
            ws2.cell(row=r, column=c).fill = alt_fill

widths2 = [26, 14, 14, 14, 20, 16, 16, 16]
for i, w in enumerate(widths2, 1):
    ws2.column_dimensions[get_column_letter(i)].width = w


# ═══════════════════════════════════════════
# SHEET 3: CIF SCOPE DEFINITION
# ═══════════════════════════════════════════
ws3 = wb.create_sheet("CIF Scope & Requirements")
ws3.sheet_properties.tabColor = ACCENT_ORANGE
COLS3 = 4

title_bar(ws3, 1, "  CIF SCOPE DEFINITION — What Must Be Included in CIF Price", COLS3)
note_row(ws3, 2, "  This defines the minimum scope that all CIF quotations must cover to be comparable.", COLS3)

r = 4
section_label(ws3, r, "INCLUDED in CIF Price (Mandatory)", COLS3)
r += 1
headers3 = ["Component", "Description", "Specification", "Status"]
for i, h in enumerate(headers3, 1):
    ws3.cell(row=r, column=i, value=h)
style_header_row(ws3, r, COLS3)

included = [
    ("BESS Containers", "LFP battery containers, fully integrated", "~5 MWh per 20ft HC container", "Required"),
    ("LFP Battery Cells", "Tier-1 LFP cells (EVE 314Ah or equivalent)", "EVE Energy, CATL, BYD, or CALB — must be BNEF Tier-1", "Required"),
    ("Battery Management System", "Container-level BMS", "Cell/module/rack/container levels, IEC 62619", "Required"),
    ("Grid-Forming PCS", "Bi-directional inverter, grid-forming capable", "Kehua BCS1000K-C-HUD / BCS1250K-C-HUD or equivalent, ≥98% eff.", "Required"),
    ("MV Transformer", "Step-up transformer (MV Skid)", "Oil-immersed or dry-type, matched to grid voltage", "Required"),
    ("MV Switchgear", "Ring Main Unit / switchgear", "22kV rated, Schneider/ABB or equivalent", "Required"),
    ("Fire Suppression", "In-container fire suppression", "Aerosol + gas detection + early warning", "Required"),
    ("Liquid Cooling System", "Active thermal management", "45kW cooling capacity per container", "Required"),
    ("C5 Corrosion Coating", "Marine/coastal corrosion protection", "ISO 12944-5 C5-M standard", "Required"),
    ("Communication Protocols", "SCADA/EMS integration ready", "Modbus TCP, IEC 60870-5-104, IEC 61850", "Required"),
    ("Training & Commissioning", "On-site commissioning support", "Factory training + site commissioning crew", "Required"),
    ("Shipping", "CIF to destination port", "CIF Limassol, marine cargo insurance included", "Required"),
    ("5-Year Base Warranty", "Standard defect/performance warranty", "5 years excl. consumables", "Required"),
    ("Grid Code Certification", "EU grid code compliance", "EN 50549-2 (or TÜV equivalent)", "Required"),
]

for idx, (comp, desc, spec, status) in enumerate(included):
    r += 1
    ws3.cell(row=r, column=1, value=comp)
    ws3.cell(row=r, column=2, value=desc)
    ws3.cell(row=r, column=3, value=spec)
    ws3.cell(row=r, column=4, value=status)
    style_data_row(ws3, r, COLS3, alt=(idx % 2 == 1))
    ws3.cell(row=r, column=1).font = Font(name="Calibri", bold=True, size=11)
    ws3.cell(row=r, column=1).alignment = Alignment(horizontal="left", vertical="center")
    ws3.cell(row=r, column=2).alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws3.cell(row=r, column=3).alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

r += 2
section_label(ws3, r, "EXCLUDED from CIF (EPC Scope — Provided Separately)", COLS3)
r += 1
for i, h in enumerate(["Item", "Description", "Typical Cost", "Responsibility"], 1):
    ws3.cell(row=r, column=i, value=h)
style_header_row(ws3, r, COLS3)

excluded = [
    ("DC Cabling", "Inter-container DC connections", "Varies by layout", "EPC / Lighthief"),
    ("AC Cabling (LV)", "LV cabling from PCS to transformer", "~€1,400/MW", "EPC / Lighthief"),
    ("MV Cabling", "MV cables from transformer to grid PCC", "~€3,500/feeder", "EPC / Lighthief"),
    ("Civil Works", "Concrete pads, foundations, fencing", "~€2,000/MWh", "EPC / Lighthief"),
    ("Earthworks & Drainage", "Site preparation and drainage", "Included in civil works", "EPC / Lighthief"),
    ("EMS / SCADA", "Energy Management System", "Quoted separately", "EPC / Lighthief"),
    ("Lightning Protection", "External LPS and SPDs", "Quoted separately", "Client / EPC"),
    ("Import Duties", "Cyprus customs (weighted ~2.66%)", "2.66% of CIF", "EPC / Lighthief"),
    ("Inland Transport", "Port to site delivery", "~€2,360/container", "EPC / Lighthief"),
]

for idx, (item, desc, cost, resp) in enumerate(excluded):
    r += 1
    ws3.cell(row=r, column=1, value=item)
    ws3.cell(row=r, column=2, value=desc)
    ws3.cell(row=r, column=3, value=cost)
    ws3.cell(row=r, column=4, value=resp)
    style_data_row(ws3, r, COLS3, alt=(idx % 2 == 1))
    ws3.cell(row=r, column=1).font = Font(name="Calibri", bold=True, size=11)
    ws3.cell(row=r, column=1).alignment = Alignment(horizontal="left", vertical="center")
    ws3.cell(row=r, column=2).alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws3.cell(row=r, column=3).alignment = Alignment(horizontal="left", vertical="center")

r += 2
section_label(ws3, r, "Key Technical Requirements", COLS3)
r += 1
for i, h in enumerate(["Parameter", "Minimum Requirement", "Preferred", "Notes"], 1):
    ws3.cell(row=r, column=i, value=h)
style_header_row(ws3, r, COLS3)

tech_reqs = [
    ("Cell Chemistry", "LFP (Lithium Iron Phosphate)", "LFP", "NMC not accepted for utility-scale"),
    ("Cell Supplier", "Tier-1 (EVE, CATL, BYD, CALB, Hithium)", "EVE 314Ah", "Must be BNEF Tier 1"),
    ("Round-Trip Efficiency", "≥86% (AC-AC full system)", "≥87%", "Including auxiliary consumption"),
    ("PCS Model", "Kehua C-HUD series or equivalent", "BCS1250K-C-HUD / BCS1000K-C-HUD", "Grid-forming capable, ≥98% eff."),
    ("Grid-Forming Capability", "Required — Virtual Synchronous Generator", "Required", "VSG mode, frequency/voltage support"),
    ("Cycle Life", "≥6,000 at 90% DoD", "≥7,000", "To 70% SOH EOL"),
    ("SOH Guarantee Year 5", "≥85%", "≥85%", ""),
    ("SOH Guarantee Year 10", "≥79%", "≥80%", ""),
    ("SOH Guarantee Year 15", "≥70%", "≥70%", ""),
    ("Operating Temperature", "-10°C to +55°C", "-20°C to +55°C", "Cyprus ambient 0°C to 42°C"),
    ("Container Rating", "IP55 minimum", "IP55", "Coastal environment"),
    ("Production Lead Time", "≤90 days", "75 days", "From advance payment"),
    ("Shipping Time", "≤60 days", "45 days", "China to Limassol CIF"),
]

for idx, (param, min_req, pref, note) in enumerate(tech_reqs):
    r += 1
    ws3.cell(row=r, column=1, value=param)
    ws3.cell(row=r, column=2, value=min_req)
    ws3.cell(row=r, column=3, value=pref)
    ws3.cell(row=r, column=4, value=note)
    style_data_row(ws3, r, COLS3, alt=(idx % 2 == 1))
    ws3.cell(row=r, column=1).font = Font(name="Calibri", bold=True, size=11)
    ws3.cell(row=r, column=1).alignment = Alignment(horizontal="left", vertical="center")
    ws3.cell(row=r, column=2).alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws3.cell(row=r, column=3).alignment = Alignment(horizontal="left", vertical="center")
    ws3.cell(row=r, column=4).alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

widths3 = [30, 40, 35, 35]
for i, w in enumerate(widths3, 1):
    ws3.column_dimensions[get_column_letter(i)].width = w


# ═══════════════════════════════════════════
# SHEET 4: PRICING CURVE SUMMARY
# ═══════════════════════════════════════════
ws4 = wb.create_sheet("Pricing Curve")
ws4.sheet_properties.tabColor = "8E44AD"
COLS4 = 7

title_bar(ws4, 1, "  CIF PRICING CURVE — €/kWh vs System Size", COLS4)
note_row(ws4, 2, "  Shows how CIF price decreases with system size. Based on actual portfolio data (Q1 2026) + target improvement.", COLS4)

r = 4
section_label(ws4, r, "CIF €/kWh by Total System MWh (All Durations)", COLS4)
r += 1
headers4 = ["System MWh", "Actual CIF €/kWh (avg)", "Target CIF €/kWh", "Saving €/kWh",
            "Saving %", "Market Benchmark $/kWh", "Notes"]
for i, h in enumerate(headers4, 1):
    ws4.cell(row=r, column=i, value=h)
style_header_row(ws4, r, COLS4)

curve_data = [
    (4,   121.87, 112.00, "Smallest system in portfolio"),
    (5,   116.82, 108.00, "1-1.5 MW systems"),
    (8,   104.22, 96.00,  "2.5 MW systems"),
    (10,  103.00, 95.00,  "3-5 MW systems (avg)"),
    (15,  103.84, 95.00,  "4-5 MW systems"),
    (20,   95.00, 88.00,  "5-7 MW systems (avg)"),
    (25,   99.50, 92.00,  "7.5-8 MW systems"),
    (30,   90.77, 84.00,  "7 MW / 30 MWh"),
    (35,   94.09, 87.00,  "8 MW / 35 MWh"),
    (40,   94.25, 87.00,  "12 MW systems"),
    (60,   82.03, 76.00,  "8 MW / 60 MWh"),
    (100,  87.54, 80.00,  "25 MW / 100 MWh"),
    (200,  None,  74.00,  "Target for future large projects"),
    (400,  None,  70.00,  "Target for utility-scale 8h"),
]

market_benchmarks = {
    4: 140, 5: 135, 8: 125, 10: 120, 15: 115, 20: 110,
    25: 108, 30: 105, 35: 103, 40: 100, 60: 95, 100: 90,
    200: 82, 400: 78,
}

for idx, (mwh, actual, target, note) in enumerate(curve_data):
    r += 1
    ws4.cell(row=r, column=1, value=mwh)
    if actual is not None:
        ws4.cell(row=r, column=2, value=actual).number_format = currency_kwh_fmt
        ws4.cell(row=r, column=4, value=round(actual - target, 2)).number_format = currency_kwh_fmt
        ws4.cell(row=r, column=5, value=round((actual - target) / actual, 4)).number_format = pct_fmt
    else:
        ws4.cell(row=r, column=2, value="N/A")
        ws4.cell(row=r, column=4, value="N/A")
        ws4.cell(row=r, column=5, value="N/A")
    ws4.cell(row=r, column=3, value=target).number_format = currency_kwh_fmt
    benchmark = market_benchmarks.get(mwh)
    if benchmark:
        ws4.cell(row=r, column=6, value=f"~${benchmark}").alignment = Alignment(horizontal="center")
    ws4.cell(row=r, column=7, value=note)
    style_data_row(ws4, r, COLS4, alt=(idx % 2 == 1))

r += 2
note_row(ws4, r, "Market benchmarks based on Ember (Oct 2025), BNEF, Saudi/Italy auction data. Converted at ~1.05 EUR/USD.", COLS4)
r += 1
note_row(ws4, r, "Current CIF scope includes PCS (Kehua C-HUD grid-forming), MV transformer, switchgear, BMS, EVE LFP cells, cooling, fire suppression — more scope than typical equipment-only benchmarks.", COLS4)

widths4 = [14, 24, 22, 18, 14, 24, 40]
for i, w in enumerate(widths4, 1):
    ws4.column_dimensions[get_column_letter(i)].width = w


# ── Freeze panes on all sheets ──
for ws in [ws1, ws2, ws3, ws4]:
    ws.sheet_view.showGridLines = False

# ── Save ──
output_path = "docs/internal/BESS-CIF-Price-Targets-Mar2026.xlsx"
wb.save(output_path)
print(f"✅ Created: {output_path}")
