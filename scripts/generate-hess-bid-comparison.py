"""HESS transformer bid comparison + landed-cost / TCO workbook.

Builds an Excel calculator in HV Transformer/comparison/ with:
  - one column per supplier (1 Poland, 2 China/7sun, 3 Turkey)
  - landed DAP-site cost = ex-works/CIF + duty + freight + port handling + inland haulage
  - capitalised losses (no-load + load) added to give Total Cost of Ownership (TCO)
  - compliance flags (T2 included, Tier 2 Ecodesign, short-circuit type test, warranty)

Inputs are live Excel formulas so numbers can be dropped in on bid day (17 Jun) and the
landed cost + TCO recompute automatically. Pre-loads the known Supplier-2 (China) figure.
INTERNAL — contains prices; never send to client or suppliers.
"""
from __future__ import annotations

import sys
from pathlib import Path

import openpyxl
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

try:
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
except Exception:  # noqa: BLE001
    pass

BASE = Path(
    r"L:\My Drive\LINYANG\BESS CLIENTS\Individual_60-120-standalone\HV Transformer"
)
OUT = BASE / "comparison" / "bid-comparison-T1-T2-jun2026.xlsx"

NAVY = "1A365D"
GOLD = "C9A432"
WHITE = "FFFFFF"
GREY_BG = "F0F4F8"
AMBER = "FEF3C7"
GREEN = "E3F2E3"
THIN = Side(style="thin", color="BFC9D4")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

# Supplier columns: C, D, E  (B = parameter, A = item)
COL = {"S1": "C", "S2": "D", "S3": "E"}  # S1 Poland, S2 China/7sun, S3 Turkey

# Row map (1-based). Built sequentially in build().
def build():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Bid comparison"
    ws.column_dimensions["A"].width = 4
    ws.column_dimensions["B"].width = 46
    for c in ("C", "D", "E"):
        ws.column_dimensions[c].width = 24
    ws.column_dimensions["F"].width = 40

    def title(r, text, span=("A", "F"), fill=NAVY, color=WHITE, size=12, h=24):
        ws.merge_cells(f"{span[0]}{r}:{span[1]}{r}")
        c = ws[f"{span[0]}{r}"]
        c.value = text
        c.font = Font(bold=True, color=color, size=size)
        c.fill = PatternFill("solid", fgColor=fill)
        c.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[r].height = h

    def section(r, text):
        ws.merge_cells(f"A{r}:F{r}")
        c = ws[f"A{r}"]
        c.value = text
        c.font = Font(bold=True, color=NAVY, size=10)
        c.fill = PatternFill("solid", fgColor=GOLD)
        for col in ("A", "B", "C", "D", "E", "F"):
            ws[f"{col}{r}"].border = BORDER

    def row(r, item, param, vals=("", "", ""), note="", bold=False, fill=None, num="#,##0"):
        ws[f"A{r}"] = item
        ws[f"B{r}"] = param
        ws[f"C{r}"], ws[f"D{r}"], ws[f"E{r}"] = vals
        ws[f"F{r}"] = note
        for col in ("A", "B", "C", "D", "E", "F"):
            cell = ws[f"{col}{r}"]
            cell.border = BORDER
            cell.alignment = Alignment(
                horizontal="left" if col in ("A", "B", "F") else "right",
                vertical="center", wrap_text=(col in ("B", "F")),
            )
            cell.font = Font(size=9, bold=bold, color="000000")
            if fill:
                cell.fill = PatternFill("solid", fgColor=fill)
            if col in ("C", "D", "E") and isinstance(cell.value, (int, float)):
                cell.number_format = num

    title(1, "HESS PSEVDAS — TRANSFORMER BID COMPARISON  (INTERNAL — contains prices)")
    title(2, "Landed DAP-site cost + capitalised losses (TCO). Fill yellow cells on bid day; totals auto-compute.",
          fill=GREY_BG, color="404040", size=9, h=20)

    # Header
    r = 4
    ws[f"A{r}"], ws[f"B{r}"] = "", "Parameter"
    ws[f"C{r}"], ws[f"D{r}"], ws[f"E{r}"] = "Supplier-1 (PL)", "Supplier-2 (CN/7sun)", "Supplier-3 (TR)"
    ws[f"F{r}"] = "Notes"
    for col in ("A", "B", "C", "D", "E", "F"):
        cell = ws[f"{col}{r}"]
        cell.font = Font(bold=True, color=WHITE, size=10)
        cell.fill = PatternFill("solid", fgColor=NAVY)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER

    YEL = AMBER
    section(5, "1.  COMMERCIAL — as quoted")
    row(6, "1.1", "Status", ("AWAITING BID", "BID RECEIVED", "AWAITING BID"), "")
    row(7, "1.2", "Currency as quoted", ("", "USD", ""), "Normalise to EUR below", fill=YEL)
    row(8, "1.3", "Incoterm as quoted", ("", "CIF Limassol", ""), "Target: DAP site", fill=YEL)
    row(9, "1.4", "T1 main price (as quoted)", ("", 1280000, ""), "S2 = 1.28M USD CIF", fill=YEL)
    row(10, "1.5", "T2 earthing/aux price (as quoted)", ("", "NOT QUOTED", ""), "Must obtain separately", fill=YEL)
    row(11, "1.6", "FX rate to EUR (per 1 unit)", ("", 0.92, ""), "USD→EUR; set on bid day", fill=YEL, num="0.0000")
    row(12, "1.7", "T1 price in EUR", (
        "=IF(C9=\"\",\"\",C9*IF(C7=\"USD\",C11,1))",
        "=IF(D9=\"\",\"\",D9*IF(D7=\"USD\",D11,1))",
        "=IF(E9=\"\",\"\",E9*IF(E7=\"USD\",E11,1))",
    ), "Auto", bold=True)
    row(13, "1.8", "T2 price in EUR", (
        "=IFERROR(C10*IF(C7=\"USD\",C11,1),0)",
        "=IFERROR(D10*IF(D7=\"USD\",D11,1),0)",
        "=IFERROR(E10*IF(E7=\"USD\",E11,1),0)",
    ), "0 if not quoted", bold=True)

    section(14, "2.  LANDED COST ADDERS to reach DAP Plot 26 Psevdas  (EUR)")
    row(15, "2.1", "EU import duty (China origin only — verify HS 8504.23)", ("", 0, ""), "Intra-EU (PL/TR EU) = 0; CN verify rate", fill=YEL)
    row(16, "2.2", "Sea/road freight & insurance to Limassol (if not in Incoterm)", (0, 0, 0), "0 if CIF/CIP already includes", fill=YEL)
    row(17, "2.3", "Port handling / THC / clearance at Limassol", ("", "", ""), "", fill=YEL)
    row(18, "2.4", "Inland haulage + offloading Limassol→Psevdas (~40 km, abnormal load)", ("", "", ""), "Same for all", fill=YEL)
    row(19, "2.5", "Erection / commissioning supervision (if priced as extra)", ("", "", ""), "", fill=YEL)
    row(20, "2.6", "Adders subtotal", (
        "=SUM(C15:C19)", "=SUM(D15:D19)", "=SUM(E15:E19)",
    ), "Auto", bold=True)
    row(21, "2.7", "LANDED DAP-SITE COST (T1+T2+adders)", (
        "=IF(C12=\"\",\"\",C12+C13+C20)",
        "=IF(D12=\"\",\"\",D12+D13+D20)",
        "=IF(E12=\"\",\"\",E12+E13+E20)",
    ), "Auto", bold=True, fill=GREEN)

    section(22, "3.  CAPITALISED LOSSES (TCO)  — producer to state losses; factors editable")
    row(23, "3.1", "No-load loss P0 (kW)", ("", "", ""), "From offer", fill=YEL)
    row(24, "3.2", "Load loss Pk (kW)", ("", "", ""), "From offer", fill=YEL)
    row(25, "3.3", "Capitalisation €/kW — no-load", (8000, 8000, 8000), "Edit (typical EU A0)", fill=YEL)
    row(26, "3.4", "Capitalisation €/kW — load", (2000, 2000, 2000), "Edit (typical EU B0)", fill=YEL)
    row(27, "3.5", "Capitalised loss cost", (
        "=IF(C23=\"\",\"\",C23*C25+C24*C26)",
        "=IF(D23=\"\",\"\",D23*D25+D24*D26)",
        "=IF(E23=\"\",\"\",E23*E25+E24*E26)",
    ), "Auto", bold=True)
    row(28, "3.6", "TOTAL COST OF OWNERSHIP (landed + losses)", (
        "=IF(C21=\"\",\"\",C21+IFERROR(C27,0))",
        "=IF(D21=\"\",\"\",D21+IFERROR(D27,0))",
        "=IF(E21=\"\",\"\",E21+IFERROR(E27,0))",
    ), "Auto — ranking metric", bold=True, fill=GREEN)

    section(29, "4.  COMPLIANCE FLAGS  (Y / N / partial)")
    flags = [
        ("4.1", "T2 earthing/aux INCLUDED", "", "N", ""),
        ("4.2", "YNd11 (not Dyn11)", "", "", ""),
        ("4.3", "uk = 21%", "", "", ""),
        ("4.4", "LV insulation 170/70 (36 kV class)", "", "", ""),
        ("4.5", "Tier 2 Ecodesign / PEI declared", "", "", ""),
        ("4.6", "Independent short-circuit type test (KEMA/CESI/IPH)", "", "", ""),
        ("4.7", "CE / EU DoC", "", "", ""),
    ]
    rr = 30
    for item, param, a, bcn, ce in flags:
        row(rr, item, param, (a, bcn, ce), "", fill=GREY_BG)
        rr += 1

    section(rr, "5.  DELIVERY & WARRANTY")
    rr += 1
    row(rr, "5.1", "Lead time ex-works (weeks)", ("", "", ""), "", fill=YEL); rr += 1
    row(rr, "5.2", "Total delivery to Limassol (weeks)", ("", "", ""), "vs Jan 2027 slot", fill=YEL); rr += 1
    row(rr, "5.3", "On-site date achievable", ("", "", ""), "Target Q1 2027", fill=YEL); rr += 1
    row(rr, "5.4", "Warranty (target: 24 commissioning / 36-60 delivery, whichever first)",
        ("", "24 commiss. / 30 deliv. — SHORT", ""), "S2 below target — negotiate", fill=YEL); rr += 1

    ws[f"B{9}"].comment = Comment("Supplier-2 (China/7sun): 1,280,000 USD CIF Limassol, presumed T1 only.", "Lighthief")
    ws.freeze_panes = "A5"

    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)
    print(f"Wrote {OUT.relative_to(BASE)}")


if __name__ == "__main__":
    build()
