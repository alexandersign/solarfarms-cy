import openpyxl, sys
sys.stdout.reconfigure(encoding="utf-8", errors="replace")

wb = openpyxl.load_workbook(
    r"L:\My Drive\LINYANG\BESS CLIENTS\Individual_60-120-standalone\clarification-prefilled-v3.xlsx",
    data_only=True,
)
ws = wb.active
for r in range(1, ws.max_row + 1):
    q = ws.cell(r, 3).value
    bid = ws.cell(r, 4).value
    resp = ws.cell(r, 5).value
    if resp:
        print(f"\n{'='*100}\nROW {r}")
        if q:
            print(f"SPEC CLAUSE: {str(q).strip()[:200]}")
        if bid:
            print(f"BIDDER ASKS: {str(bid).strip()[:200]}")
        print(f"OUR ANSWER:\n{str(resp).strip()}")
