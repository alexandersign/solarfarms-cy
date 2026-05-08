"""
Lighthief EPC Confirmed Adders v5 — Generator
Produces: docs/internal/Lighthief-EPC-Confirmed-Adders-v5-May2026.xlsx

CIF sources: LY202601271 (Jan 2026) + LY202602111 (Feb 2026)
EMS: Disperon v3 Scenario C — hardware-only internal rate (EUBESS commissions)
Crane: €2,500/container (confirmed A. Soulis Mar 2026)
SCADA Local: flat €15,000/park (corrected from tiered; v3 pricing confirmed)
SCADA Global: €60,000/group (first park in each client group only)
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter
import os

OUT_PATH = os.path.join(os.path.dirname(__file__), '..', 'docs', 'internal',
                        'Lighthief-EPC-Confirmed-Adders-v5-May2026.xlsx')

# ─── CONFIRMED CIF FROM LINYANG QUOTATIONS ───────────────────────────────────
# Source: LY202601271 (Jan 2026) and LY202602111 (Feb 2026)
# For parks not in quotation: use nearest same-config park from same quotations
CIF = {
    # Config notation: (MW, MWh, BESS, MV, CNT)
    # QUOTATION   Park                CIF           Source
    'Galascope 1':           (5.0, 20, 4, 1, 5, 1_848_712.43, 'LY202601271 — Famagusta 2 match (5MW/20MWh)'),
    'Galascope 2':           (2.5, 10, 2, 1, 3,   974_457.00, 'LY202601271 — Dianary 1 match (2.5MW/10MWh)'),
    'Esperia Famagusta':     (6.5, 20, 4, 1, 5, 1_951_710.58, 'LY202601271 — exact'),
    'Esperia Famagusta 2':   (5.0, 20, 4, 1, 5, 1_848_712.43, 'LY202601271 — exact'),
    'Esperia Frenaros':      (25.0,100,20, 3,23, 8_782_302.00, 'LY202601271 — exact'),
    'Esperia Limassol':      (8.0, 60,12, 2,14, 4_937_363.00, 'LY202601271 — exact'),
    'Esperia Tseri':         (7.0, 20, 4, 1, 5, 2_001_736.00, 'LY202601271 — exact'),
    'Esperia Tseri 2a':      (2.5,7.5, 2, 1, 3,   871_307.00, 'LY202601271 — exact'),
    'Esperia Tseri 2b':      (7.5, 25, 5, 2, 7, 2_476_649.00, 'LY202601271 — exact'),
    'Esperia Tseri 2c':      (6.0, 20, 4, 1, 5, 1_907_319.00, 'LY202601271 — exact'),
    'Esperia Tseri 3':       (4.5, 15, 3, 1, 4, 1_562_533.00, 'LY202601271 — exact'),
    'L&T Sun Energy':        (5.0, 15, 3, 1, 4, 1_592_098.00, 'LY202601271 — exact'),
    'L&T Solar Power':       (1.5,  5, 1, 1, 2,   585_976.00, 'LY202601271 — exact'),
    'L&T Res Systems':       (1.0,  4, 1, 1, 2,   509_419.00, 'LY202601271 — exact'),
    'L&T Energia':           (1.5,  5, 1, 1, 2,   585_976.00, 'LY202601271 — exact'),
    'L&T PV Tech':           (1.0,  4, 1, 1, 2,   509_419.00, 'LY202601271 — exact'),
    'AGM Sunfield 1':        (5.0, 15, 3, 1, 4, 1_592_098.00, 'LY202601271 — exact'),
    'AGM Sunfield 2':        (1.5,  5, 1, 1, 2,   585_976.00, 'LY202601271 — exact'),
    'AGM Sunfield 3':        (1.0,  4, 1, 1, 2,   509_419.00, 'LY202601271 — exact'),
    'AGM Lightpower':        (8.0, 24, 5, 2, 7, 2_523_652.00, 'LY202601271 — exact'),
    'TBC 5MWh (Timotheos)':  (1.5,  5, 1, 1, 2,   585_976.00, 'LY202601271 — L&T Solar Power match (1.5MW/5MWh)'),
    'Solar Breeze':          (1.51, 5, 1, 1, 2,   585_976.00, 'LY202601271 — Classone Solar Breeze (1.5MW/5MWh)'),
    'Solar Garden':          (3.29,10, 2, 1, 3, 1_007_809.00, 'LY202601271 — Classone Solar Garden (3.3MW/10MWh)'),
    'Paphos 1':              (2.5,7.5, 2, 1, 3,   871_307.00, 'LY202601271 — exact'),
    'Paphos 2':              (2.5,7.5, 2, 1, 3,   871_307.00, 'LY202601271 — exact'),
    'Paphos 3':              (1.5,  5, 1, 1, 2,   541_046.00, 'LY202601271 — exact'),
    'My Sun Park':           (7.7, 25, 5, 2, 7, 2_523_652.00, 'LY202601271 — exact'),
    'Anarita 1':             (5.0, 20, 4, 1, 5, 1_848_712.43, 'LY202601271 — Famagusta 2 match (5MW/20MWh)'),
    'Anarita 2':             (5.0, 20, 4, 1, 5, 1_848_712.43, 'LY202601271 — Famagusta 2 match (5MW/20MWh)'),
    'Aeolian Dynamics':      (6.5, 20, 4, 2, 6, 1_951_710.58, 'LY202601271 — Esperia Famagusta match (6.5MW/20MWh, T4+T1)'),
}

# ─── DISPERON v3 EMS — SCENARIO C: Hardware-only (EUBESS installs internally) ─
# Source: voltusv3pricing CSV — "System + Hardware" column only
EMS_HW = {
    (1,1): 11_263, (2,1): 15_141, (3,1): 19_019, (3,2): 19_019,
    (4,1): 22_897, (4,2): 22_897, (5,2): 26_775, (6,2): 30_653,
    (7,2): 34_531, (8,2): 45_794, (12,2): 61_305, (20,3): 92_329,
}
SCADA_LOCAL     = 15_000   # per park, flat (v3 corrected — NOT tiered)
SCADA_GLOBAL    = 60_000   # per group, first park only
SCADA_L_MAINT   = 3_000    # per park per year (annual recurring)
SCADA_G_MAINT   = 12_000   # per group per year (annual recurring)
EMS_SUB_CLIENT  = 400      # €/MWh/yr charged to client (SHA §6.7)
EMS_SUB_VOLTUS  = 0.20     # 20% to Voltus; EUBESS retains 80%

# ─── ADDER RATES (v5 confirmed) ──────────────────────────────────────────────
IMPORT_DUTY_PCT   = 0.0266   # 2.66% of CIF
PORT_LANDING      = 600      # €/container (ECTL)
CUSTOMS           = 85       # €/park
CRANE             = 2_500    # €/container (A. Soulis, confirmed Mar 2026)
LV_CABLE_BESS     = 1_400    # €/BESS container (avg)
MV_CABLE_MV       = 3_500    # €/MV skid
MV_TERM_MV        = 2_200    # €/MV skid
PROT_ENG_SMALL    = 5_000    # ≤3 containers
PROT_ENG_LARGE    = 6_000    # ≥4 containers
REMOTE_TRIP       = 3_000    # €/park (physical RTU hardware for DSO compliance)
UPS_AUX           = 2_000    # €/park
# DEHN — simplified rates (confirmed quoted)
LPS_PER_BESS      = 429.41
LPS_PER_MV        = 617.52
SPD_DC_BESS       = 199.28
SPD_LV_MV         = 336.47
SPD_MV_MV         = 498.92
SPD_AUX           = 163.74
SPD_COMMS         = 134.37
EARTH_PER_STR     = 793.59   # per BESS+MV structure (= per park approx)
INSTALL_LABOR     = 1_600    # €/park (StrikeRA)
CIVIL_PER_MWH     = 2_000    # €/MWh (Kamil confirmed)
INSURANCE_PCT     = 0.0075   # 0.75% of CIF (EPC CAR/EAR budget)
DOCS_COMPL        = 7_000    # €/park

# ─── CLIENT PRICING ──────────────────────────────────────────────────────────
CLIENT_PRICE = {
    'Galascope 1':          2_258_900,
    'Galascope 2':          1_206_300,
    'Esperia Famagusta':    2_500_334,
    'Esperia Famagusta 2':  2_258_900,
    'Esperia Frenaros':    11_079_435,
    'Esperia Limassol':     6_003_120,
    'Esperia Tseri':        2_559_071,
    'Esperia Tseri 2a':     1_159_991,
    'Esperia Tseri 2b':     3_169_128,
    'Esperia Tseri 2c':     2_444_867,
    'Esperia Tseri 3':      2_008_827,
    'L&T Sun Energy':       1_961_880,
    'L&T Solar Power':        796_807,
    'L&T Res Systems':        704_706,
    'L&T Energia':            796_807,
    'L&T PV Tech':            704_706,
    'AGM Sunfield 1':       1_961_880,
    'AGM Sunfield 2':         796_807,
    'AGM Sunfield 3':         704_706,
    'AGM Lightpower':       3_225_985,
    'TBC 5MWh (Timotheos)':   800_000,
    'Solar Breeze':           795_443,
    'Solar Garden':         1_321_976,
    'Paphos 1':             1_159_302,
    'Paphos 2':             1_159_302,
    'Paphos 3':               742_767,
    'My Sun Park':          3_220_675,
    'Anarita 1':            2_380_000,
    'Anarita 2':            2_380_000,
    'Aeolian Dynamics':     2_660_000,
}

# ─── GROUP ASSIGNMENTS ───────────────────────────────────────────────────────
GROUPS = {
    'Galascope 1':          ('Esperia/Galascope', 'ESP', 1, 'Famagusta', 'confirmed'),
    'Galascope 2':          ('Esperia/Galascope', 'ESP', 1, 'Famagusta', 'confirmed'),
    'Esperia Famagusta':    ('Esperia Energy',    'ESP', 2, 'Famagusta', 'confirmed'),
    'Esperia Famagusta 2':  ('Esperia Energy',    'ESP', 2, 'Famagusta', 'confirmed'),
    'Esperia Frenaros':     ('Esperia Energy',    'ESP', 2, 'Famagusta', 'confirmed'),
    'Esperia Limassol':     ('Esperia Energy',    'ESP', 2, 'Limassol',  'confirmed'),
    'Esperia Tseri':        ('Esperia Tseri',     'TSR', 3, 'Nicosia',   'confirmed'),
    'Esperia Tseri 2a':     ('Esperia Tseri',     'TSR', 3, 'Nicosia',   'confirmed'),
    'Esperia Tseri 2b':     ('Esperia Tseri',     'TSR', 3, 'Nicosia',   'confirmed'),
    'Esperia Tseri 2c':     ('Esperia Tseri',     'TSR', 3, 'Nicosia',   'confirmed'),
    'Esperia Tseri 3':      ('Esperia Tseri',     'TSR', 3, 'Nicosia',   'confirmed'),
    'L&T Sun Energy':       ('Timotheos',         'TIM', 4, 'Limassol',  'pending-50%'),
    'L&T Solar Power':      ('Timotheos',         'TIM', 4, 'Limassol',  'pending-50%'),
    'L&T Res Systems':      ('Timotheos',         'TIM', 4, 'TBC',       'pending-50%'),
    'L&T Energia':          ('Timotheos',         'TIM', 4, 'Limassol',  'pending-50%'),
    'L&T PV Tech':          ('Timotheos',         'TIM', 4, 'TBC',       'pending-50%'),
    'AGM Sunfield 1':       ('Timotheos',         'TIM', 4, 'Nicosia',   'pending-50%'),
    'AGM Sunfield 2':       ('Timotheos',         'TIM', 4, 'TBC',       'pending-50%'),
    'AGM Sunfield 3':       ('Timotheos',         'TIM', 4, 'TBC',       'pending-50%'),
    'AGM Lightpower':       ('Timotheos',         'TIM', 4, 'TBC',       'pending-50%'),
    'TBC 5MWh (Timotheos)': ('Timotheos',         'TIM', 4, 'TBC',       'pending-50%'),
    'Solar Breeze':         ('Lampros',           'LAM', 5, 'Limassol',  'pending-50%'),
    'Solar Garden':         ('Lampros',           'LAM', 5, 'Limassol',  'pending-50%'),
    'Paphos 1':             ('Kerasi',            'KER', 6, 'Paphos',    'unconfirmed'),
    'Paphos 2':             ('Kerasi',            'KER', 6, 'Paphos',    'unconfirmed'),
    'Paphos 3':             ('Kerasi',            'KER', 6, 'Paphos',    'unconfirmed'),
    'My Sun Park':          ('Karis',             'KAR', 7, 'TBC',       'unconfirmed'),
    'Anarita 1':            ('Spanercom',         'ANA', 8, 'Paphos',    'high-85%'),
    'Anarita 2':            ('Spanercom',         'ANA', 8, 'Paphos',    'high-85%'),
    'Aeolian Dynamics':     ('Aeolian (standalone)', 'AEO', 9, 'Larnaca', 'high-80%'),
}

def calc_adders(park, mw, mwh, bess, mv, cnt, cif):
    d = round(cif * IMPORT_DUTY_PCT, 2)
    port = cnt * PORT_LANDING
    cust = CUSTOMS
    crane = cnt * CRANE
    lv = bess * LV_CABLE_BESS
    mv_c = mv * MV_CABLE_MV
    mv_t = mv * MV_TERM_MV
    prot = PROT_ENG_LARGE if cnt >= 4 else PROT_ENG_SMALL
    rt = REMOTE_TRIP
    ups = UPS_AUX
    lps = round(bess * LPS_PER_BESS + mv * LPS_PER_MV, 2)
    spd = round(bess * SPD_DC_BESS + mv * (SPD_LV_MV + SPD_MV_MV) + SPD_AUX + SPD_COMMS, 2)
    earth = round(EARTH_PER_STR, 2)   # 1 structure per park
    labor = INSTALL_LABOR
    civil = mwh * CIVIL_PER_MWH
    ins = round(cif * INSURANCE_PCT, 2)
    docs = DOCS_COMPL
    phys_total = d+port+cust+crane+lv+mv_c+mv_t+prot+rt+ups+lps+spd+earth+labor+civil+ins+docs
    return {
        'duty':d,'port':port,'cust':cust,'crane':crane,'lv':lv,
        'mv_c':mv_c,'mv_t':mv_t,'prot':prot,'rt':rt,'ups':ups,
        'lps':lps,'spd':spd,'earth':earth,'labor':labor,'civil':civil,
        'ins':ins,'docs':docs,'phys_total':phys_total
    }

# ─── BUILD WORKBOOK ──────────────────────────────────────────────────────────
wb = openpyxl.Workbook()

# Colour palette
H_BG   = '1E3A5F'  # dark navy — header bg
H_FG   = 'FFFFFF'
S_BG   = 'E8F0FE'  # light blue — sub-header
A_BG   = 'FFF8E1'  # amber — physical adders
E_BG   = 'E3F2FD'  # blue — EMS
R_BG   = 'E8F5E9'  # green — revenue/margin
T_BG   = 'ECEFF1'  # grey — totals
CONF   = 'C8E6C9'  # green  — confirmed
HIGH   = 'FFF9C4'  # yellow — high probability
PEND   = 'FFE0B2'  # orange — pending
UNCON  = 'FFCDD2'  # red    — unconfirmed

def hdr(ws, row, col, val, bg=H_BG, fg=H_FG, bold=True, wrap=True, align='center'):
    c = ws.cell(row=row, column=col, value=val)
    c.font = Font(bold=bold, color=fg, size=9)
    c.fill = PatternFill('solid', fgColor=bg)
    c.alignment = Alignment(horizontal=align, vertical='center', wrap_text=wrap)
    return c

def cell(ws, row, col, val, fmt=None, bg=None, bold=False, align='right'):
    c = ws.cell(row=row, column=col, value=val)
    c.font = Font(bold=bold, size=9)
    if fmt: c.number_format = fmt
    if bg: c.fill = PatternFill('solid', fgColor=bg)
    c.alignment = Alignment(horizontal=align, vertical='center')
    return c

EUR = '#,##0'
EUR2 = '#,##0.00'
PCT = '0.00%'
INT = '#,##0'

# ═══════════════════════════════════════════════════════════════════
# SHEET 1: Park Cost Breakdown
# ═══════════════════════════════════════════════════════════════════
ws = wb.active
ws.title = 'Park Cost Breakdown'
ws.freeze_panes = 'D4'

# Row 1: Title
ws.merge_cells('A1:AM1')
t = ws['A1']
t.value = 'LIGHTHIEF EPC — CONFIRMED COST BREAKDOWN v5.0  |  May 2026  |  CIF: LY202601271 (Jan 2026) + LY202602111 (Feb 2026)  |  EMS: Disperon v3 Scenario C  |  Crane: €2,500/cnt  |  SCADA: €15K flat'
t.font = Font(bold=True, color='FFFFFF', size=10)
t.fill = PatternFill('solid', fgColor=H_BG)
t.alignment = Alignment(horizontal='center', vertical='center')
ws.row_dimensions[1].height = 18

# Row 2: Section headers
COLS = [
    # Identity (A-H)
    ('No','','A'),('Park','','B'),('Group','','C'),('District','','D'),
    ('Status','','E'),('MW','','F'),('MWh','','G'),('BESS','','H'),('MV','','I'),('CNT','','J'),
    # CIF (K-M)
    ('CIF','€','K'),('CIF Source','','L'),('€/kWh','','M'),
    # Physical Adders (N-AD)
    ('Import\nDuty 2.66%','€','N'),('Port\nECTL','€','O'),('Customs\nInterfreight','€','P'),
    ('Crane\n€2.5K/cnt','€','Q'),('LV Cable','€','R'),('MV Cable','€','S'),('MV Term','€','T'),
    ('Prot\nEng','€','U'),('RTU\nComms','€','V'),('UPS\nAux','€','W'),
    ('LPS\nDEHN','€','X'),('SPD\nDEHN','€','Y'),('Earthing\nDEHN','€','Z'),
    ('Install\nLabor','€','AA'),('Civil\nWorks','€','AB'),('Insurance\n0.75%CIF','€','AC'),
    ('Docs\nCompl','€','AD'),('Phys\nAdders','€','AE'),
    # EMS Scenario C (AF-AJ)
    ('EMS HW\nVoltus','€','AF'),('SCADA\nLocal','€','AG'),('SCADA\nGlobal','€','AH'),
    ('EMS\nTotal','€','AI'),
    # Totals (AJ-AM)
    ('Installed\nCost','€','AJ'),('Client\nRevenue','€','AK'),('Margin\n€','€','AL'),('Margin\n%','%','AM'),
]

ws.row_dimensions[2].height = 36
ws.row_dimensions[3].height = 14

group_scada_done = set()
all_park_data = []

for pname in CIF:
    mw, mwh, bess, mv, cnt, cif, cif_src = CIF[pname]
    gname, gcode, gnum, district, status = GROUPS[pname]
    rev = CLIENT_PRICE[pname]
    add = calc_adders(pname, mw, mwh, bess, mv, cnt, cif)
    hw_cost = EMS_HW.get((bess, mv), EMS_HW.get((bess, 1), 11_263))
    scada_g = SCADA_GLOBAL if gname not in group_scada_done else 0
    if gname not in group_scada_done:
        group_scada_done.add(gname)
    ems_total = hw_cost + SCADA_LOCAL + scada_g
    installed = cif + add['phys_total'] + ems_total
    margin = rev - installed
    margin_pct = margin / rev if rev else 0
    all_park_data.append({
        'name': pname, 'group': gname, 'gcode': gcode, 'district': district,
        'status': status, 'mw': mw, 'mwh': mwh, 'bess': bess, 'mv': mv, 'cnt': cnt,
        'cif': cif, 'cif_src': cif_src, 'rate': cif/mwh/1000,
        'add': add, 'hw': hw_cost, 'scada_l': SCADA_LOCAL, 'scada_g': scada_g,
        'ems': ems_total, 'installed': installed, 'rev': rev,
        'margin': margin, 'margin_pct': margin_pct,
    })

# Write headers row 2
col_labels = [
    'No','Park Name','Client Group','District','Signing Status',
    'MW','MWh','BESS Cnts','MV Skids','Total Units',
    'CIF (€)','CIF Quotation Source','€/kWh',
    'Import Duty\n(2.66%)','Port Landing\n(ECTL)','Customs\n(Interfreight)',
    'Crane+Transport\n(€2,500/cnt)','LV Cabling','MV Cabling','MV Terminations',
    'Protection\nEngineering','RTU/Comms\n(DSO)','UPS &\nAuxiliary',
    'LPS DEHN','SPD DEHN','Earthing\nGrid',
    'Install\nLabour','Civil Works\n(€2K/MWh)','Insurance\n(0.75% CIF)',
    'Docs &\nCompliance','PHYSICAL\nADDERS TOTAL',
    'EMS Hardware\n(Disperon v3)','SCADA Local\n(€15K flat)','SCADA Global\n(€60K/group)',
    'EMS TOTAL',
    'INSTALLED\nCOST','CLIENT\nREVENUE','MARGIN €','MARGIN %',
]

status_colors = {
    'confirmed': CONF, 'high-85%': HIGH, 'high-80%': HIGH,
    'pending-50%': PEND, 'unconfirmed': UNCON,
}
section_bgs = {
    range(1,11): 'F5F5F5',   # identity
    range(11,14): 'FFF3E0',  # CIF — light orange
    range(14,32): A_BG,      # physical adders
    range(32,36): E_BG,      # EMS
    range(36,40): R_BG,      # totals
}

def get_col_bg(col_idx):
    c = col_idx  # 1-based
    if c <= 10: return 'F0F4FF'
    if c <= 13: return 'FFF3E0'
    if c <= 31: return A_BG
    if c <= 35: return E_BG
    return R_BG

for ci, label in enumerate(col_labels, 1):
    bg = get_col_bg(ci)
    hdr(ws, 2, ci, label, bg=H_BG if ci <= 10 else ('E65100' if ci <= 13 else ('0D47A1' if ci >= 32 else ('27AE60' if ci >= 36 else '5D4037'))), fg='FFFFFF')

# Data rows (start row 3)
for ri, p in enumerate(all_park_data, 1):
    row = ri + 2
    status_bg = status_colors.get(p['status'], 'FFFFFF')
    a = p['add']

    vals = [
        ri, p['name'], p['group'], p['district'], p['status'],
        p['mw'], p['mwh'], p['bess'], p['mv'], p['cnt'],
        p['cif'], p['cif_src'], p['cif']/p['mwh']/1000,
        a['duty'], a['port'], a['cust'], a['crane'], a['lv'],
        a['mv_c'], a['mv_t'], a['prot'], a['rt'], a['ups'],
        a['lps'], a['spd'], a['earth'], a['labor'], a['civil'],
        a['ins'], a['docs'], a['phys_total'],
        p['hw'], p['scada_l'], p['scada_g'], p['ems'],
        p['installed'], p['rev'], p['margin'], p['margin_pct'],
    ]
    fmts = [
        INT, None, None, None, None,
        '#,##0.0', '#,##0.0', INT, INT, INT,
        EUR2, None, '#,##0.00',
        EUR2,EUR2,EUR2,EUR2,EUR2,EUR2,EUR2,EUR2,EUR2,EUR2,
        EUR2,EUR2,EUR2,EUR2,EUR2,EUR2,EUR2,EUR,
        EUR,EUR,EUR,EUR,
        EUR,EUR,EUR,PCT,
    ]
    aligns = ['right','left','left','left','left'] + ['right']*35
    for ci, (v, f, al) in enumerate(zip(vals, fmts, aligns), 1):
        bg = status_bg if ci == 5 else ('FFFDE7' if ci == 34 and p['scada_g'] > 0 else None)
        bold = ci >= 36
        c = ws.cell(row=row, column=ci, value=v)
        c.font = Font(bold=bold, size=9)
        if f: c.number_format = f
        if bg: c.fill = PatternFill('solid', fgColor=bg)
        c.alignment = Alignment(horizontal=al, vertical='center')

# Column widths
widths = [4,28,22,12,14, 5,5,5,4,4, 13,34,7,
          10,8,8,10,8,8,8,8,8,7,8,8,8,8,10,10,8,12,
          10,9,10,10, 13,13,11,8]
for ci, w in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(ci)].width = w
ws.row_dimensions[2].height = 40

# ═══════════════════════════════════════════════════════════════════
# SHEET 2: Summary by Group
# ═══════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet('Group Summary')
ws2.freeze_panes = 'B3'

hdrs2 = ['Group','Status','Parks','MWh','Containers','CIF','Phys Adders',
         'EMS (ScenC)','Installed Cost','Revenue','Margin €','Margin %',
         'Annual EMS Sub\n(client pays)','EUBESS retains\n80%/yr','15yr Sub\nretained']
for ci, h in enumerate(hdrs2, 1):
    hdr(ws2, 1, ci, h)

groups_data = {}
for p in all_park_data:
    g = p['group']
    if g not in groups_data:
        groups_data[g] = {'status': p['status'], 'parks':0,'mwh':0,'cnt':0,
                          'cif':0,'phys':0,'ems':0,'installed':0,'rev':0,'margin':0}
    d = groups_data[g]
    d['parks'] += 1; d['mwh'] += p['mwh']; d['cnt'] += p['cnt']
    d['cif'] += p['cif']; d['phys'] += p['add']['phys_total']
    d['ems'] += p['ems']; d['installed'] += p['installed']
    d['rev'] += p['rev']; d['margin'] += p['margin']

for ri, (g, d) in enumerate(groups_data.items(), 2):
    sub_yr = d['mwh'] * EMS_SUB_CLIENT
    eubess_yr = sub_yr * 0.80
    sub_15 = eubess_yr * 15
    sbg = status_colors.get(d['status'], 'FFFFFF')
    vals2 = [g, d['status'], d['parks'], d['mwh'], d['cnt'],
             d['cif'], d['phys'], d['ems'], d['installed'],
             d['rev'], d['margin'], d['margin']/d['rev'],
             sub_yr, eubess_yr, sub_15]
    fmts2 = [None,None,INT,'#,##0.0',INT,EUR,EUR,EUR,EUR,EUR,EUR,PCT,EUR,EUR,EUR]
    for ci, (v, f) in enumerate(zip(vals2, fmts2), 1):
        c = ws2.cell(row=ri, column=ci, value=v)
        c.font = Font(size=9, bold=(ci >= 10))
        if f: c.number_format = f
        if ci == 2: c.fill = PatternFill('solid', fgColor=sbg)
    ws2.row_dimensions[ri].height = 14

# Totals row
tr = len(groups_data) + 2
hdr(ws2, tr, 1, 'PORTFOLIO TOTAL', bg='1E3A5F')
totals = {k: sum(d[k] for d in groups_data.values())
          for k in ['parks','mwh','cnt','cif','phys','ems','installed','rev','margin']}
sub_tot = totals['mwh'] * EMS_SUB_CLIENT
sub_ret = sub_tot * 0.80
t_vals = ['', totals['parks'], totals['mwh'], totals['cnt'],
          totals['cif'], totals['phys'], totals['ems'], totals['installed'],
          totals['rev'], totals['margin'], totals['margin']/totals['rev'],
          sub_tot, sub_ret, sub_ret*15]
t_fmts = [None,INT,'#,##0.0',INT,EUR,EUR,EUR,EUR,EUR,EUR,PCT,EUR,EUR,EUR]
for ci, (v, f) in enumerate(zip(t_vals, t_fmts), 2):
    c = ws2.cell(row=tr, column=ci, value=v)
    c.font = Font(bold=True, size=10, color='FFFFFF')
    c.fill = PatternFill('solid', fgColor='1E3A5F')
    if f: c.number_format = f
ws2.column_dimensions['A'].width = 28
for ci in range(2,16): ws2.column_dimensions[get_column_letter(ci)].width = 14
ws2.row_dimensions[tr].height = 16

# ═══════════════════════════════════════════════════════════════════
# SHEET 3: Adder Rates Reference
# ═══════════════════════════════════════════════════════════════════
ws3 = wb.create_sheet('Adder Rates v5')
ref_rows = [
    ('LOGISTICS & IMPORT','','',''),
    ('Import Duty (weighted HS codes)','2.66% × CIF','Interfreight / Cyprus Customs','CONFIRMED'),
    ('Port Landing (ECTL THC + agent)','€600 / container','ECTL Limassol','QUOTED'),
    ('Customs Clearance','€85 / park','Interfreight','QUOTED'),
    ('Crane + Inland Transport','€2,500 / container  ← UPDATED Mar 2026','A. Soulis','CONFIRMED'),
    ('','','',''),
    ('ELECTRICAL','','',''),
    ('LV Cabling (DC side)','€1,400 / BESS container (avg)','Internal estimate','ESTIMATE'),
    ('MV Cabling (external)','€3,500 / MV skid','Internal estimate','ESTIMATE'),
    ('MV Terminations','€2,200 / MV skid','Internal estimate','ESTIMATE'),
    ('Protection Engineering','€5,000 (≤3 cnt) / €6,000 (≥4 cnt) per park','Internal estimate','ESTIMATE'),
    ('RTU/Comms (DSO remote trip) ← RENAMED','€3,000 / park flat  [physical hardware only, NOT SCADA]','Internal estimate','ESTIMATE'),
    ('UPS & Auxiliary','€2,000 / park','Internal estimate','ESTIMATE'),
    ('','','',''),
    ('LIGHTNING & SURGE (DEHN)','','',''),
    ('LPS per BESS container','€429.41','DEHN Quotation','QUOTED'),
    ('LPS per MV skid','€617.52','DEHN Quotation','QUOTED'),
    ('SPD (DC) per BESS container','€199.28','DEHN Quotation','QUOTED'),
    ('SPD (LV) per MV skid','€336.47','DEHN Quotation','QUOTED'),
    ('SPD (MV) per MV skid','€498.92','DEHN Quotation','QUOTED'),
    ('SPD (Aux) per park','€163.74','DEHN Quotation','QUOTED'),
    ('SPD (Comms) per structure','€134.37','DEHN Quotation','QUOTED'),
    ('Site Earthing Grid per park','€793.59','DEHN Quotation','QUOTED'),
    ('Install Labour (StrikeRA)','€1,600 / park flat','StrikeRA','CONFIRMED'),
    ('','','',''),
    ('CIVIL WORKS','','',''),
    ('All civil (platforms, trenches, fencing)','€2,000 / MWh all-in  (Kamil confirmed Feb 2026)','Kamil / subcontractors','CONFIRMED'),
    ('','','',''),
    ('INSURANCE & COMPLIANCE','','',''),
    ('EPC CAR/EAR Insurance budget','0.75% of CIF  [NOTE: was labelled "Marine" in v4 — corrected]','Marsh (budget)','PENDING PLACEMENT'),
    ('Docs & Statutory Compliance (EAC/CERA)','€7,000 / park','Internal estimate','ESTIMATE'),
    ('','','',''),
    ('EMS / SCADA — Disperon v3 SCENARIO C','','',''),
    ('EMS Hardware (System + Hardware only)','By (BESS, MV) config — see below','Voltus/Disperon','QUOTED (v3)'),
    ('  BESS=1, MV=1 (2 units)','€11,263','Voltus v3','QUOTED'),
    ('  BESS=2, MV=1 (3 units)','€15,141','Voltus v3','QUOTED'),
    ('  BESS=3, MV=1 (4 units)','€19,019','Voltus v3','QUOTED'),
    ('  BESS=4, MV=1 (5 units)','€22,897','Voltus v3','QUOTED'),
    ('  BESS=5, MV=2 (7 units)','€26,775','Voltus v3','QUOTED'),
    ('  BESS=6, MV=2 (8 units)','€30,653','Voltus v3','QUOTED'),
    ('  BESS=7, MV=2 (9 units)','€34,531','Voltus v3','QUOTED'),
    ('  BESS=12, MV=2 (14 units)','€61,305','Voltus v3','QUOTED'),
    ('  BESS=20, MV=3 (23 units)','€92,329','Voltus v3','QUOTED'),
    ('SCADA Local (on-premise station)','€15,000 / park  FLAT — all sizes  ← CORRECTED v4 used tiered','Voltus v3','QUOTED'),
    ('SCADA Global (group platform)','€60,000 / group (first park only; remaining parks = €0)','Voltus v3','QUOTED'),
    ('','','',''),
    ('ANNUAL RECURRING (post-commissioning)','','',''),
    ('EMS Subscription (client-facing)','€400 / MWh / year  (SHA §6.7 confirmed)','Lighthief EUBESS SHA v5','CONFIRMED'),
    ('  → Voltus share (20%)','€80 / MWh / year','SHA §6.7','CONFIRMED'),
    ('  → EUBESS retains (80%)','€320 / MWh / year','SHA §6.7','CONFIRMED'),
    ('SCADA Local maintenance','€3,000 / park / year','Voltus v3','QUOTED'),
    ('SCADA Global maintenance','€12,000 / group / year','Voltus v3','QUOTED'),
    ('','','',''),
    ('EXCLUDED / CLIENT-PAID','','',''),
    ('Protection Relay Testing (DSO)','€1,250 / container  [per Dino confirmation]','Client-paid','CONFIRMED EXCLUDED'),
    ('Electrical Drawings / As-Built','€5,000–€15,000 / site','Client-paid','CONFIRMED EXCLUDED'),
    ('VAT (19%)','Charged on all invoices, VAT-registered client','Cyprus Revenue','EXCLUDED'),
]
hdr(ws3, 1, 1, 'Item'); hdr(ws3, 1, 2, 'Rate / Formula'); hdr(ws3, 1, 3, 'Supplier / Source'); hdr(ws3, 1, 4, 'Status')
for ri, (item, rate, src, status) in enumerate(ref_rows, 2):
    is_section = (rate == '' and src == '')
    for ci, v in enumerate([item, rate, src, status], 1):
        c = ws3.cell(row=ri, column=ci, value=v)
        c.font = Font(bold=is_section, size=9, color='FFFFFF' if is_section else '000000')
        if is_section:
            c.fill = PatternFill('solid', fgColor='1E3A5F')
        elif '← CORRECTED' in (rate or '') or '← UPDATED' in (rate or '') or '← RENAMED' in (rate or ''):
            c.fill = PatternFill('solid', fgColor='FFF9C4')
        c.alignment = Alignment(wrap_text=True, vertical='center')
ws3.column_dimensions['A'].width = 42
ws3.column_dimensions['B'].width = 50
ws3.column_dimensions['C'].width = 28
ws3.column_dimensions['D'].width = 22

wb.save(OUT_PATH)
print(f"Saved: {OUT_PATH}")
print(f"Parks: {len(all_park_data)}")
print(f"\nPortfolio summary:")
for g, d in groups_data.items():
    print(f"  {g:<30} {d['parks']} parks  {d['mwh']:.1f} MWh  CIF €{d['cif']:>12,.0f}  margin €{d['margin']:>9,.0f} ({d['margin']/d['rev']*100:.1f}%)")
ttl_margin = sum(d['margin'] for d in groups_data.values())
ttl_rev = sum(d['rev'] for d in groups_data.values())
ttl_cif = sum(d['cif'] for d in groups_data.values())
print(f"\n  {'TOTAL':<30} {sum(d['parks'] for d in groups_data.values())} parks  {sum(d['mwh'] for d in groups_data.values()):.1f} MWh  CIF €{ttl_cif:>12,.0f}  margin €{ttl_margin:>9,.0f} ({ttl_margin/ttl_rev*100:.1f}%)")
print(f"\n  Annual EMS sub to client: €{sum(d['mwh'] for d in groups_data.values())*400:,.0f}/yr  |  EUBESS retains 80%: €{sum(d['mwh'] for d in groups_data.values())*400*0.8:,.0f}/yr")
