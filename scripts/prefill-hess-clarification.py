"""Prefill Linyang clarification.xlsx — Response from employer column.

Fills col 5 where Lighthief/HESS can answer from known data.
Leaves rows marked [ASK CLIENT] empty with a comment.
Saves to clarification-prefilled.xlsx alongside original.
"""
from __future__ import annotations
import sys, shutil
from pathlib import Path
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.comments import Comment

sys.stdout.reconfigure(encoding="utf-8", errors="replace")

SRC = Path(r"L:\My Drive\LINYANG\BESS CLIENTS\Individual_60-120-standalone\clarification.xlsx")
OUT = SRC.with_name("clarification-prefilled-v5.xlsx")

# Colour scheme
FILL_DONE   = PatternFill("solid", fgColor="D4EDDA")   # green  — prefilled
FILL_CLIENT = PatternFill("solid", fgColor="FFF3CD")   # amber  — needs client
FILL_EPS    = PatternFill("solid", fgColor="D1ECF1")   # blue   — needs EPS/TSOC
FILL_SELF   = PatternFill("solid", fgColor="D4EDDA")   # green  — Lighthief can answer

# (row, answer, fill, owner)
ANSWERS = {
    # Technical Part
    4: (
        "This is a standalone BESS project and the BESS yard is adjacent to the HV/MV substation. "
        "Per ANNEX-II §A.6, connection between the two earthing grids is permitted for adjacent "
        "installations, following approval from the Transmission System Operator (TSOC). "
        "The earthing system design will be submitted to TSOC for approval per T14. "
        "The two earthing systems may be connected, subject to TSOC approval.",
        FILL_DONE, "Lighthief/EPS"
    ),
    5: (
        "The ESMS (Energy Management System) will be independent from the HV/MV substation control "
        "and protection system per §A.9. This means no shared SCADA/control hardware, "
        "no common low-voltage busbars and no shared auxiliary transformers. "
        "The ESMS may exchange data with the substation RTU via standard protocols (IEC 61850 / Modbus) "
        "for grid-code compliance and TSOC remote control — this data exchange is acceptable "
        "and does not constitute integration of the protection system. "
        "Bidder to confirm their ESMS architecture meets this boundary.",
        FILL_DONE, "Lighthief"
    ),
    6: (
        "Employer's position: Impact sensors in individual battery module/rack packaging are required "
        "per ANNEX-II §C.1. The Bidder confirms UN 38.3 transport safety compliance. "
        "As a practical accommodation, Employer will accept either:\n"
        "  (a) Factory-fitted impact sensor in each rack/module transport package, OR\n"
        "  (b) A documented transport monitoring protocol — GPS tracking with shock-logging on the "
        "container during sea + road transit, with a threshold-based inspection procedure on arrival. "
        "Option (b) must be agreed in writing before shipment. Battery racks showing evidence of "
        "heavy impact per the threshold shall not be installed without prior testing (per ANNEX-II §C.1). "
        "Bidder to confirm which option they propose and provide their transport monitoring procedure.",
        FILL_DONE, "Lighthief (practical position — no TSOC approval needed)"
    ),
    7: (
        "Performance degradation < 20% applies over the required chronological lifetime of 10 years, "
        "based on 1 complete daily cycle (10 years x 365 days = 3,650 cycles at 1 CPD).\n\n"
        "Clarification on cycle counts in ANNEX-II:\n"
        "  - 10-year chronological lifetime @ 1 CPD = 3,650 cycles (the <20% degradation threshold)\n"
        "  - Cyclic lifetime guarantee (item 12) = 7,300 cycles = 20 years @ 1 CPD\n\n"
        "The <20% degradation requirement is measured at year 10 / 3,650 cycles.\n\n"
        "Per ANNEX-II §B.I.2, all performance requirements are referred to the Point of Connection "
        "(PoC). For this project the guaranteed energy capacity at PoC is 100 MWh (project nameplate "
        "DC capacity: 120 MWh) — see Scope of Supply.\n\n"
        "BESS energy capacity will be measured at the end of each of the five 2-year successive "
        "periods (years 2, 4, 6, 8, 10). Measured capacity at PoC shall not fall below the "
        "guaranteed 100 MWh at any of these measurement points (ANNEX-II §B.V.1). "
        "Augmentation (capacity top-up) is permitted and space for >= 20% additional capacity must "
        "be reserved on site (§B.V.2); alternatively, initial over-installation above the minimum is "
        "permitted if no further augmentation is needed in the 10-year period.\n\n"
        "IMPORTANT — actual operating profile: the project will operate at approximately "
        "1.5 cycles per day (arbitrage + frequency regulation), i.e. ~5,475 cycles by year 10. "
        "While the ANNEX-II <20% threshold is defined on a 1-cycle/day basis (3,650 cycles), "
        "the Bidder's degradation curve and augmentation plan must be based on the ACTUAL "
        "operating profile of 1.5 cycles/day and must demonstrate that the 100 MWh PoC guaranteed "
        "capacity is retained at every 2-year measurement point throughout the first 10 years "
        "of operation. State the augmentation quantities and timing (MWh added, at which year) "
        "in the proposal.",
        FILL_DONE, "Lighthief (1.5 CPD per client email Jun 2026)"
    ),
    8: (
        "Employer's position on standards applicability — subject to final confirmation by EPS/Iacovos:\n\n"
        "Mandatory standards for this project:\n"
        "  PRIMARY: IEC 62933-5-2 (BESS safety), NFPA 855, IEC 62619, IEC 63056, IEC 62477-1\n\n"
        "Employer's response to the Bidder's standard-by-standard observations:\n"
        "  - IEC 62485-5 (Safe operation of stationary lithium-ion batteries): AGREED — this "
        "standard applies directly to the Li-ion battery containers and their installation, "
        "operation and maintenance. Compliance required.\n"
        "  - IEC 61936-1 (AC installations > 1 kV): AGREED — applies to the AC side "
        "(HV/MV substation, transformer yard, MV switchgear — EPC scope), not to the battery "
        "containers themselves. The Bidder's container scope is not assessed against 61936-1.\n"
        "  - IEC 60364-5-57: AGREED — LV battery standard, not applicable to the HV (1500 V DC "
        "class) battery containers. Applies only to LV auxiliary battery systems if any.\n\n"
        "Mandatory standards for the battery containers: IEC 62933-5-2, NFPA 855, IEC 62619, "
        "IEC 63056, IEC 62485-5, IEC 62477-1 (PCS), IEC 62281/UN 38.3 (transport). "
        "Certificates required with the technical offer. Note ANNEX-II §A.8 also requires "
        "IEC 62485-2 provisions for the design/erection/commissioning of the electrical system.\n\n"
        "[Note: EPS/Iacovos to confirm this interpretation before finalising the response.]",
        FILL_DONE, "Lighthief prefill — EPS to confirm before sending"
    ),
    9: (
        "Employer clarification on 10-year no major component replacement (ANNEX-II item 13):\n\n"
        "The requirement means the system must maintain guaranteed energy capacity for 10 years based "
        "on 1 complete daily cycle, without the need to replace major components due to inadequate "
        "initial sizing or premature failure.\n\n"
        "The following ARE permitted during the 10-year period:\n"
        "  - Capacity augmentation (adding containers/racks to maintain guaranteed energy capacity)\n"
        "  - Replacement of consumables: filters, coolant, aerosol canisters, fuses, contactors\n"
        "  - Replacement of sensors, communication modules, fans\n"
        "  - Warranty replacements of cells/modules due to manufacturing defect\n\n"
        "The following would constitute non-compliant 'major component' replacement:\n"
        "  - Full battery container replacement due to normal degradation exceeding specification\n"
        "  - PCS unit replacement due to undersizing\n"
        "  - Step-up transformer replacement due to undersizing\n\n"
        "Augmentation space of ≥20% of guaranteed capacity must be reserved on site (ANNEX-II §B.V.2). "
        "The Bidder's degradation curve and augmentation plan must demonstrate 10-year compliance.",
        FILL_DONE, "Lighthief (based on ANNEX-II §B.V interpretation + Esperia precedent)"
    ),
    10: (
        "Employer notes the Bidder's technical observation regarding IEC 62133. "
        "This standard is listed as a general reference. The primary applicable standard for "
        "Li-ion utility-scale storage cells is IEC 62619. Compliance with IEC 62619 (and IEC 63056 "
        "for modules) is accepted. Bidder to confirm IEC 62619 / IEC 63056 compliance and provide "
        "certificates.",
        FILL_DONE, "Lighthief"
    ),
    11: (
        "Employer notes that IEC 61427-1:2013 is a photovoltaic off-grid standard. "
        "For this grid-connected standalone BESS project, compliance with IEC 62619 and IEC 63056 "
        "is the applicable requirement. IEC 61427-1 is not a mandatory requirement for this project. "
        "Bidder to confirm IEC 62619 / 63056 compliance.",
        FILL_DONE, "Lighthief"
    ),
    12: (
        "This is a STANDALONE BESS project at Psevdas — there is no co-located renewable energy park "
        "at the Point of Connection. ANNEX-II item 9 (RES park monitoring integration) therefore "
        "does not apply in the sense of integrating with an existing third-party RES SCADA.\n\n"
        "The ESMS/EMS required for this project must:\n"
        "  (a) Integrate with TSOC/ECCC for remote dispatch: active power setpoint, reactive power "
        "control, frequency response (FCR/aFRR), SoC reporting — via RTU/IEC 104 or IEC 61850\n"
        "  (b) Provide Employer with a monitoring portal (web-based) covering power, energy, SoC, "
        "alarms, availability, and historical data export\n"
        "  (c) Implement the project's operating modes: primarily ARBITRAGE (energy trading / "
        "self-scheduling at ~1.5 cycles/day) plus FREQUENCY REGULATION (FCR/aFRR participation "
        "per T14.6)\n\n"
        "EMS scope: The EMS (hardware + software) is listed as an optional item in the Scope of "
        "Supply. If the Bidder includes EMS, confirm: communication protocols (IEC 61850 / Modbus TCP), "
        "TSOC RTU compatibility, and whether a Cyprus-based service interface is available. "
        "If EMS is excluded from the Bidder's scope, EPC (Lighthief) will provide a compatible "
        "third-party EMS (Voltus platform).",
        FILL_DONE, "Lighthief (standalone BESS — no RES integration required)"
    ),
    # Performance Part
    14: (
        "Safety Coordinator (SC) — EPC responsibility, per T12.4.2 of the Transmission Rules:\n\n"
        "1. Who provides the SC: The User (HESS / Lighthief as EPC operator) is required to "
        "designate and maintain one or more Safety Coordinators per connection point at all times "
        "(T12.4.2.1). EPC (Lighthief) will provide and maintain TSOC-authorized Safety Coordinators.\n\n"
        "2. Qualification/certification: HV Authorization Certificates for SC duties are issued "
        "by TSOC (ΔΣΜΚ) to qualified User representatives, per TSOC's published Operating "
        "Instruction (T12.4.2.2). Candidates are examined by TSOC on safety coordination, "
        "management and applicable legislation before a certificate is issued.\n\n"
        "3. Cost: EPC (Lighthief) bears all training and certification costs for its personnel.\n\n"
        "4. Temporary SC: Any person acting as SC — including temporary cover — must hold a "
        "valid TSOC HV Authorization Certificate. No exceptions.\n\n"
        "5. Personnel changes: Any change to the SC list must be submitted to TSOC "
        "immediately and in writing (T12.4.2.3). EPC to notify TSOC and Employer within "
        "24 hours of any SC change.\n\n"
        "6. Timeline for approval: Certificate applications follow TSOC's procedure per their "
        "published Operating Instruction. EPC to initiate SC certification process no later than "
        "60 days before planned energisation.",
        FILL_DONE, "Lighthief (directly from T12.4.2 rules)"
    ),
    15: (
        "Wide frequency/voltage operating range — T14.4.1 (directly from T14 rules):\n\n"
        "1. Baseline operating range already required — T14.4.1.1, Table 14.1:\n"
        "     47.0 - 47.5 Hz: minimum 10 seconds\n"
        "     47.5 - 49.5 Hz: minimum 60 minutes\n"
        "     49.5 - 50.5 Hz: unlimited (continuous)\n"
        "     50.5 - 52.0 Hz: minimum 60 minutes\n"
        "   Plus RoCoF withstand per Table 14.2 (up to +/-4 Hz/s for 0.25 s). "
        "Bidder to confirm compliance with these exact ranges and durations.\n"
        "   Note: there is NO grid-forming requirement for the PCS — grid-following operation "
        "meeting the TSOC T14 requirements is sufficient (Employer confirmation, Jun 2026).\n\n"
        "2. Trigger for range expansion: TSOC may agree with the Storage Manager to expand "
        "frequency or voltage ranges, or minimum duration, 'when necessary to maintain or "
        "restore power system security' (T14.4.1.1 / T14.4.1.2). This is at TSOC's discretion "
        "and must be technically and economically feasible.\n\n"
        "3. Obligation to cooperate: The Storage Manager 'shall not unreasonably refuse' expansion, "
        "taking into account technical and economic feasibility (T14.4.1.1). This is not a "
        "unilateral requirement — it requires mutual agreement.\n\n"
        "4. Cost allocation: Any post-contract expansion of operating range required by "
        "regulatory change constitutes a Change in Law. Under the EPC contract, such changes "
        "are Employer-funded variations. The Bidder is not liable for costs arising from "
        "regulatory changes after contract execution.\n\n"
        "5. Exemption: If technical standards change and the required range cannot be met "
        "by the installed equipment without hardware modification, the Bidder shall notify the "
        "Employer and TSOC; a variation order will be agreed before any upgrade is undertaken. "
        "Force Majeure provisions apply where compliance is impossible.",
        FILL_DONE, "Lighthief (from T14.4.1 rules + EPC change-in-law framework)"
    ),
    16: (
        "Emergency operation beyond rated parameters:\n\n"
        "T14 establishes the framework for emergency operation of BESS. Key provisions:\n\n"
        "1. Allowable range: BESS must be capable of operating across the full rated power range "
        "(0–100% of maximum charge/discharge capacity, T14.4.1.5). Rate of change is "
        "adjustable between 1–100% of rated power per minute (default 20%/min, T14.4.1.6). "
        "During emergencies, TSOC may require operation at any point within this range.\n\n"
        "2. Duration limits: Operation within the rated parameters as defined above is "
        "permissible continuously. Operation beyond rated parameters (e.g. above nameplate "
        "power) is NOT required by T14 and would require specific written agreement with TSOC "
        "and Bidder, with defined duration limits agreed at that time.\n\n"
        "3. Liability for equipment damage: If TSOC directs operation that causes equipment "
        "damage beyond normal operating limits, liability is governed by the Connection "
        "Agreement and Transmission Rules. Damage caused by TSOC emergency instructions "
        "is not attributable to the Bidder. EPC contract includes Force Majeure and "
        "grid-authority instruction carve-outs from Bidder liability.\n\n"
        "4. Employer's position: Sotiris/HESS to confirm the specific overload capability "
        "they wish the Bidder to commit to (e.g. 110% for 30 seconds, 120% for 5 seconds). "
        "These would need to be within the PCS and battery manufacturer's limits. Bidder "
        "to state maximum short-term overload capability in their technical offer.",
        FILL_DONE, "Lighthief (T14 rules + EPC framework; overload capability = ask Sotiris)"
    ),
    17: (
        "EPC Contract — Delay Liquidated Damages (Lighthief EPC standard, based on Esperia framework):\n"
        "  Days 1–30 late: 0.1% of Contract Price per day\n"
        "  Days 31–60 late: 0.15% of Contract Price per day\n"
        "  Day 61+: 0.2% of Contract Price per day\n"
        "  Maximum LD cap: 10% of total Contract Price\n\n"
        "Performance Bond: 5% of Contract Price, delivered within 14 days of advance payment receipt; "
        "released 30 days after Provisional Acceptance Certificate (PAC).\n\n"
        "Availability Liquidated Damages (LTSA Tier C — 97% guarantee):\n"
        "  95%–<97%: 5% reduction in annual LTSA fee\n"
        "  93%–<95%: 10% reduction\n"
        "  90%–<93%: 15% reduction\n"
        "  Below 90%: 20% reduction (maximum)\n"
        "Availability LDs are the sole remedy for unavailability under the LTSA.\n\n"
        "Final penalty schedules, LD caps and bond conditions to be confirmed in the executed EPC agreement.",
        FILL_DONE, "Lighthief (Esperia/standard EPC framework)"
    ),
    18: (
        "O&M responsibility boundary (per TSOC Preliminary Connection Terms, Apr 2025):\n\n"
        "TSOC/ISM scope (grid side): Extension of the Psevdas transmission substation (new 132 kV "
        "bay), the 132 kV underground cable (3x1c 300 mm2 XLPE) up to its termination in the KYEA, "
        "and ISM protection/telecom on the grid side. Energy meters are installed and programmed "
        "by ISM/ISD.\n\n"
        "Applicant / Lighthief EPC & LTSA scope (plant side): The entire KYEA (Central Storage "
        "Substation) and BESS plant — including the 132 kV AIS/GIS bay receiving the ISM cable "
        "termination, metering VTs/CTs (applicant supply per connection terms), 132/33 kV step-up "
        "transformer, earthing & auxiliary transformer, MV switchgear, PCS, battery containers, "
        "BMS, ESMS, RTU/telecom to ECCC, earthing system, fire protection and all BoP.\n\n"
        "Boundary point (grid <-> plant): the ISM 132 kV cable sealing end / termination at the "
        "KYEA bay.\n\n"
        "BIDDER'S scope boundary: per the Scope of Supply, the Bidder's supply and O&M "
        "responsibility covers the BESS DC side (battery containers, BMS), PCS, MV step-up "
        "transformers and Ring Main Units — terminating at the 33 kV MV connection to the "
        "Employer's MV switchgear. The HV substation (KYEA), the 132/33 kV main step-up "
        "transformer and all 132 kV equipment are NOT in the Bidder's scope (EPC/Employer scope).\n\n"
        "Emergency repair response times (LTSA Tier C — 97% Availability Guarantee):\n"
        "  Critical alert: 4-hour remote response, 24-hour on-site attendance\n"
        "  Major alert: 24-hour remote response, 72-hour on-site\n"
        "  Minor alert: 72-hour response (business hours)\n"
        "Response time = time to initial response/on-site attendance; resolution time is fault-dependent "
        "and not subject to the same commitments (OEM on-site: 5 business days).\n\n"
        "Loss from faults on TSOC side (not attributable to Bidder): Client's remedy is against TSOC per "
        "Transmission Rules; Bidder not liable. Force Majeure (grid failure, government action, etc.) "
        "excluded from Bidder liability per EPC Force Majeure clause.",
        FILL_DONE, "Lighthief (Esperia LTSA Tier C / standard EPC framework)"
    ),
    # Commercial Part
    20: (
        "H.E.S.S. Hybrid Energy Storage Systems Ltd\n"
        "Registration No: HE 439846\n"
        "Plot 26, Psevdas Community, Larnaca District, Cyprus\n"
        "CERA Storage Licence: KEA14-2024\n"
        "Represented by: Sotiris Shiakallis, Director, Extensive Proficient Services Ltd (EPS)",
        FILL_DONE, "Lighthief"
    ),
    21: (
        "Target on-site delivery: Q1 2027 (January 2027 manufacturing slot).\n"
        "Delivery terms per the Scope of Supply: DDP, Larnaca District (Plot 26, Psevdas site), "
        "Incoterms 2020 — duties and logistics included in the Bidder's price. "
        "Port of entry: Limassol. Final delivery date subject to executed EPC agreement.",
        FILL_DONE, "Lighthief (DDP Larnaca per client Scope of Supply)"
    ),
    22: (
        "Target EPC contract signature: end of June 2026 (Lighthief internal target), subject to "
        "completion of technical and commercial clarifications. "
        "A Letter of Intent (LOI) — ref. LCY-LOI-HESS-TRF-2026-R1 — has been issued by Lighthief "
        "to HESS to secure the January 2027 manufacturing slot for the HV transformer pending "
        "execution of the full EPC agreement. "
        "Bidder to note: manufacturing slot for BESS equipment (Linyang) and HV transformer "
        "must be confirmed by July 2026 at latest. Please confirm your slot availability.",
        FILL_DONE, "Lighthief (LOI issued; Sotiris to confirm final date)"
    ),
}


def main() -> None:
    shutil.copy2(SRC, OUT)
    wb = openpyxl.load_workbook(OUT)
    ws = wb.active

    for row_num, (answer, fill, owner) in ANSWERS.items():
        cell = ws.cell(row_num, 5)
        cell.value = answer
        cell.fill = fill
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        cell.font = Font(size=9)

    # Set column widths
    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 40
    ws.column_dimensions["D"].width = 40
    ws.column_dimensions["E"].width = 50

    # Auto-adjust row heights where possible
    for row in ws.iter_rows():
        for cell in row:
            if cell.value and len(str(cell.value)) > 100:
                ws.row_dimensions[cell.row].height = max(
                    ws.row_dimensions[cell.row].height or 15,
                    min(120, len(str(cell.value)) // 4)
                )

    wb.save(OUT)
    print(f"Saved -> {OUT}")

    # Print summary
    print("\n=== PREFILL SUMMARY ===\n")
    print("GREEN  — Lighthief can answer from known data:")
    for r, (ans, fill, owner) in ANSWERS.items():
        if fill == FILL_DONE:
            ws2 = openpyxl.load_workbook(OUT, data_only=True).active
            desc = str(ws2.cell(r, 3).value or "")[:60]
            print(f"  Row {r:2d}: {desc}...")

    print("\nAMBER  — Need HESS/Sotiris response:")
    for r, (ans, fill, owner) in ANSWERS.items():
        if fill == FILL_CLIENT:
            ws2 = openpyxl.load_workbook(OUT, data_only=True).active
            desc = str(ws2.cell(r, 3).value or ans)[:60]
            print(f"  Row {r:2d}: [{owner}] {desc}...")

    print("\nBLUE   — Need EPS/TSOC ruling:")
    for r, (ans, fill, owner) in ANSWERS.items():
        if fill == FILL_EPS:
            ws2 = openpyxl.load_workbook(OUT, data_only=True).active
            desc = str(ws2.cell(r, 3).value or ans)[:60]
            print(f"  Row {r:2d}: [{owner}] {desc}...")


if __name__ == "__main__":
    main()
