"""Prefill Polish producer PT questionnaire from HESS project data + pandapower study."""
from __future__ import annotations

import json
from pathlib import Path

import openpyxl

_HVT = Path(
    r"L:\My Drive\LINYANG\BESS CLIENTS\Individual_60-120-standalone\HV Transformer"
)
# Supplier-1 = Polish producer (form: PT Technical information). See _supplier-key.md.
SRC = _HVT / "Supplier-1" / "source" / "PT Technical information.xlsx"
OUT = _HVT / "Supplier-1" / "T1-main" / "PT-T1-main-63MVA-HESS-prefilled.xlsx"
OUT_FALLBACK = OUT.with_name("PT-T1-main-63MVA-HESS-prefilled-rev2.xlsx")
RESULTS_JSON = (
    Path(__file__).resolve().parent.parent
    / "docs"
    / "dso"
    / "analysis"
    / "hess-pandapower-results.json"
)
IK_HV_LABEL = "31.5 kA / 1 s (Client GDI Table 3, 'other substations'; confirm Psevdas not rated higher)"
IK_LV_LABEL = "20 kA / 3 s (system design level, GDI Table 3); calculated prospective ~5.9 kA (transformer + BESS PCS)"

PROJECT = "HESS — Power On BESS, Psevdas Larnaca (Hybrid Energy Storage Systems Ltd)"
STANDARDS = (
    "Tier 2 Ecodesign EU 2019/1783; EN 60076-2, EN 60076-10, EN 60076-5; "
    "IEC 61869; EN 60137; EN 60214; TSOC Transmission Rules incl. T14"
)
TEMP_LIMITS = (
    "Altitude ~264 m (≤1000 m); max air 50°C; design ambient 45°C; "
    "daily avg ≤30°C; annual avg ≤20°C; Mediterranean climate"
)
SEISMIC = "CYS EN 1998-1 Zone II, agR = 0.23 g (plot 26 — Zone I/II boundary; confirm with EPS civil)"
CREEPAGE = "≥35 (heavy pollution, composite per client spec)"

# IEC 60137 typical max bushing current ratings (A) — next standard above In
IEC_BUSHING_STD_A = (200, 250, 300, 400, 630, 800, 1000, 1250, 1600, 2000, 2500, 3150)


def next_std_bushing_a(in_a: float) -> int:
    for rating in IEC_BUSHING_STD_A:
        if rating >= in_a:
            return rating
    return IEC_BUSHING_STD_A[-1]


def load_results() -> dict:
    if not RESULTS_JSON.is_file():
        raise FileNotFoundError(f"Run scripts/hess-pandapower-protection.py first — missing {RESULTS_JSON}")
    return json.loads(RESULTS_JSON.read_text(encoding="utf-8"))


def set_by_label(ws, label: str, value: str, col: int = 4) -> None:
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
        if row[1].value and str(row[1].value).strip() == label:
            row[col - 1].value = value
            return
    raise KeyError(f"Label not found: {label!r}")


def set_by_prefix(ws, prefix: str, value: str, col: int = 4) -> None:
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
        if row[0].value and str(row[0].value).strip() == prefix:
            row[col - 1].value = value
            return
    raise KeyError(f"Prefix not found: {prefix!r}")


def set_bushing_sc(ws, section_prefix: str, ka: str) -> None:
    set_by_prefix(ws, f"{section_prefix}.15", ka)


def add_assumptions_sheet(
    wb: openpyxl.Workbook, ik_33: float, vr_design: float, vr_unity: float, vr_react: float
) -> None:
    if "Assumptions" in wb.sheetnames:
        del wb["Assumptions"]
    ws = wb.create_sheet("Assumptions", 0)
    rows = [
        ("HESS Psevdas — transformer questionnaire prefill (Jun 2026)", ""),
        ("Status key", "CONFIRMED = from client/POS docs | PROVISIONAL = Lighthief study, pending ISM"),
        ("", ""),
        ("CONFIRMED", "63 MVA, 132/33 kV, YNd11, OLTC, outdoor, Tier 2, qty 1"),
        ("CONFIRMED", "uk 21% per Requirements §17 (confirmed in client docs)"),
        ("CONFIRMED", "UGC 300 mm² XLPE 3×1c; no surge arresters on transformer (KYEA AIS)"),
        ("CONFIRMED", "Altitude 264 m; pollution Class III; composite ≥35 mm/kV"),
        ("CONFIRMED", "Step-up on EAE side of POS SLD (above DSMK/ISM boundary)"),
        ("CONFIRMED", "HV/LV bushing rated current: 300 A @ 132 kV, 1250 A @ 33 kV (IEC std above 275/1102 A)"),
        ("CONFIRMED", "Impact recorders: 2 off with GPS (Poland–Cyprus transport)"),
        ("NOTE — uk/V-reg", "uk=21% firmly specified (item 17). V-reg ~"
                      f"{vr_design}% at the specified 0.9 PF lag (item 1b), ~{vr_unity}% unity PF, ~{vr_react}% full-reactive. "
                      "Wide OLTC (+12.5/-18.75%) sized for it. Confirm 21% intended."),
        ("", ""),
        ("CONFIRMED — Ik HV", IK_HV_LABEL),
        ("CONFIRMED — Ik MV", IK_LV_LABEL),
        ("PROVISIONAL", "CT 400/1 @ 132 kV; 1600/1 @ 33 kV — 132 kV ALF per ISM bay; count may rise to 9 if HVN CT for REF"),
        ("PROVISIONAL", SEISMIC + "; importance factor 1.4 (essential facility)"),
        ("", ""),
        ("CONFIRMED — LV insulation", "36 kV class: 170 kV BIL / 70 kV AC (client clarification Jun 2026 — replaces erroneous 125/50)"),
        ("CONFIRMED — MV labels", "Cable/fault tables showing '22 kV' are legacy — LV system is 33 kV throughout"),
        ("CONFIRMED — earthing trf", "Solid earthing (no NER); zig-zag ≥20 kA / 3 s (or 25 kA / 1 s) per EN 60076-5"),
        ("CONFIRMED — aux winding", "≥315 kVA @ 400/230 V recommended minimum; final kVA TBC by designer station load calc"),
        ("PROVISIONAL — 132 kV protection", "Tender on standard TSO bay (distance, backup OC/EF, busbar, breaker fail, metering). "
                                            "Detailed ISM bay SLD from designer (Iacovos) — CT count/ALF may refine when issued"),
        ("PRODUCER SCOPE", "Dual WTI (HV+LV), online DGA monitoring, EN 60296 oil, MR/ABB vacuum OLTC, EN 795 anchorage, anti-vib pads"),
        ("PRODUCER OFFER", "Losses, PEI %, weights, dimensions, OLTC/bushing make — leave blank in offer column"),
    ]
    for r, (a, b) in enumerate(rows, 1):
        ws.cell(r, 1, a)
        ws.cell(r, 2, b)
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 90


def main() -> None:
    res = load_results()
    hv = res["protection_indicative"]["hv_132"]
    lv = res["protection_indicative"]["lv_33"]
    ik_33_design = round(res["short_circuit"]["33_kv_combined_ka"], 1)
    fault = res["assumptions"]
    vr_design = fault.get("voltage_regulation_pct_design_0p9pf", 0)
    vr_unity = fault.get("voltage_regulation_pct_unity_pf", 0)
    vr_react = fault.get("voltage_regulation_pct_full_reactive", 0)

    wb = openpyxl.load_workbook(SRC)
    gen = wb["PT-General"]
    tech = wb["PT-Technical"]

    add_assumptions_sheet(wb, ik_33_design, vr_design, vr_unity, vr_react)

    for ws in (gen, tech):
        ws["B7"] = PROJECT
        ws["B8"] = (
            "Jun 2026 — Lighthief prefill FOR CLIENT REVIEW. "
            "See Assumptions tab. Do not release to producer until ISM items confirmed."
        )
        ws["B9"] = "1"

    set_by_label(gen, "Delivery terms (FOB/CIF/DAP)", "DAP site, Plot 26 Psevdas, Larnaca, Cyprus")
    set_by_label(gen, "Supervision of Assembly and Commissioning (on-site)", "Yes — Lighthief EPC scope")
    set_by_label(gen, "Port location", "Limassol (Cyprus customs clearance)")
    set_by_label(gen, "Site Location", "Plot 26, Psevdas Community, Larnaca District, Cyprus")
    set_by_label(gen, "Expect On-site Delivery Time", "2027-Q1 on-site (Jan 2027 factory slot)")
    set_by_label(gen, "Equipment warranty period", "24")

    set_by_label(tech, "Type (Oil-immersed / Dry type)", "Oil-immersed")
    set_by_label(tech, "Installation Type (indoor / outdoor)", "Outdoor")
    set_by_label(tech, "Testing and manufacturing standards", STANDARDS)
    set_by_label(tech, "Altitude (above sea level)", "264")
    set_by_label(tech, "Limits of ambient Temperature", TEMP_LIMITS)
    set_by_label(tech, "Cooling", "ONAN / ONAF")
    set_by_label(tech, "Rated Frequency", "50")
    set_by_prefix(tech, "2.5.1", "63")
    set_by_prefix(tech, "2.5.3", "63")
    set_by_prefix(tech, "2.6.1", "132")
    set_by_prefix(tech, "2.6.2", "33")
    set_by_prefix(tech, "2.7.1", "145")
    set_by_prefix(tech, "2.7.2", "38")
    set_by_prefix(tech, "2.7.3", "36")
    set_by_prefix(tech, "2.8.1.1", "230")
    set_by_prefix(tech, "2.8.1.2", "38")
    set_by_prefix(tech, "2.8.1.3", "70")
    set_by_prefix(tech, "2.8.2.1", "550")
    set_by_prefix(tech, "2.8.2.2", "95")
    set_by_prefix(tech, "2.8.2.3", "170")
    set_by_label(tech, "Vector Group", "YNd11")
    set_by_label(tech, "Tap changer type", "On-load (OLTC)")
    set_by_label(tech, "Peak Efficiency Index (PEI)", "Tier 2 Ecodesign — state PEI % in offer")
    set_by_prefix(tech, "2.13.1", "≤60 dB(A) outdoor (Requirements §19; indoor ≤50 N/A)")
    set_by_prefix(tech, "2.13.2", "≤80")
    set_by_prefix(tech, "2.14.1", "21 — CONFIRMED per Requirements §17 (client to confirm mandatory vs 18–20%)")
    set_by_prefix(tech, "2.15.1", "55")
    set_by_prefix(tech, "2.15.2", "60")
    set_by_prefix(tech, "2.15.3", "73")
    set_by_prefix(tech, "2.16.1", "230 Vac panel; 110 Vac 50 Hz AVR reference")
    set_by_prefix(tech, "2.16.2", "110 Vdc control / supervisory")
    set_by_label(tech, "Location (HV side / LV side)", "HV side")
    set_by_label(tech, "Regulation Range", "+12.5% / −18.75% on nominal 132 kV")
    set_by_label(tech, "Number of steps", "25 @ 1.25% per step")
    set_by_label(tech, "Motor Drive mechanism", "On-load local, remote and supervisory control")

    for prefix in ("4.1", "4.2", "4.3"):
        set_by_prefix(tech, f"{prefix}.7", "Composite (silicone)")
        set_by_prefix(tech, f"{prefix}.10", "145" if prefix != "4.3" else "36")
        set_by_prefix(tech, f"{prefix}.13", CREEPAGE)

    hv_bushing_a = next_std_bushing_a(hv["in_a"])
    lv_bushing_a = next_std_bushing_a(lv["in_a"])
    set_by_prefix(tech, "4.1.9", str(hv_bushing_a))
    set_by_prefix(tech, "4.2.9", str(hv_bushing_a))
    set_by_prefix(tech, "4.3.9", str(lv_bushing_a))
    set_bushing_sc(tech, "4.1", IK_HV_LABEL)
    set_bushing_sc(tech, "4.2", IK_HV_LABEL)
    set_bushing_sc(tech, "4.3", IK_LV_LABEL)

    ct_spec = (
        f"132 kV: {hv['ct_ratio']} A — {hv['ct_cores']} [ALF TBC — ISM connection bay]; "
        f"33 kV: {lv['ct_ratio']} A — {lv['ct_cores']}"
    )
    set_by_prefix(
        tech,
        "5.1",
        "6 (3 per phase × 2 windings); +3 if ISM requires HVN CT for REF → 9 total",
    )
    set_by_label(tech, "Transformation ratio", ct_spec)
    set_by_label(tech, "Burden", "30 VA (metering) / 20 VA (protection)")
    set_by_label(tech, "Accuracy class", "0.2 (metering) / 5P20 (protection)")
    set_by_label(tech, "Safety Factor (Fs) / Accuracy limit factor (n)", "ALF 20 (5P20 cores)")

    set_by_prefix(tech, "6.10.1", "Yes")
    set_by_prefix(tech, "6.10.2", "Yes")
    set_by_prefix(tech, "6.10.5", "2 (50% CMR each)")
    set_by_prefix(tech, "6.10.7", "Yes — 2 off with GPS (one per major transport lift)")
    set_by_prefix(tech, "6.10.8", "Yes")
    set_by_prefix(tech, "6.10.9", "Yes")
    set_by_prefix(tech, "6.10.10", "Yes")
    set_by_prefix(tech, "6.10.12", "IP55 marshalling kiosk")
    set_by_prefix(tech, "6.10.13", "No — surge arresters in KYEA AIS scope")

    try:
        wb.save(OUT)
        print(f"Wrote {OUT}")
    except PermissionError:
        wb.save(OUT_FALLBACK)
        print(f"Wrote {OUT_FALLBACK} (close open workbook to overwrite {OUT.name})")


if __name__ == "__main__":
    main()
