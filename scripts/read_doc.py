#!/usr/bin/env python3
"""
read_doc.py — universal document reader for scanned PDFs and Excel files.
Uses PyMuPDF for text PDFs, pdf2image+pytesseract for scanned PDFs,
and openpyxl for Excel files.

Usage:
  python3 scripts/read_doc.py "path/to/file.pdf"
  python3 scripts/read_doc.py "path/to/file.xlsx"
  python3 scripts/read_doc.py "path/to/file.pdf" --pages 1-3
  python3 scripts/read_doc.py "path/to/file.pdf" --out output.txt
"""

import sys
import os
import argparse

def read_pdf(path: str, pages: list[int] | None = None) -> str:
    import fitz  # PyMuPDF

    doc = fitz.open(path)
    total = doc.page_count
    target_pages = pages if pages else list(range(total))

    # Try text extraction first
    text_blocks = []
    for i in target_pages:
        if i >= total:
            continue
        page = doc[i]
        text = page.get_text().strip()
        text_blocks.append((i + 1, text))

    total_text = "\n".join(t for _, t in text_blocks)
    if len(total_text.strip()) > 50:
        # Text-based PDF — good extraction
        lines = []
        for page_num, text in text_blocks:
            lines.append(f"\n--- Page {page_num} ---\n{text}")
        return "\n".join(lines)

    # Scanned PDF — fall back to OCR
    print(f"  [OCR mode — scanned PDF, {total} pages]", file=sys.stderr)
    from pdf2image import convert_from_path
    import pytesseract

    result = []
    images = convert_from_path(path, dpi=300)
    for i, img in enumerate(images):
        page_num = i + 1
        if pages and i not in pages:
            continue
        print(f"  OCR page {page_num}/{len(images)}...", file=sys.stderr)
        text = pytesseract.image_to_string(img, lang="eng+ell")
        result.append(f"\n--- Page {page_num} ---\n{text.strip()}")

    return "\n".join(result)


def read_excel(path: str) -> str:
    import openpyxl

    wb = openpyxl.load_workbook(path, data_only=True)
    output = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        output.append(f"\n=== Sheet: {sheet_name} ===")
        for row in ws.iter_rows(values_only=True):
            # Skip completely empty rows
            if all(v is None for v in row):
                continue
            formatted = "\t".join("" if v is None else str(v) for v in row)
            output.append(formatted)
    return "\n".join(output)


def parse_pages(page_str: str) -> list[int]:
    """Parse '1-3,5,7' into [0,1,2,4,6] (0-indexed)."""
    indices = []
    for part in page_str.split(","):
        part = part.strip()
        if "-" in part:
            start, end = part.split("-")
            indices.extend(range(int(start) - 1, int(end)))
        else:
            indices.append(int(part) - 1)
    return indices


def main():
    parser = argparse.ArgumentParser(description="Read PDF or Excel files including scanned/OCR PDFs.")
    parser.add_argument("file", help="Path to PDF or Excel file")
    parser.add_argument("--pages", help="Pages to extract, e.g. '1-3,5' (PDF only)", default=None)
    parser.add_argument("--out", help="Output file path (default: stdout)", default=None)
    args = parser.parse_args()

    path = args.file
    if not os.path.exists(path):
        print(f"Error: file not found: {path}", file=sys.stderr)
        sys.exit(1)

    ext = os.path.splitext(path)[1].lower()
    pages = parse_pages(args.pages) if args.pages else None

    print(f"Reading: {os.path.basename(path)}", file=sys.stderr)

    if ext in (".pdf",):
        content = read_pdf(path, pages)
    elif ext in (".xlsx", ".xls", ".xlsm"):
        content = read_excel(path)
    else:
        print(f"Error: unsupported file type '{ext}'. Supported: .pdf, .xlsx, .xls", file=sys.stderr)
        sys.exit(1)

    if args.out:
        with open(args.out, "w", encoding="utf-8") as f:
            f.write(content)
        print(f"Written to: {args.out}", file=sys.stderr)
    else:
        print(content)


if __name__ == "__main__":
    main()
