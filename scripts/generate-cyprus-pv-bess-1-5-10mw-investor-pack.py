#!/usr/bin/env python3
"""
Cyprus PV + 4h BESS Investor Pack — 1 / 5 / 10 MW
=================================================
Two PV configurations per size:
  A) Fixed Bifacial                        — 1,850 kWh/kWp (client-tested)
  B) Single-Axis Tracker + Bifacial + White Albedo — 2,450 kWh/kWp (client-tested)

Outputs:
  1. Excel financial model with yellow INPUT cells (modifiable)
  2. Client-facing HTML one-pager (A4 print)

SSOT sources (all figures codebase-grounded — do not invent):
  lib/deals/rtb-deal-types.ts          — DAM daytime €140.88, RTE 86.32%, capture 87.4%,
                                          curtailment 50%, PV EPC €720k/MWp, BESS EPC €127k/MWh
  lib/portfolio-data.ts                — LTSA Tier C €1,740/MWh/yr
  lib/constants.ts                     — COMPANY_DATA, RTB €350k/MW
  docs/internal/konia-epc-cost-model.md — tracker adder +€90k/MW; tracker vs fixed O&M
  market/data/market-data.json         — evening discharge €210/MWh (18:00–21:00 avg,
                                          Oct 2025–Jul 2026, deduped half-hourly MCP)

Revenue model: merchant hybrid (curtailment recovery)
  Uncurtailed PV → DAM daytime  |  Curtailed PV → BESS (€0 charge) → evening discharge
  BESS cannot yet buy from DAM in Cyprus — capacity / ancillary = €0 in base case.

Run: python3 scripts/generate-cyprus-pv-bess-1-5-10mw-investor-pack.py
"""

from __future__ import annotations

import base64
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ── paths ─────────────────────────────────────────────────────────────────────
ROOT = Path(__file__).resolve().parent.parent
OUT_DIR = ROOT / "docs" / "clients" / "Investor_Generic_PV_BESS_1_5_10MW"
OUT_DIR.mkdir(parents=True, exist_ok=True)
TEASER_DIR = ROOT / "docs" / "teasers"
TEASER_DIR.mkdir(parents=True, exist_ok=True)
PUBLIC_DIR = ROOT / "public" / "lighthief-cyprus" / "teasers" / "cyprus-pv-bess-1-5-10mw"
PUBLIC_DIR.mkdir(parents=True, exist_ok=True)

XLSX_OUT = OUT_DIR / "cyprus-pv-bess-1-5-10mw-financial-model.xlsx"
HTML_OUT = OUT_DIR / "cyprus-pv-bess-1-5-10mw-investor-onepager.html"
HTML_TEASER = TEASER_DIR / "cyprus-pv-bess-1-5-10mw-investor-onepager.html"
HTML_PUBLIC = PUBLIC_DIR / "cyprus-pv-bess-1-5-10mw-investor-onepager.html"

LOGO_PATH = ROOT / "public" / "images" / "logo" / "lighthief-logo-200.png"
if not LOGO_PATH.exists():
    LOGO_PATH = ROOT / "public" / "images" / "lighthief-logo.png"

# ── brand ─────────────────────────────────────────────────────────────────────
NAVY = "1A365D"
GOLD = "C9A432"
WHITE = "FFFFFF"
GREY = "404040"
INPUT_HEX = "FFF2CC"
CALC_HEX = "E2EFDA"
HDR_FILL = PatternFill("solid", fgColor=NAVY)
INPUT_FILL = PatternFill("solid", fgColor=INPUT_HEX)
CALC_FILL = PatternFill("solid", fgColor=CALC_HEX)
TOTAL_FILL = PatternFill("solid", fgColor="D6E4F0")
thin = Border(
    left=Side(style="thin", color="BFBFBF"),
    right=Side(style="thin", color="BFBFBF"),
    top=Side(style="thin", color="BFBFBF"),
    bottom=Side(style="thin", color="BFBFBF"),
)

# ── SSOT defaults (mirrored into Inputs sheet) ────────────────────────────────
DEFAULTS = {
    # Technical (shared)
    "curtailment_pct": 0.50,        # BESS_DEFAULTS.curtailmentPct
    "capture_pct": 0.874,           # BESS_DEFAULTS.capturePct (Galascope 365-day)
    "rte": 0.8632,                  # BESS_DEFAULTS.rteAcAc
    "full_cycle_days": 280,         # BESS_DEFAULTS.fullCycleDaysPerYear
    "duration_h": 4,
    # Prices
    "dam_day": 140.88,              # DAM.daytimeEURPerMWh (06–17h). Full 9.5m dataset = €123.37
    "dam_evening": 210.0,           # market-data.json 18:00–21:00 avg, Oct'25–Jul'26 (deduped)
    # CAPEX unit rates
    "bess_epc_per_mwh": 127_000,    # LH_EPC.bessPerMWh (official SSOT)
    "rtb_per_mw": 350_000,          # constants INVESTMENT_SIZES rtbCost
    "connection_per_mw": 0,         # EAC terms — site-specific (flexible teaser often €80k/MW)
    "other_capex": 0,               # MV cable / permitting / contingency
    # OPEX
    "bess_ltsa_per_mwh": 1_740,     # LTSA.tierC.ratePerMWh
    "scada_pa": 5_000,
    "admin_pa": 10_000,
    "land_lease_1": 8_000,
    "land_lease_5": 25_000,
    "land_lease_10": 45_000,
    "ins_pct": 0.005,
    # Finance
    "aggregator_pct": 0.10,         # PV_DEFAULTS.aggregatorFeePct
    "cit_pct": 0.15,                # PV_DEFAULTS.citPct
    "ltv": 0.0,                     # base = 100% equity; try 0.65 for levered case
    "loan_rate": 0.055,
    "loan_years": 12,
    "da_epc_years": 20,
    "da_dev_years": 15,
    "horizon_years": 20,
    "pv_degradation": 0.005,
    "bess_degradation": 0.025,
    # Extended warranty (disclosure — paid to OEM from Yr 6)
    "ext_war_6_10": 1_661.68,
    "ext_war_11_15": 2_083.72,
}

# PV configurations — yield & EPC & O&M are config-specific, all codebase-grounded
PV_CONFIGS = {
    "fixed": {
        "label": "Fixed Bifacial",
        "short": "Fixed",
        "yield": 1850,              # client-tested Year-1
        "pv_epc": 720_000,          # LH_EPC.pvPerMWp
        "pv_om": 15_000,            # PV_DEFAULTS.omPerMWPerYear
        "note": "Fixed-tilt bifacial (1,850 kWh/kWp)",
    },
    "tracker": {
        "label": "Single-Axis Tracker · Bifacial · White Albedo",
        "short": "Tracker",
        "yield": 2450,              # client-tested Year-1
        "pv_epc": 810_000,          # €720k + €90k/MW tracker adder (konia-epc-cost-model Rev C)
        "pv_om": 20_000,            # €15k × Konia tracker/fixed O&M ratio (8,900/6,600 ≈ 1.35)
        "note": "Single-axis tracker, bifacial, white albedo (2,450 kWh/kWp)",
    },
}

SIZES = [
    {"key": "1MW", "label": "1 MW + 4 MWh", "mw": 1.0, "mwh": 4.0, "land_key": "land_lease_1"},
    {"key": "5MW", "label": "5 MW + 20 MWh", "mw": 5.0, "mwh": 20.0, "land_key": "land_lease_5"},
    {"key": "10MW", "label": "10 MW + 40 MWh", "mw": 10.0, "mwh": 40.0, "land_key": "land_lease_10"},
]

CONTACT = {
    "director": "Alexander Papacosta",
    "title": "Cyprus Director",
    "phone": "+357 99 164 158",
    "email": "office@lighthief.com",
    "company": "Lighthief Cyprus Ltd",
    "reg": "HE 477423",
    "website": "solarfarms.cy",
    "office_phone": "+357 77 77 00 50",
    "address": "15 Agaritsis, Nektaria Court, Office 201, 3045 Zakaki, Limassol, Cyprus",
}

REF = "INV-CY-PVBESS-1-5-10-2026"
DATE_STR = "July 2026"


def annuity_factor(rate: float, years: int) -> float:
    if rate <= 0 or years <= 0:
        return 0.0
    return rate * (1 + rate) ** years / ((1 + rate) ** years - 1)


def calc_scenario(size: dict, cfg: dict, d: dict) -> dict:
    mw, mwh = size["mw"], size["mwh"]
    land = d[size["land_key"]]

    pv_epc = mw * cfg["pv_epc"]
    bess_epc = mwh * d["bess_epc_per_mwh"]
    rtb = mw * d["rtb_per_mw"]
    conn = mw * d["connection_per_mw"]
    other = d["other_capex"]
    capex = pv_epc + bess_epc + rtb + conn + other

    annual = mw * cfg["yield"]
    uncurt = annual * (1 - d["curtailment_pct"])
    curt = annual * d["curtailment_pct"]
    charged = min(curt * d["capture_pct"], mwh * d["full_cycle_days"])
    discharged = charged * d["rte"]

    solar_rev = uncurt * d["dam_day"]
    bess_rev = discharged * d["dam_evening"]
    gross = solar_rev + bess_rev
    agg = gross * d["aggregator_pct"]
    net_rev = gross - agg

    pv_om = mw * cfg["pv_om"]
    bess_om = mwh * d["bess_ltsa_per_mwh"]
    insurance = capex * d["ins_pct"]
    opex = pv_om + bess_om + d["scada_pa"] + d["admin_pa"] + land + insurance
    ebitda = net_rev - opex

    debt = capex * d["ltv"]
    equity = capex - debt
    svc = debt * annuity_factor(d["loan_rate"], d["loan_years"]) if debt else 0.0

    da = (pv_epc + bess_epc) / d["da_epc_years"] + (rtb + conn) / d["da_dev_years"]
    taxable = max(0.0, ebitda - svc - da)
    tax = taxable * d["cit_pct"]
    fcf = ebitda - svc - tax

    cashflows = [-equity]
    for y in range(1, int(d["horizon_years"]) + 1):
        deg_pv = (1 - d["pv_degradation"]) ** (y - 1)
        deg_b = (1 - d["bess_degradation"]) ** (y - 1)
        g = solar_rev * deg_pv + bess_rev * min(deg_pv, deg_b)
        n = g * (1 - d["aggregator_pct"])
        o = opex
        if 6 <= y <= 10:
            o += mwh * d["ext_war_6_10"]
        elif y >= 11:
            o += mwh * d["ext_war_11_15"]
        e = n - o
        svc_y = svc if y <= d["loan_years"] else 0.0
        tax_y = max(0.0, e - svc_y - da) * d["cit_pct"]
        cashflows.append(e - svc_y - tax_y)

    npv = npv_at(cashflows, 0.08)
    irr = irr_newton(cashflows)
    payback_cum = payback_years(cashflows)
    simple_payback = (equity / fcf) if fcf > 0 else None
    cash_yield = fcf / equity if equity else 0.0

    return {
        "size_key": size["key"],
        "cfg_key": "fixed" if cfg is PV_CONFIGS["fixed"] else "tracker",
        "cfg_short": cfg["short"],
        "cfg_label": cfg["label"],
        "label": size["label"],
        "mw": mw,
        "mwh": mwh,
        "yield": cfg["yield"],
        "pv_epc": pv_epc,
        "bess_epc": bess_epc,
        "rtb": rtb,
        "conn": conn,
        "capex": capex,
        "equity": equity,
        "debt": debt,
        "annual_mwh": annual,
        "uncurt_mwh": uncurt,
        "discharged_mwh": discharged,
        "solar_rev": solar_rev,
        "bess_rev": bess_rev,
        "gross": gross,
        "net_rev": net_rev,
        "pv_om": pv_om,
        "bess_om": bess_om,
        "opex": opex,
        "ebitda": ebitda,
        "svc": svc,
        "da": da,
        "tax": tax,
        "fcf": fcf,
        "cash_yield": cash_yield,
        "payback": simple_payback,
        "payback_cum": payback_cum,
        "irr": irr,
        "npv8": npv,
        "cashflows": cashflows,
        "dscr": (ebitda / svc) if svc else None,
    }


def npv_at(cfs: list[float], rate: float) -> float:
    return sum(cf / ((1 + rate) ** t) for t, cf in enumerate(cfs))


def irr_newton(cfs: list[float], guess: float = 0.10) -> float | None:
    r = guess
    for _ in range(100):
        npv = 0.0
        deriv = 0.0
        for t, cf in enumerate(cfs):
            npv += cf / ((1 + r) ** t)
            if t:
                deriv -= t * cf / ((1 + r) ** (t + 1))
        if abs(deriv) < 1e-12:
            return None
        r2 = r - npv / deriv
        if abs(r2 - r) < 1e-8:
            return r2 if -0.99 < r2 < 5 else None
        r = r2
    return None


def payback_years(cfs: list[float]) -> float | None:
    cum = 0.0
    for t, cf in enumerate(cfs):
        prev = cum
        cum += cf
        if t > 0 and cum >= 0:
            if cf == 0:
                return float(t)
            return (t - 1) + (-prev / cf)
    return None


def eur(n: float) -> str:
    if abs(n) >= 1_000_000:
        return f"€{n/1e6:.2f}M"
    if abs(n) >= 10_000:
        return f"€{n/1_000:,.0f}k"
    return f"€{n:,.0f}"


def pct(n: float | None) -> str:
    return "n/a" if n is None else f"{n*100:.1f}%"


# ═══════════════════════════════════════════════════════════════════════════════
# EXCEL
# ═══════════════════════════════════════════════════════════════════════════════

def _font(bold=False, size=11, color="000000"):
    return Font(name="Calibri", bold=bold, size=size, color=color)


def _input_cell(ws, row, label, value, unit, source, name_map=None, key=None, is_pct=False):
    ws.cell(row=row, column=1, value=label).font = _font()
    cell = ws.cell(row=row, column=2, value=value)
    cell.fill = INPUT_FILL
    cell.font = _font(bold=True, color="9C0006")
    cell.border = thin
    cell.alignment = Alignment(horizontal="right")
    cell.number_format = "0.00%" if is_pct else "#,##0.00"
    ws.cell(row=row, column=3, value=unit).font = _font(size=10, color=GREY)
    ws.cell(row=row, column=4, value=source).font = _font(size=9, color=GREY)
    if name_map is not None and key:
        name_map[key] = row
    return cell


def build_excel(results: list[dict], d: dict) -> None:
    wb = Workbook()
    name_row: dict[str, int] = {}

    # ── Inputs ───────────────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Inputs"
    ws["A1"] = "Cyprus PV + 4h BESS — Assumptions (YELLOW = editable)"
    ws["A1"].font = Font(name="Calibri", bold=True, size=14, color=GOLD)
    ws["A2"] = f"{REF} · {DATE_STR} · Lighthief Cyprus Ltd · Non-binding indicative model"
    ws["A2"].font = _font(size=10, color=GREY)
    ws["A3"] = "Edit any yellow cell — Detail_* sheets recalc via live formulas. Re-run script to refresh Summary/Cashflow."
    ws["A3"].font = _font(size=10, color=NAVY)

    for col, h in enumerate(["Parameter", "Value", "Unit", "Source / Note"], 1):
        c = ws.cell(row=5, column=col, value=h)
        c.fill = HDR_FILL
        c.font = _font(bold=True, color=WHITE)
        c.border = thin

    r = 6

    def section(title):
        nonlocal r
        ws.cell(row=r, column=1, value=title).font = Font(name="Calibri", bold=True, size=11, color=GOLD)
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
        r += 1

    fx, tk = PV_CONFIGS["fixed"], PV_CONFIGS["tracker"]

    section("PV CONFIGURATION A — FIXED BIFACIAL")
    _input_cell(ws, r, "Fixed — specific yield", fx["yield"], "kWh/kWp/yr", "Client-tested Year-1", name_row, "yield_fixed"); r += 1
    _input_cell(ws, r, "Fixed — PV EPC turnkey", fx["pv_epc"], "€/MWp", "LH_EPC.pvPerMWp", name_row, "pvepc_fixed"); r += 1
    _input_cell(ws, r, "Fixed — PV O&M", fx["pv_om"], "€/MW/yr", "PV_DEFAULTS.omPerMWPerYear", name_row, "pvom_fixed"); r += 1

    section("PV CONFIGURATION B — TRACKER + BIFACIAL + WHITE ALBEDO")
    _input_cell(ws, r, "Tracker — specific yield", tk["yield"], "kWh/kWp/yr", "Client-tested Year-1", name_row, "yield_tracker"); r += 1
    _input_cell(ws, r, "Tracker — PV EPC turnkey", tk["pv_epc"], "€/MWp", "€720k + €90k/MW tracker adder (konia-epc-cost-model)", name_row, "pvepc_tracker"); r += 1
    _input_cell(ws, r, "Tracker — PV O&M", tk["pv_om"], "€/MW/yr", "€15k × Konia tracker/fixed O&M ratio", name_row, "pvom_tracker"); r += 1

    section("BESS & TECHNICAL")
    _input_cell(ws, r, "BESS duration", d["duration_h"], "hours", "4-hour system (MW = MWh/4)", name_row, "duration_h"); r += 1
    _input_cell(ws, r, "Curtailment baseline", d["curtailment_pct"], "%", "BESS_DEFAULTS 2027 baseline", name_row, "curtailment_pct", is_pct=True); r += 1
    _input_cell(ws, r, "BESS capture of curtailment", d["capture_pct"], "%", "Galascope 365-day 87.4%", name_row, "capture_pct", is_pct=True); r += 1
    _input_cell(ws, r, "RTE AC–AC", d["rte"], "%", "BESS_DEFAULTS.rteAcAc", name_row, "rte", is_pct=True); r += 1
    _input_cell(ws, r, "Full-cycle days / year", d["full_cycle_days"], "days", "Curtailment-active days Cyprus", name_row, "full_cycle_days"); r += 1

    section("PRICES (market data)")
    _input_cell(ws, r, "DAM daytime (uncurtailed PV)", d["dam_day"], "€/MWh", "DAM.daytimeEURPerMWh 06–17h; full dataset €123.37", name_row, "dam_day"); r += 1
    _input_cell(ws, r, "Evening discharge (BESS)", d["dam_evening"], "€/MWh", "market-data.json 18–21h avg Oct'25–Jul'26; winter €184", name_row, "dam_evening"); r += 1

    section("CAPEX (shared) & OPEX")
    _input_cell(ws, r, "BESS EPC installed", d["bess_epc_per_mwh"], "€/MWh", "LH_EPC.bessPerMWh", name_row, "bess_epc"); r += 1
    _input_cell(ws, r, "RTB / development", d["rtb_per_mw"], "€/MW", "constants INVESTMENT_SIZES", name_row, "rtb_per_mw"); r += 1
    _input_cell(ws, r, "EAC connection terms", d["connection_per_mw"], "€/MW", "Site-specific; flexible teaser €80k/MW", name_row, "connection_per_mw"); r += 1
    _input_cell(ws, r, "Other CAPEX (flat)", d["other_capex"], "€", "MV cable / permitting / contingency", name_row, "other_capex"); r += 1
    _input_cell(ws, r, "BESS LTSA Tier C (Yr 1–5)", d["bess_ltsa_per_mwh"], "€/MWh/yr", "LTSA.tierC.ratePerMWh", name_row, "ltsa"); r += 1
    _input_cell(ws, r, "SCADA / EMS", d["scada_pa"], "€/yr", "", name_row, "scada"); r += 1
    _input_cell(ws, r, "Admin", d["admin_pa"], "€/yr", "", name_row, "admin"); r += 1
    _input_cell(ws, r, "Land lease — 1 MW", d["land_lease_1"], "€/yr", "Indicative; site-specific", name_row, "land_lease_1"); r += 1
    _input_cell(ws, r, "Land lease — 5 MW", d["land_lease_5"], "€/yr", "Indicative; site-specific", name_row, "land_lease_5"); r += 1
    _input_cell(ws, r, "Land lease — 10 MW", d["land_lease_10"], "€/yr", "Indicative; site-specific", name_row, "land_lease_10"); r += 1
    _input_cell(ws, r, "Insurance", d["ins_pct"], "% of CAPEX", "0.5% CAPEX", name_row, "ins_pct", is_pct=True); r += 1

    section("FINANCE")
    _input_cell(ws, r, "Aggregator / market access", d["aggregator_pct"], "%", "PV_DEFAULTS.aggregatorFeePct", name_row, "aggregator_pct", is_pct=True); r += 1
    _input_cell(ws, r, "Cyprus CIT", d["cit_pct"], "%", "From 1 Jan 2026", name_row, "cit_pct", is_pct=True); r += 1
    _input_cell(ws, r, "LTV (senior debt)", d["ltv"], "% of CAPEX", "0 = 100% equity base; try 65%", name_row, "ltv", is_pct=True); r += 1
    _input_cell(ws, r, "Loan interest rate", d["loan_rate"], "%", "Indicative Cyprus commercial", name_row, "loan_rate", is_pct=True); r += 1
    _input_cell(ws, r, "Loan tenor", d["loan_years"], "years", "", name_row, "loan_years"); r += 1
    _input_cell(ws, r, "DA — EPC years", d["da_epc_years"], "years", "Tax depreciation", name_row, "da_epc_years"); r += 1
    _input_cell(ws, r, "DA — development years", d["da_dev_years"], "years", "Tax depreciation", name_row, "da_dev_years"); r += 1

    section("SCENARIO SIZES (editable)")
    _input_cell(ws, r, "1 MW — PV capacity", 1.0, "MWp", "", name_row, "mw_1"); r += 1
    _input_cell(ws, r, "1 MW — BESS energy", 4.0, "MWh", "4h → power = MWh/4", name_row, "mwh_1"); r += 1
    _input_cell(ws, r, "5 MW — PV capacity", 5.0, "MWp", "", name_row, "mw_5"); r += 1
    _input_cell(ws, r, "5 MW — BESS energy", 20.0, "MWh", "", name_row, "mwh_5"); r += 1
    _input_cell(ws, r, "10 MW — PV capacity", 10.0, "MWp", "", name_row, "mw_10"); r += 1
    _input_cell(ws, r, "10 MW — BESS energy", 40.0, "MWh", "", name_row, "mwh_10"); r += 1

    r += 1
    ws.cell(row=r, column=1, value="Legend:").font = _font(bold=True, color=NAVY)
    ws.cell(row=r, column=2, value="Yellow = editable input").fill = INPUT_FILL
    ws.cell(row=r, column=3, value="Green = calculated").fill = CALC_FILL

    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 13
    ws.column_dimensions["C"].width = 13
    ws.column_dimensions["D"].width = 60
    ws.freeze_panes = "A6"

    def ref(key: str) -> str:
        return f"Inputs!$B${name_row[key]}"

    # ── Summary ──────────────────────────────────────────────────────────────
    sm = wb.create_sheet("Summary")
    sm["A1"] = "Indicative Comparison — 1 / 5 / 10 MW PV + 4-hour BESS (Cyprus merchant hybrid)"
    sm["A1"].font = Font(name="Calibri", bold=True, size=14, color=GOLD)
    sm["A2"] = "Base case = 100% equity. Evening discharge from live market data. Two PV configs."
    sm["A2"].font = _font(size=10, color=GREY)

    metrics = [
        ("PV specific yield (kWh/kWp)", "yield", "0"),
        ("BESS energy (MWh)", "mwh", "0"),
        ("Total CAPEX (€)", "capex", "#,##0"),
        ("  PV EPC", "pv_epc", "#,##0"),
        ("  BESS EPC", "bess_epc", "#,##0"),
        ("  RTB / development", "rtb", "#,##0"),
        ("Annual PV production (MWh)", "annual_mwh", "#,##0"),
        ("Gross Y1 revenue (€)", "gross", "#,##0"),
        ("  Uncurtailed solar", "solar_rev", "#,##0"),
        ("  BESS discharge", "bess_rev", "#,##0"),
        ("OPEX Y1 (€)", "opex", "#,##0"),
        ("EBITDA Y1 (€)", "ebitda", "#,##0"),
        ("After-tax FCF Y1 (€)", "fcf", "#,##0"),
        ("Cash yield on equity Y1", "cash_yield", "0.0%"),
        ("Simple equity payback (yr)", "payback", "0.0"),
        ("Cumulative payback w/ deg. (yr)", "payback_cum", "0.0"),
        ("Equity IRR 20y (deg.+OEM war.)", "irr", "0.0%"),
        ("NPV @ 8% (€)", "npv8", "#,##0"),
    ]

    row0 = 4
    for cfg_key in ("fixed", "tracker"):
        block = [r for r in results if r["cfg_key"] == cfg_key]
        cfg = PV_CONFIGS[cfg_key]
        sm.cell(row=row0, column=1, value=f"■ Configuration: {cfg['label']}").font = Font(
            name="Calibri", bold=True, size=12, color=NAVY)
        sm.merge_cells(start_row=row0, start_column=1, end_row=row0, end_column=4)
        hdr = row0 + 1
        sm.cell(row=hdr, column=1, value="Metric").fill = HDR_FILL
        sm.cell(row=hdr, column=1).font = _font(bold=True, color=WHITE)
        for j, r_ in enumerate(block):
            c = sm.cell(row=hdr, column=2 + j, value=r_["label"])
            c.fill = HDR_FILL
            c.font = _font(bold=True, color=WHITE)
            c.alignment = Alignment(horizontal="center")
        for i, (label, key, fmt) in enumerate(metrics):
            rr = hdr + 1 + i
            sm.cell(row=rr, column=1, value=label).border = thin
            for j, r_ in enumerate(block):
                cell = sm.cell(row=rr, column=2 + j)
                cell.border = thin
                cell.fill = CALC_FILL
                v = r_[key]
                cell.value = v if v is not None else "n/a"
                if v is not None:
                    cell.number_format = fmt
                if key in {"capex", "fcf", "cash_yield", "irr"}:
                    cell.font = _font(bold=True, color=NAVY)
        row0 = hdr + 1 + len(metrics) + 2

    sm.column_dimensions["A"].width = 32
    for col in range(2, 5):
        sm.column_dimensions[get_column_letter(col)].width = 17

    nr = row0
    sm.cell(row=nr, column=1, value="Notes").font = Font(name="Calibri", bold=True, size=11, color=GOLD)
    notes = [
        "Revenue = uncurtailed PV at DAM daytime + curtailed energy stored (charge €0) and discharged at evening price.",
        "Evening discharge €210/MWh = 18:00–21:00 average from market/data/market-data.json (Oct 2025–Jul 2026, deduped).",
        "  Seasonality: winter (Oct–Feb) €179–194, spring (Mar–May) €240–260. Winter-only ≈ €184 as conservative sensitivity.",
        "Fixed PV EPC €720k/MWp (LH_EPC). Tracker €810k/MWp = €720k + €90k/MW adder (konia-epc-cost-model).",
        "BESS EPC €127k/MWh (LH_EPC). LTSA €1,740/MWh/yr (portfolio-data). PV O&M €15k fixed / €20k tracker per MW/yr.",
        "Cyprus rule: BESS cannot yet buy from DAM — no arbitrage. Capacity/ancillary = €0 (upside).",
        "20y IRR/NPV include PV & BESS degradation and OEM extended warranty from Year 6. Indicative, non-binding, ex-VAT.",
        f"Contact: {CONTACT['director']}, {CONTACT['title']} · {CONTACT['phone']} · {CONTACT['email']}",
    ]
    for i, n in enumerate(notes):
        sm.cell(row=nr + 1 + i, column=1, value=n).font = _font(size=9, color=GREY)
        sm.merge_cells(start_row=nr + 1 + i, start_column=1, end_row=nr + 1 + i, end_column=4)

    # ── Detail sheets (live formulas) — one per size × config ────────────────
    size_keys = {"1MW": ("mw_1", "mwh_1", "land_lease_1"),
                 "5MW": ("mw_5", "mwh_5", "land_lease_5"),
                 "10MW": ("mw_10", "mwh_10", "land_lease_10")}

    for r_ in results:
        sk = r_["size_key"]
        ck = r_["cfg_key"]
        mw_k, mwh_k, land_k = size_keys[sk]
        yk = f"yield_{ck}"
        pvepc_k = f"pvepc_{ck}"
        pvom_k = f"pvom_{ck}"

        ws = wb.create_sheet(f"D_{sk}_{PV_CONFIGS[ck]['short']}")
        ws["A1"] = f"{r_['label']} · {PV_CONFIGS[ck]['label']} — live formulas (reads Inputs)"
        ws["A1"].font = Font(name="Calibri", bold=True, size=12, color=GOLD)

        def put(row, lab, formula, fmt="#,##0", bold=False):
            ws.cell(row=row, column=1, value=lab)
            c = ws.cell(row=row, column=2, value=formula)
            c.fill = CALC_FILL
            c.number_format = fmt
            c.border = thin
            if bold:
                c.font = _font(bold=True, color=NAVY)

        put(2, "PV capacity (MWp)", f"={ref(mw_k)}", "0.00")
        put(3, "BESS energy (MWh)", f"={ref(mwh_k)}", "0.00")
        put(4, "BESS power (MW)", f"=B3/{ref('duration_h')}", "0.00")
        put(5, "Specific yield (kWh/kWp)", f"={ref(yk)}", "0")

        ws.cell(row=7, column=1, value="CAPEX").font = Font(name="Calibri", bold=True, color=GOLD)
        put(8, "PV EPC", f"=B2*{ref(pvepc_k)}")
        put(9, "BESS EPC", f"=B3*{ref('bess_epc')}")
        put(10, "RTB / development", f"=B2*{ref('rtb_per_mw')}")
        put(11, "EAC connection", f"=B2*{ref('connection_per_mw')}")
        put(12, "Other CAPEX", f"={ref('other_capex')}")
        put(13, "Total CAPEX", "=B8+B9+B10+B11+B12", bold=True)
        put(14, "Senior debt", f"=B13*{ref('ltv')}")
        put(15, "Equity", "=B13-B14", bold=True)

        ws.cell(row=17, column=1, value="ENERGY & REVENUE Y1").font = Font(name="Calibri", bold=True, color=GOLD)
        put(18, "Annual PV production (MWh)", f"=B2*B5", "#,##0.0")
        put(19, "Uncurtailed (MWh)", f"=B18*(1-{ref('curtailment_pct')})", "#,##0.0")
        put(20, "Curtailed (MWh)", f"=B18*{ref('curtailment_pct')}", "#,##0.0")
        put(21, "BESS charged (MWh)", f"=MIN(B20*{ref('capture_pct')},B3*{ref('full_cycle_days')})", "#,##0.0")
        put(22, "BESS discharged (MWh)", f"=B21*{ref('rte')}", "#,##0.0")
        put(23, "Solar revenue (€)", f"=B19*{ref('dam_day')}")
        put(24, "BESS revenue (€)", f"=B22*{ref('dam_evening')}")
        put(25, "Gross revenue (€)", "=B23+B24", bold=True)
        put(26, "Aggregator fee (€)", f"=B25*{ref('aggregator_pct')}")
        put(27, "Net revenue (€)", "=B25-B26")

        ws.cell(row=29, column=1, value="OPEX & RETURNS Y1").font = Font(name="Calibri", bold=True, color=GOLD)
        put(30, "PV O&M", f"=B2*{ref(pvom_k)}")
        put(31, "BESS LTSA", f"=B3*{ref('ltsa')}")
        put(32, "SCADA", f"={ref('scada')}")
        put(33, "Admin", f"={ref('admin')}")
        put(34, "Land lease", f"={ref(land_k)}")
        put(35, "Insurance", f"=B13*{ref('ins_pct')}")
        put(36, "Total OPEX", "=B30+B31+B32+B33+B34+B35")
        put(37, "EBITDA", "=B27-B36", bold=True)
        put(38, "Debt service",
            f"=IF(B14=0,0,B14*({ref('loan_rate')}*(1+{ref('loan_rate')})^{ref('loan_years')}/((1+{ref('loan_rate')})^{ref('loan_years')}-1)))")
        put(39, "D&A", f"=(B8+B9)/{ref('da_epc_years')}+(B10+B11)/{ref('da_dev_years')}")
        put(40, "Taxable income", "=MAX(0,B37-B38-B39)")
        put(41, "CIT", f"=B40*{ref('cit_pct')}")
        put(42, "After-tax FCF", "=B37-B38-B41", bold=True)
        put(43, "Cash yield on equity", "=IF(B15=0,0,B42/B15)", "0.0%", bold=True)
        put(44, "Simple payback (yr)", '=IF(B42<=0,"n/a",B15/B42)', "0.00")
        put(45, "DSCR", '=IF(B38=0,"n/a (unlevered)",B37/B38)', "0.00")

        ws.cell(row=47, column=1, value="Cross-check (Python Y1 FCF):")
        ws.cell(row=47, column=2, value=r_["fcf"]).number_format = "#,##0"
        ws.cell(row=47, column=3, value="= B42 at default Inputs").font = _font(size=9, color=GREY)

        ws.column_dimensions["A"].width = 30
        ws.column_dimensions["B"].width = 16
        ws.column_dimensions["C"].width = 34

    # ── Cashflow_20y ─────────────────────────────────────────────────────────
    cf = wb.create_sheet("Cashflow_20y")
    cf["A1"] = "20-Year Equity Cashflow (computed from Inputs defaults; re-run script to refresh)"
    cf["A1"].font = Font(name="Calibri", bold=True, size=12, color=GOLD)
    cf["A2"] = "Includes PV & BESS degradation and OEM extended warranty from Year 6."
    cf["A2"].font = _font(size=9, color=GREY)

    cf.cell(row=4, column=1, value="Year").fill = HDR_FILL
    cf.cell(row=4, column=1).font = _font(bold=True, color=WHITE)
    for j, r_ in enumerate(results):
        c = cf.cell(row=4, column=2 + j, value=f"{r_['label']} · {r_['cfg_short']}")
        c.fill = HDR_FILL
        c.font = _font(bold=True, color=WHITE)
        c.alignment = Alignment(horizontal="center", wrap_text=True)

    max_years = int(d["horizon_years"])
    for t in range(0, max_years + 1):
        row = 5 + t
        cf.cell(row=row, column=1, value=t if t else "0 (invest)")
        for j, r_ in enumerate(results):
            c = cf.cell(row=row, column=2 + j, value=r_["cashflows"][t])
            c.number_format = "#,##0"
            c.fill = CALC_FILL
            c.border = thin

    base = 5 + max_years + 2
    for off, (lab, key, fmt) in enumerate([("Equity IRR", "irr", "0.0%"),
                                           ("NPV @ 8%", "npv8", "#,##0"),
                                           ("Simple payback (yr)", "payback", "0.0")]):
        cf.cell(row=base + off, column=1, value=lab).font = _font(bold=True)
        for j, r_ in enumerate(results):
            c = cf.cell(row=base + off, column=2 + j, value=r_[key])
            c.number_format = fmt
            c.fill = TOTAL_FILL

    cf.column_dimensions["A"].width = 14
    for col in range(2, 2 + len(results)):
        cf.column_dimensions[get_column_letter(col)].width = 16

    # ── Sources ──────────────────────────────────────────────────────────────
    src = wb.create_sheet("Sources")
    src["A1"] = "Single Source of Truth — validation map"
    src["A1"].font = Font(name="Calibri", bold=True, size=12, color=GOLD)
    for col, h in enumerate(["Item", "File", "Value / note"], 1):
        c = src.cell(row=3, column=col, value=h)
        c.fill = HDR_FILL
        c.font = _font(bold=True, color=WHITE)
    rows = [
        ("Evening discharge €210", "market/data/market-data.json", "18–21h avg Oct'25–Jul'26 (13,582 deduped prints)"),
        ("DAM daytime €140.88", "lib/deals/rtb-deal-types.ts → DAM", "06–17h; full dataset 06–17h = €123.37"),
        ("RTE / capture / curtailment", "rtb-deal-types.ts → BESS_DEFAULTS", "86.32% / 87.4% / 50%"),
        ("Fixed PV EPC €720k/MWp", "rtb-deal-types.ts → LH_EPC.pvPerMWp", "Bifacial fixed, ~5MWp scale"),
        ("Tracker PV EPC €810k/MWp", "docs/internal/konia-epc-cost-model.md", "€720k + €90k/MW tracker adder"),
        ("BESS EPC €127k/MWh", "rtb-deal-types.ts → LH_EPC.bessPerMWh", "Tier-1 LFP installed"),
        ("PV O&M €15k / €20k", "PV_DEFAULTS + konia O&M ratio", "Fixed €15k; tracker +~35%"),
        ("BESS LTSA €1,740/MWh/yr", "lib/portfolio-data.ts → LTSA.tierC", "Tier C, Yr 1–5"),
        ("RTB €350k/MW", "lib/constants.ts → INVESTMENT_SIZES", ""),
        ("CIT 15% / aggregator 10%", "PV_DEFAULTS", "From 1 Jan 2026"),
        ("Yields 1,850 / 2,450", "Client-tested Year-1", "Fixed bifacial / tracker+bifacial+albedo"),
        ("Contact", "lib/constants.ts → COMPANY_DATA", f"{CONTACT['email']} / {CONTACT['phone']}"),
    ]
    for i, (a, b, cval) in enumerate(rows):
        src.cell(row=4 + i, column=1, value=a)
        src.cell(row=4 + i, column=2, value=b)
        src.cell(row=4 + i, column=3, value=cval)
    src.column_dimensions["A"].width = 28
    src.column_dimensions["B"].width = 44
    src.column_dimensions["C"].width = 52
    src.cell(row=4 + len(rows) + 1, column=1,
             value=f"{CONTACT['company']} · {CONTACT['reg']} · {CONTACT['website']} · "
                   f"{CONTACT['director']} · {CONTACT['phone']} · {CONTACT['email']}").font = _font(size=9, color=GREY)

    wb.save(XLSX_OUT)
    print(f"Wrote {XLSX_OUT}")


# ═══════════════════════════════════════════════════════════════════════════════
# HTML ONE-PAGER
# ═══════════════════════════════════════════════════════════════════════════════

CSS = """
*{margin:0;padding:0;box-sizing:border-box}
:root{--navy:#1A365D;--gold:#C9A432;--bg:#F8FAFC;--border:#E2E8F0;--text:#1A202C;--muted:#64748B;--green:#059669}
body{font-family:'Segoe UI',system-ui,sans-serif;background:#E2E8F0;padding:16px;font-size:9.5pt;color:var(--text)}
.page{background:#fff;width:210mm;max-width:100%;margin:0 auto;padding:11mm 13mm;box-shadow:0 4px 20px rgba(0,0,0,.1)}
.header{display:flex;justify-content:space-between;align-items:center;border-bottom:3px solid var(--navy);padding-bottom:8px;margin-bottom:9px}
.logo{height:32px}
.doc-label{text-align:right;font-size:7pt;color:var(--muted);line-height:1.6}
.doc-label strong{color:var(--navy);font-size:7.5pt}
h1{font-size:14pt;font-weight:800;color:var(--navy);margin-bottom:2px}
.subtitle{font-size:8.5pt;color:var(--muted);margin-bottom:7px}
.badges{margin-bottom:8px}
.badge{display:inline-block;font-size:6.5pt;font-weight:700;padding:2px 8px;border-radius:10px;text-transform:uppercase;letter-spacing:.3px;margin-right:4px;background:#EFF6FF;color:var(--navy)}
.badge-g{background:#DCFCE7;color:#166534}
.cfg-title{font-size:9.5pt;font-weight:800;color:#fff;background:var(--navy);padding:4px 9px;border-radius:4px 4px 0 0;margin-top:9px}
.cfg-title span{font-weight:500;font-size:7.5pt;color:#CBD5E1}
h2{font-size:8.5pt;font-weight:700;color:var(--gold);border-bottom:2px solid var(--gold);padding-bottom:2px;display:inline-block;margin:8px 0 5px}
table{width:100%;border-collapse:collapse;font-size:7.5pt;margin-bottom:4px}
th{background:#2B4C7E;color:#fff;padding:4px 6px;text-align:left;font-size:7pt}
td{padding:3px 6px;border-bottom:1px solid var(--border)}
.r{text-align:right}
.total td{background:#EFF6FF;font-weight:700}
.hi td{background:#ECFDF5;font-weight:700;color:#065F46}
.two{display:grid;grid-template-columns:1fr 1fr;gap:12px}
ul.bullets{font-size:7.3pt;line-height:1.65}
ul.bullets li{list-style:none;padding-left:12px;position:relative;margin-bottom:2px}
ul.bullets li::before{content:"·";position:absolute;left:0;color:var(--gold);font-weight:900;font-size:12pt;line-height:1}
.contact{background:#EFF6FF;border:1px solid #BFDBFE;border-radius:4px;padding:7px 10px;font-size:7.5pt;margin-top:6px}
.note{font-size:6.2pt;color:var(--muted);margin-top:6px;line-height:1.45;padding-top:5px;border-top:1px solid var(--border)}
.footer{margin-top:6px;padding-top:5px;border-top:2px solid var(--navy);font-size:6pt;color:var(--muted);text-align:center}
.print-btn{position:fixed;top:12px;right:12px;background:var(--navy);color:#fff;border:none;padding:8px 14px;border-radius:4px;cursor:pointer;font-size:12px;z-index:100}
@media print{.print-btn{display:none}html,body{background:#fff;padding:0}.page{box-shadow:none;padding:9mm 11mm;width:auto}}
"""


def build_html(results: list[dict], d: dict) -> None:
    logo_src = (f"data:image/png;base64,{base64.b64encode(LOGO_PATH.read_bytes()).decode()}"
                if LOGO_PATH.exists() else "")

    def e(n: float) -> str:
        return f"&euro;{n/1e6:.2f}M" if abs(n) >= 1_000_000 else f"&euro;{n/1_000:,.0f}k"

    def returns_table(cfg_key: str) -> str:
        block = [r for r in results if r["cfg_key"] == cfg_key]
        cfg = PV_CONFIGS[cfg_key]
        head = "".join(f'<th class="r">{r["label"]}</th>' for r in block)
        rowdefs = [
            ("Total CAPEX", lambda r: e(r["capex"]), "total"),
            ("Annual production", lambda r: f'{r["annual_mwh"]:,.0f} MWh', ""),
            ("Gross revenue Y1", lambda r: e(r["gross"]), ""),
            ("  — solar / BESS", lambda r: f'{e(r["solar_rev"])} / {e(r["bess_rev"])}', ""),
            ("EBITDA Y1", lambda r: e(r["ebitda"]), ""),
            ("After-tax FCF Y1", lambda r: e(r["fcf"]), ""),
            ("Cash yield Y1", lambda r: pct(r["cash_yield"]), "hi"),
            ("Simple payback", lambda r: f'{r["payback"]:.1f} yr' if r["payback"] else "n/a", "hi"),
        ]
        body = ""
        for lab, fn, cls in rowdefs:
            cells = "".join(f'<td class="r">{fn(r)}</td>' for r in block)
            body += f'<tr class="{cls}"><td>{lab}</td>{cells}</tr>\n'
        return f"""
  <div class="cfg-title">{cfg['label']} <span>&mdash; {cfg['yield']:,} kWh/kWp · PV EPC &euro;{cfg['pv_epc']/1000:.0f}k/MWp</span></div>
  <table>
    <thead><tr><th>Metric</th>{head}</tr></thead>
    <tbody>{body}</tbody>
  </table>"""

    fixed_tbl = returns_table("fixed")
    tracker_tbl = returns_table("tracker")

    html = f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1.0">
<title>Cyprus PV + 4h BESS — 1 / 5 / 10 MW Investor One-Pager | Lighthief</title>
<style>{CSS}</style>
</head>
<body>
<button class="print-btn" onclick="window.print()">Print / PDF</button>
<div class="page">

  <div class="header">
    {"<img class='logo' src='" + logo_src + "' alt='Lighthief Cyprus'>" if logo_src else "<strong style='color:#1A365D;font-size:14pt'>Lighthief Cyprus</strong>"}
    <div class="doc-label">
      <strong>INVESTOR ONE-PAGER &mdash; CONFIDENTIAL</strong><br>
      {REF} &middot; {DATE_STR}<br>
      Non-binding &middot; 100% equity basis
    </div>
  </div>

  <h1>Cyprus Solar + 4-Hour BESS</h1>
  <div class="subtitle">Indicative investment cases &mdash; 1 / 5 / 10 MW co-located hybrid · two PV configurations · merchant curtailment recovery</div>
  <div class="badges">
    <span class="badge badge-g">Live DAM market data</span>
    <span class="badge">4-hour LFP BESS</span>
    <span class="badge">Fixed &amp; Tracker options</span>
    <span class="badge">Lighthief EPC + LTSA</span>
  </div>

  {fixed_tbl}
  {tracker_tbl}

  <div class="two">
    <div>
      <h2>How revenue works</h2>
      <ul class="bullets">
        <li><strong>Uncurtailed PV</strong> sold on TSOC DAM daytime (~&euro;{d['dam_day']:.0f}/MWh)</li>
        <li><strong>Curtailed PV</strong> (~{d['curtailment_pct']*100:.0f}%) charges the BESS at &euro;0 energy cost</li>
        <li><strong>Evening discharge</strong> at &euro;{d['dam_evening']:.0f}/MWh &mdash; the actual 18:00&ndash;21:00 DAM average from our market data</li>
        <li>RTE {d['rte']*100:.1f}% · curtailment capture {d['capture_pct']*100:.1f}% (Galascope-calibrated)</li>
        <li>No DAM arbitrage yet under Cyprus rules &mdash; capacity products = upside</li>
      </ul>
    </div>
    <div>
      <h2>Key assumptions (all in-house rates)</h2>
      <ul class="bullets">
        <li>Fixed PV EPC &euro;720k/MWp · Tracker &euro;810k/MWp</li>
        <li>BESS EPC &euro;127k/MWh installed 4h LFP</li>
        <li>RTB / development &euro;350k/MW</li>
        <li>PV O&amp;M &euro;15k (fixed) / &euro;20k (tracker) per MW/yr</li>
        <li>BESS LTSA &euro;1,740/MWh/yr · aggregator 10% · CIT 15%</li>
        <li>100% equity base case (model supports leverage)</li>
      </ul>
      <div class="contact">
        <strong>{CONTACT['director']}</strong> &mdash; {CONTACT['title']}<br>
        {CONTACT['phone']} &middot; {CONTACT['email']}<br>
        {CONTACT['website']} &middot; Office {CONTACT['office_phone']}
      </div>
    </div>
  </div>

  <div class="note">
    Indicative, non-binding Year-1 figures for discussion only (100% equity, ex-VAT). Evening discharge &euro;{d['dam_evening']:.0f}/MWh is the
    18:00&ndash;21:00 DAM average over Oct 2025&ndash;Jul 2026 (winter ~&euro;184, spring &euro;240&ndash;260). Land lease, EAC connection, and grid works are
    site-specific and will adjust CAPEX and returns. 20-year IRR / NPV with degradation and OEM warranty from Year 6 are in the accompanying Excel model
    (yellow input cells fully editable): <em>cyprus-pv-bess-1-5-10mw-financial-model.xlsx</em>
  </div>

  <div class="footer">
    {CONTACT['company']} &middot; Company No. {CONTACT['reg']} &middot; {CONTACT['website']}<br>
    {CONTACT['address']}
  </div>
</div>
</body>
</html>
"""
    for path in (HTML_OUT, HTML_TEASER, HTML_PUBLIC):
        path.write_text(html, encoding="utf-8")
        print(f"Wrote {path}")


def main():
    d = DEFAULTS
    results = []
    for cfg_key in ("fixed", "tracker"):
        for size in SIZES:
            results.append(calc_scenario(size, PV_CONFIGS[cfg_key], d))

    print("\n=== Validated base case (100% equity) ===")
    for r in results:
        pb = f"{r['payback']:.1f}y" if r["payback"] else "n/a"
        print(f"{r['cfg_short']:8s} {r['label']:16s}: CAPEX {eur(r['capex']):>8} | "
              f"Gross {eur(r['gross']):>8} | FCF {eur(r['fcf']):>8} | "
              f"Yield {pct(r['cash_yield']):>6} | PB {pb:>5} | 20yIRR {pct(r['irr']):>6}")

    build_excel(results, d)
    build_html(results, d)
    print("\nDone.")


if __name__ == "__main__":
    main()
