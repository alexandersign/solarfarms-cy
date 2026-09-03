"""
Cyprus PV (single-axis trackers) + 4h BESS — German investor pack
Generates:
  - HTML teaser (English)
  - Excel workbook with yellow INPUT cells and live FORMULAS

Locked assumptions (Jul 2026):
  - Sizes: 1 / 5 / 10 MWp PV with 4h BESS (4 / 20 / 40 MWh)
  - Modules: 700 Wp bifacial · single-axis trackers · white albedo
  - Specific yield: 2,400 kWh/kWp
  - RTB: €550k/MW + EAC connection €80k/MW
  - Curtailment cases: 50% and 65%
  - Cyprus-sited · generic German institutional investor · English

Sources (SSOT):
  docs/internal/solarpark-epc.md          — PV/BESS client EPC, tracker premium
  lib/portfolio-data.ts                   — RTE 86.32%, LTSA €1,740/MWh
  lib/deals/rtb-deal-types.ts             — DAM, capture, cycle-day capacity
  lib/market/cyprus-tsoc-dam-sample.ts    — daytime €140.88 / evening €182.99
"""

from __future__ import annotations

import math
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
OUT_DIR = ROOT / "docs" / "clients" / "German_Institutional_PV_BESS_Jul2026"
OUT_DIR.mkdir(parents=True, exist_ok=True)

HTML_OUT = OUT_DIR / "cyprus-pv-tracker-bess-investor-teaser-jul2026.html"
XLSX_OUT = OUT_DIR / "cyprus-pv-tracker-bess-investor-model-jul2026.xlsx"

# ── Brand ─────────────────────────────────────────────────────────────────────
NAVY = "1A365D"
NAVY2 = "2B5FA0"
GOLD = "C9A432"
WHITE = "FFFFFF"
GREY = "404040"
INPUT_BG = "FFF9E6"   # yellow — editable inputs
CALC_BG = "EEF6EE"    # green — formula outputs
LGREY = "EEF3FB"
MGREY = "D0DDEF"

# ── SSOT constants ────────────────────────────────────────────────────────────
RTE = 0.8632
CAPTURE = 0.874          # Galascope-calibrated (rtb-deal-types BESS_DEFAULTS)
FULL_CYCLE_DAYS = 280
DAM_DAY = 140.88
DAM_EVE = 182.99
AGG_FEE = 0.10
CIT = 0.15
LTV = 0.65
EQUITY_PCT = 0.35
LOAN_RATE = 0.055
LOAN_YEARS = 12
DA_YEARS = 20
PV_OM_PER_MW = 15_000
BESS_LTSA_PER_MWH = 1_740
LAND_PER_MW = 5_000
OTHER_FIXED = 15_000
OTHER_PCT_CAPEX = 0.005
YIELD_KWH_KWP = 2_400
RTB_PER_MW = 550_000
EAC_PER_MW = 80_000
TRACKER_PREMIUM_PER_MW = 100_000   # mid of €80–120k/MW (solarpark-epc §7.3)
PERMITTING = 35_000
MV_CABLE = 20_000                  # illustrative 1 km

_r, _n = LOAN_RATE, LOAN_YEARS
ANNUITY = _r * (1 + _r) ** _n / ((1 + _r) ** _n - 1)

# Fixed-tilt PV client EPC (solarpark-epc.md) — tracker added separately
SCENARIOS = [
    {
        "label": "1 MWp + 4 MWh",
        "mw": 1.0,
        "mwh": 4.0,
        "pv_base": 730_077,       # client fixed-tilt
        "bess_eur_mwh": 168_584,  # 4 MWh tier
    },
    {
        "label": "5 MWp + 20 MWh",
        "mw": 5.0,
        "mwh": 20.0,
        "pv_base": 3_200_385,
        "bess_eur_mwh": 111_900,  # confirmed Galascope 5MW/20MWh
    },
    {
        "label": "10 MWp + 40 MWh",
        "mw": 10.0,
        "mwh": 40.0,
        "pv_base": 6_100_770,
        "bess_eur_mwh": 109_797,  # solarpark-epc 40 MWh client
    },
]

CURTAILMENT_CASES = [
    {"key": "base", "label": "50% curtailment (base)", "pct": 0.50},
    {"key": "high", "label": "65% curtailment (stress)", "pct": 0.65},
]


def annuity_factor(rate: float, years: int) -> float:
    return rate * (1 + rate) ** years / ((1 + rate) ** years - 1)


def calc(s: dict, curt_pct: float) -> dict:
    mw, mwh = s["mw"], s["mwh"]
    kwp = mw * 1_000

    pv_epc = s["pv_base"] + mw * TRACKER_PREMIUM_PER_MW
    bess_epc = mwh * s["bess_eur_mwh"]
    rtb = mw * RTB_PER_MW
    eac = mw * EAC_PER_MW
    capex = pv_epc + bess_epc + rtb + eac + PERMITTING + MV_CABLE

    prod = kwp * YIELD_KWH_KWP / 1_000  # MWh
    uncurt = prod * (1 - curt_pct)
    curt = prod * curt_pct
    ideal = curt * CAPTURE
    max_ch = mwh * FULL_CYCLE_DAYS
    charged = min(ideal, max_ch)
    discharged = charged * RTE

    solar_rev = uncurt * DAM_DAY
    bess_rev = discharged * DAM_EVE
    gross = solar_rev + bess_rev
    net_rev = gross * (1 - AGG_FEE)

    pv_om = mw * PV_OM_PER_MW
    bess_om = mwh * BESS_LTSA_PER_MWH
    land = mw * LAND_PER_MW
    other = OTHER_FIXED + capex * OTHER_PCT_CAPEX
    opex = pv_om + bess_om + land + other
    ebitda = net_rev - opex

    equity = capex * EQUITY_PCT
    debt = capex * LTV
    svc = debt * ANNUITY
    dscr = ebitda / svc if svc else 0
    da = capex / DA_YEARS
    pretax = ebitda - svc
    ebt = pretax - da
    tax = max(0.0, ebt * CIT)
    net_cash = pretax - tax

    upay = capex / ebitda if ebitda > 0 else 99.0
    epay = equity / net_cash if net_cash > 0 else 99.0
    cash_yield = net_cash / equity if equity else 0
    ebitda_yield = ebitda / capex if capex else 0

    return {
        "mw": mw, "mwh": mwh, "label": s["label"],
        "pv_epc": pv_epc, "bess_epc": bess_epc, "bess_eur_mwh": s["bess_eur_mwh"],
        "rtb": rtb, "eac": eac, "permitting": PERMITTING, "cable": MV_CABLE,
        "capex": capex,
        "prod": prod, "uncurt": uncurt, "curt": curt,
        "charged": charged, "discharged": discharged,
        "cap_constrained": ideal > max_ch + 1e-6,
        "solar_rev": solar_rev, "bess_rev": bess_rev,
        "gross": gross, "net_rev": net_rev,
        "pv_om": pv_om, "bess_om": bess_om, "land": land, "other": other,
        "opex": opex, "ebitda": ebitda,
        "equity": equity, "debt": debt, "svc": svc, "dscr": dscr,
        "da": da, "tax": tax, "net_cash": net_cash,
        "upay": upay, "epay": epay,
        "cash_yield": cash_yield, "ebitda_yield": ebitda_yield,
        "curt_pct": curt_pct,
    }


def eur(n: float, dp: int = 0) -> str:
    if dp == 0:
        return f"€{int(round(n)):,}"
    return f"€{n:,.{dp}f}"


def eur_m(n: float) -> str:
    if abs(n) >= 1_000_000:
        return f"€{n / 1_000_000:.2f}M"
    return f"€{n / 1_000:.0f}k"


# ══════════════════════════════════════════════════════════════════════════════
# HTML TEASER
# ══════════════════════════════════════════════════════════════════════════════

def make_html(results_50: list[dict], results_65: list[dict]) -> None:
    def rows_capex(rs: list[dict]) -> str:
        lines = [
            ("PV EPC (client fixed-tilt + €100k/MW tracker)", [eur(r["pv_epc"]) for r in rs]),
            ("BESS EPC — Linyang LFP 4h turnkey", [f"{eur(r['bess_epc'])} ({eur(r['bess_eur_mwh']/1000,0)}k/MWh)" for r in rs]),
            ("RTB acquisition (€550k/MW)", [eur(r["rtb"]) for r in rs]),
            ("EAC grid connection (€80k/MW)", [eur(r["eac"]) for r in rs]),
            ("Permitting + MV cable (illustrative)", [eur(r["permitting"] + r["cable"]) for r in rs]),
            ("Total CAPEX (ex-VAT)", [eur(r["capex"]) for r in rs]),
        ]
        html = ""
        for i, (lab, vals) in enumerate(lines):
            cls = ' class="total"' if i == len(lines) - 1 else ""
            html += f"<tr{cls}><td>{lab}</td>" + "".join(f'<td class="r">{v}</td>' for v in vals) + "</tr>\n"
        return html

    def rows_ops(rs: list[dict], curt_label: str) -> str:
        lines = [
            (f"Gross PV production (2,400 kWh/kWp)", [f"{r['prod']:,.0f} MWh" for r in rs]),
            (f"Uncurtailed solar → DAM day (€{DAM_DAY:.2f}/MWh)", [eur(r["solar_rev"]) for r in rs]),
            (f"BESS discharge → DAM evening (€{DAM_EVE:.2f}/MWh)", [eur(r["bess_rev"]) for r in rs]),
            ("Gross energy revenue", [eur(r["gross"]) for r in rs]),
            ("Net revenue (after 10% aggregator)", [eur(r["net_rev"]) for r in rs]),
            ("OPEX (PV O&M + BESS LTSA + land + insurance)", [eur(r["opex"]) for r in rs]),
            ("EBITDA", [eur(r["ebitda"]) for r in rs]),
            ("Unlevered payback", [f"{r['upay']:.1f} yr" for r in rs]),
            ("Equity (35%) / Debt (65% @ 5.5% · 12y)", [f"{eur_m(r['equity'])} / {eur_m(r['debt'])}" for r in rs]),
            ("DSCR", [f"{r['dscr']:.2f}×" for r in rs]),
            ("Net cash to equity (post 15% CIT)", [eur(r["net_cash"]) for r in rs]),
            ("Equity cash yield / equity payback", [f"{r['cash_yield']*100:.1f}% · {r['epay']:.1f} yr" for r in rs]),
        ]
        html = f'<tr class="sec"><td colspan="4">{curt_label}</td></tr>\n'
        for i, (lab, vals) in enumerate(lines):
            bold = i in (6, 11)
            cls = ' class="hi"' if bold else ""
            html += f"<tr{cls}><td>{lab}</td>" + "".join(f'<td class="r">{v}</td>' for v in vals) + "</tr>\n"
        return html

    labels = [s["label"] for s in SCENARIOS]
    col_heads = "".join(f"<th>{l}</th>" for l in labels)

    html = f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>Cyprus PV Trackers + 4h BESS — Investor Teaser | July 2026</title>
<style>
:root {{
  --primary: #1A365D;
  --primary-light: #2B5FA0;
  --accent: #C9A432;
  --accent-dark: #9C7D22;
  --white: #FFFFFF;
  --grey-text: #404040;
  --body-bg: #F0F4F8;
  --border: #D0DDEF;
}}
* {{ margin:0; padding:0; box-sizing:border-box; }}
body {{
  font-family: 'Segoe UI', system-ui, sans-serif;
  font-size: 13px;
  color: var(--grey-text);
  background: var(--body-bg);
  padding: 24px;
  line-height: 1.55;
}}
.page {{
  max-width: 920px;
  margin: 0 auto;
  background: #fff;
  padding: 28px 32px 36px;
  border: 1px solid var(--border);
}}
.header {{
  display: flex;
  justify-content: space-between;
  align-items: flex-end;
  background: linear-gradient(135deg, var(--primary) 0%, var(--primary-light) 100%);
  color: #fff;
  padding: 18px 20px;
  margin: -28px -32px 22px;
}}
.brand {{ font-size: 22px; font-weight: 700; color: var(--accent); letter-spacing: .04em; }}
.brand small {{ display:block; font-size: 11px; color: #A8B8CC; font-weight: 400; margin-top: 4px; }}
.doc-meta {{ text-align: right; font-size: 11px; color: #A8B8CC; line-height: 1.6; }}
.doc-meta strong {{ color: #fff; }}
h1 {{ color: var(--accent); font-size: 22px; margin: 8px 0 6px; }}
.lead {{ color: var(--primary); font-size: 14px; font-weight: 600; margin-bottom: 10px; }}
.badge {{
  display: inline-block;
  background: #FFF8E6;
  color: var(--accent-dark);
  font-size: 10px;
  font-weight: 700;
  letter-spacing: .06em;
  padding: 4px 10px;
  margin-bottom: 8px;
}}
.kpis {{
  display: grid;
  grid-template-columns: repeat(4, 1fr);
  gap: 10px;
  margin: 16px 0 20px;
}}
.kpi {{
  background: var(--primary);
  color: #fff;
  padding: 12px 10px;
  text-align: center;
}}
.kpi .v {{ font-size: 18px; font-weight: 700; color: var(--accent); }}
.kpi .l {{ font-size: 9px; text-transform: uppercase; letter-spacing: .05em; color: #A8B8CC; margin-top: 4px; }}
h2 {{
  color: var(--accent);
  font-size: 15px;
  margin: 22px 0 8px;
  padding-bottom: 4px;
  border-bottom: 2px solid var(--primary);
  display: inline-block;
}}
.two {{ display: grid; grid-template-columns: 1fr 1fr; gap: 12px; margin: 10px 0; }}
.box {{ border: 1px solid var(--border); padding: 12px 14px; }}
.box h3 {{ color: var(--primary); font-size: 12px; margin-bottom: 6px; }}
.box ul {{ margin-left: 16px; }}
.box li {{ margin-bottom: 3px; }}
table {{ width: 100%; border-collapse: collapse; font-size: 11.5px; margin: 8px 0 14px; }}
th {{ background: var(--primary); color: #fff; padding: 7px 8px; text-align: left; font-weight: 600; }}
th.r, td.r {{ text-align: right; }}
td {{ padding: 6px 8px; border-bottom: 1px solid var(--border); }}
tr:nth-child(even) td {{ background: #F8FAFC; }}
tr.total td {{ background: var(--primary) !important; color: #fff; font-weight: 700; }}
tr.hi td {{ background: #EEF6EE !important; font-weight: 700; color: var(--primary); }}
tr.sec td {{ background: #EEF3FB !important; color: var(--primary); font-weight: 700; font-size: 10px; letter-spacing: .04em; text-transform: uppercase; }}
.callout {{
  background: #FFF8E6;
  border-left: 4px solid var(--accent);
  padding: 10px 14px;
  margin: 12px 0;
  font-size: 12px;
}}
.note {{ font-size: 10px; color: #6B7C8F; margin: 6px 0; }}
.footer {{
  margin-top: 24px;
  padding: 14px 16px;
  background: var(--primary);
  color: #A8B8CC;
  font-size: 11px;
  display: flex;
  justify-content: space-between;
  gap: 16px;
}}
.footer strong {{ color: #fff; }}
.footer a {{ color: var(--accent); text-decoration: none; }}
.print-btn {{
  position: fixed; top: 14px; right: 14px;
  background: var(--primary); color: #fff; border: none;
  padding: 10px 16px; cursor: pointer; font-size: 12px; z-index: 10;
}}
@media print {{
  .print-btn {{ display: none; }}
  body {{ background: #fff; padding: 0; }}
  .page {{ border: none; max-width: none; }}
}}
@media (max-width: 720px) {{
  .kpis, .two {{ grid-template-columns: 1fr 1fr; }}
  .header {{ flex-direction: column; align-items: flex-start; gap: 10px; }}
}}
</style>
</head>
<body>
<button class="print-btn" type="button" onclick="window.print()">Print / PDF</button>
<div class="page">
  <div class="header">
    <div>
      <div class="brand">LIGHTHIEF<small>Cyprus Ltd · HE 477423 · solarfarms.cy</small></div>
    </div>
    <div class="doc-meta">
      <strong>INVESTMENT TEASER</strong><br>
      Ref: LH-CY-PVBESS-TRK-JUL2026<br>
      July 2026 · Strictly Confidential<br>
      Prepared for: German Institutional Investor (generic)
    </div>
  </div>

  <div class="badge">CYPRUS RTB · SINGLE-AXIS TRACKERS · 4-HOUR BESS HYBRID</div>
  <h1>Cyprus PV + Battery Storage Investment</h1>
  <p class="lead">1 / 5 / 10 MWp single-axis tracker PV with 4-hour LFP BESS — merchant DAM + curtailment capture</p>
  <p class="note">Indicative economics for Cyprus-sited Ready-to-Build parks. Not a binding offer. Figures ex-VAT; site-specific grid and land terms subject to diligence.</p>

  <div class="kpis">
    <div class="kpi"><div class="v">2,400</div><div class="l">kWh/kWp yield<br>trackers · 700 Wp bifacial · white albedo</div></div>
    <div class="kpi"><div class="v">50–65%</div><div class="l">Curtailment cases<br>Cyprus 2027 operating range</div></div>
    <div class="kpi"><div class="v">€183</div><div class="l">DAM evening €/MWh<br>TSOC sample Oct 2025–Feb 2026</div></div>
    <div class="kpi"><div class="v">86.32%</div><div class="l">BESS AC–AC RTE<br>Linyang LFP full system</div></div>
  </div>

  <h2>Investment thesis</h2>
  <div class="two">
    <div class="box">
      <h3>Why Cyprus</h3>
      <ul>
        <li>Island balancing zone with liquid TSOC day-ahead market</li>
        <li>Persistent evening vs daytime price spread (~€183 vs ~€141/MWh)</li>
        <li>High solar resource; tracker + bifacial + white albedo at 2,400 kWh/kWp</li>
        <li>Eurozone · EU rule of law · 15% CIT (from 1 Jan 2026)</li>
      </ul>
    </div>
    <div class="box">
      <h3>Why this configuration</h3>
      <ul>
        <li>4-hour BESS sized 4 MWh per MWp — matches hybrid curtailment capture</li>
        <li>Daytime uncurtailed solar sold at DAM; curtailed energy shifted to evening peak</li>
        <li>Turnkey EPC (PV trackers + BESS) with RTB ticket + EAC connection priced separately</li>
        <li>Editable Excel accompanies this teaser for sensitivity work</li>
      </ul>
    </div>
  </div>

  <div class="callout">
    <strong>Technology.</strong> 700 Wp N-type TOPCon bifacial modules on single-axis trackers over white ground (albedo ~0.70).
    BESS: Linyang LFP, ~5.015 MWh containers, AC–AC RTE 86.32%, LTSA €1,740/MWh/year (97% availability target).
    Revenue model applies a 87.4% curtailment-capture factor and a 280 full-cycle-day capacity constraint (Galascope-calibrated).
  </div>

  <h2>Capital expenditure (ex-VAT)</h2>
  <table>
    <thead>
      <tr><th>Component</th>{col_heads}</tr>
    </thead>
    <tbody>
      {rows_capex(results_50)}
    </tbody>
  </table>
  <p class="note">PV EPC = solarpark client fixed-tilt schedule + €100k/MW tracker premium (market mid of €80–120k/MW). BESS: 4 MWh @ €168.6k/MWh; 20 MWh @ €111.9k/MWh (confirmed); 40 MWh @ €109.8k/MWh. RTB €550k/MW + EAC connection €80k/MW.</p>

  <h2>Year-1 operations &amp; returns</h2>
  <table>
    <thead>
      <tr><th>Metric</th>{col_heads}</tr>
    </thead>
    <tbody>
      {rows_ops(results_50, "Case A — 50% curtailment (base)")}
      {rows_ops(results_65, "Case B — 65% curtailment (stress)")}
    </tbody>
  </table>

  <h2>Indicative timeline</h2>
  <table>
    <thead><tr><th>Step</th><th>Timing</th><th>Notes</th></tr></thead>
    <tbody>
      <tr><td>RTB + EPC contract</td><td>Month 0</td><td>SPV acquisition / EPC award</td></tr>
      <tr><td>OEM production</td><td>+90 days</td><td>Modules / BESS manufacturing</td></tr>
      <tr><td>CIF Limassol</td><td>+~50 days</td><td>Ocean freight</td></tr>
      <tr><td>Install &amp; commission</td><td>+60–90 days</td><td>Civil, MV, commissioning</td></tr>
      <tr><td>PAC / COD</td><td>~Month 9–11</td><td>Subject to grid readiness</td></tr>
    </tbody>
  </table>

  <h2>Key risks &amp; diligence items</h2>
  <div class="two">
    <div class="box">
      <h3>Market / operations</h3>
      <ul>
        <li>Curtailment trajectory (modelled 50% base / 65% stress)</li>
        <li>DAM evening price persistence vs sample window</li>
        <li>Aggregator / offtake fee (modelled 10%)</li>
        <li>BESS cycle-day utilisation vs weather / grid events</li>
      </ul>
    </div>
    <div class="box">
      <h3>Project / delivery</h3>
      <ul>
        <li>Final EAC connection cost vs €80k/MW placeholder</li>
        <li>Land lease / title inside or outside RTB ticket</li>
        <li>MV distance (cable line is illustrative €20k)</li>
        <li>Debt terms (model assumes 65% LTV · 5.5% · 12 years)</li>
      </ul>
    </div>
  </div>

  <p class="note">Companion workbook: <strong>{XLSX_OUT.name}</strong> — yellow cells are inputs; green cells are formulas. Change curtailment, yield, prices, or CapEx unit rates and all returns recalculate.</p>

  <div class="footer">
    <div>
      <strong>Alexander Papacosta</strong> · Cyprus Director<br>
      Phone: <a href="tel:+35799164158">+357 99 164 158</a><br>
      Email: <a href="mailto:office@lighthief.com">office@lighthief.com</a><br>
      Office: +357 77 77 00 50
    </div>
    <div style="text-align:right">
      Lighthief Cyprus Ltd · HE 477423<br>
      15 Agaritsis, Nektaria Court, Office 201<br>
      3045 Zakaki, Limassol, Cyprus · <a href="https://solarfarms.cy">solarfarms.cy</a>
    </div>
  </div>
</div>
</body>
</html>
"""
    HTML_OUT.write_text(html, encoding="utf-8")
    print(f"  HTML -> {HTML_OUT}")


# ══════════════════════════════════════════════════════════════════════════════
# EXCEL — formula-driven
# ══════════════════════════════════════════════════════════════════════════════

def make_xlsx() -> None:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.formatting.rule import FormulaRule
    from openpyxl.utils import get_column_letter
    from openpyxl.workbook.defined_name import DefinedName

    wb = openpyxl.Workbook()

    fill = lambda h: PatternFill("solid", fgColor=h)
    thin = Border(
        left=Side(style="thin", color=MGREY),
        right=Side(style="thin", color=MGREY),
        top=Side(style="thin", color=MGREY),
        bottom=Side(style="thin", color=MGREY),
    )
    font = lambda **kw: Font(name="Calibri", **kw)

    def style_input(c):
        c.fill = fill(INPUT_BG)
        c.font = font(size=11, color=GREY, bold=True)
        c.border = thin
        c.alignment = Alignment(horizontal="right")

    def style_calc(c):
        c.fill = fill(CALC_BG)
        c.font = font(size=11, color=NAVY)
        c.border = thin
        c.alignment = Alignment(horizontal="right")

    def style_label(c, bold=False):
        c.font = font(size=11, color=GREY, bold=bold)
        c.border = thin
        c.alignment = Alignment(horizontal="left", indent=1)

    def header_row(ws, row, values, bg=NAVY):
        for i, v in enumerate(values, 1):
            c = ws.cell(row=row, column=i, value=v)
            c.fill = fill(bg)
            c.font = font(size=11, bold=True, color=GOLD if i == 1 else WHITE)
            c.border = thin
            c.alignment = Alignment(horizontal="center" if i > 1 else "left", vertical="center")

    def section(ws, row, text, ncols=5):
        c = ws.cell(row=row, column=1, value=text.upper())
        c.fill = fill(LGREY)
        c.font = font(size=10, bold=True, color=NAVY)
        c.border = thin
        for col in range(2, ncols + 1):
            cc = ws.cell(row=row, column=col)
            cc.fill = fill(LGREY)
            cc.border = thin

    # ── Sheet: Assumptions ────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Assumptions"
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 44
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 16
    ws.column_dimensions["E"].width = 42

    ws["A1"] = "LIGHTHIEF CYPRUS — PV TRACKERS + 4h BESS INVESTOR MODEL"
    ws["A1"].font = font(size=14, bold=True, color=GOLD)
    ws["A1"].fill = fill(NAVY)
    for col in range(1, 6):
        ws.cell(row=1, column=col).fill = fill(NAVY)
    ws.merge_cells("A1:E1")
    ws.row_dimensions[1].height = 28

    ws["A2"] = "Yellow cells = INPUTS (edit these). Green cells on Model sheets = FORMULAS. Ref: LH-CY-PVBESS-TRK-JUL2026 · July 2026 · Confidential"
    ws["A2"].font = font(size=9, italic=True, color="A8B8CC")
    ws["A2"].fill = fill(NAVY)
    for col in range(1, 6):
        ws.cell(row=2, column=col).fill = fill(NAVY)
    ws.merge_cells("A2:E2")

    # Global market / tech
    r = 4
    section(ws, r, "Global market & technology inputs", 5)
    globals_ = [
        ("Yield_kWh_per_kWp", YIELD_KWH_KWP, "B2", "Single-axis + 700 Wp bifacial + white albedo"),
        ("DAM_Day_EUR_per_MWh", DAM_DAY, "B3", "TSOC daytime 06–17h sample"),
        ("DAM_Eve_EUR_per_MWh", DAM_EVE, "B4", "TSOC evening peak sample"),
        ("RTE_AC_AC", RTE, "B5", "Linyang full-system AC–AC"),
        ("Capture_Pct", CAPTURE, "B6", "Share of curtailed energy captured (pre capacity cap)"),
        ("Full_Cycle_Days", FULL_CYCLE_DAYS, "B7", "Max BESS full cycles / year"),
        ("Aggregator_Fee_Pct", AGG_FEE, "B8", "Market access / aggregator"),
        ("CIT_Pct", CIT, "B9", "Cyprus CIT from 1 Jan 2026"),
        ("LTV_Pct", LTV, "B10", "Senior debt / total CapEx"),
        ("Equity_Pct", EQUITY_PCT, "B11", "Must equal 1 − LTV"),
        ("Loan_Rate", LOAN_RATE, "B12", "Nominal p.a."),
        ("Loan_Years", LOAN_YEARS, "B13", "Annuity term"),
        ("DA_Years", DA_YEARS, "B14", "Straight-line depreciation years"),
        ("PV_OM_EUR_per_MW", PV_OM_PER_MW, "B15", "PV O&M"),
        ("BESS_LTSA_EUR_per_MWh", BESS_LTSA_PER_MWH, "B16", "LTSA Tier C"),
        ("Land_EUR_per_MW", LAND_PER_MW, "B17", "Indicative lease"),
        ("Other_Fixed_EUR", OTHER_FIXED, "B18", "Admin / monitoring"),
        ("Other_Pct_of_Capex", OTHER_PCT_CAPEX, "B19", "Insurance proxy"),
        ("RTB_EUR_per_MW", RTB_PER_MW, "B20", "User-confirmed RTB ticket"),
        ("EAC_EUR_per_MW", EAC_PER_MW, "B21", "Connection terms (separate)"),
        ("Tracker_Premium_EUR_per_MW", TRACKER_PREMIUM_PER_MW, "B22", "Market mid €80–120k/MW"),
        ("Permitting_EUR", PERMITTING, "B23", "Flat per park"),
        ("MV_Cable_EUR", MV_CABLE, "B24", "Illustrative 1 km"),
        ("Curtailment_Base_Pct", 0.50, "B25", "Case A"),
        ("Curtailment_Stress_Pct", 0.65, "B26", "Case B"),
    ]

    # Place values starting row 5 in column B, with named ranges
    ws["A4"]  # already section
    start = 5
    named_cells = {}
    for i, (name, val, _alias, note) in enumerate(globals_):
        row = start + i
        ws.cell(row=row, column=1, value=name.replace("_", " "))
        style_label(ws.cell(row=row, column=1))
        cell = ws.cell(row=row, column=2, value=val)
        style_input(cell)
        if "Pct" in name or name in ("RTE_AC_AC", "Loan_Rate", "Other_Pct_of_Capex", "Capture_Pct", "LTV_Pct", "Equity_Pct", "CIT_Pct", "Aggregator_Fee_Pct"):
            cell.number_format = "0.00%"
            # store as true percent for openpyxl percent format — values already 0-1
        elif name in ("Yield_kWh_per_kWp", "Full_Cycle_Days", "Loan_Years", "DA_Years"):
            cell.number_format = "0"
        else:
            cell.number_format = '€#,##0.00'
        ws.cell(row=row, column=5, value=note).font = font(size=9, italic=True, color="6B7C8F")
        # Named range
        ref = f"Assumptions!$B${row}"
        try:
            wb.defined_names.add(DefinedName(name=name, attr_text=ref))
        except Exception:
            pass
        named_cells[name] = f"$B${row}"

    # Fix percent inputs: openpyxl with number_format 0.00% expects 0.5 for 50%
    # Already using 0.50 style floats — good.

    # Scenario block
    r0 = start + len(globals_) + 2
    section(ws, r0, "Scenario sizing & unit CapEx (per column)", 5)
    header_row(ws, r0 + 1, ["Parameter", "1 MWp", "5 MWp", "10 MWp", "Note"])

    scen_rows = {
        "MW": (r0 + 2, [1, 5, 10], "0.00"),
        "MWh": (r0 + 3, [4, 20, 40], "0.00"),
        "PV_Base_EPC_EUR": (r0 + 4, [730_077, 3_200_385, 6_100_770], '€#,##0'),
        "BESS_EUR_per_MWh": (r0 + 5, [168_584, 111_900, 109_797], '€#,##0'),
    }
    labels_map = {
        "MW": "PV capacity (MWp)",
        "MWh": "BESS energy (MWh)",
        "PV_Base_EPC_EUR": "PV base EPC € (fixed-tilt client)",
        "BESS_EUR_per_MWh": "BESS client €/MWh",
    }
    for key, (row, vals, fmt) in scen_rows.items():
        ws.cell(row=row, column=1, value=labels_map[key])
        style_label(ws.cell(row=row, column=1), bold=True)
        for i, v in enumerate(vals):
            c = ws.cell(row=row, column=2 + i, value=v)
            style_input(c)
            c.number_format = fmt
        note = {
            "MW": "Edit to resize parks",
            "MWh": "4h duration = 4 × MWp",
            "PV_Base_EPC_EUR": "solarpark-epc client schedule",
            "BESS_EUR_per_MWh": "4 MWh tier / confirmed 20 MWh / 40 MWh schedule",
        }[key]
        ws.cell(row=row, column=5, value=note).font = font(size=9, italic=True, color="6B7C8F")

    # Store scenario row numbers for Model sheet
    MW_ROW = scen_rows["MW"][0]
    MWH_ROW = scen_rows["MWh"][0]
    PVBASE_ROW = scen_rows["PV_Base_EPC_EUR"][0]
    BESSRATE_ROW = scen_rows["BESS_EUR_per_MWh"][0]

    # Annuity factor (formula)
    ann_row = r0 + 7
    ws.cell(row=ann_row, column=1, value="Annuity factor (loan)")
    style_label(ws.cell(row=ann_row, column=1), bold=True)
    # = rate*(1+rate)^n / ((1+rate)^n - 1)
    rate_cell = named_cells["Loan_Rate"]
    years_cell = named_cells["Loan_Years"]
    ws.cell(
        row=ann_row,
        column=2,
        value=f"={rate_cell}*(1+{rate_cell})^{years_cell}/((1+{rate_cell})^{years_cell}-1)",
    )
    style_calc(ws.cell(row=ann_row, column=2))
    ws.cell(row=ann_row, column=2).number_format = "0.0000"
    wb.defined_names.add(DefinedName(name="Annuity_Factor", attr_text=f"Assumptions!$B${ann_row}"))

    legend_row = ann_row + 2
    ws.cell(row=legend_row, column=1, value="Legend").font = font(bold=True, color=NAVY)
    ws.cell(row=legend_row + 1, column=1, value="INPUT").fill = fill(INPUT_BG)
    ws.cell(row=legend_row + 1, column=2, value="Editable assumption")
    ws.cell(row=legend_row + 2, column=1, value="FORMULA").fill = fill(CALC_BG)
    ws.cell(row=legend_row + 2, column=2, value="Calculated — do not overwrite")

    # ── Sheet: Model ──────────────────────────────────────────────────────────
    wm = wb.create_sheet("Model")
    wm.sheet_view.showGridLines = False
    wm.column_dimensions["A"].width = 48
    for col in range(2, 8):
        wm.column_dimensions[get_column_letter(col)].width = 15

    wm["A1"] = "FULL PROJECT MODEL — FORMULA DRIVEN FROM Assumptions"
    wm["A1"].font = font(size=13, bold=True, color=GOLD)
    wm["A1"].fill = fill(NAVY)
    for col in range(1, 8):
        wm.cell(row=1, column=col).fill = fill(NAVY)
    wm.merge_cells("A1:G1")

    # Column layout: B/C = 1MW base/stress, D/E = 5MW, F/G = 10MW
    # Simpler: two blocks — Case 50% (cols B-D) and Case 65% (cols E-G)
    header_row(wm, 3, ["Metric", "1 MWp · 50%", "5 MWp · 50%", "10 MWp · 50%",
                       "1 MWp · 65%", "5 MWp · 65%", "10 MWp · 65%"])

    # Map: case columns
    # Cols 2,3,4 = scenarios 0,1,2 at base curt
    # Cols 5,6,7 = scenarios 0,1,2 at stress curt
    # Scenario source cols on Assumptions: B=1MW, C=5MW, D=10MW → cols 2,3,4

    def asum(row):
        return f"Assumptions!{row}"

    def A(name):
        return f"Assumptions!{named_cells[name]}"

    # Build metric rows with formulas
    # For each output column j (2..7):
    #   scen_idx = (j-2) % 3   → 0,1,2,0,1,2
    #   curt = base if j < 5 else stress
    #   scen_col on Assumptions = 2 + scen_idx  → B,C,D

    metrics = []  # (label, formula_builder(scen_col_letter, curt_ref), fmt)

    def build_model_rows():
        rows = {}
        r = 4
        section(wm, r, "Capacity", 7)
        r = 5
        # MW
        wm.cell(row=r, column=1, value="PV capacity (MWp)")
        style_label(wm.cell(row=r, column=1))
        for j in range(2, 8):
            scen_col = get_column_letter(2 + (j - 2) % 3)
            c = wm.cell(row=r, column=j, value=f"=Assumptions!{scen_col}{MW_ROW}")
            style_calc(c)
            c.number_format = "0.00"
        rows["mw"] = r

        r = 6
        wm.cell(row=r, column=1, value="BESS energy (MWh)")
        style_label(wm.cell(row=r, column=1))
        for j in range(2, 8):
            scen_col = get_column_letter(2 + (j - 2) % 3)
            c = wm.cell(row=r, column=j, value=f"=Assumptions!{scen_col}{MWH_ROW}")
            style_calc(c)
            c.number_format = "0.00"
        rows["mwh"] = r

        r = 7
        wm.cell(row=r, column=1, value="Curtailment %")
        style_label(wm.cell(row=r, column=1))
        for j in range(2, 8):
            curt_ref = A("Curtailment_Base_Pct") if j < 5 else A("Curtailment_Stress_Pct")
            c = wm.cell(row=r, column=j, value=f"={curt_ref}")
            style_calc(c)
            c.number_format = "0.0%"
        rows["curt"] = r

        r = 9
        section(wm, r, "Capital expenditure", 7)

        r = 10
        wm.cell(row=r, column=1, value="PV EPC (base + tracker premium)")
        style_label(wm.cell(row=r, column=1))
        for j in range(2, 8):
            scen_col = get_column_letter(2 + (j - 2) % 3)
            # = PV_Base + MW * Tracker_Premium
            f = (
                f"=Assumptions!{scen_col}{PVBASE_ROW}"
                f"+Assumptions!{scen_col}{MW_ROW}*{A('Tracker_Premium_EUR_per_MW')}"
            )
            c = wm.cell(row=r, column=j, value=f)
            style_calc(c)
            c.number_format = '€#,##0'
        rows["pv_epc"] = r

        r = 11
        wm.cell(row=r, column=1, value="BESS EPC")
        style_label(wm.cell(row=r, column=1))
        for j in range(2, 8):
            scen_col = get_column_letter(2 + (j - 2) % 3)
            f = f"=Assumptions!{scen_col}{MWH_ROW}*Assumptions!{scen_col}{BESSRATE_ROW}"
            c = wm.cell(row=r, column=j, value=f)
            style_calc(c)
            c.number_format = '€#,##0'
        rows["bess_epc"] = r

        r = 12
        wm.cell(row=r, column=1, value="RTB acquisition")
        style_label(wm.cell(row=r, column=1))
        for j in range(2, 8):
            scen_col = get_column_letter(2 + (j - 2) % 3)
            f = f"=Assumptions!{scen_col}{MW_ROW}*{A('RTB_EUR_per_MW')}"
            c = wm.cell(row=r, column=j, value=f)
            style_calc(c)
            c.number_format = '€#,##0'
        rows["rtb"] = r

        r = 13
        wm.cell(row=r, column=1, value="EAC connection terms")
        style_label(wm.cell(row=r, column=1))
        for j in range(2, 8):
            scen_col = get_column_letter(2 + (j - 2) % 3)
            f = f"=Assumptions!{scen_col}{MW_ROW}*{A('EAC_EUR_per_MW')}"
            c = wm.cell(row=r, column=j, value=f)
            style_calc(c)
            c.number_format = '€#,##0'
        rows["eac"] = r

        r = 14
        wm.cell(row=r, column=1, value="Permitting + MV cable")
        style_label(wm.cell(row=r, column=1))
        for j in range(2, 8):
            f = f"={A('Permitting_EUR')}+{A('MV_Cable_EUR')}"
            c = wm.cell(row=r, column=j, value=f)
            style_calc(c)
            c.number_format = '€#,##0'
        rows["other_capex"] = r

        r = 15
        wm.cell(row=r, column=1, value="Total CAPEX")
        style_label(wm.cell(row=r, column=1), bold=True)
        for j in range(2, 8):
            cl = get_column_letter(j)
            f = f"={cl}{rows['pv_epc']}+{cl}{rows['bess_epc']}+{cl}{rows['rtb']}+{cl}{rows['eac']}+{cl}{rows['other_capex']}"
            c = wm.cell(row=r, column=j, value=f)
            style_calc(c)
            c.font = font(size=11, bold=True, color=NAVY)
            c.number_format = '€#,##0'
            c.fill = fill(LGREY)
        rows["capex"] = r

        r = 17
        section(wm, r, "Production & revenue (Year 1)", 7)

        r = 18
        wm.cell(row=r, column=1, value="Gross PV production (MWh)")
        style_label(wm.cell(row=r, column=1))
        for j in range(2, 8):
            cl = get_column_letter(j)
            f = f"={cl}{rows['mw']}*1000*{A('Yield_kWh_per_kWp')}/1000"
            c = wm.cell(row=r, column=j, value=f)
            style_calc(c)
            c.number_format = '#,##0.0'
        rows["prod"] = r

        r = 19
        wm.cell(row=r, column=1, value="Uncurtailed solar (MWh)")
        style_label(wm.cell(row=r, column=1))
        for j in range(2, 8):
            cl = get_column_letter(j)
            f = f"={cl}{rows['prod']}*(1-{cl}{rows['curt']})"
            c = wm.cell(row=r, column=j, value=f)
            style_calc(c)
            c.number_format = '#,##0.0'
        rows["uncurt"] = r

        r = 20
        wm.cell(row=r, column=1, value="Curtailed energy (MWh)")
        style_label(wm.cell(row=r, column=1))
        for j in range(2, 8):
            cl = get_column_letter(j)
            f = f"={cl}{rows['prod']}*{cl}{rows['curt']}"
            c = wm.cell(row=r, column=j, value=f)
            style_calc(c)
            c.number_format = '#,##0.0'
        rows["curted"] = r

        r = 21
        wm.cell(row=r, column=1, value="BESS charged (MWh) — min(ideal, capacity)")
        style_label(wm.cell(row=r, column=1))
        for j in range(2, 8):
            cl = get_column_letter(j)
            # MIN(curtailed * capture, mwh * cycle_days)
            f = (
                f"=MIN({cl}{rows['curted']}*{A('Capture_Pct')},"
                f"{cl}{rows['mwh']}*{A('Full_Cycle_Days')})"
            )
            c = wm.cell(row=r, column=j, value=f)
            style_calc(c)
            c.number_format = '#,##0.0'
        rows["charged"] = r

        r = 22
        wm.cell(row=r, column=1, value="BESS discharged (MWh)")
        style_label(wm.cell(row=r, column=1))
        for j in range(2, 8):
            cl = get_column_letter(j)
            f = f"={cl}{rows['charged']}*{A('RTE_AC_AC')}"
            c = wm.cell(row=r, column=j, value=f)
            style_calc(c)
            c.number_format = '#,##0.0'
        rows["discharged"] = r

        r = 23
        wm.cell(row=r, column=1, value="Solar revenue (€)")
        style_label(wm.cell(row=r, column=1))
        for j in range(2, 8):
            cl = get_column_letter(j)
            f = f"={cl}{rows['uncurt']}*{A('DAM_Day_EUR_per_MWh')}"
            c = wm.cell(row=r, column=j, value=f)
            style_calc(c)
            c.number_format = '€#,##0'
        rows["solar_rev"] = r

        r = 24
        wm.cell(row=r, column=1, value="BESS revenue (€)")
        style_label(wm.cell(row=r, column=1))
        for j in range(2, 8):
            cl = get_column_letter(j)
            f = f"={cl}{rows['discharged']}*{A('DAM_Eve_EUR_per_MWh')}"
            c = wm.cell(row=r, column=j, value=f)
            style_calc(c)
            c.number_format = '€#,##0'
        rows["bess_rev"] = r

        r = 25
        wm.cell(row=r, column=1, value="Gross energy revenue")
        style_label(wm.cell(row=r, column=1), bold=True)
        for j in range(2, 8):
            cl = get_column_letter(j)
            f = f"={cl}{rows['solar_rev']}+{cl}{rows['bess_rev']}"
            c = wm.cell(row=r, column=j, value=f)
            style_calc(c)
            c.number_format = '€#,##0'
        rows["gross"] = r

        r = 26
        wm.cell(row=r, column=1, value="Net revenue (after aggregator)")
        style_label(wm.cell(row=r, column=1), bold=True)
        for j in range(2, 8):
            cl = get_column_letter(j)
            f = f"={cl}{rows['gross']}*(1-{A('Aggregator_Fee_Pct')})"
            c = wm.cell(row=r, column=j, value=f)
            style_calc(c)
            c.font = font(size=11, bold=True, color="1E7D4A")
            c.number_format = '€#,##0'
        rows["net_rev"] = r

        r = 28
        section(wm, r, "OPEX & EBITDA", 7)

        r = 29
        wm.cell(row=r, column=1, value="PV O&M")
        style_label(wm.cell(row=r, column=1))
        for j in range(2, 8):
            cl = get_column_letter(j)
            f = f"={cl}{rows['mw']}*{A('PV_OM_EUR_per_MW')}"
            c = wm.cell(row=r, column=j, value=f)
            style_calc(c)
            c.number_format = '€#,##0'
        rows["pv_om"] = r

        r = 30
        wm.cell(row=r, column=1, value="BESS LTSA")
        style_label(wm.cell(row=r, column=1))
        for j in range(2, 8):
            cl = get_column_letter(j)
            f = f"={cl}{rows['mwh']}*{A('BESS_LTSA_EUR_per_MWh')}"
            c = wm.cell(row=r, column=j, value=f)
            style_calc(c)
            c.number_format = '€#,##0'
        rows["bess_om"] = r

        r = 31
        wm.cell(row=r, column=1, value="Land lease")
        style_label(wm.cell(row=r, column=1))
        for j in range(2, 8):
            cl = get_column_letter(j)
            f = f"={cl}{rows['mw']}*{A('Land_EUR_per_MW')}"
            c = wm.cell(row=r, column=j, value=f)
            style_calc(c)
            c.number_format = '€#,##0'
        rows["land"] = r

        r = 32
        wm.cell(row=r, column=1, value="Insurance + admin")
        style_label(wm.cell(row=r, column=1))
        for j in range(2, 8):
            cl = get_column_letter(j)
            f = f"={A('Other_Fixed_EUR')}+{cl}{rows['capex']}*{A('Other_Pct_of_Capex')}"
            c = wm.cell(row=r, column=j, value=f)
            style_calc(c)
            c.number_format = '€#,##0'
        rows["other"] = r

        r = 33
        wm.cell(row=r, column=1, value="Total OPEX")
        style_label(wm.cell(row=r, column=1), bold=True)
        for j in range(2, 8):
            cl = get_column_letter(j)
            f = f"={cl}{rows['pv_om']}+{cl}{rows['bess_om']}+{cl}{rows['land']}+{cl}{rows['other']}"
            c = wm.cell(row=r, column=j, value=f)
            style_calc(c)
            c.number_format = '€#,##0'
        rows["opex"] = r

        r = 34
        wm.cell(row=r, column=1, value="EBITDA")
        style_label(wm.cell(row=r, column=1), bold=True)
        for j in range(2, 8):
            cl = get_column_letter(j)
            f = f"={cl}{rows['net_rev']}-{cl}{rows['opex']}"
            c = wm.cell(row=r, column=j, value=f)
            style_calc(c)
            c.font = font(size=11, bold=True, color="1E7D4A")
            c.number_format = '€#,##0'
            c.fill = fill(CALC_BG)
        rows["ebitda"] = r

        r = 36
        section(wm, r, "Returns (levered)", 7)

        r = 37
        wm.cell(row=r, column=1, value="Equity required")
        style_label(wm.cell(row=r, column=1))
        for j in range(2, 8):
            cl = get_column_letter(j)
            f = f"={cl}{rows['capex']}*{A('Equity_Pct')}"
            c = wm.cell(row=r, column=j, value=f)
            style_calc(c)
            c.number_format = '€#,##0'
        rows["equity"] = r

        r = 38
        wm.cell(row=r, column=1, value="Senior debt")
        style_label(wm.cell(row=r, column=1))
        for j in range(2, 8):
            cl = get_column_letter(j)
            f = f"={cl}{rows['capex']}*{A('LTV_Pct')}"
            c = wm.cell(row=r, column=j, value=f)
            style_calc(c)
            c.number_format = '€#,##0'
        rows["debt"] = r

        r = 39
        wm.cell(row=r, column=1, value="Annual debt service")
        style_label(wm.cell(row=r, column=1))
        for j in range(2, 8):
            cl = get_column_letter(j)
            f = f"={cl}{rows['debt']}*Annuity_Factor"
            c = wm.cell(row=r, column=j, value=f)
            style_calc(c)
            c.number_format = '€#,##0'
        rows["svc"] = r

        r = 40
        wm.cell(row=r, column=1, value="DSCR")
        style_label(wm.cell(row=r, column=1), bold=True)
        for j in range(2, 8):
            cl = get_column_letter(j)
            f = f"=IF({cl}{rows['svc']}=0,0,{cl}{rows['ebitda']}/{cl}{rows['svc']})"
            c = wm.cell(row=r, column=j, value=f)
            style_calc(c)
            c.number_format = '0.00"×"'
        rows["dscr"] = r

        r = 41
        wm.cell(row=r, column=1, value="Depreciation")
        style_label(wm.cell(row=r, column=1))
        for j in range(2, 8):
            cl = get_column_letter(j)
            f = f"={cl}{rows['capex']}/{A('DA_Years')}"
            c = wm.cell(row=r, column=j, value=f)
            style_calc(c)
            c.number_format = '€#,##0'
        rows["da"] = r

        r = 42
        wm.cell(row=r, column=1, value="Tax (15% CIT on EBT)")
        style_label(wm.cell(row=r, column=1))
        for j in range(2, 8):
            cl = get_column_letter(j)
            # MAX(0, (EBITDA - debt_service - DA) * CIT)
            f = (
                f"=MAX(0,({cl}{rows['ebitda']}-{cl}{rows['svc']}-{cl}{rows['da']})"
                f"*{A('CIT_Pct')})"
            )
            c = wm.cell(row=r, column=j, value=f)
            style_calc(c)
            c.number_format = '€#,##0'
        rows["tax"] = r

        r = 43
        wm.cell(row=r, column=1, value="Net cash to equity")
        style_label(wm.cell(row=r, column=1), bold=True)
        for j in range(2, 8):
            cl = get_column_letter(j)
            f = f"={cl}{rows['ebitda']}-{cl}{rows['svc']}-{cl}{rows['tax']}"
            c = wm.cell(row=r, column=j, value=f)
            style_calc(c)
            c.font = font(size=11, bold=True, color="1E7D4A")
            c.number_format = '€#,##0'
        rows["net_cash"] = r

        r = 44
        wm.cell(row=r, column=1, value="Unlevered payback (years)")
        style_label(wm.cell(row=r, column=1), bold=True)
        for j in range(2, 8):
            cl = get_column_letter(j)
            f = f"=IF({cl}{rows['ebitda']}<=0,99,{cl}{rows['capex']}/{cl}{rows['ebitda']})"
            c = wm.cell(row=r, column=j, value=f)
            style_calc(c)
            c.number_format = '0.0'
        rows["upay"] = r

        r = 45
        wm.cell(row=r, column=1, value="Equity payback (years)")
        style_label(wm.cell(row=r, column=1), bold=True)
        for j in range(2, 8):
            cl = get_column_letter(j)
            f = f"=IF({cl}{rows['net_cash']}<=0,99,{cl}{rows['equity']}/{cl}{rows['net_cash']})"
            c = wm.cell(row=r, column=j, value=f)
            style_calc(c)
            c.font = font(size=11, bold=True, color="9C7D22")
            c.number_format = '0.0'
            c.fill = fill(INPUT_BG)  # highlight key KPI (still formula)
        rows["epay"] = r

        r = 46
        wm.cell(row=r, column=1, value="Equity cash yield")
        style_label(wm.cell(row=r, column=1), bold=True)
        for j in range(2, 8):
            cl = get_column_letter(j)
            f = f"=IF({cl}{rows['equity']}=0,0,{cl}{rows['net_cash']}/{cl}{rows['equity']})"
            c = wm.cell(row=r, column=j, value=f)
            style_calc(c)
            c.number_format = '0.0%'
        rows["eyield"] = r

        r = 48
        wm.cell(row=r, column=1, value="Notes").font = font(bold=True, color=NAVY)
        wm.cell(
            row=r + 1,
            column=1,
            value=(
                "BESS charged = MIN(curtailed × capture%, MWh × full-cycle days). "
                "At 2,400 kWh/kWp and 4h sizing, higher curtailment increases BESS share until the cycle-day cap binds. "
                "All figures ex-VAT. Not a binding offer."
            ),
        ).font = font(size=9, italic=True, color="6B7C8F")
        wm.merge_cells(start_row=r + 1, start_column=1, end_row=r + 1, end_column=7)

        return rows

    build_model_rows()

    # ── Sheet: Summary ────────────────────────────────────────────────────────
    ws_s = wb.create_sheet("Summary", 0)
    ws_s.sheet_view.showGridLines = False
    ws_s.column_dimensions["A"].width = 36
    for col in range(2, 5):
        ws_s.column_dimensions[get_column_letter(col)].width = 18

    ws_s["A1"] = "INVESTOR SUMMARY — CYPRUS PV TRACKERS + 4h BESS"
    ws_s["A1"].font = font(size=14, bold=True, color=GOLD)
    ws_s["A1"].fill = fill(NAVY)
    for col in range(1, 5):
        ws_s.cell(row=1, column=col).fill = fill(NAVY)
    ws_s.merge_cells("A1:D1")

    ws_s["A2"] = "Values linked to Model sheet (50% curtailment base case). Edit Assumptions to refresh."
    ws_s["A2"].font = font(size=9, italic=True, color="A8B8CC")
    ws_s["A2"].fill = fill(NAVY)
    for col in range(1, 5):
        ws_s.cell(row=2, column=col).fill = fill(NAVY)

    header_row(ws_s, 4, ["KPI (50% curtailment)", "1 MWp + 4 MWh", "5 MWp + 20 MWh", "10 MWp + 40 MWh"])

    summary_links = [
        ("Total CAPEX", 15, '€#,##0'),
        ("EBITDA Y1", 34, '€#,##0'),
        ("Net cash to equity Y1", 43, '€#,##0'),
        ("Unlevered payback (yr)", 44, "0.0"),
        ("Equity payback (yr)", 45, "0.0"),
        ("Equity cash yield", 46, "0.0%"),
        ("DSCR", 40, '0.00"×"'),
    ]
    for i, (lab, model_row, fmt) in enumerate(summary_links):
        row = 5 + i
        ws_s.cell(row=row, column=1, value=lab)
        style_label(ws_s.cell(row=row, column=1), bold=True)
        for j, model_col in enumerate([2, 3, 4]):  # 50% columns
            cl = get_column_letter(model_col)
            c = ws_s.cell(row=row, column=2 + j, value=f"=Model!{cl}{model_row}")
            style_calc(c)
            c.number_format = fmt

    header_row(ws_s, 14, ["KPI (65% curtailment stress)", "1 MWp + 4 MWh", "5 MWp + 20 MWh", "10 MWp + 40 MWh"])
    for i, (lab, model_row, fmt) in enumerate(summary_links):
        row = 15 + i
        ws_s.cell(row=row, column=1, value=lab)
        style_label(ws_s.cell(row=row, column=1), bold=True)
        for j, model_col in enumerate([5, 6, 7]):  # 65% columns
            cl = get_column_letter(model_col)
            c = ws_s.cell(row=row, column=2 + j, value=f"=Model!{cl}{model_row}")
            style_calc(c)
            c.number_format = fmt

    ws_s["A24"] = "Contact"
    ws_s["A24"].font = font(bold=True, color=NAVY)
    ws_s["A25"] = "Alexander Papacosta · Cyprus Director · +357 99 164 158 · office@lighthief.com"
    ws_s["A26"] = "Lighthief Cyprus Ltd · HE 477423 · solarfarms.cy"
    ws_s["A25"].font = font(size=10, color=GREY)
    ws_s["A26"].font = font(size=10, color=GREY)

    # ── Sources sheet ─────────────────────────────────────────────────────────
    wsrc = wb.create_sheet("Sources")
    wsrc.column_dimensions["A"].width = 28
    wsrc.column_dimensions["B"].width = 72
    wsrc["A1"] = "Data sources (internal SSOT)"
    wsrc["A1"].font = font(size=12, bold=True, color=GOLD)
    wsrc["A1"].fill = fill(NAVY)
    wsrc["B1"].fill = fill(NAVY)
    sources = [
        ("PV / BESS client EPC", "docs/internal/solarpark-epc.md"),
        ("Confirmed 5MW/20MWh BESS", "lib/portfolio-data.ts CLIENT_PRICING €111,900/MWh"),
        ("RTE / LTSA", "lib/portfolio-data.ts — 86.32% RTE · LTSA Tier C €1,740/MWh"),
        ("DAM prices", "lib/market/cyprus-tsoc-dam-sample.ts / rtb-deal-types DAM"),
        ("Capture & cycle days", "lib/deals/rtb-deal-types.ts BESS_DEFAULTS"),
        ("Tracker premium", "solarpark-epc.md §7.3 — €80–120k/MW (model uses €100k)"),
        ("Yield", "Investor assumption — 2,400 kWh/kWp trackers + white albedo"),
        ("RTB + connection", "Investor assumption — €550k/MW RTB + €80k/MW EAC"),
    ]
    for i, (a, b) in enumerate(sources, 3):
        wsrc.cell(row=i, column=1, value=a).font = font(bold=True, color=NAVY)
        wsrc.cell(row=i, column=2, value=b).font = font(color=GREY)

    wb.save(XLSX_OUT)
    print(f"  XLSX -> {XLSX_OUT}")


def main():
    print("Generating Cyprus PV Trackers + 4h BESS German investor pack...")
    r50 = [calc(s, 0.50) for s in SCENARIOS]
    r65 = [calc(s, 0.65) for s in SCENARIOS]

    print("\n-- CapEx check --")
    for r in r50:
        print(f"  {r['label']}: CapEx {eur(r['capex'])} | EBITDA@50% {eur(r['ebitda'])} | "
              f"eq.payback {r['epay']:.1f}y | DSCR {r['dscr']:.2f}x")

    print("\n-- Stress 65% --")
    for r in r65:
        print(f"  {r['label']}: EBITDA {eur(r['ebitda'])} | eq.payback {r['epay']:.1f}y | "
              f"cash yield {r['cash_yield']*100:.1f}%")

    make_html(r50, r65)
    make_xlsx()
    print("\nDone.")


if __name__ == "__main__":
    main()
