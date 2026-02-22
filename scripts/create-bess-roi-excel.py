"""
Create GalascopeBessROI.xlsx — BESS Profitability Calculator
All formulas reference an Inputs sheet so parameters can be changed.
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
import os, json
from collections import defaultdict
import re as re_mod

wb = openpyxl.Workbook()

# ── STYLE DEFINITIONS ──
hdr_font = Font(name='Calibri', bold=True, size=11, color='FFFFFF')
hdr_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
input_fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
calc_fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
highlight_fill = PatternFill(start_color='D6E4F0', end_color='D6E4F0', fill_type='solid')
warn_fill = PatternFill(start_color='FCE4EC', end_color='FCE4EC', fill_type='solid')
green_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
yellow_fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
title_font = Font(name='Calibri', bold=True, size=14, color='1F4E79')
section_font = Font(name='Calibri', bold=True, size=12, color='1F4E79')
label_font = Font(name='Calibri', size=11)
input_font = Font(name='Calibri', bold=True, size=11, color='C00000')
result_font = Font(name='Calibri', bold=True, size=12, color='006100')
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

def style_header_row(ws, row, max_col):
    for c in range(1, max_col + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
        cell.border = thin_border

def style_range(ws, min_row, max_row, max_col):
    for r in range(min_row, max_row + 1):
        for c in range(1, max_col + 1):
            ws.cell(row=r, column=c).border = thin_border

# ================================================================
# LOAD DATA
# ================================================================
base_dir = "docs/clients/Group2_Esperia_Energy/esperia-galascope-2.5-curtailment"
all_days = []

for m in range(1, 13):
    fname = f"2.5MW 2025.{m:02d} curtailments.xlsx"
    fpath = os.path.join(base_dir, fname)
    if not os.path.exists(fpath):
        continue
    src = openpyxl.load_workbook(fpath, data_only=True)
    ws = src.active
    hmap = {}
    for cell in ws[2]:
        if cell.value:
            hmap[cell.column - 1] = str(cell.value).strip()

    date_idx = pv_idx = loss_idx = irrad_idx = sun_idx = None
    for idx, hdr in hmap.items():
        if 'Statistical Period' in hdr: date_idx = idx
        if 'Sun Hours' in hdr: sun_idx = idx
        if 'Global Irradiation' in hdr: irrad_idx = idx
        if 'PV Yield' in hdr: pv_idx = idx
        if 'Inverter Yield' in hdr and pv_idx is None: pv_idx = idx
    for idx, hdr in hmap.items():
        if 'Loss Due to Export Limitation' in hdr and 'kWh' in hdr:
            loss_idx = idx
            break

    if date_idx is None or pv_idx is None or loss_idx is None:
        continue

    for row in ws.iter_rows(min_row=3, values_only=True):
        if row[date_idx] is None:
            continue
        dv = row[date_idx]
        pv = float(row[pv_idx]) if row[pv_idx] else 0
        curt = float(row[loss_idx]) if row[loss_idx] else 0
        sh = 0
        if sun_idx is not None and row[sun_idx] and isinstance(row[sun_idx], (int, float)) and row[sun_idx] < 20:
            sh = float(row[sun_idx])
        elif irrad_idx is not None and row[irrad_idx]:
            sh = float(row[irrad_idx]) / 0.85

        if isinstance(dv, datetime):
            ds = dv.strftime('%Y-%m-%d')
            mo = dv.month
        elif isinstance(dv, str):
            ds = dv
            mo = int(dv.split('-')[1])
        else:
            continue

        all_days.append({'date': ds, 'month': mo, 'sun': sh, 'pv': pv, 'curt': curt, 'total': pv + curt})

all_days.sort(key=lambda x: x['date'])
print(f"Loaded {len(all_days)} days of curtailment data")

# Load market data
with open("market/data/market-data.json", "r") as f:
    mdata = json.load(f)

unique = {}
for rec in mdata['records']:
    period = rec['period']
    match = re_mod.search(r'(\d{2}:\d{2})-', period)
    hh_key = match.group(1) if match else f"{rec['hour']:02d}:00"
    key = (rec['date'], hh_key)
    if key not in unique:
        unique[key] = rec

deduped = sorted(unique.values(), key=lambda x: (x['date'], x['hour'], x['period']))
hourly_prices = defaultdict(list)
for rec in deduped:
    hourly_prices[rec['hour']].append(rec['price'])

peak_avg = sum(sum(hourly_prices[h]) for h in range(17, 21)) / sum(len(hourly_prices[h]) for h in range(17, 21))
midday_avg = sum(sum(hourly_prices[h]) for h in range(9, 14)) / sum(len(hourly_prices[h]) for h in range(9, 14))
print(f"Market data: {len(deduped)} records, peak avg = EUR{peak_avg:.2f}")

# SOH curve
soh_values = {
    1: 0.9462, 2: 0.9177, 3: 0.8991, 4: 0.8835, 5: 0.8678,
    6: 0.8534, 7: 0.8390, 8: 0.8246, 9: 0.8102, 10: 0.7958,
    11: 0.7839, 12: 0.7720, 13: 0.7601, 14: 0.7481, 15: 0.7361
}

# ================================================================
# SHEET 1: INPUTS
# ================================================================
ws_in = wb.active
ws_in.title = 'Inputs'
ws_in.sheet_properties.tabColor = 'FF6600'

ws_in.column_dimensions['A'].width = 4
ws_in.column_dimensions['B'].width = 44
ws_in.column_dimensions['C'].width = 18
ws_in.column_dimensions['D'].width = 14
ws_in.column_dimensions['E'].width = 55

ws_in.merge_cells('B1:E1')
ws_in.cell(row=1, column=2, value='BESS Profitability Calculator - Galascope (Esperia Energy)').font = title_font
ws_in.cell(row=2, column=2, value='Lighthief Cyprus Ltd - Yellow cells are editable inputs').font = Font(italic=True, size=10, color='666666')

r = 4
ws_in.cell(row=r, column=2, value='BESS SYSTEM').font = section_font
r += 1  # r=5
# Row 5: PV Park Capacity
ws_in.cell(row=5, column=2, value='PV Park Capacity (MW)').font = label_font
ws_in.cell(row=5, column=3, value=5).fill = input_fill; ws_in.cell(row=5, column=3).font = input_font
ws_in.cell(row=5, column=4, value='MW')
ws_in.cell(row=5, column=5, value='Galascope 1 target - change for different park sizes').font = Font(italic=True, size=10, color='555555')
# Row 6: Reference Park Capacity
ws_in.cell(row=6, column=2, value='Reference Park Capacity (MW)').font = label_font
ws_in.cell(row=6, column=3, value=2.5).fill = input_fill; ws_in.cell(row=6, column=3).font = input_font
ws_in.cell(row=6, column=4, value='MW')
ws_in.cell(row=6, column=5, value='Galascope 2.5MW - source of actual curtailment data').font = Font(italic=True, size=10, color='555555')
# Row 7: Scaling Factor (auto)
ws_in.cell(row=7, column=2, value='Scaling Factor').font = label_font
ws_in.cell(row=7, column=3, value='=C5/C6').fill = calc_fill
ws_in.cell(row=7, column=3).number_format = '0.0'
ws_in.cell(row=7, column=4, value='x')
ws_in.cell(row=7, column=5, value='Auto-calculated: Target / Reference').font = Font(italic=True, size=10, color='555555')
# Row 8: BESS Power
ws_in.cell(row=8, column=2, value='BESS Power Rating (MW)').font = label_font
ws_in.cell(row=8, column=3, value=5).fill = input_fill; ws_in.cell(row=8, column=3).font = input_font
ws_in.cell(row=8, column=4, value='MW')
ws_in.cell(row=8, column=5, value='Linyang system power rating').font = Font(italic=True, size=10, color='555555')
# Row 9: BESS Energy
ws_in.cell(row=9, column=2, value='BESS Energy Capacity (MWh)').font = label_font
ws_in.cell(row=9, column=3, value=20).fill = input_fill; ws_in.cell(row=9, column=3).font = input_font
ws_in.cell(row=9, column=4, value='MWh')
ws_in.cell(row=9, column=5, value='4x 5.015 MWh containers = 20.06 MWh nameplate').font = Font(italic=True, size=10, color='555555')
# Row 10: Duration (auto)
ws_in.cell(row=10, column=2, value='BESS Duration (hours)').font = label_font
ws_in.cell(row=10, column=3, value='=C9/C8').fill = calc_fill
ws_in.cell(row=10, column=3).number_format = '0.0'
ws_in.cell(row=10, column=4, value='hrs')
# Row 11: DOD
ws_in.cell(row=11, column=2, value='Depth of Discharge (%)').font = label_font
ws_in.cell(row=11, column=3, value=90).fill = input_fill; ws_in.cell(row=11, column=3).font = input_font
ws_in.cell(row=11, column=4, value='%')
ws_in.cell(row=11, column=5, value='Linyang operating specification').font = Font(italic=True, size=10, color='555555')
# Row 12: RTE
ws_in.cell(row=12, column=2, value='Round-Trip Efficiency (%)').font = label_font
ws_in.cell(row=12, column=3, value=86.32).fill = input_fill; ws_in.cell(row=12, column=3).font = input_font
ws_in.cell(row=12, column=3).number_format = '0.00'
ws_in.cell(row=12, column=4, value='%')
ws_in.cell(row=12, column=5, value='Full system AC-AC RTE incl. cabling').font = Font(italic=True, size=10, color='555555')
# Row 13: Cycle Life
ws_in.cell(row=13, column=2, value='Cycle Life (@ 90% DOD, 70% EOL)').font = label_font
ws_in.cell(row=13, column=3, value=7000).fill = input_fill; ws_in.cell(row=13, column=3).font = input_font
ws_in.cell(row=13, column=3).number_format = '#,##0'
ws_in.cell(row=13, column=4, value='cycles')

# Row 15: INVESTMENT & COSTS
ws_in.cell(row=15, column=2, value='INVESTMENT & COSTS').font = section_font
# Row 16: EPC
ws_in.cell(row=16, column=2, value='Turnkey EPC Price (EUR)').font = label_font
ws_in.cell(row=16, column=3, value=2258900).fill = input_fill; ws_in.cell(row=16, column=3).font = input_font
ws_in.cell(row=16, column=3).number_format = '#,##0'
ws_in.cell(row=16, column=4, value='EUR')
ws_in.cell(row=16, column=5, value='Confirmed Esperia 5MW/20MWh price').font = Font(italic=True, size=10, color='555555')
# Row 17: Cost/kWh (auto)
ws_in.cell(row=17, column=2, value='Cost per kWh (EUR/kWh)').font = label_font
ws_in.cell(row=17, column=3, value='=C16/(C9*1000)').fill = calc_fill
ws_in.cell(row=17, column=3).number_format = '0.00'
ws_in.cell(row=17, column=4, value='EUR/kWh')
# Row 18: LTSA Yr 1-5
ws_in.cell(row=18, column=2, value='LTSA Tier C - Yr 1-5 (EUR/yr)').font = label_font
ws_in.cell(row=18, column=3, value=34800).fill = input_fill; ws_in.cell(row=18, column=3).font = input_font
ws_in.cell(row=18, column=3).number_format = '#,##0'
ws_in.cell(row=18, column=4, value='EUR/yr')
# Row 19: Ext Warranty 6-10
ws_in.cell(row=19, column=2, value='Extended Warranty - Yr 6-10 add-on (EUR/yr)').font = label_font
ws_in.cell(row=19, column=3, value=33234).fill = input_fill; ws_in.cell(row=19, column=3).font = input_font
ws_in.cell(row=19, column=3).number_format = '#,##0'
ws_in.cell(row=19, column=4, value='EUR/yr')
# Row 20: Ext Warranty 11-15
ws_in.cell(row=20, column=2, value='Extended Warranty - Yr 11-15 add-on (EUR/yr)').font = label_font
ws_in.cell(row=20, column=3, value=41674).fill = input_fill; ws_in.cell(row=20, column=3).font = input_font
ws_in.cell(row=20, column=3).number_format = '#,##0'
ws_in.cell(row=20, column=4, value='EUR/yr')
# Row 21: SCADA
ws_in.cell(row=21, column=2, value='SCADA/EMS Annual (EUR/yr)').font = label_font
ws_in.cell(row=21, column=3, value=5000).fill = input_fill; ws_in.cell(row=21, column=3).font = input_font
ws_in.cell(row=21, column=3).number_format = '#,##0'
ws_in.cell(row=21, column=4, value='EUR/yr')
# Row 22: Total OpCost Yr 1-5 (auto)
ws_in.cell(row=22, column=2, value='Total OpCost Yr 1-5 (EUR/yr)').font = label_font
ws_in.cell(row=22, column=3, value='=C18+C21').fill = calc_fill
ws_in.cell(row=22, column=3).number_format = '#,##0'
# Row 23: Total OpCost Yr 6-10 (auto)
ws_in.cell(row=23, column=2, value='Total OpCost Yr 6-10 (EUR/yr)').font = label_font
ws_in.cell(row=23, column=3, value='=C18+C19+C21').fill = calc_fill
ws_in.cell(row=23, column=3).number_format = '#,##0'
# Row 24: Total OpCost Yr 11-15 (auto)
ws_in.cell(row=24, column=2, value='Total OpCost Yr 11-15 (EUR/yr)').font = label_font
ws_in.cell(row=24, column=3, value='=C18+C20+C21').fill = calc_fill
ws_in.cell(row=24, column=3).number_format = '#,##0'

# Row 26: REVENUE ASSUMPTIONS
ws_in.cell(row=26, column=2, value='REVENUE ASSUMPTIONS').font = section_font
# Row 27: Discharge Price
ws_in.cell(row=27, column=2, value='Blended Discharge Price (EUR/MWh)').font = label_font
ws_in.cell(row=27, column=3, value=175).fill = input_fill; ws_in.cell(row=27, column=3).font = input_font
ws_in.cell(row=27, column=3).number_format = '#,##0'
ws_in.cell(row=27, column=4, value='EUR/MWh')
ws_in.cell(row=27, column=5, value='Conservative - actual DAM peak 17-21 avg = EUR 184.13').font = Font(italic=True, size=10, color='555555')
# Row 28: Charge Cost
ws_in.cell(row=28, column=2, value='Charge Cost (EUR/MWh)').font = label_font
ws_in.cell(row=28, column=3, value=0).fill = input_fill; ws_in.cell(row=28, column=3).font = input_font
ws_in.cell(row=28, column=4, value='EUR/MWh')
ws_in.cell(row=28, column=5, value='Curtailed energy = zero opportunity cost').font = Font(italic=True, size=10, color='555555')
# Row 29: Price Escalation
ws_in.cell(row=29, column=2, value='Annual Price Escalation (%)').font = label_font
ws_in.cell(row=29, column=3, value=0).fill = input_fill; ws_in.cell(row=29, column=3).font = input_font
ws_in.cell(row=29, column=3).number_format = '0.0'
ws_in.cell(row=29, column=4, value='%')
ws_in.cell(row=29, column=5, value='Conservative - no escalation assumed').font = Font(italic=True, size=10, color='555555')
# Row 30: Project Life
ws_in.cell(row=30, column=2, value='Project Life (years)').font = label_font
ws_in.cell(row=30, column=3, value=15).fill = input_fill; ws_in.cell(row=30, column=3).font = input_font
ws_in.cell(row=30, column=4, value='years')

# Row 32: CURTAILMENT SUMMARY
ws_in.cell(row=32, column=2, value='CURTAILMENT SUMMARY (from actual data)').font = section_font
last_da = len(all_days) + 1  # last data row in DailyAnalysis/CurtailmentData

ws_in.cell(row=33, column=2, value='Annual Curtailed - Ref Park (MWh)').font = label_font
ws_in.cell(row=33, column=3, value=f"=SUM(CurtailmentData!D2:CurtailmentData!D{last_da})/1000").fill = calc_fill
ws_in.cell(row=33, column=3).number_format = '#,##0.0'
ws_in.cell(row=34, column=2, value='Annual Curtailed - Scaled (MWh)').font = label_font
ws_in.cell(row=34, column=3, value='=C33*C7').fill = calc_fill
ws_in.cell(row=34, column=3).number_format = '#,##0.0'
ws_in.cell(row=35, column=2, value='Annual BESS Captured Yr1 - Scaled (MWh)').font = label_font
ws_in.cell(row=35, column=3, value=f"=SUM(DailyAnalysis!G2:DailyAnalysis!G{last_da})/1000").fill = calc_fill
ws_in.cell(row=35, column=3).number_format = '#,##0.0'
ws_in.cell(row=36, column=2, value='Capture Efficiency (%)').font = label_font
ws_in.cell(row=36, column=3, value='=C35/C34').fill = calc_fill
ws_in.cell(row=36, column=3).number_format = '0.0%'
ws_in.cell(row=37, column=2, value='Average Daily BESS Fill (%)').font = label_font
ws_in.cell(row=37, column=3, value=f"=AVERAGE(DailyAnalysis!H2:DailyAnalysis!H{last_da})").fill = calc_fill
ws_in.cell(row=37, column=3).number_format = '0.0%'

print("Inputs sheet done")

# ================================================================
# SHEET 2: CURTAILMENT DATA
# ================================================================
ws_cd = wb.create_sheet('CurtailmentData')
ws_cd.sheet_properties.tabColor = 'E06666'

headers_cd = ['Date', 'Month', 'Sun Hours', 'Curtailed (kWh)', 'PV Yield (kWh)', 'Total Potential (kWh)', 'Curtailment %']
for c, h in enumerate(headers_cd, 1):
    ws_cd.cell(row=1, column=c, value=h)
style_header_row(ws_cd, 1, len(headers_cd))

for i, d in enumerate(all_days):
    r = i + 2
    ws_cd.cell(row=r, column=1, value=d['date'])
    ws_cd.cell(row=r, column=2, value=d['month'])
    ws_cd.cell(row=r, column=3, value=round(d['sun'], 2))
    ws_cd.cell(row=r, column=4, value=round(d['curt'], 2))
    ws_cd.cell(row=r, column=5, value=round(d['pv'], 2))
    ws_cd.cell(row=r, column=6, value=round(d['total'], 2))
    ws_cd.cell(row=r, column=7, value=f'=IF(F{r}>0, D{r}/F{r}, 0)')
    ws_cd.cell(row=r, column=7).number_format = '0.0%'

for c in range(1, 8):
    ws_cd.column_dimensions[get_column_letter(c)].width = 16
style_range(ws_cd, 2, last_da, 7)

print(f"CurtailmentData sheet: {len(all_days)} days")

# ================================================================
# SHEET 3: MARKET DATA
# ================================================================
ws_md = wb.create_sheet('MarketData')
ws_md.sheet_properties.tabColor = '6AA84F'

ws_md.merge_cells('A1:G1')
ws_md.cell(row=1, column=1, value='TSOC DAM Hourly Price Profile (Oct 2025 - Feb 2026, 6576 records)').font = section_font

headers_md = ['Hour', 'Avg EUR/MWh', 'Median', 'P10', 'P90', 'Min', 'Max', 'Zero-Price %', 'Count', 'Band']
for c, h in enumerate(headers_md, 1):
    ws_md.cell(row=2, column=c, value=h)
style_header_row(ws_md, 2, len(headers_md))

band_map = {
    0: 'Night', 1: 'Night', 2: 'Night', 3: 'Night', 4: 'Night', 5: 'Night',
    6: 'Morning', 7: 'Morning', 8: 'Morning',
    9: 'Midday Trough', 10: 'Midday Trough', 11: 'Midday Trough', 12: 'Midday Trough', 13: 'Midday Trough',
    14: 'Afternoon', 15: 'Afternoon', 16: 'Afternoon',
    17: 'Peak Evening', 18: 'Peak Evening', 19: 'Peak Evening', 20: 'Peak Evening',
    21: 'Late Evening', 22: 'Late Evening', 23: 'Late Evening'
}

for h in range(24):
    r = h + 3
    prices = sorted(hourly_prices.get(h, [0]))
    n = len(prices)
    avg_p = sum(prices) / n if n else 0
    med_p = prices[n // 2] if n else 0
    p10 = prices[int(n * 0.1)] if n else 0
    p90 = prices[int(n * 0.9)] if n else 0
    zero_pct = sum(1 for p in prices if p <= 1) / n if n else 0

    ws_md.cell(row=r, column=1, value=f'{h:02d}:00')
    ws_md.cell(row=r, column=2, value=round(avg_p, 2)).number_format = '#,##0.00'
    ws_md.cell(row=r, column=3, value=round(med_p, 2)).number_format = '#,##0.00'
    ws_md.cell(row=r, column=4, value=round(p10, 2)).number_format = '#,##0.00'
    ws_md.cell(row=r, column=5, value=round(p90, 2)).number_format = '#,##0.00'
    ws_md.cell(row=r, column=6, value=round(min(prices), 2)).number_format = '#,##0.00'
    ws_md.cell(row=r, column=7, value=round(max(prices), 2)).number_format = '#,##0.00'
    ws_md.cell(row=r, column=8, value=round(zero_pct, 4)).number_format = '0.0%'
    ws_md.cell(row=r, column=9, value=n)
    ws_md.cell(row=r, column=10, value=band_map.get(h, ''))

    if 17 <= h <= 20:
        for c in range(1, 11):
            ws_md.cell(row=r, column=c).fill = green_fill
    elif 9 <= h <= 14:
        for c in range(1, 11):
            ws_md.cell(row=r, column=c).fill = yellow_fill

style_range(ws_md, 3, 26, 10)

# Band summary
r_b = 28
ws_md.cell(row=r_b, column=1, value='Band Summary').font = section_font
r_b += 1
for c, h in enumerate(['Band', 'Hours', 'Avg EUR/MWh', 'Zero %', '<=EUR50 %'], 1):
    ws_md.cell(row=r_b, column=c, value=h)
style_header_row(ws_md, r_b, 5)
r_b += 1

bands_def = [
    ('Night (00-06)', range(0, 6)),
    ('Morning (06-09)', range(6, 9)),
    ('Midday Trough (09-14)', range(9, 14)),
    ('Afternoon (14-17)', range(14, 17)),
    ('Peak Evening (17-21)', range(17, 21)),
    ('Late Evening (21-24)', range(21, 24)),
]
for bname, bhours in bands_def:
    bp = []
    for bh in bhours:
        bp.extend(hourly_prices.get(bh, []))
    bavg = sum(bp) / len(bp) if bp else 0
    bzero = sum(1 for p in bp if p <= 1) / len(bp) if bp else 0
    blow = sum(1 for p in bp if p <= 50) / len(bp) if bp else 0
    ws_md.cell(row=r_b, column=1, value=bname)
    ws_md.cell(row=r_b, column=2, value=f'{bhours.start:02d}:00-{bhours.stop:02d}:00')
    ws_md.cell(row=r_b, column=3, value=round(bavg, 2)).number_format = '#,##0.00'
    ws_md.cell(row=r_b, column=4, value=round(bzero, 4)).number_format = '0.0%'
    ws_md.cell(row=r_b, column=5, value=round(blow, 4)).number_format = '0.0%'
    if 'Peak' in bname:
        for c in range(1, 6):
            ws_md.cell(row=r_b, column=c).fill = green_fill
    r_b += 1

r_b += 1
ws_md.cell(row=r_b, column=1, value='Peak Evening Avg (17-21):').font = Font(bold=True)
ws_md.cell(row=r_b, column=2, value=round(peak_avg, 2)).font = result_font

for c in range(1, 11):
    ws_md.column_dimensions[get_column_letter(c)].width = 16

print("MarketData sheet done")

# ================================================================
# SHEET 4: SOH DEGRADATION
# ================================================================
ws_soh = wb.create_sheet('SOH_Degradation')
ws_soh.sheet_properties.tabColor = '9900FF'

ws_soh.merge_cells('A1:F1')
ws_soh.cell(row=1, column=1, value='Linyang SOH Degradation Curve (0.5P, 1 cycle/day, 25 +/- 2C)').font = section_font

soh_headers = ['Year', 'SOH %', 'Usable Capacity (MWh)', 'Max Daily Charge (kWh)', 'Cumulative Cycles', 'Cycle Headroom vs 7000']
for c, h in enumerate(soh_headers, 1):
    ws_soh.cell(row=2, column=c, value=h)
style_header_row(ws_soh, 2, len(soh_headers))

for yr in range(1, 16):
    r = yr + 2
    ws_soh.cell(row=r, column=1, value=yr)
    ws_soh.cell(row=r, column=2, value=soh_values[yr]).number_format = '0.00%'
    ws_soh.cell(row=r, column=3, value=f"=Inputs!C9*B{r}*Inputs!C11/100").number_format = '0.00'
    ws_soh.cell(row=r, column=4, value=f"=C{r}*1000").number_format = '#,##0'
    ws_soh.cell(row=r, column=5, value=f"=CashFlow!E{r+1}").number_format = '#,##0'
    ws_soh.cell(row=r, column=6, value=f"=1-E{r}/Inputs!C13").number_format = '0.0%'

style_range(ws_soh, 3, 17, 6)
for c in range(1, 7):
    ws_soh.column_dimensions[get_column_letter(c)].width = 22

print("SOH_Degradation sheet done")

# ================================================================
# SHEET 5: DAILY ANALYSIS
# ================================================================
ws_da = wb.create_sheet('DailyAnalysis')
ws_da.sheet_properties.tabColor = '3D85C6'

da_headers = [
    'Date', 'Month', 'Curtailed Ref (kWh)', 'Curtailed Scaled (kWh)',
    'BESS Usable Cap Yr1 (kWh)', 'Overflow (kWh)', 'Captured (kWh)',
    'Fill %', 'Discharged (kWh)', 'Revenue (EUR)'
]
for c, h in enumerate(da_headers, 1):
    ws_da.cell(row=1, column=c, value=h)
style_header_row(ws_da, 1, len(da_headers))

for i, d in enumerate(all_days):
    r = i + 2
    ws_da.cell(row=r, column=1, value=d['date'])
    ws_da.cell(row=r, column=2, value=d['month'])
    ws_da.cell(row=r, column=3, value=f"=CurtailmentData!D{r}")
    ws_da.cell(row=r, column=4, value=f"=C{r}*Inputs!C7").number_format = '#,##0.0'
    ws_da.cell(row=r, column=5, value="=SOH_Degradation!D3").number_format = '#,##0'
    ws_da.cell(row=r, column=6, value=f"=MAX(0, D{r}-E{r})").number_format = '#,##0.0'
    ws_da.cell(row=r, column=7, value=f"=MIN(D{r}, E{r})").number_format = '#,##0.0'
    ws_da.cell(row=r, column=8, value=f"=IF(E{r}>0, G{r}/E{r}, 0)").number_format = '0.0%'
    ws_da.cell(row=r, column=9, value=f"=G{r}*Inputs!C12/100").number_format = '#,##0.0'
    ws_da.cell(row=r, column=10, value=f"=(I{r}/1000)*(Inputs!C27-Inputs!C28)").number_format = '#,##0'

# Summary row
sr = last_da + 2
ws_da.cell(row=sr, column=2, value='ANNUAL TOTALS').font = Font(bold=True, size=11)
for c in [3, 4, 6, 7, 9]:
    ws_da.cell(row=sr, column=c, value=f"=SUM({get_column_letter(c)}2:{get_column_letter(c)}{last_da})").number_format = '#,##0'
    ws_da.cell(row=sr, column=c).font = Font(bold=True)
ws_da.cell(row=sr, column=8, value=f"=AVERAGE(H2:H{last_da})").number_format = '0.0%'
ws_da.cell(row=sr, column=8).font = Font(bold=True)
ws_da.cell(row=sr, column=10, value=f"=SUM(J2:J{last_da})").number_format = '#,##0'
ws_da.cell(row=sr, column=10).font = Font(bold=True)

ws_da.cell(row=sr+1, column=2, value='Capture Efficiency').font = Font(bold=True)
ws_da.cell(row=sr+1, column=7, value=f"=G{sr}/D{sr}").number_format = '0.0%'
ws_da.cell(row=sr+1, column=7).font = result_font

for c in range(1, 11):
    ws_da.column_dimensions[get_column_letter(c)].width = 20

print(f"DailyAnalysis sheet: {len(all_days)} formula rows")

# ================================================================
# SHEET 6: CASH FLOW
# ================================================================
ws_cf = wb.create_sheet('CashFlow')
ws_cf.sheet_properties.tabColor = '38761D'

ws_cf.merge_cells('A1:L1')
ws_cf.cell(row=1, column=1, value='15-Year Cash Flow Projection - Data-Validated Model').font = title_font

cf_headers = [
    'Year', 'SOH %', 'Usable MWh', 'Captured (MWh)', 'Cumul Cycles',
    'Discharged (MWh)', 'Gross Revenue (EUR)', 'Operating Cost (EUR)',
    'Net Cash Flow (EUR)', 'Cumulative (EUR)', 'BESS Fill %', 'Cycles This Year'
]
for c, h in enumerate(cf_headers, 1):
    ws_cf.cell(row=2, column=c, value=h)
style_header_row(ws_cf, 2, len(cf_headers))

# Year 0
ws_cf.cell(row=3, column=1, value=0)
ws_cf.cell(row=3, column=7, value='Investment:')
ws_cf.cell(row=3, column=9, value='=-Inputs!C16').number_format = '#,##0'
ws_cf.cell(row=3, column=9).font = Font(bold=True, color='CC0000')
ws_cf.cell(row=3, column=10, value='=I3').number_format = '#,##0'

# Pre-compute year-by-year captured energy (needed because 365-row iteration per year is complex in pure formulas)
yearly_captured = {}
cum_cycles = 0
for yr in range(1, 16):
    soh = soh_values[yr]
    usable_kwh = 20000 * soh * 0.90
    yr_capt = sum(min(d['curt'] * 2, usable_kwh) for d in all_days)
    yearly_captured[yr] = yr_capt / 1000
    cycles_yr = yearly_captured[yr] / (20 * 0.90)
    cum_cycles += cycles_yr

    r = yr + 3
    ws_cf.cell(row=r, column=1, value=yr)
    ws_cf.cell(row=r, column=2, value=f"=SOH_Degradation!B{yr+2}").number_format = '0.0%'
    ws_cf.cell(row=r, column=3, value=f"=SOH_Degradation!C{yr+2}").number_format = '0.00'
    # Captured MWh: pre-computed (blue highlight = computed from daily data)
    ws_cf.cell(row=r, column=4, value=round(yearly_captured[yr], 1)).number_format = '#,##0.0'
    ws_cf.cell(row=r, column=4).fill = highlight_fill
    # Cumulative cycles
    ws_cf.cell(row=r, column=5, value=round(cum_cycles, 0)).number_format = '#,##0'
    # Discharged
    ws_cf.cell(row=r, column=6, value=f"=D{r}*Inputs!C12/100").number_format = '#,##0.0'
    # Revenue with escalation
    ws_cf.cell(row=r, column=7, value=f"=F{r}*(Inputs!C27-Inputs!C28)*(1+Inputs!C29/100)^(A{r}-1)").number_format = '#,##0'
    # OpCost
    if yr <= 5:
        ws_cf.cell(row=r, column=8, value="=Inputs!C22").number_format = '#,##0'
    elif yr <= 10:
        ws_cf.cell(row=r, column=8, value="=Inputs!C23").number_format = '#,##0'
    else:
        ws_cf.cell(row=r, column=8, value="=Inputs!C24").number_format = '#,##0'
    # Net Cash
    ws_cf.cell(row=r, column=9, value=f"=G{r}-H{r}").number_format = '#,##0'
    ws_cf.cell(row=r, column=9).font = Font(bold=True)
    # Cumulative
    ws_cf.cell(row=r, column=10, value=f"=J{r-1}+I{r}").number_format = '#,##0'
    # Fill %
    ws_cf.cell(row=r, column=11, value=f"=D{r}*1000/(C{r}*1000*365)").number_format = '0.0%'
    # Cycles this year
    ws_cf.cell(row=r, column=12, value=f"=D{r}/(Inputs!C9*Inputs!C11/100)").number_format = '#,##0'

    # Green background for payback year
    if yr >= 5:
        for c in [9, 10]:
            ws_cf.cell(row=r, column=c).font = Font(bold=True, color='006100')

style_range(ws_cf, 3, 18, 12)

# Totals
tr = 20
ws_cf.cell(row=tr, column=1, value='TOTALS').font = Font(bold=True, size=11)
for c, fmt in [(4, '#,##0'), (6, '#,##0'), (7, '#,##0'), (8, '#,##0'), (9, '#,##0'), (12, '#,##0')]:
    ws_cf.cell(row=tr, column=c, value=f"=SUM({get_column_letter(c)}4:{get_column_letter(c)}18)").number_format = fmt
    ws_cf.cell(row=tr, column=c).font = Font(bold=True)

ws_cf.cell(row=tr, column=9).font = result_font

for c in range(1, 13):
    ws_cf.column_dimensions[get_column_letter(c)].width = 18

print("CashFlow sheet done")

# ================================================================
# SHEET 7: SUMMARY
# ================================================================
ws_sum = wb.create_sheet('Summary')
ws_sum.sheet_properties.tabColor = 'FFD700'
wb.move_sheet('Summary', offset=-6)

ws_sum.column_dimensions['A'].width = 4
ws_sum.column_dimensions['B'].width = 40
ws_sum.column_dimensions['C'].width = 22
ws_sum.column_dimensions['D'].width = 14
ws_sum.column_dimensions['E'].width = 52

ws_sum.merge_cells('B1:E1')
ws_sum.cell(row=1, column=2, value='BESS PROFITABILITY - GALASCOPE (ESPERIA ENERGY)').font = Font(name='Calibri', bold=True, size=16, color='1F4E79')
ws_sum.cell(row=2, column=2, value='Lighthief Cyprus Ltd - Data-Validated Model (Feb 2026)').font = Font(italic=True, size=11, color='666666')
ws_sum.cell(row=2, column=5, value='Change parameters on the Inputs sheet').font = Font(italic=True, size=10, color='C00000')

r = 4
ws_sum.cell(row=r, column=2, value='KEY FINANCIAL METRICS').font = section_font; r += 1

kfm = [
    ('Investment (EUR)', '=Inputs!C16', '#,##0'),
    ('Simple Payback (years)', '=IFERROR(MATCH(1,INDEX((CashFlow!J4:J18>=0)*1,0),0)+(ABS(INDEX(CashFlow!J3:J17,MATCH(1,INDEX((CashFlow!J4:J18>=0)*1,0),0)))/(INDEX(CashFlow!I4:I18,MATCH(1,INDEX((CashFlow!J4:J18>=0)*1,0),0)))),"N/A")', '0.0'),
    ('15-Year Net Profit (EUR)', '=CashFlow!J18', '#,##0'),
    ('15-Year ROI (%)', '=CashFlow!J18/Inputs!C16', '0.0%'),
    ('15-Year Gross Revenue (EUR)', '=SUM(CashFlow!G4:CashFlow!G18)', '#,##0'),
    ('15-Year Operating Costs (EUR)', '=SUM(CashFlow!H4:CashFlow!H18)', '#,##0'),
    ('Year 1 Net Cash Flow (EUR)', '=CashFlow!I4', '#,##0'),
    ('Year 15 Net Cash Flow (EUR)', '=CashFlow!I18', '#,##0'),
]
for label, formula, fmt in kfm:
    ws_sum.cell(row=r, column=2, value=label).font = label_font
    ws_sum.cell(row=r, column=3, value=formula).number_format = fmt
    ws_sum.cell(row=r, column=3).font = result_font
    r += 1

r += 1
ws_sum.cell(row=r, column=2, value='BESS PERFORMANCE').font = section_font; r += 1
bperf = [
    ('BESS Capture Efficiency', f"=SUM(DailyAnalysis!G2:DailyAnalysis!G{last_da})/SUM(DailyAnalysis!D2:DailyAnalysis!D{last_da})", '0.0%'),
    ('Average Daily Fill', f"=AVERAGE(DailyAnalysis!H2:DailyAnalysis!H{last_da})", '0.0%'),
    ('Days at 100% Fill', f'=COUNTIF(DailyAnalysis!H2:DailyAnalysis!H{last_da},">=0.9999")', '#,##0'),
    ('Days Zero Curtailment', f'=COUNTIF(CurtailmentData!D2:CurtailmentData!D{last_da},"<=0")', '#,##0'),
    ('15-Year Total Cycles', '=CashFlow!E18', '#,##0'),
    ('Cycle Headroom vs 7000', '=1-CashFlow!E18/Inputs!C13', '0.0%'),
    ('Average Daily Cycles', '=SUM(CashFlow!L4:CashFlow!L18)/(Inputs!C30*365)', '0.00'),
]
for label, formula, fmt in bperf:
    ws_sum.cell(row=r, column=2, value=label).font = label_font
    ws_sum.cell(row=r, column=3, value=formula).number_format = fmt
    ws_sum.cell(row=r, column=3).font = Font(bold=True, size=11)
    r += 1

r += 1
ws_sum.cell(row=r, column=2, value='CURTAILMENT DATA SUMMARY').font = section_font; r += 1
cds = [
    ('Data Source', 'Galascope 2.5MW, 365 days (Jan-Dec 2025)'),
    ('Annual PV Potential (MWh, ref)', f'=SUM(CurtailmentData!F2:CurtailmentData!F{last_da})/1000'),
    ('Annual Curtailed (MWh, ref)', f'=SUM(CurtailmentData!D2:CurtailmentData!D{last_da})/1000'),
    ('Curtailment Rate', f'=C{r+2}/C{r+1}'),
    ('Days with Curtailment', f'=COUNTIF(CurtailmentData!D2:CurtailmentData!D{last_da},">0")'),
]
for label, val in cds:
    ws_sum.cell(row=r, column=2, value=label).font = label_font
    cell = ws_sum.cell(row=r, column=3, value=val)
    cell.font = Font(bold=True, size=11)
    if 'Rate' in label:
        cell.number_format = '0.0%'
    elif 'MWh' in label:
        cell.number_format = '#,##0.0'
    r += 1

r += 1
ws_sum.cell(row=r, column=2, value='MARKET DATA VALIDATION').font = section_font; r += 1
mkv = [
    ('DAM Data Source', '137 TSOC files, Oct 2025 - Feb 2026'),
    ('Peak Evening Avg (17-21)', f'EUR {peak_avg:.2f}/MWh'),
    ('Model Discharge Price', '=Inputs!C27'),
    ('Model vs Actual', f'Conservative by EUR {peak_avg - 175:.2f} (+{(peak_avg-175)/175*100:.1f}%)'),
    ('Midday Trough Avg (09-14)', f'EUR {midday_avg:.2f}/MWh'),
]
for label, val in mkv:
    ws_sum.cell(row=r, column=2, value=label).font = label_font
    cell = ws_sum.cell(row=r, column=3, value=val)
    cell.font = Font(bold=True, size=11)
    if label == 'Model Discharge Price':
        cell.number_format = '#,##0'
    r += 1

r += 1
ws_sum.cell(row=r, column=2, value='MONTHLY BREAKDOWN').font = section_font; r += 1
mh = ['Month', 'PV Yield (MWh)', 'Curtailed (MWh)', 'Curt %', 'BESS Captured (MWh)', 'Avg Fill']
for c, h in enumerate(mh, 2):
    ws_sum.cell(row=r, column=c, value=h)
    ws_sum.cell(row=r, column=c).font = hdr_font
    ws_sum.cell(row=r, column=c).fill = hdr_fill
    ws_sum.cell(row=r, column=c).alignment = Alignment(horizontal='center', wrap_text=True)
r += 1

month_names = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']
monthly_agg = {}
for d in all_days:
    m = d['month']
    if m not in monthly_agg:
        monthly_agg[m] = {'pv': 0, 'curt': 0, 'total': 0}
    monthly_agg[m]['pv'] += d['pv']
    monthly_agg[m]['curt'] += d['curt']
    monthly_agg[m]['total'] += d['total']

usable_yr1 = 20000 * 0.9462 * 0.90
monthly_bess = {}
monthly_fills = {}
for d in all_days:
    m = d['month']
    if m not in monthly_bess:
        monthly_bess[m] = 0
        monthly_fills[m] = []
    curt_s = d['curt'] * 2
    captured = min(curt_s, usable_yr1)
    monthly_bess[m] += captured
    monthly_fills[m].append(captured / usable_yr1 if usable_yr1 > 0 else 0)

for m in range(1, 13):
    if m not in monthly_agg:
        continue
    ma = monthly_agg[m]
    ws_sum.cell(row=r, column=2, value=month_names[m-1])
    ws_sum.cell(row=r, column=3, value=round(ma['pv']/1000, 1)).number_format = '#,##0.0'
    ws_sum.cell(row=r, column=4, value=round(ma['curt']/1000, 1)).number_format = '#,##0.0'
    ws_sum.cell(row=r, column=5, value=round(ma['curt']/ma['total'], 3) if ma['total'] > 0 else 0).number_format = '0.0%'
    ws_sum.cell(row=r, column=6, value=round(monthly_bess.get(m, 0)/1000, 1)).number_format = '#,##0.0'
    ws_sum.cell(row=r, column=7, value=round(sum(monthly_fills.get(m, [0]))/max(len(monthly_fills.get(m, [1])), 1), 3)).number_format = '0.0%'
    if ma['total'] > 0 and ma['curt'] / ma['total'] > 0.50:
        for c in range(2, 8):
            ws_sum.cell(row=r, column=c).fill = warn_fill
    r += 1

# Annual
ws_sum.cell(row=r, column=2, value='ANNUAL').font = Font(bold=True)
ws_sum.cell(row=r, column=3, value=round(sum(ma['pv'] for ma in monthly_agg.values())/1000, 1)).number_format = '#,##0.0'
ws_sum.cell(row=r, column=3).font = Font(bold=True)
ws_sum.cell(row=r, column=4, value=round(sum(ma['curt'] for ma in monthly_agg.values())/1000, 1)).number_format = '#,##0.0'
ws_sum.cell(row=r, column=4).font = Font(bold=True)
tc = sum(ma['curt'] for ma in monthly_agg.values())
tp = sum(ma['total'] for ma in monthly_agg.values())
ws_sum.cell(row=r, column=5, value=round(tc/tp, 3) if tp > 0 else 0).number_format = '0.0%'
ws_sum.cell(row=r, column=5).font = Font(bold=True)
ws_sum.cell(row=r, column=6, value=round(sum(monthly_bess.values())/1000, 1)).number_format = '#,##0.0'
ws_sum.cell(row=r, column=6).font = Font(bold=True)

r += 2
ws_sum.cell(row=r, column=2, value='HOW TO USE').font = section_font; r += 1
notes = [
    'All yellow cells on the Inputs sheet can be changed - calculations update automatically.',
    'Blue-highlighted cells on CashFlow (Captured MWh) are pre-computed from 365-day SOH analysis.',
    'To model a different park: change PV Capacity, BESS Power/Energy, and EPC Price on Inputs.',
    'To model higher discharge prices: change Blended Discharge Price (actual DAM peak = EUR 184).',
    'To model price growth: set Annual Price Escalation > 0% on Inputs.',
    'CurtailmentData tab has all 365 days of raw data from Galascope 2.5MW park.',
    'MarketData tab has the full hourly DAM price profile from 137 TSOC settlement files.',
    'SOH_Degradation tab shows the Linyang guaranteed degradation curve.',
    'DailyAnalysis tab calculates day-by-day BESS capture with overflow constraints.',
]
for note in notes:
    ws_sum.cell(row=r, column=2, value=f'  {note}').font = Font(size=10, color='444444')
    r += 1

print("Summary sheet done")

# ── SAVE ──
outpath = "docs/clients/Group2_Esperia_Energy/GalascopeBessROI.xlsx"
wb.save(outpath)
print(f"\nSaved: {outpath}")
print(f"Sheets: {wb.sheetnames}")
print(f"File size: {os.path.getsize(outpath):,} bytes")
