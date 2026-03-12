"""
Build combined market pricing + curtailment Excel file.

Data Sources:
1. TSOC DAM (Day-Ahead Market) half-hourly Excel reports from market/excel/
   - Source: https://tsoc.org.cy/competitive-electricity-market/mms-reports/day-ahead-market-daily-activity-reports-en/
   - Period: Oct 1, 2025 onwards (open market start)
   - Granularity: 48 half-hourly periods per day
   - Fields: Period, Energy Type (OIL/SOLAR/ALL), MCP (€/MWh), Sale/Purchase Volume (MWh)

2. Esperia Galascope 2.5MW PV Plant curtailment reports
   - Source: Dino / Esperia Energy (Galascope monitoring portal)
   - Period: Jan 2025 - Dec 2025
   - Granularity: Daily
   - Fields: Date, Irradiation, Sun Hours, PV Yield (kWh), Export Limitation Loss (kWh)
"""

import os
import json
from datetime import datetime, timedelta
from collections import OrderedDict

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter

BASE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
EXCEL_DIR = os.path.join(BASE, "market", "excel")
CURTAILMENT_DIR = os.path.join(
    BASE,
    "docs", "clients", "Group2_Esperia_Energy",
    "esperia-galascope-2.5-curtailment",
)
OUTPUT = os.path.join(BASE, "market", "marketprice-curtailment.xlsx")

# ── Styles ──

HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
HEADER_FILL2 = PatternFill(start_color="548235", end_color="548235", fill_type="solid")
HEADER_FILL3 = PatternFill(start_color="BF8F00", end_color="BF8F00", fill_type="solid")
LIGHT_BLUE = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
LIGHT_GREEN = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
LIGHT_YELLOW = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
RED_FILL = PatternFill(start_color="FCE4EC", end_color="FCE4EC", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)
NUM_2DP = numbers.FORMAT_NUMBER_COMMA_SEPARATED1  # #,##0.00
NUM_0DP = "#,##0"


def style_header(ws, row, max_col, fill=HEADER_FILL):
    for c in range(1, max_col + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = HEADER_FONT
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = THIN_BORDER


def auto_width(ws, max_col, min_width=10, max_width=25):
    for c in range(1, max_col + 1):
        col_letter = get_column_letter(c)
        max_len = min_width
        for row in ws.iter_rows(min_col=c, max_col=c, values_only=False):
            for cell in row:
                if cell.value:
                    max_len = max(max_len, min(len(str(cell.value)), max_width))
        ws.column_dimensions[col_letter].width = max_len + 2


# ── 1. Parse TSOC Excel files ──

def parse_tsoc_files():
    """Parse all TSOC DAM Excel files into half-hourly records."""
    records = OrderedDict()  # key: (date_str, period_str) -> record

    files = sorted(
        f for f in os.listdir(EXCEL_DIR)
        if f.endswith(".xlsx") and not f.startswith("._")
    )
    print(f"Parsing {len(files)} TSOC DAM Excel files...")

    for fname in files:
        fpath = os.path.join(EXCEL_DIR, fname)
        try:
            wb = openpyxl.load_workbook(fpath, read_only=True, data_only=True)
        except Exception as e:
            print(f"  SKIP {fname}: {e}")
            continue

        ws = wb.active
        if ws is None or ws.max_row is None or ws.max_row < 3:
            wb.close()
            continue

        for row in ws.iter_rows(min_row=3, values_only=True):
            if not row or not row[0]:
                continue

            period_str = str(row[0]).strip()
            energy_type = str(row[1]).strip() if len(row) > 1 and row[1] else ""
            mcp = row[2] if len(row) > 2 else None
            sale_vol = row[3] if len(row) > 3 else None
            buy_vol = row[4] if len(row) > 4 else None
            curtailment_flag = row[5] if len(row) > 5 else None

            if mcp is None:
                continue
            try:
                mcp = float(mcp)
            except (ValueError, TypeError):
                continue

            # Extract date and hour from period like "03/01/2026 00:00-03/01/2026 00:30"
            parts = period_str.split(" ")
            if len(parts) < 2:
                continue
            date_str = parts[0]  # DD/MM/YYYY
            time_str = parts[1].split("-")[0] if "-" in parts[1] else parts[1]

            # Parse to sortable date
            try:
                dt = datetime.strptime(date_str, "%d/%m/%Y")
                iso_date = dt.strftime("%Y-%m-%d")
            except ValueError:
                continue

            key = (iso_date, period_str, energy_type)

            sale_vol_f = 0.0
            buy_vol_f = 0.0
            try:
                sale_vol_f = float(sale_vol) if sale_vol else 0.0
            except (ValueError, TypeError):
                pass
            try:
                buy_vol_f = float(buy_vol) if buy_vol else 0.0
            except (ValueError, TypeError):
                pass

            curt = None
            if curtailment_flag is not None:
                try:
                    curt = int(float(curtailment_flag))
                except (ValueError, TypeError):
                    curt = None

            records[key] = {
                "iso_date": iso_date,
                "date_str": date_str,
                "period": period_str,
                "time": time_str,
                "energy_type": energy_type,
                "mcp": mcp,
                "sale_vol": sale_vol_f,
                "buy_vol": buy_vol_f,
                "curtailment_flag": curt,
            }

        wb.close()

    all_recs = sorted(records.values(), key=lambda r: (r["iso_date"], r["time"], r["energy_type"]))
    print(f"  Total half-hourly records: {len(all_recs)}")

    # Also build a deduplicated "ALL" energy type view for hourly summary
    all_only = [r for r in all_recs if r["energy_type"] == "ALL"]
    print(f"  ALL-type records (unique half-hours): {len(all_only)}")

    dates = sorted(set(r["iso_date"] for r in all_recs))
    print(f"  Date range: {dates[0]} to {dates[-1]} ({len(dates)} days)")

    return all_recs, all_only, dates


# ── 2. Parse curtailment files ──

def parse_curtailment_files():
    """Parse all Galascope 2.5MW curtailment monthly Excel files."""
    records = []

    months = sorted(
        f for f in os.listdir(CURTAILMENT_DIR)
        if f.endswith(".xlsx") and not f.startswith("._")
    )
    print(f"\nParsing {len(months)} curtailment files...")

    for fname in months:
        fpath = os.path.join(CURTAILMENT_DIR, fname)
        try:
            wb = openpyxl.load_workbook(fpath, read_only=True, data_only=True)
        except Exception as e:
            print(f"  SKIP {fname}: {e}")
            continue

        ws = wb.active
        if ws is None:
            wb.close()
            continue

        # Find the header row to identify column positions
        headers = []
        header_row = 2  # typically row 2
        for row in ws.iter_rows(min_row=1, max_row=5, values_only=True):
            if row and row[0] and "period" in str(row[0]).lower():
                headers = [str(c).strip() if c else "" for c in row]
                break
            header_row += 1

        # Find relevant column indices
        pv_yield_col = None
        loss_col = None
        irrad_col = None
        sun_hours_col = None

        for i, h in enumerate(headers):
            hl = h.lower()
            if "loss" in hl and "export" in hl:
                loss_col = i
            elif "pv yield" in hl or (("yield" in hl or "pv" in hl) and "inverter" not in hl and "theoretical" not in hl and "total" not in hl and "specific" not in hl):
                pv_yield_col = i
            elif "irradiation" in hl or "irrad" in hl:
                irrad_col = i
            elif "sun hour" in hl:
                sun_hours_col = i

        if loss_col is None:
            # Try last column as curtailment
            for i, h in enumerate(headers):
                if h and "loss" in h.lower():
                    loss_col = i

        for row in ws.iter_rows(min_row=3, values_only=True):
            if not row or not row[0]:
                continue

            date_val = row[0]
            if isinstance(date_val, datetime):
                date_str = date_val.strftime("%Y-%m-%d")
            elif isinstance(date_val, str) and len(date_val) >= 10:
                date_str = date_val[:10]
            else:
                continue

            # Skip formula rows (summary)
            if isinstance(date_val, str) and "=" in date_val:
                continue

            pv_yield = None
            loss = None
            irrad = None
            sun_h = None

            try:
                if pv_yield_col is not None and row[pv_yield_col] is not None:
                    val = row[pv_yield_col]
                    if not isinstance(val, str) or "=" not in val:
                        pv_yield = float(val)
            except (ValueError, TypeError, IndexError):
                pass

            try:
                if loss_col is not None and len(row) > loss_col and row[loss_col] is not None:
                    val = row[loss_col]
                    if not isinstance(val, str) or "=" not in val:
                        loss = float(val)
            except (ValueError, TypeError, IndexError):
                pass

            try:
                if irrad_col is not None and row[irrad_col] is not None:
                    irrad = float(row[irrad_col])
            except (ValueError, TypeError, IndexError):
                pass

            try:
                if sun_hours_col is not None and row[sun_hours_col] is not None:
                    sun_h = float(row[sun_hours_col])
            except (ValueError, TypeError, IndexError):
                pass

            records.append({
                "date": date_str,
                "pv_yield_kwh": pv_yield,
                "curtailment_kwh": loss if loss and loss > 0 else 0.0,
                "irradiation": irrad,
                "sun_hours": sun_h,
            })

        wb.close()
        print(f"  Parsed: {fname} ({sum(1 for r in records if r['date'].startswith(fname.split(' ')[1].split('.')[0][:4]))} total)")

    records.sort(key=lambda r: r["date"])
    print(f"  Total curtailment days: {len(records)}")
    if records:
        print(f"  Date range: {records[0]['date']} to {records[-1]['date']}")
    return records


# ── 3. Build Excel ──

def build_excel(dam_records, dam_all, dam_dates, curtailment_records):
    print(f"\nBuilding Excel workbook...")
    wb = openpyxl.Workbook()

    # ── Sheet 1: Half-Hourly DAM Pricing (ALL energy type) ──
    ws1 = wb.active
    ws1.title = "DAM Half-Hourly Pricing"

    # Title row
    ws1.merge_cells("A1:H1")
    ws1["A1"] = "Cyprus Day-Ahead Market (DAM) - Half-Hourly Market Clearing Price"
    ws1["A1"].font = Font(name="Calibri", bold=True, size=14, color="2F5496")

    ws1.merge_cells("A2:H2")
    ws1["A2"] = f"Source: TSOC (tsoc.org.cy) | Open Market Period: Oct 2025 onwards | Data: {dam_dates[0]} to {dam_dates[-1]} | {len(dam_all)} periods"
    ws1["A2"].font = Font(name="Calibri", italic=True, size=10, color="808080")

    headers1 = [
        "Date", "Period", "Time (Start)", "Time (End)",
        "MCP (€/MWh)", "Total Sale Volume (MWh)",
        "Total Purchase Volume (MWh)", "Hour (0-23)"
    ]
    for c, h in enumerate(headers1, 1):
        ws1.cell(row=4, column=c, value=h)
    style_header(ws1, 4, len(headers1))

    row_num = 5
    for r in dam_all:
        period = r["period"]
        # Extract start and end times
        time_parts = period.split("-")
        start_time = ""
        end_time = ""
        if len(time_parts) >= 2:
            sp = time_parts[0].strip().split(" ")
            start_time = sp[-1] if len(sp) > 1 else sp[0]
            ep = time_parts[1].strip().split(" ")
            end_time = ep[-1] if len(ep) > 1 else ep[0]

        try:
            hour = int(start_time.split(":")[0])
        except (ValueError, IndexError):
            hour = 0

        ws1.cell(row=row_num, column=1, value=r["iso_date"])
        ws1.cell(row=row_num, column=2, value=period)
        ws1.cell(row=row_num, column=3, value=start_time)
        ws1.cell(row=row_num, column=4, value=end_time)
        ws1.cell(row=row_num, column=5, value=r["mcp"]).number_format = NUM_2DP
        ws1.cell(row=row_num, column=6, value=r["sale_vol"]).number_format = NUM_2DP
        ws1.cell(row=row_num, column=7, value=r["buy_vol"]).number_format = NUM_2DP
        ws1.cell(row=row_num, column=8, value=hour)

        # Alternate row shading
        if row_num % 2 == 0:
            for c in range(1, len(headers1) + 1):
                ws1.cell(row=row_num, column=c).fill = LIGHT_BLUE
        for c in range(1, len(headers1) + 1):
            ws1.cell(row=row_num, column=c).border = THIN_BORDER

        row_num += 1

    auto_width(ws1, len(headers1))
    ws1.freeze_panes = "A5"
    ws1.auto_filter.ref = f"A4:H{row_num - 1}"
    print(f"  Sheet 1: {row_num - 5} half-hourly pricing rows")

    # ── Sheet 2: Daily Curtailment (Galascope 2.5MW) ──
    ws2 = wb.create_sheet("Daily Curtailment (Galascope)")

    ws2.merge_cells("A1:G1")
    ws2["A1"] = "Esperia Galascope 2.5MW PV Plant - Daily Curtailment (Export Limitation Loss)"
    ws2["A1"].font = Font(name="Calibri", bold=True, size=14, color="548235")

    ws2.merge_cells("A2:G2")
    ws2["A2"] = "Source: Dino / Esperia Energy (Galascope monitoring portal) | Plant: 2,512.3 kWp | Jan-Dec 2025"
    ws2["A2"].font = Font(name="Calibri", italic=True, size=10, color="808080")

    headers2 = [
        "Date", "PV Yield (kWh)", "Curtailment Loss (kWh)",
        "Total Production (kWh)", "Curtailment %",
        "Irradiation (kWh/m²)", "Sun Hours (H)"
    ]
    for c, h in enumerate(headers2, 1):
        ws2.cell(row=4, column=c, value=h)
    style_header(ws2, 4, len(headers2), fill=HEADER_FILL2)

    row_num = 5
    for r in curtailment_records:
        pv = r["pv_yield_kwh"] or 0
        curt = r["curtailment_kwh"] or 0
        total = pv + curt
        curt_pct = (curt / total * 100) if total > 0 else 0

        ws2.cell(row=row_num, column=1, value=r["date"])
        ws2.cell(row=row_num, column=2, value=pv).number_format = NUM_2DP
        ws2.cell(row=row_num, column=3, value=curt).number_format = NUM_2DP
        ws2.cell(row=row_num, column=4, value=total).number_format = NUM_2DP
        ws2.cell(row=row_num, column=5, value=curt_pct / 100).number_format = "0.0%"
        ws2.cell(row=row_num, column=6, value=r["irradiation"]).number_format = "0.000" if r["irradiation"] else NUM_2DP
        ws2.cell(row=row_num, column=7, value=r["sun_hours"]).number_format = "0.00" if r["sun_hours"] else NUM_2DP

        # Highlight high curtailment days
        if curt_pct > 50:
            for c in range(1, len(headers2) + 1):
                ws2.cell(row=row_num, column=c).fill = RED_FILL
        elif row_num % 2 == 0:
            for c in range(1, len(headers2) + 1):
                ws2.cell(row=row_num, column=c).fill = LIGHT_GREEN

        for c in range(1, len(headers2) + 1):
            ws2.cell(row=row_num, column=c).border = THIN_BORDER

        row_num += 1

    # Summary row
    last_data = row_num - 1
    ws2.cell(row=row_num, column=1, value="TOTALS").font = Font(bold=True)
    ws2.cell(row=row_num, column=2, value=f"=SUM(B5:B{last_data})").number_format = NUM_0DP
    ws2.cell(row=row_num, column=3, value=f"=SUM(C5:C{last_data})").number_format = NUM_0DP
    ws2.cell(row=row_num, column=4, value=f"=SUM(D5:D{last_data})").number_format = NUM_0DP
    ws2.cell(row=row_num, column=5, value=f"=C{row_num}/D{row_num}").number_format = "0.0%"
    for c in range(1, len(headers2) + 1):
        ws2.cell(row=row_num, column=c).font = Font(bold=True)
        ws2.cell(row=row_num, column=c).border = THIN_BORDER

    auto_width(ws2, len(headers2))
    ws2.freeze_panes = "A5"
    ws2.auto_filter.ref = f"A4:G{last_data}"
    print(f"  Sheet 2: {last_data - 4} curtailment days")

    # ── Sheet 3: Daily Summary (Pricing + Curtailment merged) ──
    ws3 = wb.create_sheet("Daily Summary (Combined)")

    ws3.merge_cells("A1:L1")
    ws3["A1"] = "Combined Daily View: DAM Pricing + Curtailment Data"
    ws3["A1"].font = Font(name="Calibri", bold=True, size=14, color="BF8F00")

    ws3.merge_cells("A2:L2")
    ws3["A2"] = "Pricing from TSOC DAM (Oct 2025+) | Curtailment from Galascope 2.5MW (Jan-Dec 2025) | Overlap: Oct-Dec 2025"
    ws3["A2"].font = Font(name="Calibri", italic=True, size=10, color="808080")

    headers3 = [
        "Date", "Day of Week",
        "Avg MCP (€/MWh)", "Min MCP (€/MWh)", "Max MCP (€/MWh)",
        "Solar Hours MCP (€/MWh)\n(06:00-18:00)",
        "Peak Hours MCP (€/MWh)\n(17:00-21:00)",
        "Arbitrage Spread\n(Peak - Solar)",
        "PV Yield (kWh)", "Curtailment Loss (kWh)",
        "Curtailment %", "Data Sources"
    ]
    for c, h in enumerate(headers3, 1):
        ws3.cell(row=4, column=c, value=h)
    style_header(ws3, 4, len(headers3), fill=HEADER_FILL3)

    # Build daily pricing aggregates
    daily_pricing = {}
    for r in dam_all:
        d = r["iso_date"]
        if d not in daily_pricing:
            daily_pricing[d] = {"prices": [], "solar_prices": [], "peak_prices": []}
        daily_pricing[d]["prices"].append(r["mcp"])
        try:
            h = int(r["time"].split(":")[0])
        except (ValueError, IndexError):
            h = 0
        if 6 <= h <= 18:
            daily_pricing[d]["solar_prices"].append(r["mcp"])
        if 17 <= h <= 21:
            daily_pricing[d]["peak_prices"].append(r["mcp"])

    # Build curtailment lookup
    curt_lookup = {r["date"]: r for r in curtailment_records}

    # Get all unique dates
    all_dates = sorted(set(list(daily_pricing.keys()) + list(curt_lookup.keys())))

    row_num = 5
    for d in all_dates:
        dp = daily_pricing.get(d)
        cr = curt_lookup.get(d)

        try:
            dt = datetime.strptime(d, "%Y-%m-%d")
            dow = dt.strftime("%A")
        except ValueError:
            dow = ""

        sources = []
        if dp:
            sources.append("TSOC DAM")
        if cr:
            sources.append("Galascope")

        avg_mcp = sum(dp["prices"]) / len(dp["prices"]) if dp else None
        min_mcp = min(dp["prices"]) if dp else None
        max_mcp = max(dp["prices"]) if dp else None
        solar_mcp = (sum(dp["solar_prices"]) / len(dp["solar_prices"])) if dp and dp["solar_prices"] else None
        peak_mcp = (sum(dp["peak_prices"]) / len(dp["peak_prices"])) if dp and dp["peak_prices"] else None
        spread = (peak_mcp - solar_mcp) if peak_mcp is not None and solar_mcp is not None else None

        pv_yield = cr["pv_yield_kwh"] if cr else None
        curt_kwh = cr["curtailment_kwh"] if cr else None
        total = (pv_yield or 0) + (curt_kwh or 0)
        curt_pct = (curt_kwh / total) if total > 0 and curt_kwh else None

        ws3.cell(row=row_num, column=1, value=d)
        ws3.cell(row=row_num, column=2, value=dow)
        if avg_mcp is not None:
            ws3.cell(row=row_num, column=3, value=avg_mcp).number_format = NUM_2DP
        if min_mcp is not None:
            ws3.cell(row=row_num, column=4, value=min_mcp).number_format = NUM_2DP
        if max_mcp is not None:
            ws3.cell(row=row_num, column=5, value=max_mcp).number_format = NUM_2DP
        if solar_mcp is not None:
            ws3.cell(row=row_num, column=6, value=solar_mcp).number_format = NUM_2DP
        if peak_mcp is not None:
            ws3.cell(row=row_num, column=7, value=peak_mcp).number_format = NUM_2DP
        if spread is not None:
            ws3.cell(row=row_num, column=8, value=spread).number_format = NUM_2DP
        if pv_yield is not None:
            ws3.cell(row=row_num, column=9, value=pv_yield).number_format = NUM_2DP
        if curt_kwh is not None:
            ws3.cell(row=row_num, column=10, value=curt_kwh).number_format = NUM_2DP
        if curt_pct is not None:
            ws3.cell(row=row_num, column=11, value=curt_pct).number_format = "0.0%"
        ws3.cell(row=row_num, column=12, value=" + ".join(sources))

        # Color based on data availability
        if dp and cr:
            fill = LIGHT_YELLOW  # both sources
        elif dp:
            fill = LIGHT_BLUE  # pricing only
        elif cr:
            fill = LIGHT_GREEN  # curtailment only
        else:
            fill = None

        if fill and row_num % 2 == 0:
            for c in range(1, len(headers3) + 1):
                ws3.cell(row=row_num, column=c).fill = fill

        for c in range(1, len(headers3) + 1):
            ws3.cell(row=row_num, column=c).border = THIN_BORDER

        row_num += 1

    auto_width(ws3, len(headers3), min_width=12, max_width=30)
    ws3.freeze_panes = "A5"
    ws3.auto_filter.ref = f"A4:L{row_num - 1}"
    print(f"  Sheet 3: {row_num - 5} daily combined rows")

    # ── Sheet 4: Data Sources & Notes ──
    ws4 = wb.create_sheet("Data Sources & Notes")

    notes = [
        ("Data Sources & Methodology", None),
        ("", None),
        ("SOURCE 1: TSOC Day-Ahead Market (DAM) Reports", HEADER_FILL),
        ("URL", "https://tsoc.org.cy/competitive-electricity-market/mms-reports/day-ahead-market-daily-activity-reports-en/"),
        ("Provider", "Cyprus Transmission System Operator (TSOC)"),
        ("Market", "Cyprus Competitive Electricity Market (Open Market)"),
        ("Start Date", "October 1, 2025 (open market launch)"),
        ("Granularity", "48 half-hourly periods per day (00:00-00:30, 00:30-01:00, ..., 23:30-00:00)"),
        ("Key Field", "MCP (Market Clearing Price) in €/MWh - same for all energy types per period"),
        ("Energy Types", "OIL, SOLAR, ALL, - (purchase). Sheet 1 shows ALL type for unique pricing."),
        ("Files Parsed", f"{len([f for f in os.listdir(EXCEL_DIR) if f.endswith('.xlsx') and not f.startswith('._')])} daily reports"),
        ("Date Coverage", f"{dam_dates[0]} to {dam_dates[-1]}"),
        ("", None),
        ("SOURCE 2: Esperia Galascope 2.5MW Curtailment Reports", HEADER_FILL2),
        ("Provider", "Dino / Esperia Energy (Galascope PV monitoring portal)"),
        ("Plant", "Galascope 2.5MW PV installation (2,512.3 kWp string capacity)"),
        ("Granularity", "Daily totals"),
        ("Key Field", "Loss Due to Export Limitation (kWh) = energy curtailed by grid operator"),
        ("Date Coverage", "January 2025 - December 2025"),
        ("Files Parsed", f"{len([f for f in os.listdir(CURTAILMENT_DIR) if f.endswith('.xlsx') and not f.startswith('._')])} monthly reports"),
        ("", None),
        ("IMPORTANT NOTES", HEADER_FILL3),
        ("1. Pre-Oct 2025", "No hourly DAM pricing exists before Oct 2025 (market was not open). Curtailment data for Jan-Sep 2025 has no matching market price."),
        ("2. Overlap Period", "Oct-Dec 2025: Both pricing and curtailment data available. See 'Daily Summary' sheet."),
        ("3. Jan-Feb 2026", "DAM pricing available but no Galascope curtailment data (reports not yet received)."),
        ("4. Curtailment", "Export limitation loss represents energy the PV plant could have produced but was curtailed by the DSO due to grid constraints."),
        ("5. Curtailment %", "Calculated as: Curtailment Loss / (PV Yield + Curtailment Loss) * 100"),
        ("6. Arbitrage Spread", "Peak MCP (17:00-21:00) minus Solar Hours MCP (06:00-18:00). Positive = BESS revenue opportunity."),
        ("7. Missing Data", f"TSOC data ends at {dam_dates[-1]}. Run 'npx ts-node scripts/fetch-tsoc-playwright.ts --backfill' to fetch newer reports."),
    ]

    for i, (label, value) in enumerate(notes, 1):
        if value is None and label:
            ws4.cell(row=i, column=1, value=label).font = Font(bold=True, size=13)
            ws4.merge_cells(f"A{i}:B{i}")
        elif isinstance(value, PatternFill):
            ws4.cell(row=i, column=1, value=label).font = HEADER_FONT
            ws4.cell(row=i, column=1).fill = value
            ws4.cell(row=i, column=2).fill = value
            ws4.merge_cells(f"A{i}:B{i}")
        else:
            ws4.cell(row=i, column=1, value=label).font = Font(bold=True)
            ws4.cell(row=i, column=2, value=value)

    ws4.column_dimensions["A"].width = 25
    ws4.column_dimensions["B"].width = 100

    # ── Save ──
    wb.save(OUTPUT)
    print(f"\nSaved: {OUTPUT}")
    print(f"  File size: {os.path.getsize(OUTPUT) / (1024*1024):.1f} MB")


# ── Main ──

if __name__ == "__main__":
    dam_records, dam_all, dam_dates = parse_tsoc_files()
    curtailment_records = parse_curtailment_files()
    build_excel(dam_records, dam_all, dam_dates, curtailment_records)
