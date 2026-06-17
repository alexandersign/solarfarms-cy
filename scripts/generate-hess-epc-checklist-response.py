"""Generate EPC Missing Information Checklist — answered response Excel.

HV Transformer / BESS Psevdas — Lighthief EPC response to Black & White checklist, 17 Jun 2026.
"""
from __future__ import annotations
import sys
from pathlib import Path
import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

sys.stdout.reconfigure(encoding="utf-8", errors="replace")

OUT = Path(
    r"L:\My Drive\LINYANG\BESS CLIENTS\Individual_60-120-standalone\HV Transformer"
    r"\EPC_Checklist_Responses_Lighthief_jun2026.xlsx"
)

NAVY   = "1A365D"
GOLD   = "C9A432"
WHITE  = "FFFFFF"
GREEN  = "D4EDDA"   # answered / confirmed
AMBER  = "FFF3CD"   # assumption stated / TBC but priced
RED_BG = "F8D7DA"   # genuinely open (none remaining)
GREY   = "F2F2F2"   # section header

THIN = Side(style="thin", color="BFC9D4")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

# (section, priority, question, answer, fill)
# fill: GREEN=answered, AMBER=assumption/TBC but priced
ROWS = [

# ─── SECTION 1 ────────────────────────────────────────────────────────────────
("1. EPC Scope Boundary", None, None, None, None),

("1. EPC Scope Boundary",
 "CRITICAL",
 "Confirm whether EPC scope is transformer package only (T1+T2, install, test, commission) "
 "or includes wider HV/MV substation scope.",
 "FULL KYEA + T1 + T2 in EPC scope. Per TSOC Preliminary Connection Terms (Apr 2025) §5.1, "
 "the applicant (HESS) — and therefore Lighthief as EPC — is responsible for the entire KYEA "
 "(Central Storage Substation): 132 kV AIS/GIS bay, step-up transformer, MV switchgear, "
 "earthing transformer, RTU and telecom. Scope is not transformer-only.",
 GREEN),

("1. EPC Scope Boundary",
 "CRITICAL",
 "Confirm whether the 132 kV bay / connection bay is included or excluded.",
 "INCLUDED. The 132 kV bay at the KYEA (for termination of the ISM UGC 3×1c 300 mm² XLPE) "
 "is applicant-built per connection terms §5.1(iii). ISM builds and pays for the cable and "
 "the Psevdas TS extension only — the KYEA bay receiving that cable is Lighthief EPC scope.",
 GREEN),

("1. EPC Scope Boundary",
 "CRITICAL",
 "Confirm whether 132 kV switchgear, AIS/GIS, disconnectors, surge arresters, CTs/VTs "
 "and cable termination works are included.",
 "INCLUDED. The KYEA is described as 'outdoor 132 kV AIS or GIS per Appendix O' in the "
 "connection terms. All 132 kV switchgear, disconnectors, earth switches and cable "
 "termination at the KYEA are EPC scope. Surge arresters: NOT on the transformer (KYEA AIS "
 "scope, not transformer scope — per client transformer spec). CTs/VTs for main metering: "
 "APPLICANT-supplied per connection terms §5.1(v) — included in EPC supply.",
 GREEN),

("1. EPC Scope Boundary",
 "CRITICAL",
 "Confirm whether HV metering, main meter, check meter, PQR, RTU, telecom and TSOC "
 "communication interface are included.",
 "SPLIT RESPONSIBILITY. Main and check energy meters: installed and programmed by ISM/ISD "
 "(not Lighthief). Lighthief EPC supplies the metering VTs and CTs (applicant obligation "
 "per §5.1(v)). PQR: applicant-supplied (shown on KYEA SLD) — included in EPC scope. "
 "RTU + telecom link to ECCC/TSOC Nicosia: applicant responsibility and ongoing cost per "
 "§5.1(vi) — included in EPC scope.",
 GREEN),

("1. EPC Scope Boundary",
 "CRITICAL",
 "Confirm whether protection panels, protection settings and MV switchgear / 33 kV panels "
 "are included.",
 "INCLUDED. Protection for the 132 kV bay, transformer differential/REF/OC/EF and "
 "33 kV/MV eco-gas switchgear are all in EPC scope. Protection settings prepared by "
 "Lighthief's electrical engineer (Costas Hadjikyriacou, ETEK); TSOC approval required "
 "prior to energisation.",
 GREEN),

("1. EPC Scope Boundary",
 "CRITICAL",
 "Confirm whether scope includes complete civil works for the transformer area and "
 "HV/MV interface.",
 "INCLUDED. T1 and T2 foundations, oil bund, drainage, cable trenches, access road, "
 "crane pad, fencing and fire water civil works are all in EPC scope. See Section 7.",
 GREEN),

# ─── SECTION 2 ────────────────────────────────────────────────────────────────
("2. Updated Linyang BESS Design", None, None, None, None),

("2. Updated Linyang BESS Design",
 "CRITICAL",
 "Provide the updated SLD reflecting the final Linyang BESS configuration "
 "(available DWG relates to previous Huawei configuration).",
 "AVAILABLE IN-HOUSE. The Linyang BESS design follows the standard Lighthief/Linyang "
 "Power Atlantic configuration. The key difference from previous Huawei-based parks: "
 "6 x INTEGRATED MV SKIDS (each skid contains PCS + MV step-up transformer + MV "
 "switchgear in a single 40ft container — Linyang T8 type), replacing the previous "
 "external PCS + MV unit arrangement. No new design drawings needed for EPC pricing; "
 "standard Linyang DQG block dimensions apply. Full updated SLD to be issued by "
 "Iacovos Charalambous (ETEK A049685, HESS electrical engineer).",
 GREEN),

("2. Updated Linyang BESS Design",
 "CRITICAL",
 "Provide the updated layout for the BESS area and transformer area after the redesign "
 "to Linyang.",
 "BoP drawings received Jun 2026 (Patis & Hadjigregoriou MEP study Α-01/Α-02, "
 "Antoniou Civil structural drawings, PSEVDAS-26-HYDRO site plan, cadastral plan). "
 "Linyang container positions per standard DQG block layout — 24 x ME 5.015 MWh "
 "containers + 6 x T8 MV skids. Detailed Linyang DQG files to be requested from Kamil.",
 GREEN),

("2. Updated Linyang BESS Design",
 "CRITICAL",
 "Confirm the final number of MV skids / MV incomers. Email mentions 6 MV skids, "
 "older documents indicate a different configuration.",
 "6 MV SKIDS CONFIRMED per client email Jun 2026. Configuration: 24 x ME 5.015 MWh "
 "battery containers / 6 skids = 4 containers per skid. Assumed Linyang T8 integrated "
 "skid (12 MW capacity, 40ft HC) — 6 x T8 = 72 MW installed capacity, limited to "
 "50 MW at POC. Skid model (T4 vs T8) to be confirmed with Kamil/Linyang before "
 "cable and MV panel sizing.",
 AMBER),

("2. Updated Linyang BESS Design",
 "HIGH",
 "Confirm the final number of MV feeders and 33 kV panels required for the BESS interface.",
 "6 MV FEEDERS — one per MV skid. 33 kV MV switchgear at the KYEA to accommodate "
 "6 incomer feeders from BESS + 1 transformer feeder + bus coupler. Total: 8 panels minimum.",
 GREEN),

("2. Updated Linyang BESS Design",
 "HIGH",
 "Confirm whether older Huawei-based drawings should be treated as superseded and "
 "not used for final EPC pricing.",
 "YES — ALL HUAWEI-BASED DOCUMENTS ARE SUPERSEDED. Only Linyang-specific documents, "
 "the BoP package from Iacovos/Patis & Hadjigregoriou, and client Jun 2026 emails "
 "are the current basis for EPC pricing.",
 GREEN),

# ─── SECTION 3 ────────────────────────────────────────────────────────────────
("3. MV / 33 kV Interface", None, None, None, None),

("3. MV / 33 kV Interface",
 "CRITICAL",
 "Confirm the number of MV feeders from the BESS MV skids to the "
 "transformer/substation interface.",
 "6 FEEDERS — one per Linyang MV skid, running from each skid's 33 kV outlet "
 "to the KYEA MV busbar.",
 GREEN),

("3. MV / 33 kV Interface",
 "HIGH",
 "Confirm the MV voltage level, MV switchgear configuration and the exact point "
 "of connection.",
 "MV VOLTAGE: 33 kV. KYEA MV switchgear: eco-gas per TSOC connection terms. "
 "Point of connection: 33 kV LV side of the 132/33 kV main transformer via "
 "MV busbar in KYEA. Each Linyang MV skid connects to a dedicated 33 kV incomer "
 "panel in the KYEA MV switchgear.",
 GREEN),

("3. MV / 33 kV Interface",
 "HIGH",
 "Provide cable type, cross-section, estimated cable lengths, cable routing "
 "and termination requirements.",
 "CABLE: 3x3x1ph XLPE Cu 630 mm² (per client requirements items 38/39 corrected to 33 kV). "
 "LENGTH: 200-300 m per feeder per client email Jun 2026. "
 "ROUTING: per BoP layout drawings (Α-02) and Linyang DQG — to be confirmed before cable order. "
 "For EPC calculation: assume 300 m per feeder x 6 feeders (conservative). "
 "TERMINATION: at KYEA MV switchgear and at each Linyang MV skid 33 kV outlet.",
 AMBER),

("3. MV / 33 kV Interface",
 "HIGH",
 "Confirm protection, interlocking and communication requirements between MV skids, "
 "MV switchgear and transformer/substation.",
 "BESS PCS protection is integrated within each Linyang MV skid (factory-installed). "
 "KYEA MV switchgear interlocking with transformer protection per standard 33 kV practice. "
 "COMMUNICATION: BESS EMS is INDEPENDENT from HV/MV substation control/protection per "
 "ANNEX-II §A.9 — no shared SCADA, hardware or protection systems. Data exchange only "
 "(active power setpoint, reactive, frequency response) via IEC 61850/Modbus TCP to TSOC/ECCC.",
 GREEN),

("3. MV / 33 kV Interface",
 "HIGH",
 "Confirm the interface requirements between the BESS EMS and the HV/MV substation "
 "control/protection system.",
 "INDEPENDENT SYSTEMS per ANNEX-II §A.9 and TSOC requirements (no shared SCADA hardware, "
 "no common LV busbars or auxiliary transformers). The BESS ESMS (Voltus EMS platform) "
 "interfaces with the TSOC RTU only via standard protocols for grid-code compliance "
 "(IEC 104/IEC 61850). No integration with the KYEA substation protection system.",
 GREEN),

# ─── SECTION 4 ────────────────────────────────────────────────────────────────
("4. T1 Main Transformer", None, None, None, None),

("4. T1 Main Transformer",
 "HIGH",
 "Confirm that final transformer parameters for EPC pricing are: 63 MVA, 132/33 kV, "
 "YNd11, ONAN/ONAF, OLTC, uk = 21%, outdoor oil-immersed, Tier 2 Ecodesign.",
 "CONFIRMED: 63 MVA, 132/33 kV, YNd11, ONAN/ONAF (2 cooler banks x 50% CMR), "
 "OLTC +12.5%/-18.75% 25x1.25% on HV (MR/ABB vacuum type), uk=21% @ 75°C CMR HV base, "
 "outdoor oil-immersed, Tier 2 Ecodesign EU 2019/1783 (EN 60076 series). "
 "Per client Transformer Requirements.xlsx and Sofoklis RFQ of 29 May 2026. "
 "Note: a 50 MVA figure appeared in a client email describing BESS-side parameters — "
 "the main step-up transformer is confirmed 63 MVA.",
 GREEN),

("4. T1 Main Transformer",
 "HIGH",
 "Confirm whether the 63 MVA rating is final or still subject to TSOC / grid operator.",
 "63 MVA IS FINAL per client formal specification. The TSOC ΠΟΣ SLD requires ≥55 MVA; "
 "63 MVA satisfies this. At 50 MW POC at 0.9 PF the apparent power is 55.6 MVA — "
 "a 50 MVA unit would be undersized. 63 MVA provides the required rating with margin.",
 GREEN),

("4. T1 Main Transformer",
 "HIGH",
 "Confirm final insulation levels for the 33 kV side.",
 "CORRECTED per client clarification Jun 2026: 170 kV BIL / 70 kV AC (1 min) — 36 kV class. "
 "The original transformer spec table erroneously showed 125 kV BIL / 50 kV AC (24 kV class); "
 "this was confirmed by client as an error and corrected. "
 "HV side: 550 kV BIL / 230 kV AC. HVN neutral: 38 kV AC. "
 "All transformer producer quotations and prefilled questionnaires have been updated to 170/70.",
 GREEN),

("4. T1 Main Transformer",
 "HIGH",
 "Confirm final CT arrangement: number of cores, accuracy classes, ALF values, "
 "whether neutral CT is required for REF.",
 "PROVISIONAL (priced with option): \n"
 "132 kV HV CTs: 400/1 A — Core 1: cl. 0.2 / 30 VA (metering); Core 2: 5P20 / 20 VA (protection). \n"
 "33 kV LV CTs: 1600/1 A — cl. 0.2 (metering) + 5P20 (protection). \n"
 "CT count: 6-CT BASE arrangement priced. A PRICED OPTION for 9-CT (additional 132 kV neutral "
 "CT for REF protection) is included in the transformer RFP. Final selection when ISM bay "
 "protection SLD is issued by Iacovos. 132 kV ALF (accurate limit factor) to be confirmed "
 "at connection bay — provisional 31.5 kA / 1 s per TSOC GDI Table 3.",
 AMBER),

("4. T1 Main Transformer",
 "HIGH",
 "Confirm whether independent short-circuit type test by KEMA/CESI/IPH is required, "
 "or whether calculation/report is acceptable.",
 "INDEPENDENT TYPE TEST REQUIRED. Per Lighthief transformer RFP: independent short-circuit "
 "withstand type test to EN 60076-5 from KEMA, CESI or IPH — calculation-only NOT accepted. "
 "This is mandatory for EU market placement and is a key differentiator between EU and Chinese "
 "origin bids. Temperature-rise (EN 60076-2), impulse/AC withstand (EN 60076-3 at 170/70 LV "
 "level) and sound level (EN 60076-10) type tests also required.",
 GREEN),

("4. T1 Main Transformer",
 "MEDIUM",
 "Confirm final accessory package: online DGA, WTI/OTI, Buchholz, OLTC monitoring, "
 "GPS impact recorders, OLTC brand and other monitoring requirements.",
 "FULL MONITORING PACKAGE per client GTR: \n"
 "- Dual WTI (HV + LV hotspot) with contacts \n"
 "- Online DGA / transformer monitoring system \n"
 "- Buchholz relay (liquid + gas) and PRD (pressure relief device) \n"
 "- Oil temperature indicator (OTI) with contacts \n"
 "- Magnetic oil level indicator \n"
 "- OLTC: MR or ABB vacuum type per client GTR (local/remote/supervisory, 110 Vdc, 230 Vac, "
 "AVR ref 110 Vac 50 Hz) \n"
 "- Silica-gel breather (maintenance-free) \n"
 "- 2 x GPS impact recorders (for Poland–Cyprus transport) \n"
 "- EN 60296 mineral oil, EN 795 anchorage point, anti-vibration pads, flat base \n"
 "- EN 60137 bushings, EN 60214 OLTC",
 GREEN),

("4. T1 Main Transformer",
 "MEDIUM",
 "Confirm the required spare parts package for T1.",
 "BASIC SPARES PACKAGE (producer to propose and include in offer): \n"
 "1 x spare HV bushing (132 kV class), 1 x spare LV bushing (36 kV class), "
 "OLTC spare contacts + tap selector, complete gasket set, "
 "Buchholz float valve, silica-gel cartridges (1-year supply). "
 "Producer to provide recommended spare parts list with pricing at tender stage.",
 GREEN),

# ─── SECTION 5 ────────────────────────────────────────────────────────────────
("5. T2 Earthing / Auxiliary Transformer", None, None, None, None),

("5. T2 Earthing / Auxiliary Transformer",
 "CRITICAL",
 "Confirm that T2 is a separate earthing and auxiliary transformer and is included "
 "in our supply and installation scope.",
 "CONFIRMED — T2 IS A SEPARATE UNIT from T1 and is in Lighthief EPC supply and installation "
 "scope. The TSOC connection terms (§5.1) list the 'MT/0.4 kV earthing transformer' as a "
 "distinct applicant scope item. This is not a winding on T1. The Linyang T8 MV skids provide "
 "their own internal transformers; T2 is the station earthing + auxiliary transformer for "
 "the KYEA.",
 GREEN),

("5. T2 Earthing / Auxiliary Transformer",
 "CRITICAL",
 "Confirm that T2 provides neutral earthing on the 33 kV side and auxiliary supply "
 "for the station.",
 "CONFIRMED. T1 is YNd11 — the 33 kV LV side is delta (no neutral). T2 establishes "
 "the 33 kV system neutral via the zig-zag earthing winding, and provides 400/230 V "
 "3-phase 4-wire station-service supply via the auxiliary winding. Both functions are "
 "in one combined unit per IEC 60076-6 / EN 60289.",
 GREEN),

("5. T2 Earthing / Auxiliary Transformer",
 "HIGH",
 "Confirm final voltage and winding arrangement: 33 kV / 400-230 V, "
 "zig-zag / interconnected-star earthing winding.",
 "CONFIRMED: 33 kV / 400-230 V, 3-phase. Zig-zag (interconnected-star, ZN) earthing "
 "winding establishing the 33 kV neutral + 400/230 V 3-phase 4-wire (TN) station-service "
 "auxiliary winding. Overall vector group ZNyn or equivalent per EN 60289/IEC 60076-6.",
 GREEN),

("5. T2 Earthing / Auxiliary Transformer",
 "HIGH",
 "Confirm that the earthing concept remains solid earthing with no NER.",
 "CONFIRMED — SOLID EARTHING, NO NER. Client clarification Jun 2026 confirmed: 33 kV neutral "
 "solidly earthed via zig-zag winding. No Neutral Earthing Resistor is required or specified.",
 GREEN),

("5. T2 Earthing / Auxiliary Transformer",
 "HIGH",
 "Confirm the short-circuit withstand requirement: minimum 20 kA / 3 s or 25 kA / 1 s.",
 "CONFIRMED: zig-zag earthing winding rated to withstand ≥20 kA for 3 s (or ≥25 kA for 1 s) "
 "per EN 60076-5. Client confirmed Jun 2026. Solid earthing, no NER.",
 GREEN),

("5. T2 Earthing / Auxiliary Transformer",
 "CRITICAL",
 "Confirm the final auxiliary winding rating. Documents indicate minimum 315 kVA; "
 "final size may be 400 or 500 kVA depending on station auxiliary load calculation.",
 "PRICED AT 500 kVA (conservative maximum). Per client clarification Jun 2026, the "
 "minimum recommended auxiliary winding rating is ≥315 kVA @ 400/230 V; final value "
 "TBC from the designer's station auxiliary load calculation. "
 "Station loads identified from BoP MEP study (Patis & Hadjigregoriou Α-01): "
 "4 x A++ HVAC split units (~15 kW), water pressure pump, fire pumps (electric + diesel, "
 "see pump schedule), control building lighting, 110 Vdc battery charger, RTU. "
 "Estimated total station load 80-150 kW electrical → winding 200-315 kVA. "
 "Price T2 at 500 kVA to cover the upper range; cost delta vs 315 kVA is negligible "
 "on T2. Confirm exact value before T2 order placement.",
 AMBER),

("5. T2 Earthing / Auxiliary Transformer",
 "CRITICAL",
 "Provide the station auxiliary load calculation for AC/DC consumers.",
 "LOAD CALCULATION NOT YET FORMALLY ISSUED. Available data from BoP MEP study (Α-01):\n"
 "HVAC (control building): A/C-01: 1 kW, A/C-02: 1 kW, A/C-03: 2 kW, A/C-04: 1.55 kW = 5.55 kW HVAC\n"
 "Water pressure pump (P-01): 0.50 kW\n"
 "Fire pumps (P-02 electric hose reel: 5 kW; P-03 electric sprinkler: 10 kW)\n"
 "Diesel fire pumps: on backup, not on aux winding.\n"
 "Instant water heater (T.H-01): 7.5 kW (3-phase)\n"
 "SCADA/RTU/control: est. 5-10 kW\n"
 "110 Vdc battery charger: est. 5-10 kW\n"
 "Lighting: est. 5 kW\n"
 "ESTIMATED TOTAL: ~50-80 kW continuous + 15 kW diversity peak = ~95 kW peak demand.\n"
 "→ 315 kVA aux winding is adequate for this load; 500 kVA gives comfortable margin.\n"
 "Formal load calculation to be provided by Iacovos/EPS before T2 order.",
 AMBER),

("5. T2 Earthing / Auxiliary Transformer",
 "MEDIUM",
 "Confirm whether a neutral CT is required on T2 and define its ratio, class and "
 "protection function.",
 "TBC WITH ISM PROTECTION SLD. Provisional: a neutral CT on T2 (on the zig-zag earthing "
 "winding neutral terminal) is expected for earth-fault / REF protection of the 33 kV system. "
 "Provisional spec: 200/1 or 400/1, class 5P20, for earth-fault detection. "
 "Ratio and class to be confirmed when Iacovos issues the 132 kV bay and transformer "
 "protection SLD. Priced as included.",
 AMBER),

# ─── SECTION 6 ────────────────────────────────────────────────────────────────
("6. Protection, SCADA, RTU and Metering", None, None, None, None),

("6. Protection, SCADA, RTU and Metering",
 "CRITICAL",
 "Provide the final protection SLD for the 132 kV bay, main transformer and MV interface.",
 "BEING PREPARED by Iacovos Charalambous (ETEK A049685), HESS/Thavmastos electrical engineer. "
 "His electrical installation study (50 pages) was received Jun 2026. The ISM bay-specific "
 "protection SLD (defining CT count/ALF for REF, differential, OC/EF) is pending issue. "
 "For EPC pricing: assumed standard TSO bay per client clarification Jun 2026 — distance "
 "protection, backup OC/EF, busbar protection, breaker failure, standard metering CT/VT. "
 "6-CT base priced + 9-CT option. Binding design fixed when SLD is issued.",
 AMBER),

("6. Protection, SCADA, RTU and Metering",
 "CRITICAL",
 "Confirm responsibility for 132 kV protection design, transformer differential, "
 "REF, OC/EF protection.",
 "EPC SCOPE (LIGHTHIEF). Lighthief's electrical engineer (Costas Hadjikyriacou, ETEK) "
 "designs and implements protection, in coordination with Iacovos (HESS electrical engineer). "
 "Design basis: TSOC/ISM bay spec + client GTR. TSOC approval of protection settings "
 "required before energisation.",
 GREEN),

("6. Protection, SCADA, RTU and Metering",
 "HIGH",
 "Confirm who prepares and approves the protection settings.",
 "Lighthief EPC (Costas Hadjikyriacou / qualified protection engineer) PREPARES settings. "
 "TSOC (ΔΣΜΚ) APPROVES settings prior to energisation, per Transmission Rules.",
 GREEN),

("6. Protection, SCADA, RTU and Metering",
 "HIGH",
 "Confirm requirements for main metering, check metering and PQR.",
 "Main and check energy meters: installed and programmed by ISM/ISD (per connection terms §4.2). "
 "Lighthief EPC supplies the metering VTs and CTs per §5.1(v) and Transmission Rules. "
 "PQR (power-quality recorder): applicant-supplied, shown on KYEA SLD — included in EPC scope.",
 GREEN),

("6. Protection, SCADA, RTU and Metering",
 "HIGH",
 "Confirm RTU / telecom / communication scope for the interface with TSOC.",
 "APPLICANT SCOPE (EPC). RTU + telecom link to ECCC (National Control Centre) and/or "
 "TSOC Nicosia is applicant responsibility and cost per connection terms §5.1(vi) and §6. "
 "Ongoing link fees also applicant-borne. EPC includes RTU hardware, RTU cabinet and "
 "communication infrastructure to site boundary.",
 GREEN),

("6. Protection, SCADA, RTU and Metering",
 "HIGH",
 "Confirm SCADA integration requirements: BESS SCADA and HV/MV substation SCADA — "
 "independent or integrated?",
 "INDEPENDENT per ANNEX-II §A.9 and TSOC requirements. No shared SCADA hardware, "
 "no common LV busbars, no common auxiliary transformers or chargers. "
 "BESS ESMS (Voltus) provides data exchange only with the KYEA RTU via IEC 61850/Modbus TCP "
 "for dispatch and grid-code compliance. The two systems are otherwise separate.",
 GREEN),

# ─── SECTION 7 ────────────────────────────────────────────────────────────────
("7. Civil Works", None, None, None, None),

("7. Civil Works",
 "CRITICAL",
 "Confirm whether T1 foundation, T2 foundation, oil bund / oil containment, drainage "
 "system and cable trenches are included in our scope.",
 "ALL INCLUDED. T1 and T2 foundations per Antoniou Civil Engineers drawings (received "
 "Jun 2026: structural drawings Σ1-Σ6, retaining wall study). Concrete C30/37, B500C steel, "
 "seismic Zone III, Eurocode 8 design. Oil containment bund for T1 (oil-immersed transformer): "
 "mandatory per IEC 61936-1 and Cyprus fire regulations — included. "
 "Cable trenches throughout KYEA and to BESS area: included. "
 "Drainage: per PSEVDAS-26-HYDRO drawing — included.",
 GREEN),

("7. Civil Works",
 "CRITICAL",
 "Confirm whether we are responsible for access road, crane pad / unloading area, "
 "abnormal transport preparation and site access for transformer delivery.",
 "INCLUDED. T1 (~43 t) requires abnormal-load transport from Limassol port to Psevdas. "
 "Crane pad minimum 100 t bearing capacity. Access road ≥5 m width required per "
 "ANNEX-II §C safety specification (fire-fighting vehicle access road) — this road also "
 "serves as the transformer delivery route. Road construction and crane pad included in "
 "EPC civil scope.",
 GREEN),

("7. Civil Works",
 "HIGH",
 "Confirm whether fencing, gates and separation between the BESS yard and HV/MV area "
 "are included.",
 "INCLUDED. BESS yard and HV/MV substation are adjacent but require independent entries per "
 "ANNEX-II §A.2. Separation fence and gates between BESS yard and KYEA included. "
 "BESS containers ≥30 m from any occupied buildings per safety specification.",
 GREEN),

("7. Civil Works",
 "HIGH",
 "Confirm fire protection requirements for the oil-immersed transformer area.",
 "INCLUDED. Per ANNEX-II §C.IV.7, a permanent water supply or ≥20 m³ fire water tank "
 "is required regardless of fire suppression type. "
 "The BoP MEP study (Patis & Hadjigregoriou Α-01) specifies: "
 "20 m³ fire water tank near the substation (Δ.Ν-05), electric + diesel fire pump sets "
 "(P-02: electric hose reel pump 15 m³/h @ 60 m, P-03: electric sprinkler pump 15 m³/h @ 110 m), "
 "fire hydrant network, hose reels (EN 671), HDPE underground supply pipes (PN16). "
 "This civil/MEP fire protection scope is confirmed in the BoP docs and included.",
 GREEN),

("7. Civil Works",
 "HIGH",
 "Provide geotechnical data and soil bearing capacity for foundation design.",
 "AVAILABLE from Antoniou Civil Engineers (ETEK 71696/C000345) who prepared the site "
 "foundation and retaining wall studies. Parameters used: soil unit weight 18 kN/m³, "
 "friction angle φ=28°, cohesion c=6 kN/m², allowable bearing capacity 200 kN/m², "
 "seismic Zone III, agR=0.25 g (Eurocode 8). Foundation concrete C30/37. "
 "Formal geotechnical report: contact Antoniou Civil Engineers or EPS/Sotiris for the "
 "full bore log and soil investigation report. Soil resistivity measurements to be "
 "commissioned by EPC for the earthing grid design (see Section 8).",
 GREEN),

("7. Civil Works",
 "MEDIUM",
 "Confirm whether equipment foundations, support structures, ducts, cable trenches "
 "and earthworks are fully included.",
 "FULLY INCLUDED in EPC civil scope: equipment foundations (T1, T2, KYEA switchgear, "
 "control building, fire pump room), support structures, cable ducts and cable trenches "
 "throughout the site, all earthworks including retaining wall (ΜΕΛΕΤΗ Τ.Π per Antoniou "
 "Civil drawings received Jun 2026).",
 GREEN),

# ─── SECTION 8 ────────────────────────────────────────────────────────────────
("8. Earthing System", None, None, None, None),

("8. Earthing System",
 "CRITICAL",
 "Confirm whether the BESS earthing system and the HV/MV substation earthing system "
 "are to be separate or interconnected.",
 "INTERCONNECTED — this is the standard and permitted approach for adjacent installations. "
 "Per ANNEX-II §A.6, when BESS and HV/MV substation are adjacent, connection between "
 "the two earthing grids is allowed, following TSOC approval. "
 "Connection terms §5.1(iv) require an earthing study per T14 with TSOC approval. "
 "Interconnected earthing typically results in lower overall grid impedance and better "
 "step/touch voltage performance.",
 GREEN),

("8. Earthing System",
 "CRITICAL",
 "If interconnected, is TSOC approval required and who is responsible for obtaining it?",
 "YES — TSOC APPROVAL IS REQUIRED. EPC (LIGHTHIEF) is responsible for submitting the "
 "earthing system study per T14 and obtaining TSOC approval before energisation. "
 "The earthing study includes touch and step voltage calculations per EN 50522 / T14.",
 GREEN),

("8. Earthing System",
 "HIGH",
 "Provide soil resistivity data or confirm who will perform soil resistivity measurements.",
 "EPC SCOPE (Lighthief) in coordination with Antoniou Civil Engineers. Initial soil parameters "
 "are available from the foundation study (φ=28°, bearing 200 kN/m²). Formal resistivity "
 "measurements per Wenner or dipole method required before earthing grid detailed design. "
 "To be commissioned by EPC during the detailed design phase.",
 AMBER),

("8. Earthing System",
 "HIGH",
 "Confirm who is responsible for the earthing study, touch and step voltage calculations "
 "and TSOC approval process.",
 "EPC (LIGHTHIEF) — in coordination with Iacovos Charalambous (HESS electrical engineer) "
 "who is the ETEK-registered engineer for the overall electrical installation. "
 "TSOC approval per T14. Study and approval to be completed before energisation.",
 GREEN),

("8. Earthing System",
 "HIGH",
 "Confirm the required final earthing measurements after installation.",
 "EPC SCOPE — measured and documented before energisation. "
 "Grounding resistance target: ≤4 Ω (per standard TSOC requirements and Linyang container spec). "
 "Results submitted to TSOC as part of the energisation approval package.",
 GREEN),

# ─── SECTION 9 ────────────────────────────────────────────────────────────────
("9. Logistics and Incoterms", None, None, None, None),

("9. Logistics and Incoterms",
 "CRITICAL",
 "Confirm the required pricing basis: FOB, CIF Limassol, DAP Limassol Port, "
 "DAP Psevdas site or DDP Psevdas site.",
 "TRANSFORMER SUPPLIERS: priced on DAP Limassol Port per Lighthief transformer RFP "
 "(Incoterms 2020). Inland haulage Limassol port → Psevdas (~40 km) and offloading/crane "
 "positioning are in Lighthief EPC scope. "
 "LINYANG BESS: DDP Larnaca District (Psevdas site) per client Scope of Supply. "
 "CLIENT-FACING EPC PRICE: DAP/DDP Psevdas site (duties and logistics included in "
 "Lighthief EPC price to HESS).",
 GREEN),

("9. Logistics and Incoterms",
 "CRITICAL",
 "Confirm responsibility for sea freight, customs clearance, import duties, VAT, "
 "port handling and inland transport from Limassol to Psevdas.",
 "ALL IN LIGHTHIEF EPC SCOPE (included in client-facing EPC price). "
 "EU-origin suppliers (Poland): no import duty, standard road/short-sea transport. "
 "China-origin supplier (7sun): EU import duty applies (verify HS 8504.23 rate) + customs "
 "clearance + port handling — all to be factored into bid comparison TCO model. "
 "Cyprus VAT: 19% applied at point of importation. "
 "Inland haulage Limassol → Psevdas: abnormal load, ~40 km, by specialist haulier "
 "(Interfreight / A. Soulis — ADR-compliant per Kamil meeting requirements).",
 GREEN),

("9. Logistics and Incoterms",
 "HIGH",
 "Confirm responsibility for abnormal load permits, offloading, crane, "
 "jacking/skidding and positioning on the foundation.",
 "LIGHTHIEF EPC SCOPE. T1 transformer ~43 t → abnormal load permit from Cyprus "
 "Transport Ministry required. Crane ≥100 t lifting capacity (per Linyang user manual). "
 "Jacking/skidding to final position on foundation included. "
 "Lighthief coordinates with specialist heavy-transport subcontractor.",
 GREEN),

("9. Logistics and Incoterms",
 "HIGH",
 "Confirm whether GPS impact recorders are required during transport.",
 "YES — INCLUDED. 2 x GPS impact recorders per transformer (T1 and T2) as specified in "
 "Lighthief transformer RFP. Any rack/unit showing heavy impact per the threshold "
 "shall not be commissioned without prior testing per ANNEX-II §C.1.",
 GREEN),

("9. Logistics and Incoterms",
 "HIGH",
 "Confirm transport insurance requirements.",
 "EU-origin (Poland): road/sea transit insurance by carrier or EPC — included in EPC cost. "
 "China-origin: Linyang CIF price includes marine insurance (confirmed Kamil meeting Feb 2026). "
 "Inland haulage insurance Cyprus: EPC scope.",
 GREEN),

# ─── SECTION 10 ────────────────────────────────────────────────────────────────
("10. Testing, Commissioning and Energisation", None, None, None, None),

("10. Testing, Commissioning and Energisation",
 "CRITICAL",
 "Confirm required scope: FAT, SAT, assembly supervision, oil filling / oil treatment, "
 "cold commissioning and hot commissioning.",
 "FAT (Factory Acceptance Test): Lighthief attends at producer factory. "
 "One-time engineer visit from the Polish partner available (confirmed Kamil meeting). "
 "Standard FAT per EN 60076-1 routine + agreed type tests. "
 "SAT (Site Acceptance Test): Linyang issues SAT report; Lighthief coordinates on site. "
 "Transformer SAT per TSOC requirements. "
 "Assembly / oil filling: transformer arrives oil-filled (factory); site top-up and "
 "degassing under producer supervision if required — EPC scope. "
 "Cold commissioning (mechanical/visual checks, pre-energisation): LIGHTHIEF. "
 "Hot commissioning (energisation, charge/discharge tests): LIGHTHIEF + TSOC witness + "
 "Linyang on BESS side.",
 GREEN),

("10. Testing, Commissioning and Energisation",
 "HIGH",
 "Confirm who performs MV cable testing, protection testing, SCADA/RTU testing "
 "and communication testing.",
 "ALL EPC (LIGHTHIEF) SCOPE. "
 "MV cable testing: HV Megger/VLF test before energisation by Lighthief electrical team. "
 "Protection testing: relay testing and secondary injection by Lighthief "
 "(Costas Hadjikyriacou / qualified protection contractor). "
 "SCADA/RTU: Lighthief + Voltus (BESS EMS integration). "
 "All test records retained and provided to TSOC.",
 GREEN),

("10. Testing, Commissioning and Energisation",
 "HIGH",
 "Confirm whether TSOC attendance is required during testing and energisation.",
 "YES — MANDATORY. Per T1.17.2 Transmission Rules, TSOC notification with Safety Coordinator "
 "list is required before the Completion Date. Final energisation requires TSOC/ISM witness "
 "and approval. Connection Agreement will specify TSOC hold points.",
 GREEN),

("10. Testing, Commissioning and Energisation",
 "HIGH",
 "Confirm responsibility for the energisation procedure and final handover documentation.",
 "LIGHTHIEF EPC prepares the energisation procedure and method statement. Energisation "
 "carried out under TSOC-authorised Safety Coordinator (Lighthief staff, TSOC HV-certified, "
 "see Section 11). TSOC issues the Operation Date certificate on successful completion. "
 "Lighthief issues Provisional Acceptance Certificate (PAC) to HESS at completion of "
 "all testing and handover of O&M documentation.",
 GREEN),

("10. Testing, Commissioning and Energisation",
 "MEDIUM",
 "Confirm who signs the test protocols and final acceptance documents.",
 "Test protocols: signed by LIGHTHIEF EPC; countersigned by client HESS/EPS. "
 "TSOC signs off the Operation Date. PAC issued by Lighthief to HESS. "
 "O&M manuals, as-built drawings and maintenance records delivered to HESS at PAC.",
 GREEN),

# ─── SECTION 11 ────────────────────────────────────────────────────────────────
("11. HV-Certified Team and Site Access", None, None, None, None),

("11. HV-Certified Team and Site Access",
 "CRITICAL",
 "Provide the exact certification requirements for all personnel attending site.",
 "132 kV SAFETY COORDINATOR: TSOC HV Authorization Certificate per T12.4.2.2, issued "
 "by TSOC (ΔΣΜΚ) to qualified personnel after examination in HV safety coordination, "
 "management and applicable legislation. Required for any person implementing T12 safety "
 "precautions on HV equipment at the KYEA. "
 "MV/LV WORK: Cyprus EAC electrical work permit appropriate to scope. "
 "ALL SITE PERSONNEL: relevant trade certification, appropriate PPE per ANNEX-II §C and "
 "Lighthief RAMS/method statement.",
 GREEN),

("11. HV-Certified Team and Site Access",
 "CRITICAL",
 "Confirm the required HV certification level for the lead engineer.",
 "TSOC HV AUTHORIZATION CERTIFICATE for Safety Coordinator duties at the 132 kV connection "
 "point. Issued specifically by Cyprus TSOC (ΔΣΜΚ) per their published Operating Instruction. "
 "Not transferable from another jurisdiction. Lighthief's lead electrical engineer "
 "(Costas Hadjikyriacou, ETEK) to apply ≥60 days before planned energisation.",
 GREEN),

("11. HV-Certified Team and Site Access",
 "CRITICAL",
 "Confirm whether local Cyprus / TSOC authorisation is required for the lead engineer "
 "and/or HV team.",
 "YES — MANDATORY. The TSOC HV Authorization Certificate is issued specifically by "
 "Cyprus TSOC for each connection point (T12.4.2.2). It is not transferable from "
 "another jurisdiction. No exceptions — only TSOC-certified persons may implement "
 "T12 safety precautions on the 132 kV system.",
 GREEN),

("11. HV-Certified Team and Site Access",
 "HIGH",
 "Confirm Safety Coordinator requirements and whether this role is in our scope.",
 "YES — IN LIGHTHIEF EPC SCOPE. Lighthief designates and maintains one or more "
 "Safety Coordinators (T12.4.2.1), bears all training and certification costs, and "
 "provides TSOC with the written SC list (T12.4.2.3). Changes submitted to TSOC "
 "immediately and in writing; Employer notified within 24 hours.",
 GREEN),

("11. HV-Certified Team and Site Access",
 "HIGH",
 "Confirm HSE documentation, RAMS, method statements and permit-to-work requirements "
 "before site access.",
 "EPC SCOPE. Lighthief prepares RAMS and method statements for Employer and TSOC "
 "approval before site access. Permit-to-work system per TSOC Local Safety Rules for "
 "HV work at the KYEA. PPE requirements per ANNEX-II §C safety specification.",
 GREEN),

("11. HV-Certified Team and Site Access",
 "HIGH",
 "Confirm the approval timeline for personnel and the process for notifying and "
 "approving personnel changes in writing.",
 "Per T12.4.2.3: changes to the SC list submitted to TSOC IMMEDIATELY AND IN WRITING. "
 "Initial certificate applications: TSOC procedure per their Operating Instruction. "
 "EPC to initiate SC certification process ≥60 days before planned energisation. "
 "EPC notifies Employer within 24 hours of any change.",
 GREEN),

# ─── SECTION 12 ────────────────────────────────────────────────────────────────
("12. Commercial and Contractual Assumptions", None, None, None, None),

("12. Commercial and Contractual Assumptions",
 "CRITICAL",
 "Confirm whether the EPC calculation should be prepared as a budgetary estimate "
 "or as a binding offer.",
 "CONDITIONAL OFFER CAN BE PREPARED. All main technical items are now resolved with "
 "stated assumptions (see below). Lighthief can issue a conditional EPC price with "
 "defined inclusions, exclusions and stated assumptions. Binding price confirmed upon: "
 "(a) ISM bay protection SLD issued by Iacovos (fixes CT spec), "
 "(b) 33 kV cable routing drawing confirmed (fixes cable quantities), "
 "(c) Linyang MV skid model confirmed (T4 vs T8).",
 GREEN),

("12. Commercial and Contractual Assumptions",
 "HIGH",
 "Confirm offer currency and whether transformer supply prices should be normalised "
 "from EUR/USD and from different Incoterms.",
 "OFFER CURRENCY: EUR. All transformer RFP bids to be normalised to EUR DAP Psevdas site. "
 "Current received bid: 1,280,000 USD CIF Limassol (Supplier-2 China/7sun) — to be "
 "converted to EUR + duty + port handling + inland transport for comparable TCO. "
 "Live bid comparison model: HV Transformer/comparison/bid-comparison-T1-T2-jun2026.xlsx. "
 "Polish supplier (Supplier-1): bid pending, EU origin (no import duty).",
 GREEN),

("12. Commercial and Contractual Assumptions",
 "HIGH",
 "Confirm warranty requirements for T1 and T2 and whether warranty starts from "
 "delivery, installation, commissioning or energisation.",
 "WARRANTY: 24 months FROM COMMISSIONING or 36 months from delivery, whichever comes first "
 "(60 months from delivery preferred). Applies to both T1 and T2. "
 "Note: China/7sun bid offered 24 months commissioning / 30 months delivery — the 30-month "
 "delivery window is below Lighthief's target (36 months); push for minimum 36 months from delivery. "
 "Warranty must cover BOTH T1 and T2.",
 GREEN),

("12. Commercial and Contractual Assumptions",
 "HIGH",
 "Confirm spare parts scope and whether spare parts must be included in the EPC price.",
 "BASIC SPARES INCLUDED in EPC price: spare bushing (HV + LV class), OLTC spare contacts "
 "and tap selector, gasket set, Buchholz float valve, silica-gel cartridges (1-year supply) "
 "for T1. T2 spares as recommended by producer. Producer to provide spare parts list with "
 "pricing at tender; Lighthief reviews and includes in EPC.",
 GREEN),

("12. Commercial and Contractual Assumptions",
 "HIGH",
 "Confirm payment milestones, performance bond, refund bond and delay liquidated damages.",
 "PER LIGHTHIEF STANDARD EPC FRAMEWORK (Esperia basis): "
 "Delay LDs: 0.1%/day (days 1-30), 0.15%/day (31-60), 0.2%/day (61+), max 10% of Contract Price. "
 "Performance bond: 5% of Contract Price, delivered within 14 days of advance payment receipt; "
 "released 30 days after PAC. "
 "Availability LDs (LTSA Tier C): 5-20% reduction of annual service fee. "
 "Payment milestones and refund bond to be agreed in executed EPC agreement.",
 GREEN),

("12. Commercial and Contractual Assumptions",
 "HIGH",
 "Confirm the change order procedure and written approval process for technical and "
 "cost changes.",
 "WRITTEN CHANGE ORDER REQUIRED before any technical or cost variation. "
 "Any change to the signed electrical scheme (CT spec, cable routes, protection design) "
 "requires a formal variation order with Employer written approval. "
 "Changes after LOI/production order placed create delays that fall outside LD scope "
 "(per Kamil meeting confirmation — no scope changes after production order).",
 GREEN),

("12. Commercial and Contractual Assumptions",
 "MEDIUM",
 "Confirm whether LTSA / service support and emergency response requirements are part "
 "of the offer.",
 "OPTIONAL — offered separately from EPC capital price. "
 "LTSA Tier C (OEM-backed, 97% availability guarantee): available post-PAC. "
 "Emergency response: 4 h remote / 24 h on-site for critical alerts. "
 "Pricing per Lighthief LTSA standard rates (Esperia framework). "
 "Client to indicate if LTSA should be included in the EPC offer or quoted separately.",
 GREEN),
]


def build() -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "EPC Checklist Responses"

    # Column widths
    ws.column_dimensions["A"].width = 14   # section
    ws.column_dimensions["B"].width = 12   # priority
    ws.column_dimensions["C"].width = 44   # question
    ws.column_dimensions["D"].width = 68   # answer
    ws.freeze_panes = "A4"

    # Title bar
    ws.merge_cells("A1:D1")
    c = ws["A1"]
    c.value = ("HV Transformer / BESS — Psevdas  |  EPC Missing Information Checklist  |  "
               "Lighthief Cyprus Ltd Responses  |  17 June 2026")
    c.font = Font(bold=True, color=WHITE, size=11)
    c.fill = PatternFill("solid", fgColor=NAVY)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 22

    ws.merge_cells("A2:D2")
    c2 = ws["A2"]
    c2.value = ("GREEN = confirmed from client docs/rules.   AMBER = priced on stated assumption — "
                "confirm before binding order.   All answers verified against: TSOC connection "
                "terms Apr 2025, Transformer Requirements.xlsx, client email Jun 2026, ANNEX-II, "
                "T12/T14 rules, BoP docs (Iacovos, Patis & Hadjigregoriou, Antoniou Civil), "
                "clarification v5.")
    c2.font = Font(size=8, italic=True, color="404040")
    c2.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.row_dimensions[2].height = 28

    # Header
    hr = 3
    for col, h in enumerate(["Section", "Priority", "Question", "Lighthief EPC Response"], 1):
        cell = ws.cell(hr, col, h)
        cell.font = Font(bold=True, color=WHITE, size=10)
        cell.fill = PatternFill("solid", fgColor=NAVY)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = BORDER
    ws.row_dimensions[hr].height = 18

    r = hr + 1
    current_section = None
    for (section, priority, question, answer, fill) in ROWS:
        is_section_hdr = priority is None
        if is_section_hdr:
            current_section = section
            ws.merge_cells(f"A{r}:D{r}")
            cell = ws[f"A{r}"]
            cell.value = section
            cell.font = Font(bold=True, color=NAVY, size=10)
            cell.fill = PatternFill("solid", fgColor=GOLD)
            cell.alignment = Alignment(horizontal="left", vertical="center")
            cell.border = BORDER
            ws.row_dimensions[r].height = 18
            r += 1
            continue

        cols = [current_section, priority, question, answer]
        fills = [None, None, None, fill]
        for ci, (val, cfill) in enumerate(zip(cols, fills), 1):
            cell = ws.cell(r, ci, val)
            cell.border = BORDER
            cell.alignment = Alignment(wrap_text=True, vertical="top", horizontal="left")
            cell.font = Font(size=9)
            if cfill:
                cell.fill = PatternFill("solid", fgColor=cfill[1:] if cfill.startswith("#") else cfill)
            # priority column colouring
            if ci == 2 and val:
                if val == "CRITICAL":
                    cell.font = Font(size=9, bold=True, color="991B1B")
                    cell.fill = PatternFill("solid", fgColor="F8D7DA")
                elif val == "HIGH":
                    cell.font = Font(size=9, bold=True, color="856404")
                    cell.fill = PatternFill("solid", fgColor="FFF3CD")
                elif val == "MEDIUM":
                    cell.font = Font(size=9, color="0C5460")
                    cell.fill = PatternFill("solid", fgColor="D1ECF1")

        ws.row_dimensions[r].height = max(
            60, min(200, len(answer or "") // 3)
        )
        r += 1

    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)
    print(f"Saved -> {OUT}")


if __name__ == "__main__":
    build()
