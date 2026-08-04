"""Generate founder loan schedule Excel — Lighthief Cyprus Ltd.
Output: financial/statements/founder-loan-schedule-lighthief.xlsx
"""
from pathlib import Path
import openpyxl
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter

OUT = Path(__file__).parent / "founder-loan-schedule-lighthief.xlsx"

# ── Brand colours ──────────────────────────────────────────────────────────
NAVY   = "1A365D"
GOLD   = "C9A432"
WHITE  = "FFFFFF"
GREY   = "F0F4F8"
AMBER  = "FFF3CD"
GREEN  = "D4EDDA"
RED_BG = "F8D7DA"
LIGHT  = "EEF2F7"

def fill(hex_): return PatternFill("solid", fgColor=hex_)
def font(bold=False, color="000000", size=10, italic=False):
    return Font(bold=bold, color=color, size=size, italic=italic,
                name="Calibri")
def border_thin():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)
def border_medium_bottom():
    thin = Side(style="thin", color="CCCCCC")
    med  = Side(style="medium", color=NAVY)
    return Border(left=thin, right=thin, top=thin, bottom=med)

def header_row(ws, row, texts, widths=None):
    for col, txt in enumerate(texts, 1):
        c = ws.cell(row=row, column=col, value=txt)
        c.fill    = fill(NAVY)
        c.font    = font(bold=True, color=WHITE, size=9)
        c.alignment = Alignment(horizontal="center", vertical="center",
                                wrap_text=True)
        c.border  = border_thin()

def title_row(ws, row, text, ncols, bg=NAVY, fg=WHITE, size=11):
    ws.merge_cells(start_row=row, start_column=1,
                   end_row=row, end_column=ncols)
    c = ws.cell(row=row, column=1, value=text)
    c.fill = fill(bg)
    c.font = font(bold=True, color=fg, size=size)
    c.alignment = Alignment(horizontal="left", vertical="center",
                            indent=1)

def section_header(ws, row, text, ncols, bg=LIGHT):
    ws.merge_cells(start_row=row, start_column=1,
                   end_row=row, end_column=ncols)
    c = ws.cell(row=row, column=1, value=text)
    c.fill = fill(bg)
    c.font = font(bold=True, color=NAVY, size=10)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)

def data_row(ws, row, values, bg=WHITE, bold=False, italic=False,
             number_format=None, amber=False, green=False, red=False):
    if amber: bg = AMBER
    if green: bg = GREEN
    if red:   bg = RED_BG
    for col, val in enumerate(values, 1):
        c = ws.cell(row=row, column=col, value=val)
        c.fill   = fill(bg)
        c.font   = font(bold=bold, italic=italic, size=9)
        c.border = border_thin()
        c.alignment = Alignment(vertical="center", wrap_text=True,
                                horizontal="right" if col >= 3 else "left")
        if number_format and col >= 3 and isinstance(val, (int, float)):
            c.number_format = number_format

def total_row(ws, row, values, ncols, bg=NAVY, fg=WHITE):
    for col, val in enumerate(values, 1):
        c = ws.cell(row=row, column=col, value=val)
        c.fill   = fill(bg)
        c.font   = font(bold=True, color=fg, size=10)
        c.border = border_medium_bottom()
        c.alignment = Alignment(vertical="center",
                                horizontal="right" if col >= 3 else "left",
                                indent=1 if col == 1 else 0)
        if col >= 3 and isinstance(val, (int, float)):
            c.number_format = '#,##0.00'

def blank(ws, row, ncols):
    for col in range(1, ncols+1):
        ws.cell(row=row, column=col).fill = fill(WHITE)

# ── Build workbook ──────────────────────────────────────────────────────────
wb = openpyxl.Workbook()

# ════════════════════════════════════════════════════════════════════════════
# SHEET 1 — Full Schedule
# ════════════════════════════════════════════════════════════════════════════
ws = wb.active
ws.title = "Founder Loan Schedule"
ws.sheet_view.showGridLines = False
ws.freeze_panes = "A4"

NCOLS = 5
col_widths = [18, 46, 16, 16, 28]
for i, w in enumerate(col_widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w
ws.row_dimensions[1].height = 30
ws.row_dimensions[2].height = 16
ws.row_dimensions[3].height = 30

# ── Doc header ──────────────────────────────────────────────────────────────
title_row(ws, 1,
    "Lighthief Cyprus Ltd (HE 477423) — Founder / Director Loan & Funding Schedule",
    NCOLS, size=12)
title_row(ws, 2,
    "Prepared: 1 August 2026  |  Basis: BOC bank statement master + Arkadiusz Sybaris verbal statement  |  Status: MANAGEMENT — UNAUDITED",
    NCOLS, bg="2B5FA0", size=9)

# ── Column headers ───────────────────────────────────────────────────────────
header_row(ws, 3,
    ["Date", "Description / Narrative", "Amount (€)", "Running Total (€)", "Verification status"])

# ════════════════════════════════════════════════════════════════════════════
# SECTION A — BOC bank-labelled loans
# ════════════════════════════════════════════════════════════════════════════
r = 4
section_header(ws, r, "SECTION A — BOC Bank of Cyprus: Labelled Shareholder Loans (Arkadiusz Sybaris)", NCOLS)
r += 1

A_rows = [
    ("22 Aug 2025",  "Arkadiusz Sybaris — Shareholders loan",                       500.00),
    ("08 Sep 2025",  "Arkadiusz Sybaris — Shareholders loan to the company",        5_000.00),
    ("15 Jan 2026",  "Arkadiusz Sybaris — Shareholders loan to the company",        8_000.00),
    ("10 Feb 2026",  "Arkadiusz Sybaris — Shareholders loan to the company",        4_000.00),
    ("18 Feb 2026",  "Arkadiusz Sybaris — Shareholders loan to the company",       10_000.00),
    ("02 Jul 2026",  "Arkadiusz Sybaris — Shareholders loan (€45,000 same-day to Trikkis Energy for site prep)", 60_000.00),
    ("09 Jul 2026",  "Arkadiusz Sybaris — Shareholders loan to the company",        3_000.00),
]

running = 0.0
for date, desc, amt in A_rows:
    running += amt
    data_row(ws, r, [date, desc, amt, running, "✓ Confirmed — BOC bank statement"],
             number_format='#,##0.00', green=(amt == 60000))
    r += 1

total_row(ws, r, ["Section A total", "", 90_500.00, "", "Bank-confirmed gross"], NCOLS)
r += 1
blank(ws, r, NCOLS); r += 1

# ════════════════════════════════════════════════════════════════════════════
# SECTION B — Revolut drawdowns
# ════════════════════════════════════════════════════════════════════════════
section_header(ws, r, "SECTION B — Revolut Card Drawdowns (BOC → Revolut): Loan Returns", NCOLS, bg="FFE8E8")
r += 1

B_rows = [
    ("2025 (full year)", "Revolut card top-ups — BOC card purchase → Revolut account (Apr report: founder loan returns)", -5_203.57),
    ("2026 YTD",         "Revolut card top-ups — same treatment; accountant to confirm repayment vs opex",                -4_986.90),
]

running_b = 90_500.00
for date, desc, amt in B_rows:
    running_b += amt
    data_row(ws, r, [date, desc, amt, running_b,
                     "Provisional — accountant to classify as loan repayment or opex"],
             number_format='#,##0.00', amber=True)
    r += 1

total_row(ws, r, ["Net Section A − B", "", 80_309.53, "", "Net BOC-labelled balance"], NCOLS,
          bg="2B5FA0")
r += 1
blank(ws, r, NCOLS); r += 1

# ════════════════════════════════════════════════════════════════════════════
# SECTION C — Arek off-bank director-funded expenses
# ════════════════════════════════════════════════════════════════════════════
section_header(ws, r,
    "SECTION C — Arkadiusz Sybaris: Director-Funded Expenses Paid Outside LCY BOC Account  [VERIFY AREK]",
    NCOLS, bg=AMBER)
r += 1

note_row = r
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=NCOLS)
c = ws.cell(row=r, column=1,
    value="⚠  All figures in Section D are management estimates stated verbally by Arkadiusz Sybaris. "
          "Receipts, invoices and Arkadiusz's personal bank/card statements are required before these "
          "amounts can be included in statutory accounts or disclosed to investors as verified figures.")
c.fill = fill(AMBER)
c.font = font(bold=False, color="856404", size=9, italic=True)
c.alignment = Alignment(horizontal="left", vertical="center",
                        wrap_text=True, indent=1)
c.border = border_thin()
ws.row_dimensions[r].height = 32
r += 1

D_rows = [
    ("Mar 2025 – Aug 2026",
     "Office rent — 18 months × €7,000/month. Amount confirmed by Arkadiusz; landlord receipts to follow.",
     126_000.00,
     "Amount confirmed — landlord receipts pending"),
    ("2025 (pre-Jul)",
     "Initial staff salaries — funded directly by Arkadiusz before formal LCY payroll commenced. Headcount, months and exact amounts TBC.",
     None,
     "⚠ [VERIFY AREK] — payroll records required"),
    ("2025 – 2026",
     "Marketing and advertising — paid directly by Arkadiusz outside BOC card spend. Overlap with BOC Meta/Google card charges to be reconciled.",
     None,
     "⚠ [VERIFY AREK] — invoices required"),
    ("2025 – 2026",
     "Car wrapping — KIA and Mitsubishi L200 branding. Wrap supplier invoice required.",
     None,
     "⚠ [VERIFY AREK] — supplier invoice required"),
    ("2025 – 2026",
     "Two company vehicles: KIA (model TBC) and Mitsubishi L200. Purchase price or finance payments. Owned or financed TBC.",
     None,
     "⚠ [VERIFY AREK] — V5/logbook + purchase receipt or finance agreement"),
    ("2025 – 2026",
     "Company phones and internet — estimated €200–400/month × 18 months. Overlap with BOC utility card payments to reconcile.",
     None,
     "⚠ [VERIFY AREK] — provider invoices required"),
]

for date, desc, amt, status in D_rows:
    data_row(ws, r, [date, desc, amt if amt else "TBC", "", status],
             number_format='#,##0.00',
             green=(amt == 126_000),
             amber=(amt is None))
    r += 1

total_row(ws, r,
    ["Section C — Arek stated total (all Section C items)",
     "", "≈ 309,500", "",
     "⚠ Unverified — Arek verbal only. Office rent €126,000 confirmed in amount."],
    NCOLS, bg="856404")
r += 1
blank(ws, r, NCOLS); r += 1

# ════════════════════════════════════════════════════════════════════════════
# SECTION E — Grand Summary
# ════════════════════════════════════════════════════════════════════════════
section_header(ws, r, "SECTION D — Grand Summary", NCOLS, bg=LIGHT)
r += 1

summary_rows = [
    ("Section A — BOC gross loans",          90_500.00, "Bank-confirmed"),
    ("Section B — Revolut drawdowns (net)", -10_190.47, "Provisional"),
    ("Net BOC-labelled balance",             80_309.53, "Bank-confirmed net"),
    ("Section C — off-bank expenses",       "≈ 309,500", "[VERIFY AREK]"),
]

for label, amt, status in summary_rows:
    is_net = "Net BOC" in label
    data_row(ws, r, [label, "", amt, "", status],
             number_format='#,##0.00',
             bold=is_net,
             bg=LIGHT if not is_net else "DCE6F1")
    r += 1

# Grand total — Arek all-in
total_row(ws, r,
    ["ARKADIUSZ SYBARIS — TOTAL DIRECTOR FUNDING (ALL-IN STATED)",
     "", "≈ 400,000", "",
     "⚠ Includes Sections A + D. BOC loans are a subset, not additive. [VERIFY AREK]"],
    NCOLS)
r += 1
blank(ws, r, NCOLS); r += 1

# ── Footer note ──────────────────────────────────────────────────────────────
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=NCOLS)
c = ws.cell(row=r, column=1,
    value="Lighthief Cyprus Ltd · HE 477423 · solarfarms.cy · office@lighthief.com · "
          "Prepared by Alexander Papacosta, Cyprus Director · 1 August 2026 · MANAGEMENT — NOT AUDITED")
c.fill = fill(NAVY)
c.font = font(bold=False, color="C8D9EC", size=8)
c.alignment = Alignment(horizontal="left", vertical="center", indent=1)

# ════════════════════════════════════════════════════════════════════════════
# SHEET 2 — BOC Monthly Cash Summary (context)
# ════════════════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("Monthly Cash (BOC context)")
ws2.sheet_view.showGridLines = False

col_widths2 = [14, 16, 16, 16, 18, 16]
for i, w in enumerate(col_widths2, 1):
    ws2.column_dimensions[get_column_letter(i)].width = w

title_row(ws2, 1, "Monthly BOC Cash Summary — August 2025 to July 2026", 6, size=11)
title_row(ws2, 2, "Source: TransactionHistory_master_boc.csv (BOC a/c 357044102353)  |  1 Aug 2026", 6, bg="2B5FA0", size=9)
header_row(ws2, 3, ["Month", "Client IN (€)", "SH Loan IN (€)", "Total IN (€)", "Total OUT (€)", "Net (€)"])

monthly = [
    ("Aug 2025",   4_242.87,   500.00,  4_742.87,   1_462.68,   3_280.19),
    ("Sep 2025",  11_075.99, 5_000.00, 16_075.99,  11_749.75,   4_326.24),
    ("Oct 2025",  11_832.13,     0.00, 11_832.13,   8_611.08,   3_221.05),
    ("Nov 2025",  15_021.44,     0.00, 18_021.44,  22_519.72,  -4_498.28),
    ("Dec 2025",  18_356.19,     0.00, 20_826.19,  25_231.13,  -4_404.94),
    ("Jan 2026",   9_811.11, 8_000.00, 20_261.11,  12_777.46,   7_483.65),
    ("Feb 2026",  24_627.63,14_000.00, 38_627.63,  34_475.71,   4_151.92),
    ("Mar 2026",  17_400.16,     0.00, 18_420.16,  20_602.68,  -2_182.52),
    ("Apr 2026",  11_001.60,     0.00, 12_651.60,  23_926.41, -11_274.81),
    ("May 2026",  39_816.32,     0.00, 39_816.32,  22_562.96,  17_253.36),
    ("Jun 2026",  14_775.60,     0.00, 17_775.60,  34_398.87, -16_623.27),
    ("Jul 2026",  15_668.50,63_000.00, 78_668.50,  76_511.23,   2_157.27),
]

r2 = 4
for month, cli, sh, tin, tout, net in monthly:
    bg = WHITE if r2 % 2 == 0 else GREY
    red_month = net < 0
    for col, val in enumerate([month, cli, sh, tin, tout, net], 1):
        c = ws2.cell(row=r2, column=col, value=val)
        c.fill = fill(RED_BG if (red_month and col == 6) else bg)
        c.font = font(size=9, color="721c24" if (red_month and col == 6) else "000000")
        c.border = border_thin()
        c.alignment = Alignment(vertical="center",
                                horizontal="right" if col > 1 else "left")
        if col > 1:
            c.number_format = '#,##0.00'
    r2 += 1

# Totals
total_row(ws2, r2,
    ["TOTAL",
     sum(r[1] for r in monthly),
     sum(r[2] for r in monthly),
     sum(r[3] for r in monthly),
     sum(r[4] for r in monthly),
     sum(r[5] for r in monthly)],
    6)
r2 += 2

ws2.merge_cells(start_row=r2, start_column=1, end_row=r2, end_column=6)
c = ws2.cell(row=r2, column=1,
    value="Note: SH Loan Jul 2026 €63,000 = €60,000 (2 Jul) + €3,000 (9 Jul). "
          "€45,000 of the Jul loan was paid same-day to Trikkis Energy for site preparation.")
c.fill = fill(AMBER)
c.font = font(size=8, italic=True, color="856404")
c.alignment = Alignment(horizontal="left", wrap_text=True, indent=1)

# ── Save ─────────────────────────────────────────────────────────────────────
wb.save(OUT)
print(f"Saved: {OUT}")
